import json
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify
from sqlalchemy import text as _sa_text
from models import db, Department, ActionHistory, HazardReport, ASRReport, Hazard, Risk, Control, Action, Audit, Finding, Investigation, InvestigationEvent, MOC, MOCHazard, MOCMilestone, MOCUpdate, MOCStakeholder, SPIIndicator, SPIData, SPIEscalation, ChecklistTemplate, ChecklistTemplateItem, DistributionList, EmailLog, SurveyResponse, User, VoluntaryReport, ConfidentialReport, SafetyNewsletter, SafetyCampaign, SafetySurvey, LessonLearned, SafetyBulletin, Training, AuditPlan, AuditSchedule, AuditChecklist, AuditFinding, AuditAction, SafetyPolicy, SafetyRole, SafetyPersonnel, ERPlan, ERPDrill, ERPActivation, SMSDocument, DocumentLink, RiskOccurrence, RiskAction, RAChecklistItem, RiskAssessment, RARow, RAMitigation, RAReview, Employee, ApiToken, DeviceToken, SafetyPromoRead, SafetyPromoAck, AccountableExecutive, SRBMeeting, SRBAgendaItem, SRBAttendee, SRBDecision, RiskAcceptance, GovernanceAuditLog, ComplianceObligation
try:
    from models import SPIEventLink
except ImportError:
    SPIEventLink = None
try:
    from models import AuditVerificationItem
except ImportError:
    AuditVerificationItem = None
# Employee notification log
try:
    from models import EmployeeNotificationLog
except ImportError:
    EmployeeNotificationLog = None

# Phase 2 enforcement models
try:
    from models import (
        ReportFeedback, JustCulturePolicy, ConfidentialAccessLog,
        SafetyRecommendation, InvestigationTimeline, RegulatoryNotification,
        SoDViolationBlock, RAReviewCycle, LeadingIndicatorConfig,
    )
    _ENFORCEMENT_MODELS = True
except ImportError:
    _ENFORCEMENT_MODELS = False
    ReportFeedback = JustCulturePolicy = ConfidentialAccessLog = None
    SafetyRecommendation = InvestigationTimeline = RegulatoryNotification = None
    SoDViolationBlock = RAReviewCycle = LeadingIndicatorConfig = None
from datetime import datetime, date
import os, uuid, io, hashlib, functools
from werkzeug.security import generate_password_hash, check_password_hash as _wz_check
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import send_file, make_response, Response
from flask_wtf.csrf import CSRFProtect, generate_csrf
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address

# ── Enterprise PDF engine (ReportLab — no system deps) ───────────────────────
try:
    from reports import (
        pdf_hazard_report, pdf_asr_report, pdf_investigation,
        pdf_risk_assessment, pdf_action, pdf_moc, pdf_audit,
        pdf_erp, pdf_voluntary, pdf_confidential, pdf_training,
        pdf_audit_finding, pdf_spi_summary,
    )
    _PDF_ENGINE = True
except ImportError:
    _PDF_ENGINE = False

app = Flask(__name__)

@app.template_filter('fromjson')
def fromjson_filter(s):
    import json
    try: return json.loads(s or '{}')
    except: return {}

import smtplib, json as _json_mod
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
SMTP_HOST=os.environ.get('SMTP_HOST','')
SMTP_PORT=int(os.environ.get('SMTP_PORT',587))
SMTP_USER=os.environ.get('SMTP_USER','')
SMTP_PASSWORD=os.environ.get('SMTP_PASSWORD','')
SMTP_FROM=os.environ.get('SMTP_FROM','safety@jordanaviation.com')
SMTP_FROM_NAME=os.environ.get('SMTP_FROM_NAME','AviaS Safety')

def send_email(to_list, subject, html_body):
    if not SMTP_HOST or not SMTP_USER: return len(to_list), None
    sent=0; errs=[]
    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as srv:
            srv.starttls(); srv.login(SMTP_USER, SMTP_PASSWORD)
            for r in to_list:
                try:
                    msg=MIMEMultipart('alternative')
                    msg['Subject']=subject; msg['From']=SMTP_FROM_NAME+' <'+SMTP_FROM+'>'; msg['To']=r
                    msg.attach(MIMEText(html_body,'html'))
                    srv.sendmail(SMTP_FROM, r, msg.as_string()); sent+=1
                except Exception as ex: errs.append(str(ex)[:40])
    except Exception as ex: return 0, str(ex)
    return sent, '; '.join(errs) if errs else None

def get_recipients(dept_ids=None):
    q=DistributionList.query.filter_by(is_active=True)
    if dept_ids: q=q.filter(DistributionList.department_id.in_(dept_ids))
    return [r.email for r in q.all() if r.email]

def email_html(title, subtitle, body_html, ref='', dt=''):
    ref_line = ('<div style="color:#c9a84c;font-size:11px;margin-top:6px">Ref: '+ref+' · '+dt+'</div>') if ref else ''
    return ('<!DOCTYPE html><html><body style="background:#f0f2f8;font-family:Arial,sans-serif;padding:24px">'
            '<table width="580" style="background:#fff;border-radius:12px;overflow:hidden;margin:0 auto">'
            '<tr><td style="background:#0f1c3f;padding:18px 26px;border-bottom:3px solid #c9a84c">'
            '<span style="color:#fff;font-size:14px;font-weight:800">✈ AviaS</span></td></tr>'
            '<tr><td style="background:#0f1c3f;padding:16px 26px">'
            '<div style="color:rgba(255,255,255,.5);font-size:10px;text-transform:uppercase;letter-spacing:1px">'+subtitle+'</div>'
            '<div style="color:#fff;font-size:20px;font-weight:800;margin-top:4px">'+title+'</div>'
            +ref_line+'</td></tr>'
            '<tr><td style="padding:22px 26px;font-size:14px;color:#374151;line-height:1.7">'+body_html+'</td></tr>'
            '<tr><td style="background:#f8f9fc;padding:12px 26px;font-size:11px;color:#9ca3af">'
            'AviaS · Safety Management System · ICAO Annex 19</td></tr>'
            '</table></body></html>')

# Evidence file uploads
# Upload folder: use env var override for cloud (e.g. /tmp on Render free tier)
_default_upload = os.path.join(os.path.dirname(__file__), 'static', 'evidence')
UPLOAD_FOLDER = os.environ.get('UPLOAD_FOLDER', _default_upload)
ALLOWED_EXT   = {'pdf','docx','xlsx','png','jpg','jpeg','gif','mp4','mov'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.',1)[1].lower() in ALLOWED_EXT


def _save_upload(field_name, prefix=''):
    """Save an uploaded file, return filename or None."""
    if field_name not in request.files:
        return None
    f = request.files[field_name]
    if not f or not f.filename or not allowed_file(f.filename):
        return None
    from werkzeug.utils import secure_filename
    fname = f'{prefix}{secure_filename(f.filename)}'
    f.save(os.path.join(app.config['UPLOAD_FOLDER'], fname))
    return fname

# ─── Config ───────────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# Render PostgreSQL uses postgres:// but SQLAlchemy needs postgresql://
_db_url = os.environ.get('DATABASE_URL', f'sqlite:///{os.path.join(BASE_DIR, "sms.db")}')
if _db_url.startswith('postgres://'):
    _db_url = _db_url.replace('postgres://', 'postgresql://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = _db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
_db_opts = {'pool_pre_ping': True, 'pool_recycle': 280, 'pool_timeout': 20}
if 'postgresql' in _db_url or 'postgres' in _db_url:
    _db_opts['connect_args'] = {'connect_timeout': 10}
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = _db_opts
app.secret_key = os.environ.get('SECRET_KEY', 'jav-sms-dev-only-change-in-prod')
# IMPORTANT: Set a strong SECRET_KEY env var in production (Render dashboard)

# ── Session / cookie security ──────────────────────────────────────────────────
_is_prod = os.environ.get('FLASK_ENV', 'development') == 'production'
app.config['SESSION_COOKIE_HTTPONLY'] = True          # no JS access to cookie
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'        # CSRF mitigation
app.config['SESSION_COOKIE_SECURE']   = _is_prod      # HTTPS-only in prod
app.config['PERMANENT_SESSION_LIFETIME'] = 86400 * 7  # 7-day sessions

db.init_app(app)

# ── CSRF protection (flask-wtf) ───────────────────────────────────────────────
csrf = CSRFProtect(app)
app.config['WTF_CSRF_TIME_LIMIT'] = 3600   # 1-hour token lifetime

# ── Rate limiting (flask-limiter) ─────────────────────────────────────────────
limiter = Limiter(
    key_func=get_remote_address,
    app=app,
    default_limits=[],          # no blanket limit; apply per-route
    storage_uri='memory://',    # in-proc store — fine for 1-worker Gunicorn
)

# ── Governance Blueprint (ICAO Annex 19 §4 / Doc 9859 §§3-5) ─────────────────
from governance import gov as governance_bp
app.register_blueprint(governance_bp)

# ── Enforcement Blueprint (Phase 2 — SoD, Gates, Feedback, Timeline) ──────────
from sms_enforcement import enf as enforcement_bp
from sms_enforcement import (
    advance_report_feedback, enforce_sod, validate_action_closure,
    process_ineffective_action, initialize_investigation_timeline,
    create_regulatory_notification, log_confidential_access,
    get_or_create_feedback,
)
app.register_blueprint(enforcement_bp)

# ─── Global error handlers ────────────────────────────────────────────────────
@app.errorhandler(404)
def err_404(e):
    return render_template('error.html', code=404,
                           title='Page Not Found',
                           message='The page you requested does not exist.'), 404

@app.errorhandler(403)
def err_403(e):
    return render_template('error.html', code=403,
                           title='Access Denied',
                           message='You do not have permission to access this page.'), 403

@app.errorhandler(500)
def err_500(e):
    import traceback as _tb
    app.logger.error('500 error: %s\n%s', e, _tb.format_exc())
    return render_template('error.html', code=500,
                           title='Internal Server Error',
                           message='Something went wrong. The safety team has been notified.'), 500

# ─── Helpers ──────────────────────────────────────────────────────────────────
INTOLERABLE = {'5A','5B','5C','4A','4B','3A'}
TOLERABLE   = {'5D','5E','4C','4D','4E','3B','3C','3D','2A','2B','2C','1A'}

def get_tolerance(ri):
    if ri in INTOLERABLE: return 'INTOLERABLE'
    if ri in TOLERABLE:   return 'TOLERABLE'
    return 'ACCEPTABLE'

def new_id(prefix):
    """
    Generate standardised control number: MODULE/SMS/NN
    e.g.  HR/SMS/01  RA/SMS/03  AUD/SMS/02
    Falls back to UUID-suffix if module unknown.
    """
    MODULE_MAP = {
        'HR':   ('HR',  HazardReport),
        'ASR':  ('ASR', ASRReport),
        'HAZ':  ('HAZ', Hazard),
        'RSK':  ('RSK', Risk),
        'CTL':  ('CTL', Control),
        'ACT':  ('ACT', Action),
        'INV':  ('INV', Investigation),
        'MOC':  ('MOC', MOC),
        'AUD':  ('AUD', AuditSchedule),
        'PLAN': ('PLAN', AuditPlan),
        'FND':  ('FND', AuditFinding),
        'POL':  ('POL', SafetyPolicy),
        'ROLE': ('ROLE', SafetyRole),
        'PERS': ('PERS', SafetyPersonnel),
        'ERP':  ('ERP', ERPlan),
        'RA':   ('RA',  RiskAssessment),
        'RACT': ('RACT', RiskAction),
        'BUL':  ('BUL', SafetyBulletin),
        'DOC':  ('DOC', SMSDocument),
    }
    if prefix in MODULE_MAP:
        code, model = MODULE_MAP[prefix]
        # Use MAX existing sequence number, not COUNT, to survive deletions
        prefix_pattern = f'{code}-SMS-'
        try:
            existing_ids = db.session.execute(
                db.text(f"SELECT id FROM {model.__tablename__} WHERE id LIKE :pat"),
                {'pat': f'{prefix_pattern}%'}
            ).fetchall()
            max_seq = 0
            for (eid,) in existing_ids:
                try:
                    num = int(eid.replace(prefix_pattern, ''))
                    if num > max_seq:
                        max_seq = num
                except (ValueError, AttributeError):
                    pass
            seq = max_seq + 1
        except Exception:
            seq = model.query.count() + 1
        return f'{code}-SMS-{seq:02d}'
    # fallback
    short = str(uuid.uuid4())[:6].upper()
    return f'{prefix}-SMS-{short}'

_overdue_last_run: float = 0.0          # module-level timestamp; 0 = never run
_OVERDUE_COOLDOWN: int  = 300           # seconds between DB sweeps (5 minutes)

def check_overdue_actions():
    """Auto-mark actions as Overdue when due_date passes (only Open/In Progress).

    Uses a module-level cooldown so the DB sweep runs at most once every
    5 minutes regardless of how many page loads hit routes that call this.
    Safe for 1-worker Gunicorn (single process, no shared state issues).
    """
    global _overdue_last_run
    import time as _time
    now = _time.monotonic()
    if now - _overdue_last_run < _OVERDUE_COOLDOWN:
        return                           # still within cooldown window
    _overdue_last_run = now              # update before DB work to avoid stampede

    today = date.today().isoformat()
    try:
        actions = Action.query.filter(Action.status.in_(['Open','In Progress'])).all()
        changed = False
        for a in actions:
            if a.due_date and a.due_date < today:
                a.status = 'Overdue'
                changed = True
        if changed:
            db.session.commit()
    except Exception:
        db.session.rollback()

@app.context_processor
def inject_globals():
    # Always start from a clean session state so lazy loads in templates never
    # hit InFailedSqlTransaction from an earlier failed query in this request.
    try:
        depts = Department.query.all()
    except Exception:
        db.session.rollback()
        depts = []
    try:
        overdue = Action.query.filter_by(status='Overdue').count()
    except Exception:
        db.session.rollback()
        overdue = 0
    try:
        avi_open = AuditVerificationItem.query.filter(
            AuditVerificationItem.status.in_(['Pending', 'Scheduled', 'In Review', 'Ineffective', 'Escalated'])
        ).count()
    except Exception:
        db.session.rollback()
        avi_open = 0
    now = datetime.utcnow()
    return dict(all_departments=depts, now=now, today=date.today(),
                get_tolerance=get_tolerance,
                nav_overdue=overdue, nav_avi_open=avi_open, enumerate=enumerate)


@app.after_request
def add_cors(response):
    response.headers['Access-Control-Allow-Origin']  = '*'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization, X-Dept-Id'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS'
    return response


# ── Auth helpers ──────────────────────────────────────────────────────────────

def hash_pw(pw):
    """Hash a password using pbkdf2:sha256 (salted, compatible with all werkzeug versions).
    Explicitly forces pbkdf2 so we never get werkzeug 3.x scrypt hashes, which are
    CPU-intensive and can cause timeouts on Render free tier.
    """
    return generate_password_hash(pw, method='pbkdf2:sha256')

def _is_legacy_hash(h):
    """Return True if h is a raw SHA-256 hex digest (64 hex chars, no colon prefix)."""
    return h and len(h) == 64 and ':' not in h

def check_pw(pw, hashed):
    """Verify password against stored hash.

    Supports both legacy SHA-256 (64-char hex) and modern werkzeug pbkdf2.
    If a legacy hash matches, the caller should upgrade it via hash_pw().
    """
    if _is_legacy_hash(hashed):
        return hashlib.sha256(pw.encode()).hexdigest() == hashed
    return _wz_check(hashed, pw)

def is_logged_in():
    return session.get('admin_logged_in') is True

def log_action_history(action_id, changed_by, from_status, to_status, notes='', field='status'):
    """Write audit trail entry — never crashes the main operation."""
    try:
        db.session.add(ActionHistory(
            action_id=action_id, changed_by=changed_by,
            from_status=from_status, to_status=to_status,
            notes=notes or '', field_changed=field
        ))
    except Exception:
        pass


def require_login(f):
    """Decorator — redirects to login if not authenticated."""
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        if not is_logged_in():
            return redirect(url_for('admin_login', next=request.path))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    """Decorator — requires login AND admin role."""
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        if not is_logged_in():
            return redirect(url_for('admin_login', next=request.path))
        user = User.query.filter_by(username=session.get('admin_user'), is_active=True).first()
        if not user or user.role != 'admin':
            return render_template('error.html', code=403, title='Access Denied',
                                   message='This page requires administrator access.'), 403
        return f(*args, **kwargs)
    return decorated

def seed_admin():
    """Create default admin account if no users exist."""
    if User.query.count() == 0:
        admin = User(
            username     = 'admin',
            password_hash= hash_pw('Jordan@SMS2026'),
            full_name    = 'Safety Manager',
            role         = 'admin',
            is_active    = True,
        )
        db.session.add(admin)
        db.session.commit()
        print("✅ Default admin created — user: admin / pass: Jordan@SMS2026")

# ─── SEED DATABASE ────────────────────────────────────────────────────────────
def seed():
    """Database initialisation — creates tables only. No demo data."""
    db.create_all()
    # Seed departments only if they do not exist yet
    if not Department.query.first():
        depts = [
            Department(id=1, code='FO', name='Flight Operations'),
            Department(id=2, code='ME', name='Maintenance & Engineering'),
            Department(id=3, code='GO', name='Ground Operations'),
            Department(id=4, code='CC', name='Cabin Crew'),
            Department(id=5, code='SD', name='Safety Department'),
        ]
        for d in depts:
            db.session.add(d)
        db.session.commit()
        print('✅ Departments seeded.')
    seed_admin()
    print('✅ Database ready — no demo data loaded.')


# ═══════════════════════════════════════════════════════════════════════════════
#  PUBLIC REPORTING PORTAL  — no login required
# ═══════════════════════════════════════════════════════════════════════════════

@app.route('/')
def index():
    """Root: if logged in → admin dashboard, else → public reporting portal."""
    if is_logged_in():
        return redirect(url_for('dashboard'))
    return redirect(url_for('public_portal'))


@app.route('/portal')
def public_portal():
    """Public-facing reporting portal — passes live safety content to template."""
    from datetime import date as _date
    bulletins   = SafetyBulletin.query.filter_by(status='Active').order_by(
                      SafetyBulletin.created_at.desc()).limit(4).all()
    newsletters = SafetyNewsletter.query.filter_by(status='Published').order_by(
                      SafetyNewsletter.created_at.desc()).limit(3).all()
    campaigns   = SafetyCampaign.query.filter_by(status='Active').order_by(
                      SafetyCampaign.created_at.desc()).limit(4).all()
    surveys     = SafetySurvey.query.filter_by(status='Active').order_by(
                      SafetySurvey.created_at.desc()).limit(4).all()
    lessons     = LessonLearned.query.order_by(
                      LessonLearned.created_at.desc()).limit(4).all()
    safety_messages = [
        "Safety is everyone's responsibility — report hazards before they become incidents.",
        "Every report you make protects your colleagues and improves our operations.",
        "Speaking up for safety is an act of professionalism, not weakness.",
        "Proactive hazard identification is the foundation of aviation safety.",
        "Non-punitive reporting: your safety observations are valued and protected.",
        "One report can prevent an accident. Your voice matters.",
    ]
    import hashlib
    week_num = int(hashlib.md5(_date.today().strftime('%Y-W%V').encode()).hexdigest(), 16)
    safety_message = safety_messages[week_num % len(safety_messages)]
    return render_template('portal/portal.html',
                           bulletins=bulletins, newsletters=newsletters,
                           campaigns=campaigns, surveys=surveys,
                           lessons=lessons, safety_message=safety_message)


# ── WORKFLOW STATUS ENGINE ─────────────────────────────────────────────────────

def resolve_report_status(hazard_id=None, hr_status=None):
    """
    Enterprise Workflow Intelligence Engine.
    Dynamically calculates report lifecycle state from real workflow data.
    Returns: (status, color, stage, timeline, guidance, responsible, next_step)
    """
    STATUS_MAP = {
        'Submitted':                  ('#3b82f6', 1),
        'Under Review':               ('#f59e0b', 2),
        'Assigned':                   ('#8b5cf6', 3),
        'Investigation Open':         ('#eab308', 4),
        'Risk Assessment Open':       ('#06b6d4', 4),
        'Mitigation In Progress':     ('#06b6d4', 5),
        'Pending SAG':                ('#f97316', 5),
        'Awaiting Closure':           ('#6b7280', 6),
        'Awaiting Safety Approval':   ('#f97316', 6),
        'Closed':                     ('#22c55e', 7),
        'Rejected':                   ('#ef4444', 0),
    }

    GUIDANCE_MAP = {
        'Submitted':
            'Your report has been received. The Safety Department will begin review within 48 hours.',
        'Under Review':
            'Safety Department is reviewing your report. A workflow action will be assigned shortly.',
        'Assigned':
            'A corrective action has been assigned to the responsible department. Work is in progress.',
        'Investigation Open':
            'A formal investigation has been opened. Awaiting root cause analysis and mitigation plan.',
        'Risk Assessment Open':
            'A risk assessment is being conducted to evaluate the safety impact of this occurrence.',
        'Pending SAG':
            'Awaiting corrective action implementation by the assigned Safety Action Group (SAG).',
        'Mitigation In Progress':
            'Corrective action plan has been submitted. Implementation and verification is ongoing.',
        'Awaiting Closure':
            'All corrective actions are complete. Safety Manager must verify effectiveness before closure.',
        'Awaiting Safety Approval':
            'SAG has completed all corrective actions. Awaiting Safety Manager final review and closure approval.',
        'Closed':
            'This occurrence has been fully resolved and closed. Thank you for your safety report.',
        'Rejected':
            'This report has been reviewed and classified as not requiring further action.',
    }

    status      = hr_status or 'Submitted'
    timeline    = []
    responsible = 'Safety Department'
    next_step   = 'Awaiting safety review'

    if hazard_id:
        try:
            actions       = Action.query.filter_by(hazard_id=hazard_id).order_by(Action.created_at).all()
            investigations= Investigation.query.filter_by(hazard_id=hazard_id).all()
            risks         = Risk.query.filter_by(hazard_id=hazard_id).all()
            haz           = Hazard.query.filter_by(id=hazard_id).first()

            # ── Determine status (highest priority wins) ──────────────────────
            if haz and haz.status == 'Closed':
                status      = 'Closed'
                responsible = 'Safety Manager'
                next_step   = 'Occurrence fully closed'

            elif haz and haz.status == 'Awaiting Safety Approval':
                status      = 'Awaiting Safety Approval'
                responsible = 'Safety Manager'
                next_step   = 'Safety Manager must review evidence and approve final closure'

            elif investigations and any(i.status not in ('Closed','Completed') for i in investigations):
                status      = 'Investigation Open'
                responsible = 'Investigation Team'
                next_step   = 'Complete root cause analysis'

            elif actions:
                open_acts   = [a for a in actions if a.status != 'Closed']
                sag_acts    = [a for a in actions if a.sag_member and a.status != 'Closed']
                cap_acts    = [a for a in actions if a.status in ('CAP Submitted','Root Cause Submitted')]

                if all(a.status == 'Closed' for a in actions):
                    # Actions closed but safety approval still needed
                    status      = 'Awaiting Safety Approval'
                    responsible = 'Safety Manager'
                    next_step   = 'Safety Manager must verify effectiveness and approve closure'

                elif cap_acts:
                    status      = 'Mitigation In Progress'
                    responsible = (cap_acts[0].owner or 'Department Manager')
                    next_step   = 'Verify corrective action implementation'

                elif sag_acts:
                    status      = 'Pending SAG'
                    responsible = (sag_acts[0].sag_member or 'SAG Team')
                    next_step   = 'Implement corrective action plan'

                elif open_acts:
                    status      = 'Assigned'
                    responsible = (open_acts[0].owner or 'Department Manager')
                    next_step   = f'Complete assigned action: {(open_acts[0].description or "")[:50]}'

            elif risks and any(r.initial_risk_index for r in risks):
                status      = 'Risk Assessment Open'
                responsible = 'Safety Officer'
                next_step   = 'Complete risk assessment and assign controls'

            elif haz and haz.status == 'Under Assessment':
                status      = 'Under Review'
                responsible = 'Safety Department'
                next_step   = 'Safety review in progress — action will be assigned'

            else:
                status      = 'Under Review'
                responsible = 'Safety Department'
                next_step   = 'Awaiting safety review and workflow assignment'

            # ── Build rich timeline ───────────────────────────────────────────
            if haz:
                dt = haz.created_at.strftime('%d %b %Y') if haz.created_at else 'On submission'
                timeline.append({
                    'date':  dt,
                    'event': 'Report received — Occurrence record created in Safety Registry',
                    'icon':  '📋', 'type': 'submitted',
                })
                if haz.status in ('Under Assessment', 'Open', 'Closed'):
                    timeline.append({
                        'date':  dt,
                        'event': 'Safety Department review initiated',
                        'icon':  '🔍', 'type': 'review',
                    })

            # Bulk-load all ActionHistory for this hazard's actions in ONE query
            # (replaces N per-action queries with a single bulk fetch + in-memory grouping)
            action_ids = [a.id for a in actions]
            if action_ids:
                all_histories = ActionHistory.query.filter(
                    ActionHistory.action_id.in_(action_ids)
                ).order_by(ActionHistory.changed_at).all()
                history_by_action = {}
                for h in all_histories:
                    history_by_action.setdefault(h.action_id, []).append(h)
            else:
                history_by_action = {}

            for a in actions:
                if a.created_at:
                    dept = a.owner or 'Department'
                    timeline.append({
                        'date':  a.created_at.strftime('%d %b %Y'),
                        'event': f'Action assigned to {dept}: {(a.description or "")[:50]}',
                        'icon':  '⚡', 'type': 'action',
                    })
                if a.sag_member and a.created_at:
                    timeline.append({
                        'date':  a.created_at.strftime('%d %b %Y'),
                        'event': f'SAG assigned: {a.sag_member}',
                        'icon':  '🛡', 'type': 'sag',
                    })
                for h in history_by_action.get(a.id, []):
                    if h.changed_at and h.to_status:
                        icon = '✅' if h.to_status == 'Closed' else '🔄'
                        timeline.append({
                            'date':  h.changed_at.strftime('%d %b %Y'),
                            'event': f'Action status → {h.to_status}',
                            'icon':  icon, 'type': 'update',
                        })

            for inv in investigations:
                if inv.created_at:
                    timeline.append({
                        'date':  inv.created_at.strftime('%d %b %Y'),
                        'event': f'Investigation opened: {(inv.title or "Formal Investigation")[:50]}',
                        'icon':  '🔬', 'type': 'investigation',
                    })

            if haz and haz.status == 'Closed':
                timeline.append({
                    'date':  '—',
                    'event': 'Occurrence closed — corrective action verified effective',
                    'icon':  '✅', 'type': 'closed',
                })

            # Sort timeline by date (approximate)
            timeline = sorted(timeline, key=lambda x: x.get('date',''))

        except Exception:
            pass

    color, stage = STATUS_MAP.get(status, ('#6b7280', 1))
    guidance = GUIDANCE_MAP.get(status,
        'Your report is being processed by the Safety Management System.')

    return status, color, stage, timeline, guidance, responsible, next_step


def get_report_timeline(hazard_id, created_at, report_type='Hazard Report'):
    """Build chronological timeline for a report."""
    events = []
    if created_at:
        events.append({
            'date': created_at.strftime('%d %b %Y %H:%M'),
            'event': f'{report_type} submitted',
            'type': 'submitted', 'icon': '📝'
        })
    if hazard_id:
        try:
            haz = Hazard.query.get(hazard_id)
            if haz and haz.status == 'Under Assessment':
                events.append({'date': '—', 'event': 'Safety review started',
                    'type': 'review', 'icon': '🔍'})
            actions = Action.query.filter_by(hazard_id=hazard_id).order_by(Action.created_at).all()
            # Bulk-load histories in one query instead of N per-action queries
            _aids = [a.id for a in actions]
            _hist_map = {}
            if _aids:
                for h in ActionHistory.query.filter(
                        ActionHistory.action_id.in_(_aids)
                ).order_by(ActionHistory.changed_at).all():
                    _hist_map.setdefault(h.action_id, []).append(h)
            for a in actions:
                if a.created_at:
                    events.append({
                        'date': a.created_at.strftime('%d %b %Y'),
                        'event': f'Action assigned: {(a.description or "")[:60]}',
                        'type': 'action', 'icon': '⚡'
                    })
                for h in _hist_map.get(a.id, []):
                    if h.changed_at and h.to_status:
                        events.append({
                            'date': h.changed_at.strftime('%d %b %Y'),
                            'event': f'Action updated to: {h.to_status}',
                            'type': 'update', 'icon': '🔄'
                        })
            if haz and haz.status == 'Closed':
                events.append({'date': '—', 'event': 'Occurrence closed',
                    'type': 'closed', 'icon': '✅'})
        except Exception:
            pass
    return events


def sync_report_status(hazard_id):
    """
    Call this after ANY workflow event to persist the calculated status
    back to all linked HazardReport records. This ensures DB status
    matches real workflow state for web lists and dashboards.
    """
    if not hazard_id:
        return
    try:
        st, col, stage, tl, guid, resp, nxt = resolve_report_status(
            hazard_id=hazard_id, hr_status='Submitted')
        HazardReport.query.filter_by(hazard_id=hazard_id).update(
            {'status': st}, synchronize_session=False)
        db.session.flush()
    except Exception:
        pass


# ── OCCURRENCE REGISTRY HELPER ────────────────────────────────────────────────
def _register_occurrence(report_id, report_type, description, location='',
                          date_str='', department_id=None,
                          consequences='', reporter='Anonymous',
                          is_confidential=False, classification='Operational'):
    """
    Creates linked HazardReport + Hazard records so any report type
    appears in the central Safety Occurrence Registry (/hazard-reports).
    Wrapped in try/except — never blocks the main submission.
    """
    from datetime import date as _d
    try:
        hid = new_id('HAZ')
        hr_id = new_id('HR')
        disp_desc = '*** CONFIDENTIAL — Description restricted ***' if is_confidential else description

        # Hazard record (occurrence registry entry)
        haz = Hazard(
            id                  = hid,
            source              = report_type,
            linked_report_id    = report_id,
            department_id       = department_id,
            classification      = classification,
            type_of_activity    = report_type,
            generic_hazard      = f'{report_type}: {(disp_desc or "")[:80]}',
            specific_components = disp_desc,
            consequences        = consequences or 'Under Review',
            status              = 'Open',
        )
        db.session.add(haz)
        db.session.flush()

        # HazardReport record (shows in /hazard-reports list)
        hr = HazardReport(
            id                = hr_id,
            hazard_id         = hid,
            department_id     = department_id,
            date              = date_str or _d.today().isoformat(),
            location          = location,
            description       = disp_desc,
            classification    = classification,
            generic_hazard    = f'{report_type}: {(disp_desc or "")[:60]}',
            consequences      = consequences or 'Under Review',
            immediate_action  = '',
            reporter          = '*** CONFIDENTIAL ***' if is_confidential else reporter,
            reporter_severity = 'Medium',
            report_type       = report_type,
            status            = 'Submitted',
        )
        db.session.add(hr)
        db.session.flush()
        return hid, hr_id
    except Exception:
        return None, None


@app.route('/portal/voluntary', methods=['GET', 'POST'])
def portal_voluntary():
    """Voluntary Safety Report — open to all staff."""
    if request.method == 'POST':
        f = request.form
        from datetime import date as _date
        rnum = f'VR-{_date.today().strftime("%Y%m%d")}-{VoluntaryReport.query.count()+1:03d}'
        rpt = VoluntaryReport(
            ref_number    = rnum,
            reporter_name = f.get('reporter_name', '').strip() or 'Anonymous',
            position      = f.get('position', ''),
            department_id = int(f['department_id']) if f.get('department_id') else None,
            date          = f.get('date', _date.today().isoformat()),
            location      = f.get('location', ''),
            report_type   = f.get('report_type', 'Safety Concern'),
            description   = f.get('description', ''),
            consequences  = f.get('consequences', ''),
            suggestion    = f.get('suggestion', ''),
            status        = 'Submitted',
            is_confidential = False,
        )
        db.session.add(rpt)
        db.session.flush()
        # Register in Safety Occurrence Registry
        _register_occurrence(
            report_id     = rnum,
            report_type   = 'Voluntary',
            description   = f.get('description', ''),
            location      = f.get('location', ''),
            date_str      = f.get('date', ''),
            department_id = int(f['department_id']) if f.get('department_id') else None,
            consequences  = f.get('consequences', ''),
            reporter      = f.get('reporter_name', '') or 'Anonymous',
            classification= 'Voluntary Report',
        )
        db.session.commit()
        return render_template('portal/portal_submitted.html',
                               ref=rnum, report_type='Voluntary Safety Report')
    return render_template('portal/portal_voluntary.html')


@app.route('/portal/confidential', methods=['GET', 'POST'])
def portal_confidential():
    """Confidential Safety Report — identity protected."""
    if request.method == 'POST':
        f = request.form
        from datetime import date as _date
        rnum = f'CR-{_date.today().strftime("%Y%m%d")}-{ConfidentialReport.query.count()+1:03d}'
        rpt = ConfidentialReport(
            ref_number    = rnum,
            position      = f.get('position', ''),
            department_id = int(f['department_id']) if f.get('department_id') else None,
            date          = f.get('date', _date.today().isoformat()),
            location      = f.get('location', ''),
            report_type   = f.get('report_type', 'Safety Concern'),
            description   = f.get('description', ''),
            consequences  = f.get('consequences', ''),
            suggestion    = f.get('suggestion', ''),
            status        = 'Submitted',
        )
        db.session.add(rpt)
        db.session.flush()
        # Register in Safety Occurrence Registry (masked for confidentiality)
        _register_occurrence(
            report_id       = rnum,
            report_type     = 'Confidential',
            description     = f.get('description', ''),
            location        = f.get('location', ''),
            date_str        = f.get('date', ''),
            department_id   = int(f['department_id']) if f.get('department_id') else None,
            consequences    = f.get('consequences', ''),
            reporter        = 'Anonymous',
            is_confidential = True,
            classification  = 'Confidential Report',
        )
        db.session.commit()
        return render_template('portal/portal_submitted.html',
                               ref=rnum, report_type='Confidential Safety Report')
    return render_template('portal/portal_confidential.html')


@app.route('/portal/hazard', methods=['GET', 'POST'])
def portal_hazard():
    """Public Hazard Report form — simplified for fast submission."""
    if request.method == 'POST':
        # Reuse existing hazard_report POST logic
        return hazard_report()
    return render_template('portal/portal_hazard.html')


@app.route('/portal/asr', methods=['GET', 'POST'])
def portal_asr():
    """Public ASR form — reuses existing ASR logic."""
    if request.method == 'POST':
        return asr_report()
    return render_template('portal/portal_asr.html')


# ═══════════════════════════════════════════════════════════════════════════════
#  MOBILE REPORTING PORTAL  (PWA-style aviation reporting)
# ═══════════════════════════════════════════════════════════════════════════════

@app.route('/mobile')
def mobile_home():
    return render_template('mobile/mobile_home.html')


@app.route('/mobile/hazard', methods=['GET', 'POST'])
def mobile_hazard():
    departments = Department.query.all()
    if request.method == 'POST':
        f = request.form
        from datetime import date as _d
        err = None
        try:
            # 1. Create Hazard record
            haz = Hazard(
                id             = new_id('HAZ'),
                source         = 'Mobile Report',
                classification = f.get('classification', 'Operational'),
                generic_hazard = f.get('hazard_title', 'Mobile Report'),
                specific_components = f.get('hazard_description', ''),
                consequences   = f.get('consequences', ''),
                status         = 'Open',
                department_id  = int(f['dept_id']) if f.get('dept_id','').isdigit() else None,
            )
            db.session.add(haz)
            db.session.flush()
            # 2. Create HazardReport record
            rep = HazardReport(
                id             = new_id('HR'),
                hazard_id      = haz.id,
                department_id  = int(f['dept_id']) if f.get('dept_id','').isdigit() else None,
                date           = f.get('occurrence_date', _d.today().isoformat()),
                location       = f.get('location', ''),
                description    = f.get('hazard_description', ''),
                classification = f.get('classification', 'Operational'),
                generic_hazard = f.get('hazard_title', ''),
                consequences   = f.get('consequences', ''),
                immediate_action = f.get('immediate_action', ''),
                suggested_mitigation = f.get('suggested_mitigation', ''),
                reporter       = f.get('reporter_name', '') or identity_name or 'Anonymous',
                reporter_severity = 'Medium',
                status         = 'Submitted',
            )
            db.session.add(rep)
            # 3. Handle photo upload
            photo = request.files.get('photo')
            if photo and photo.filename:
                try:
                    from werkzeug.utils import secure_filename as _sf
                    import os
                    ext = photo.filename.rsplit('.', 1)[-1].lower() if '.' in photo.filename else ''
                    if ext in {'jpg','jpeg','png','gif','pdf','heic','bmp'}:
                        fname = _sf(f'mobile_{rep.id}_{photo.filename}')
                        updir = app.config.get('UPLOAD_FOLDER', 'uploads')
                        os.makedirs(updir, exist_ok=True)
                        photo.save(os.path.join(updir, fname))
                except Exception:
                    pass
            db.session.commit()
            return redirect(url_for('mobile_submitted', rtype='hazard', rid=rep.id))
        except Exception as e:
            db.session.rollback()
            err = str(e)[:120]
    else:
        err = None
    return render_template('mobile/mobile_hazard.html', departments=departments, error=err)


@app.route('/mobile/asr', methods=['GET', 'POST'])
def mobile_asr():
    if request.method == 'POST':
        f = request.form
        from datetime import date as _d
        err = None
        try:
            asr_id  = new_id('ASR')
            # Build event description with any dynamic category extras
            ev_desc = f.get('event_description', '')
            extras  = [f'{k.replace("extra_","").replace("_"," ").title()}: {v}'
                       for k,v in f.items() if k.startswith('extra_') and v]
            if extras:
                sep = chr(10)*2 + '--- Category Details ---' + chr(10)
                ev_desc += sep + chr(10).join(extras)
            rep = ASRReport(
                id               = asr_id,
                report_type      = f.get('report_type', 'Voluntary'),
                occurrence_type  = f.get('event_category', ''),
                captain          = f.get('reporter_name', '') or identity_name,
                captain_staff_no = f.get('staff_number', ''),
                date             = f.get('occurrence_date', _d.today().isoformat()),
                time_local       = f.get('time_local', ''),
                time_utc         = f.get('time_utc', ''),
                flight_no        = f.get('flight_number', ''),
                route_from       = f.get('route_from', ''),
                route_to         = f.get('route_to', ''),
                diverted_to      = f.get('diversion_airport', ''),
                squawk           = f.get('squawk', ''),
                aircraft_type    = f.get('aircraft_type', ''),
                registration     = f.get('registration', ''),
                pax              = int(f['pax'])        if f.get('pax','').isdigit()            else None,
                crew             = int(f['crew'])       if f.get('crew','').isdigit()           else None,
                altitude_ft      = int(f['altitude'])   if f.get('altitude','').isdigit()       else None,
                flight_phase     = f.get('flight_phase', ''),
                weather_wind     = f.get('weather_wind', ''),
                weather_vis_rvr  = f.get('visibility', ''),
                weather_clouds   = f.get('clouds', ''),
                weather_temp_c   = int(f['temperature']) if f.get('temperature','').lstrip('-').isdigit() else None,
                weather_qnh      = int(f['qnh'])        if f.get('qnh','').isdigit()            else None,
                runway           = f.get('runway', ''),
                runway_state     = f.get('runway_condition', ''),
                event_description= ev_desc,
                action_taken     = f.get('immediate_actions', ''),
                severity         = f.get('severity_level', 'C'),
            )
            db.session.add(rep)
            db.session.commit()
            return redirect(url_for('mobile_submitted', rtype='asr', rid=asr_id))
        except Exception as e:
            db.session.rollback()
            err = str(e)[:120]
        return render_template('mobile/mobile_asr.html', error=err)
    return render_template('mobile/mobile_asr.html', error=None)


@app.route('/mobile/confidential', methods=['GET', 'POST'])
def mobile_confidential():
    departments = Department.query.all()
    if request.method == 'POST':
        f = request.form
        from datetime import date as _d
        err = None
        try:
            from datetime import datetime as _dt
            seq  = ConfidentialReport.query.count() + 1
            ref  = 'CR-SMS-{:02d}'.format(seq)
            rep  = ConfidentialReport(
                ref_number    = ref,
                date          = f.get('occurrence_date', _d.today().isoformat()),
                location      = f.get('location', ''),
                description   = f.get('description', ''),
                consequences  = f.get('consequences', ''),
                suggestion    = f.get('suggestion', ''),
                department_id = int(f['dept_id']) if f.get('dept_id','').isdigit() else None,
                position      = f.get('reporter_position', ''),
                report_type   = 'Confidential',
                status        = 'Submitted',
            )
            db.session.add(rep)
            db.session.commit()
            return redirect(url_for('mobile_submitted', rtype='confidential', rid=ref))
        except Exception as e:
            db.session.rollback()
            err = str(e)[:120]
    else:
        err = None
    return render_template('mobile/mobile_confidential.html', departments=departments, error=err)


@app.route('/mobile/voluntary', methods=['GET', 'POST'])
def mobile_voluntary():
    departments = Department.query.all()
    if request.method == 'POST':
        f = request.form
        from datetime import date as _d
        err = None
        try:
            seq  = VoluntaryReport.query.count() + 1
            ref  = 'VR-SMS-{:02d}'.format(seq)
            rep  = VoluntaryReport(
                ref_number    = ref,
                date          = f.get('occurrence_date', _d.today().isoformat()),
                location      = f.get('location', ''),
                description   = f.get('description', ''),
                consequences  = f.get('consequences', ''),
                suggestion    = f.get('suggestion', ''),
                reporter_name = f.get('reporter_name', '') or 'Anonymous',
                position      = f.get('reporter_position', ''),
                department_id = int(f['dept_id']) if f.get('dept_id','').isdigit() else None,
                report_type   = 'Voluntary',
                status        = 'Submitted',
            )
            db.session.add(rep)
            db.session.commit()
            return redirect(url_for('mobile_submitted', rtype='voluntary', rid=ref))
        except Exception as e:
            db.session.rollback()
            err = str(e)[:120]
    else:
        err = None
    return render_template('mobile/mobile_voluntary.html', departments=departments, error=err)


@app.route('/mobile/submitted')
def mobile_submitted():
    return render_template('mobile/mobile_submitted.html',
                           rtype=request.args.get('rtype','report'),
                           rid=request.args.get('rid',''))


# ═══════════════════════════════════════════════════════════════════════════════
#  FLUTTER MOBILE API  — JSON endpoints for Flutter app
#  Base URL: https://jav-sms-p0c2.onrender.com/api/mobile
# ═══════════════════════════════════════════════════════════════════════════════

def api_ok(data=None, message='Success', status=200):
    return jsonify({'status': 'ok', 'message': message, 'data': data or {}}), status

def api_err(message='Error', status=400):
    return jsonify({'status': 'error', 'message': message}), status


@app.route('/api/mobile/hazard', methods=['POST', 'OPTIONS'])
@csrf.exempt
def api_mobile_hazard():
    """Flutter: Submit Hazard Report → existing Hazard workflow."""
    if request.method == 'OPTIONS':
        return api_ok()
    try:
        # Accept both JSON and form data from Flutter
        if request.is_json:
            f = request.get_json()
        else:
            f = request.form.to_dict()

        # Enrich reporter name from auth token if not provided
        token = request.headers.get('Authorization', '').replace('Bearer ', '')
        identity_name = ''
        try:
            id_data = _get_identity(token)
            if id_data:
                identity_name = id_data['name']
        except Exception:
            pass

        from datetime import date as _d
        # Resolve dept_id: payload > X-Dept-Id header > None
        dept_id_raw = str(f.get('dept_id', ''))
        if not dept_id_raw.isdigit():
            dept_id_raw = request.headers.get('X-Dept-Id', '')
        resolved_dept_id = int(dept_id_raw) if dept_id_raw.isdigit() else None

        haz = Hazard(
            id                  = new_id('HAZ'),
            source              = 'Flutter App',
            classification      = f.get('classification', 'Operational'),
            generic_hazard      = f.get('hazard_title', 'Mobile Hazard Report'),
            specific_components = f.get('hazard_description', ''),
            consequences        = f.get('consequences', ''),
            status              = 'Open',
            department_id       = resolved_dept_id,
        )
        haz.status = 'Under Assessment'
        db.session.add(haz)
        db.session.flush()

        rep = HazardReport(
            id                   = new_id('HR'),
            hazard_id            = haz.id,
            department_id        = resolved_dept_id,
            date                 = f.get('date', _d.today().isoformat()),
            location             = f.get('location', ''),
            description          = f.get('hazard_description', ''),
            classification       = f.get('classification', 'Operational'),
            generic_hazard       = f.get('hazard_title', ''),
            consequences         = f.get('consequences', ''),
            immediate_action     = f.get('immediate_action', ''),
            suggested_mitigation = f.get('suggested_mitigation', ''),
            reporter             = identity_name or f.get('reporter_name', '') or 'Anonymous',
            reporter_severity    = 'Medium',
            status               = 'Submitted',
        )
        db.session.add(rep)
        db.session.flush()

        # Phase 2: Create reporter feedback record (ICAO Annex 19 §3.1.2)
        try:
            reporter_uid = None
            if token:
                tok = ApiToken.query.filter_by(token=token).first()
                if tok:
                    reporter_uid = tok.user_id
                    rep.reporter_user_id = reporter_uid
            fb = get_or_create_feedback(rep.id, 'HazardReport', reporter_uid)
            db.session.add(fb)
        except Exception:
            pass

        db.session.commit()
        return api_ok({'report_id': rep.id, 'hazard_id': haz.id}, 'Hazard report submitted successfully', 201)
    except Exception as e:
        db.session.rollback()
        return api_err(str(e)[:120], 500)


@app.route('/api/mobile/asr', methods=['POST', 'OPTIONS'])
@csrf.exempt
def api_mobile_asr():
    """Flutter: Submit Air Safety Report → full SMS workflow (Hazard + ASR + HazardReport)."""
    if request.method == 'OPTIONS':
        return api_ok()
    try:
        f      = request.get_json() if request.is_json else request.form.to_dict()
        token  = request.headers.get('Authorization', '').replace('Bearer ', '')
        identity_name = ''
        submitter_dept_id = None  # will be resolved from token
        try:
            id_data = _get_identity(token)
            if id_data:
                identity_name = id_data['name']
                # Resolve submitter's real department from Employee/User record
                uid_str = str(id_data.get('uid', ''))
                if uid_str.startswith('emp_'):
                    _emp = Employee.query.get(int(uid_str.replace('emp_', '')))
                    if _emp: submitter_dept_id = _emp.department_id
                elif uid_str.startswith('usr_') or uid_str.isdigit():
                    _uid = int(uid_str.replace('usr_', '')) if uid_str.startswith('usr_') else int(uid_str)
                    _usr = User.query.get(_uid)
                    if _usr: submitter_dept_id = _usr.department_id
        except Exception:
            pass
        # Fall back to Flight Operations (dept 1) only if truly unknown
        dept_id = submitter_dept_id or 1
        from datetime import date as _d
        se     = f.get('severity_level', 'C')
        li     = 3
        ri     = f'{li}{se}'
        hid    = new_id('HAZ')
        asr_id = new_id('ASR')
        occ    = f.get('event_category', 'Flight Occurrence')

        # 1. Hazard record → Hazard Log + Dashboard
        haz = Hazard(
            id                  = hid,
            source              = 'ASR (Mobile)',
            linked_report_id    = asr_id,
            department_id       = dept_id,
            classification      = 'Operational',
            type_of_activity    = 'Flight Operations',
            generic_hazard      = occ,
            specific_components = f.get('event_description', ''),
            consequences        = f.get('operational_impact', 'To Be Assessed'),
            status              = 'Open',
            owner               = f.get('reporter_name', '') or 'Flight Operations Manager',
        )
        db.session.add(haz)
        db.session.flush()

        # 2. ASR record
        rep = ASRReport(
            id               = asr_id,
            report_type      = f.get('report_type', 'Voluntary'),
            occurrence_type  = occ,
            captain          = identity_name or f.get('reporter_name', '') or 'Anonymous',
            captain_staff_no = f.get('staff_number', ''),
            date             = f.get('date', _d.today().isoformat()),
            time_local       = f.get('time_local', ''),
            time_utc         = f.get('time_utc', ''),
            flight_no        = f.get('flight_number', ''),
            route_from       = f.get('route_from', ''),
            route_to         = f.get('route_to', ''),
            aircraft_type    = f.get('aircraft_type', ''),
            registration     = f.get('registration', ''),
            flight_phase     = f.get('flight_phase', ''),
            event_description= f.get('event_description', ''),
            action_taken     = f.get('immediate_actions', ''),
            severity         = se, likelihood=li, risk_index=ri, hazard_id=hid,
        )
        db.session.add(rep)
        db.session.flush()

        # 3. HazardReport → appears in Hazard Reports list
        try:
            hr = HazardReport(
                id             = new_id('HR'),
                hazard_id      = hid,
                department_id  = dept_id,
                date           = f.get('date', _d.today().isoformat()),
                location       = f'{f.get("route_from","")}-{f.get("route_to","")}',
                description    = f.get('event_description', ''),
                classification = 'Operational',
                generic_hazard = occ,
                consequences   = f.get('operational_impact', 'To Be Assessed'),
                immediate_action = f.get('immediate_actions', ''),
                reporter       = identity_name or f.get('reporter_name', '') or 'Flight Crew',
                reporter_severity = se,
                report_type    = 'ASR',
                status         = 'Submitted',
            )
            db.session.add(hr)
        except Exception:
            pass

        # Mark hazard as Under Assessment
        haz.status = 'Under Assessment'
        db.session.commit()
        return api_ok({'report_id': asr_id, 'hazard_id': hid},
                      'ASR submitted successfully', 201)
    except Exception as e:
        db.session.rollback()
        return api_err(str(e)[:120], 500)


@app.route('/api/mobile/confidential', methods=['POST', 'OPTIONS'])
@csrf.exempt
def api_mobile_confidential():
    """Flutter: Submit Confidential Report.

    Accepts all fields collected by confidential_screen:
    - description, location, consequences, suggestion
    - reporter_position (optional — blank for fully anonymous)
    - report_type: Safety Concern / Systemic Issue / Management Issue / etc.
    - dept_id: from payload OR X-Dept-Id header (auth-token fallback)
    Identity is NOT recorded for confidential reports.
    """
    if request.method == 'OPTIONS':
        return api_ok()
    try:
        f = request.get_json() if request.is_json else request.form.to_dict()
        from datetime import date as _d
        # Resolve dept_id: payload > X-Dept-Id header > token lookup > None
        dept_id_raw = str(f.get('dept_id', ''))
        if not dept_id_raw.isdigit():
            dept_id_raw = request.headers.get('X-Dept-Id', '')
        if not dept_id_raw.isdigit():
            token = request.headers.get('Authorization', '').replace('Bearer ', '')
            try:
                id_data = _get_identity(token)
                if id_data:
                    uid_str = str(id_data.get('uid', ''))
                    if uid_str.startswith('emp_'):
                        _emp = Employee.query.get(int(uid_str.replace('emp_', '')))
                        if _emp and _emp.department_id:
                            dept_id_raw = str(_emp.department_id)
                    elif uid_str.startswith('usr_') or uid_str.isdigit():
                        _uid = int(uid_str.replace('usr_', '')) if uid_str.startswith('usr_') else int(uid_str)
                        _usr = User.query.get(_uid)
                        if _usr and _usr.department_id:
                            dept_id_raw = str(_usr.department_id)
            except Exception:
                pass
        resolved_dept_id = int(dept_id_raw) if dept_id_raw.isdigit() else None
        # report_type stores the Flutter sub-category (Safety Concern, Systemic Issue, etc.)
        report_subtype = f.get('report_type', 'Confidential')
        seq = ConfidentialReport.query.count() + 1
        ref = 'CR-SMS-{:03d}'.format(seq)
        rep = ConfidentialReport(
            ref_number    = ref,
            date          = f.get('date', _d.today().isoformat()),
            location      = f.get('location', ''),
            description   = f.get('description', ''),
            consequences  = f.get('consequences', ''),
            suggestion    = f.get('suggestion', ''),
            department_id = resolved_dept_id,
            position      = f.get('reporter_position', ''),
            report_type   = report_subtype,
            status        = 'Submitted',
        )
        db.session.add(rep)
        db.session.flush()
        # Register in Safety Occurrence Registry (masked)
        _register_occurrence(
            report_id       = ref,
            report_type     = 'Confidential',
            description     = f.get('description', ''),
            location        = f.get('location', ''),
            date_str        = f.get('date', ''),
            consequences    = f.get('consequences', ''),
            reporter        = 'Anonymous',
            is_confidential = True,
            classification  = 'Confidential Report',
        )
        db.session.commit()
        return api_ok({'report_id': ref}, 'Confidential report submitted', 201)
    except Exception as e:
        db.session.rollback()
        return api_err(str(e)[:120], 500)


@app.route('/api/mobile/voluntary', methods=['POST', 'OPTIONS'])
@csrf.exempt
def api_mobile_voluntary():
    """Flutter: Submit Voluntary Safety Report.

    Accepts all fields collected by the Flutter voluntary_screen:
    - description, location, consequences, suggestion (body text fields)
    - reporter_name, reporter_position (identity — optional / anonymous)
    - report_type: Near Miss / Safety Concern / Unsafe Condition / etc.
    - dept_id: from payload OR X-Dept-Id header (auth-token fallback)
    """
    if request.method == 'OPTIONS':
        return api_ok()
    try:
        f = request.get_json() if request.is_json else request.form.to_dict()
        from datetime import date as _d
        # Resolve dept_id: payload > X-Dept-Id header > token lookup > None
        dept_id_raw = str(f.get('dept_id', ''))
        if not dept_id_raw.isdigit():
            dept_id_raw = request.headers.get('X-Dept-Id', '')
        if not dept_id_raw.isdigit():
            token = request.headers.get('Authorization', '').replace('Bearer ', '')
            try:
                id_data = _get_identity(token)
                if id_data:
                    uid_str = str(id_data.get('uid', ''))
                    if uid_str.startswith('emp_'):
                        _emp = Employee.query.get(int(uid_str.replace('emp_', '')))
                        if _emp and _emp.department_id:
                            dept_id_raw = str(_emp.department_id)
                    elif uid_str.startswith('usr_') or uid_str.isdigit():
                        _uid = int(uid_str.replace('usr_', '')) if uid_str.startswith('usr_') else int(uid_str)
                        _usr = User.query.get(_uid)
                        if _usr and _usr.department_id:
                            dept_id_raw = str(_usr.department_id)
            except Exception:
                pass
        resolved_dept_id = int(dept_id_raw) if dept_id_raw.isdigit() else None
        # report_type stores the Flutter sub-category (Near Miss, Safety Concern, etc.)
        report_subtype = f.get('report_type', 'Voluntary Safety Report')
        seq = VoluntaryReport.query.count() + 1
        ref = 'VR-SMS-{:03d}'.format(seq)
        rep = VoluntaryReport(
            ref_number    = ref,
            date          = f.get('date', _d.today().isoformat()),
            location      = f.get('location', ''),
            description   = f.get('description', ''),
            consequences  = f.get('consequences', ''),
            suggestion    = f.get('suggestion', ''),
            reporter_name = f.get('reporter_name', '') or 'Anonymous',
            position      = f.get('reporter_position', ''),
            department_id = resolved_dept_id,
            report_type   = report_subtype,
            status        = 'Submitted',
        )
        db.session.add(rep)
        db.session.flush()
        # Register in Safety Occurrence Registry
        _register_occurrence(
            report_id     = ref,
            report_type   = 'Voluntary',
            description   = f.get('description', ''),
            location      = f.get('location', ''),
            date_str      = f.get('date', ''),
            consequences  = f.get('consequences', ''),
            reporter      = f.get('reporter_name', '') or 'Anonymous',
            classification= 'Voluntary Report',
        )
        db.session.commit()
        return api_ok({'report_id': ref}, 'Voluntary report submitted', 201)
    except Exception as e:
        db.session.rollback()
        return api_err(str(e)[:120], 500)


@app.route('/api/report/detail/<rid>', methods=['GET'])
def api_report_detail(rid):
    """Flutter: Get full report detail — only own reports accessible."""
    token = request.headers.get('Authorization','').replace('Bearer ','')
    identity = _get_identity(token)
    if not identity:
        return api_err('Unauthorized', 401)
    try:
        # Try HazardReport first
        hr = HazardReport.query.get(rid)
        if hr:
            wf_status, wf_color, wf_stage, timeline, wf_guidance, wf_responsible, wf_next = resolve_report_status(
                hazard_id=hr.hazard_id, hr_status=hr.status)
            timeline = get_report_timeline(hr.hazard_id, hr.created_at,
                                           hr.report_type or 'Hazard Report')
            actions = []
            if hr.hazard_id:
                for a in Action.query.filter_by(hazard_id=hr.hazard_id).all():
                    actions.append({
                        'id': a.id, 'description': (a.description or '')[:100],
                        'status': a.status or '', 'owner': a.owner or '',
                        'due_date': a.due_date or '', 'priority': a.priority or '',
                    })
            return api_ok({
                'id': hr.id, 'type': hr.report_type or 'Hazard Report',
                'title': hr.generic_hazard or '', 'description': hr.description or '',
                'location': hr.location or '', 'date': hr.date or '',
                'status': wf_status, 'status_color': wf_color, 'stage': wf_stage,
                'guidance': wf_guidance, 'responsible': wf_responsible,
                'next_step': wf_next,
                'severity': hr.reporter_severity or 'Medium',
                'hazard_id': hr.hazard_id or '',
                'created_at': hr.created_at.isoformat() if hr.created_at else '',
                'timeline': timeline, 'actions': actions,
            }, 'Report loaded')
        # Try ASR
        asr = ASRReport.query.get(rid)
        if asr:
            wf_status, wf_color, wf_stage, timeline, wf_guidance, wf_responsible, wf_next = resolve_report_status(
                hazard_id=asr.hazard_id, hr_status='Submitted')
            timeline = get_report_timeline(asr.hazard_id, asr.created_at, 'ASR')
            return api_ok({
                'id': asr.id, 'type': 'ASR',
                'title': asr.occurrence_type or 'Air Safety Report',
                'description': asr.event_description or '',
                'location': f'{asr.route_from or ""}→{asr.route_to or ""}',
                'date': asr.date or '',
                'status': wf_status, 'status_color': wf_color, 'stage': wf_stage,
                'severity': asr.severity or 'C', 'hazard_id': asr.hazard_id or '',
                'created_at': asr.created_at.isoformat() if asr.created_at else '',
                'timeline': timeline, 'actions': [],
            }, 'Report loaded')
        return api_err('Report not found', 404)
    except Exception as e:
        return api_err(str(e)[:120], 500)


@app.route('/api/mobile/ping', methods=['GET'])
def api_ping():
    """Flutter: Health check — verify API is reachable."""
    return api_ok({'server': 'AviaS Safety Management System', 'version': '1.0'}, 'API online')


@app.route('/api/mobile/register_token', methods=['POST', 'OPTIONS'])
@csrf.exempt
def api_register_device_token():
    """Flutter Phase 5: Register FCM push-notification token for this device.

    Upserts a DeviceToken row keyed on (user_id, fcm_token) so each device
    gets at most one row. Called on login and whenever FCM rotates the token.
    """
    if request.method == 'OPTIONS':
        return api_ok()
    token = request.headers.get('Authorization', '').replace('Bearer ', '').strip()
    identity = _get_identity(token)
    if not identity:
        return api_err('Unauthorized', 401)
    try:
        data      = request.get_json(force=True, silent=True) or {}
        fcm_token = str(data.get('fcm_token', '')).strip()
        if not fcm_token:
            return api_err('fcm_token is required')
        uid = identity['uid']
        # Upsert: delete any stale row for this user/token, then insert fresh
        existing = DeviceToken.query.filter_by(user_id=uid, fcm_token=fcm_token).first()
        if not existing:
            db.session.add(DeviceToken(user_id=uid, fcm_token=fcm_token))
        else:
            existing.updated_at = datetime.utcnow()
        db.session.commit()
        return api_ok({'registered': True}, 'FCM token registered')
    except Exception as e:
        db.session.rollback()
        return api_err(f'Registration failed: {e}')


# ── EMPLOYEE AUTH APIs ────────────────────────────────────────────────────────

def _get_identity(token):
    """
    Resolve token → (account_type, full_name, employee_id, is_admin).
    Returns None if token invalid.
    """
    data = _verify_token(token)
    if not data:
        return None
    uid = str(data.get('user_id', ''))
    try:
        if uid.startswith('emp_'):
            emp = Employee.query.get(int(uid.replace('emp_', '')))
            if emp:
                return {
                    'type':     'employee',
                    'uid':      uid,
                    'name':     emp.full_name or emp.username,
                    'username': emp.username,
                    'emp_id':   emp.employee_id,
                    'is_admin': False,
                }
        else:
            usr = User.query.get(int(uid.replace('usr_', '')))
            if usr:
                return {
                    'type':     'admin',
                    'uid':      uid,
                    'name':     usr.full_name or usr.username,
                    'username': usr.username,
                    'emp_id':   '',
                    'is_admin': usr.role in ('admin', 'safety_manager', 'safety_officer'),
                }
    except Exception:
        pass
    return None


def _ensure_device_token_table():
    """Create device_tokens table if it doesn't exist yet (Phase 5 — FCM push tokens)."""
    try:
        db.session.execute(db.text('SELECT 1 FROM device_tokens LIMIT 1'))
    except Exception:
        db.session.rollback()
        try:
            db.session.execute(db.text('''
                CREATE TABLE IF NOT EXISTS device_tokens (
                    id SERIAL PRIMARY KEY,
                    user_id VARCHAR(30) NOT NULL,
                    fcm_token TEXT NOT NULL,
                    updated_at TIMESTAMP DEFAULT NOW(),
                    CONSTRAINT uq_device_token UNIQUE (user_id, fcm_token)
                )
            '''))
            db.session.execute(db.text(
                'CREATE INDEX IF NOT EXISTS ix_device_tokens_user_id ON device_tokens(user_id)'
            ))
            db.session.commit()
        except Exception:
            db.session.rollback()


def _ensure_token_table():
    """Create api_tokens table if it doesn't exist yet (safe for free-tier Render)."""
    try:
        db.session.execute(db.text('SELECT 1 FROM api_tokens LIMIT 1'))
    except Exception:
        db.session.rollback()
        try:
            db.session.execute(db.text('''
                CREATE TABLE IF NOT EXISTS api_tokens (
                    id SERIAL PRIMARY KEY,
                    token VARCHAR(64) UNIQUE NOT NULL,
                    user_id VARCHAR(30) NOT NULL,
                    username VARCHAR(80) NOT NULL,
                    expires_at TIMESTAMP NOT NULL,
                    created_at TIMESTAMP DEFAULT NOW()
                )
            '''))
            db.session.execute(db.text(
                'CREATE INDEX IF NOT EXISTS ix_api_tokens_token ON api_tokens(token)'
            ))
            db.session.commit()
        except Exception:
            db.session.rollback()

def _make_token(user_id, username):
    """Generate a secure token persisted in the api_tokens table.

    Survives Gunicorn worker restarts and Render dyno spin-downs.
    Auto-creates the table if missing (safe on free-tier Render).
    """
    import secrets
    from datetime import timedelta
    _ensure_token_table()
    _ensure_device_token_table()
    token = secrets.token_urlsafe(32)
    expires_at = datetime.utcnow() + timedelta(hours=24)
    # Lazy cleanup: remove expired tokens for this user_id before adding new one
    try:
        ApiToken.query.filter(
            ApiToken.user_id == str(user_id),
            ApiToken.expires_at < datetime.utcnow()
        ).delete(synchronize_session=False)
        db.session.flush()
    except Exception:
        db.session.rollback()
    try:
        db.session.add(ApiToken(
            token=token,
            user_id=str(user_id),
            username=username,
            expires_at=expires_at,
        ))
        db.session.commit()
    except Exception:
        db.session.rollback()
        raise
    return token

def _verify_token(token):
    """Look up token in DB and return {'user_id': ..., 'username': ...} or None."""
    if not token:
        return None
    try:
        row = ApiToken.query.filter_by(token=token).first()
        if row is None:
            return None
        if row.is_expired():
            db.session.delete(row)
            db.session.commit()
            return None
        return {'user_id': row.user_id, 'username': row.username}
    except Exception:
        return None

# NOTE: _token_store removed — tokens now persisted in api_tokens table


@app.route('/api/login', methods=['POST', 'OPTIONS'])
@csrf.exempt
@limiter.limit('10 per minute')
def api_login():
    """Flutter: Employee login → returns auth token."""
    if request.method == 'OPTIONS':
        return api_ok()
    try:
        f = request.get_json() if request.is_json else request.form.to_dict()
        username = f.get('username', '').strip()
        password = f.get('password', '')
        if not username or not password:
            return api_err('Username and password required', 400)
        # Check Employee table first (mobile users), then User table (admins)
        emp = Employee.query.filter_by(username=username, is_active=True).first()
        user = None if emp else User.query.filter_by(username=username, is_active=True).first()

        if emp:
            if not check_pw(password, emp.password_hash):
                return api_err('Invalid username or password', 401)
            # Silently upgrade legacy or scrypt hash to pbkdf2 on successful login
            if _is_legacy_hash(emp.password_hash) or emp.password_hash.startswith('scrypt:'):
                emp.password_hash = hash_pw(password)
            try:
                emp.last_login = datetime.utcnow()
                db.session.commit()
            except Exception:
                db.session.rollback()  # last_login column may not exist yet — non-fatal
            token = _make_token(f'emp_{emp.id}', emp.username)
            dept_name = ''
            if emp.department_id:
                try:
                    dept = Department.query.get(emp.department_id)
                    dept_name = dept.name if dept else ''
                except Exception: pass
            from datetime import timedelta
            expires_at = (datetime.utcnow() + timedelta(hours=24)).isoformat() + 'Z'
            return api_ok({
                'token':         token,
                'user_id':       f'emp_{emp.id}',
                'username':      emp.username,
                'full_name':     emp.full_name,
                'role':          emp.role or 'employee',
                'department':    dept_name,
                'department_id': emp.department_id,
                'employee_id':   emp.employee_id,
                'account_type':  'employee',
                'expires_at':    expires_at,
            }, 'Login successful')
        elif user:
            if not check_pw(password, user.password_hash):
                return api_err('Invalid username or password', 401)
            # Silently upgrade legacy or scrypt hash to pbkdf2 on successful login
            if _is_legacy_hash(user.password_hash) or user.password_hash.startswith('scrypt:'):
                user.password_hash = hash_pw(password)
            try:
                user.last_login = datetime.utcnow()
                db.session.commit()
            except Exception:
                db.session.rollback()  # last_login column may not exist yet — non-fatal
            token = _make_token(f'usr_{user.id}', user.username)
            dept_name = ''
            if user.department_id:
                try:
                    dept = Department.query.get(user.department_id)
                    dept_name = dept.name if dept else ''
                except Exception: pass
            from datetime import timedelta
            expires_at = (datetime.utcnow() + timedelta(hours=24)).isoformat() + 'Z'
            return api_ok({
                'token':         token,
                'user_id':       f'usr_{user.id}',
                'username':      user.username,
                'full_name':     user.full_name or user.username,
                'role':          user.role or 'admin',
                'department':    dept_name,
                'department_id': user.department_id,
                'employee_id':   '',
                'account_type':  'admin',
                'expires_at':    expires_at,
            }, 'Login successful')
        else:
            return api_err('Invalid username or password', 401)
    except Exception as e:
        return api_err(str(e)[:120], 500)


@app.route('/api/me', methods=['GET'])
def api_me():
    """Flutter: Get current user/employee profile.

    Token user_id is prefixed: 'emp_<id>' for Employee accounts,
    'usr_<id>' for admin/safety User accounts. Previously this always
    queried the User table, causing 404 for all Employee users.
    """
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    data = _verify_token(token)
    if not data:
        return api_err('Unauthorized — please login', 401)
    try:
        uid_str = str(data['user_id'])

        if uid_str.startswith('emp_'):
            # Mobile employee account
            emp_id = int(uid_str.replace('emp_', ''))
            emp = Employee.query.get(emp_id)
            if not emp:
                return api_err('Employee not found', 404)
            dept_name = ''
            if emp.department_id:
                dept = Department.query.get(emp.department_id)
                dept_name = dept.name if dept else ''
            import json as _json
            profile_image_url = _profile_image_url(emp)
            try:
                notif_prefs    = _json.loads(emp.notification_prefs or '{}')
                privacy_prefs  = _json.loads(emp.privacy_settings or '{}')
            except Exception:
                notif_prefs    = {}
                privacy_prefs  = {}
            return api_ok({
                # Identity (read-only in Flutter — managed by web admin)
                'user_id':            uid_str,
                'username':           emp.username,
                'full_name':          emp.full_name,
                'role':               emp.role or 'employee',
                'department':         dept_name,
                'department_id':      emp.department_id,
                'employee_id':        emp.employee_id,
                'email':              emp.email or '',
                'mobile':             emp.mobile or '',
                'base_station':       emp.base_station or 'AMM',
                'join_date':          emp.join_date or '',
                'employment_status':  emp.employment_status or 'Active',
                'position':           emp.position or '',
                'account_type':       'employee',
                'last_login':         emp.last_login.isoformat() if emp.last_login else '',
                'profile_image_url':  profile_image_url,
                # Preferences (Flutter-editable)
                'language_preference': emp.language_preference or 'en',
                'dark_mode':           emp.dark_mode or False,
                'notification_prefs':  notif_prefs,
                'privacy_settings':    privacy_prefs,
            }, 'Profile loaded')
        else:
            # Web admin / safety user account
            usr_id = int(uid_str.replace('usr_', '')) if uid_str.startswith('usr_') else int(uid_str)
            user = User.query.get(usr_id)
            if not user:
                return api_err('User not found', 404)
            dept_name = ''
            if user.department_id:
                dept = Department.query.get(user.department_id)
                dept_name = dept.name if dept else ''
            return api_ok({
                'user_id':       uid_str,
                'username':      user.username,
                'full_name':     user.full_name or user.username,
                'role':          user.role or 'admin',
                'department':    dept_name,
                'department_id': user.department_id,
                'employee_id':   '',
                'email':         '',
                'mobile':        '',
                'account_type':  'admin',
                'last_login':    user.last_login.isoformat() if user.last_login else '',
            }, 'Profile loaded')
    except Exception as e:
        return api_err(str(e)[:120], 500)


def _resolve_employee(data):
    """Resolve token data to (uid_str, emp, emp_name, emp_username) or None tuple."""
    uid_str = str(data['user_id'])
    emp, emp_name, emp_username = None, '', ''
    if uid_str.startswith('emp_'):
        emp = Employee.query.get(int(uid_str.replace('emp_', '')))
        if emp:
            emp_name = emp.full_name
            emp_username = emp.username
    return uid_str, emp, emp_name, emp_username


@app.route('/api/mobile/profile/full', methods=['GET'])
@csrf.exempt
def api_mobile_profile_full():
    """Employee Portal: Personal info + comprehensive safety statistics."""
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    data = _verify_token(token)
    if not data:
        return api_err('Unauthorized', 401)
    try:
        from sqlalchemy import or_
        uid_str, emp, emp_name, emp_username = _resolve_employee(data)

        reports_by_type = {'Hazard': 0, 'ASR': 0, 'Voluntary': 0}
        total_reports = 0
        if emp_name:
            reports_by_type['Hazard'] = HazardReport.query.filter(or_(
                HazardReport.reporter == emp_name,
                HazardReport.reporter_user_id == uid_str)).count()
            reports_by_type['ASR'] = ASRReport.query.filter(or_(
                ASRReport.captain == emp_name,
                ASRReport.copilot == emp_name)).count()
            reports_by_type['Voluntary'] = VoluntaryReport.query.filter(
                VoluntaryReport.reporter_name == emp_name).count()
            total_reports = sum(reports_by_type.values())

        open_actions = overdue_count = closed_actions = 0
        if emp_name:
            open_actions  = Action.query.filter(Action.owner == emp_name, Action.status.in_(['Open', 'In Progress'])).count()
            overdue_count = Action.query.filter(Action.owner == emp_name, Action.status == 'Overdue').count()
            closed_actions = Action.query.filter(Action.owner == emp_name, Action.status == 'Closed').count()

        training_count = 0
        expired_training = 0
        if emp_name:
            training_count = Training.query.filter(Training.employee_name == emp_name).count()
            expired_training = Training.query.filter(Training.employee_name == emp_name, Training.status == 'Expired').count()

        reads_count = 0
        acks_count  = 0
        if uid_str:
            all_reads   = SafetyPromoRead.query.filter_by(user_id=uid_str).all()
            reads_count = sum(1 for r in all_reads if _content_exists(r.content_type, r.content_id))
            all_acks    = SafetyPromoAck.query.filter_by(user_id=uid_str).all()
            acks_count  = sum(1 for r in all_acks  if _content_exists(r.content_type, r.content_id))

        dept_name = ''
        if emp and emp.department:
            dept_name = emp.department.name or ''

        profile_image_url = _profile_image_url(emp) if emp else None

        return api_ok({
            'employee_id':       emp.employee_id if emp else '',
            'full_name':         emp_name,
            'email':             (emp.email or '') if emp else '',
            'phone':             (emp.mobile or '') if emp else '',
            'role':              (emp.role or 'Employee') if emp else '',
            'department':        dept_name,
            'base_station':      (emp.base_station or 'AMM') if emp else 'AMM',
            'join_date':         (emp.join_date or '') if emp else '',
            'employment_status': (emp.employment_status or 'Active') if emp else 'Active',
            'position':          (emp.position or '') if emp else '',
            'profile_image_url': profile_image_url,
            'total_reports':     total_reports,
            'reports_by_type':   reports_by_type,
            'open_actions':      open_actions,
            'overdue_actions':   overdue_count,
            'closed_actions':    closed_actions,
            'training_count':    training_count,
            'expired_training':  expired_training,
            'content_reads':     reads_count,
            'content_acks':      acks_count,
        }, 'Profile loaded')
    except Exception as e:
        return api_err(str(e)[:120], 500)


@app.route('/api/mobile/profile/safety-score', methods=['GET'])
@csrf.exempt
def api_mobile_profile_safety_score():
    """Employee Portal: Compute safety participation score."""
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    data = _verify_token(token)
    if not data:
        return api_err('Unauthorized', 401)
    try:
        uid_str, emp, emp_name, _ = _resolve_employee(data)

        # Reporting activity (max 30 pts): 5 pts per report, cap 30
        from sqlalchemy import or_
        total_reports = 0
        if emp_name:
            hz  = HazardReport.query.filter(or_(HazardReport.reporter == emp_name,
                      HazardReport.reporter_user_id == uid_str)).count()
            asr = ASRReport.query.filter(or_(ASRReport.captain == emp_name,
                      ASRReport.copilot == emp_name)).count()
            vol = VoluntaryReport.query.filter(VoluntaryReport.reporter_name == emp_name).count()
            total_reports = hz + asr + vol
        reporting_score = min(30, total_reports * 5)

        # Training compliance (max 30 pts)
        training_score = 30
        if emp_name:
            expired = Training.query.filter(Training.employee_name == emp_name, Training.status == 'Expired').count()
            training_score = max(0, 30 - expired * 10)

        # Action compliance (max 25 pts): deduct for overdue
        action_score = 25
        if emp_name:
            overdue = Action.query.filter(Action.owner == emp_name, Action.status == 'Overdue').count()
            action_score = max(0, 25 - overdue * 8)

        # Content engagement (max 15 pts): reads + acks for EXISTING content only
        # Per-type breakdown for transparency
        engagement_score   = 0
        reads_by_type      = {'newsletter': 0, 'bulletin': 0, 'survey': 0, 'lesson': 0, 'other': 0}
        stale_reads        = 0
        valid_reads        = 0
        valid_acks         = 0
        if uid_str:
            read_rows = SafetyPromoRead.query.filter_by(user_id=uid_str).all()
            for r in read_rows:
                if _content_exists(r.content_type, r.content_id):
                    valid_reads += 1
                    key = r.content_type if r.content_type in reads_by_type else 'other'
                    reads_by_type[key] += 1
                else:
                    stale_reads += 1
            ack_rows = SafetyPromoAck.query.filter_by(user_id=uid_str).all()
            valid_acks = sum(
                1 for r in ack_rows if _content_exists(r.content_type, r.content_id))
            engagement_score = min(15, valid_reads * 2 + valid_acks * 3)

        total = reporting_score + training_score + action_score + engagement_score
        level = 'Excellent' if total >= 90 else ('Good' if total >= 70 else ('Fair' if total >= 50 else 'Needs Improvement'))

        return api_ok({
            'score':             total,
            'level':             level,
            'reporting_score':   reporting_score,
            'training_score':    training_score,
            'action_score':      action_score,
            'engagement_score':  engagement_score,
            # Engagement breakdown — visible in Flutter profile
            'engagement_detail': {
                'newsletter_reads':  reads_by_type['newsletter'],
                'bulletin_reads':    reads_by_type['bulletin'],
                'survey_reads':      reads_by_type['survey'],
                'lesson_reads':      reads_by_type['lesson'],
                'acknowledgements':  valid_acks,
                'stale_reads':       stale_reads,   # reads for deleted content (not counted)
                'points_from_reads': valid_reads * 2,
                'points_from_acks':  valid_acks * 3,
            },
        }, 'Safety score calculated')
    except Exception as e:
        return api_err(str(e)[:120], 500)


@app.route('/api/mobile/profile/training', methods=['GET'])
@csrf.exempt
def api_mobile_profile_training():
    """Employee Portal: Training records for current employee."""
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    data = _verify_token(token)
    if not data:
        return api_err('Unauthorized', 401)
    try:
        _, _, emp_name, _ = _resolve_employee(data)
        records = []
        if emp_name:
            trainings = Training.query.filter(
                Training.employee_name == emp_name
            ).order_by(Training.created_at.desc()).limit(20).all()
            for t in trainings:
                records.append({
                    'training_name': t.training_program or t.training_type or '',
                    'status':        t.status or 'Completed',
                    'training_date': t.training_date or '',
                    'expiry_date':   t.expiry_date or '',
                    'scheduled_date': t.scheduled_date or '',
                    'provider':      t.instructor or '',
                    'category':      t.training_type or '',
                })
        return api_ok(records, 'Training records loaded')
    except Exception as e:
        return api_err(str(e)[:120], 500)


@app.route('/api/mobile/profile/actions', methods=['GET'])
@csrf.exempt
def api_mobile_profile_actions():
    """Employee Portal: All corrective actions owned by current employee."""
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    data = _verify_token(token)
    if not data:
        return api_err('Unauthorized', 401)
    try:
        _, _, emp_name, _ = _resolve_employee(data)
        records = []
        if emp_name:
            actions = Action.query.filter(
                Action.owner == emp_name
            ).order_by(Action.due_date.asc()).limit(50).all()
            for a in actions:
                records.append({
                    'ref':      str(a.id),
                    'title':    (a.description or '')[:80],
                    'status':   a.status or 'Open',
                    'priority': a.priority or 'Medium',
                    'due_date': a.due_date or '',
                    'source':   a.source or '',
                    'linked_ref': a.linked_ref_id or a.hazard_id or '',
                })
        return api_ok(records, 'Actions loaded')
    except Exception as e:
        return api_err(str(e)[:120], 500)


@app.route('/api/mobile/profile/timeline', methods=['GET'])
@csrf.exempt
def api_mobile_profile_timeline():
    """Employee Portal: Combined safety activity timeline."""
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    data = _verify_token(token)
    if not data:
        return api_err('Unauthorized', 401)
    try:
        from sqlalchemy import or_
        uid_str, emp, emp_name, emp_username = _resolve_employee(data)
        events = []

        if emp_name:
            for h in HazardReport.query.filter(or_(
                    HazardReport.reporter == emp_name,
                    HazardReport.reporter_user_id == uid_str)).order_by(HazardReport.created_at.desc()).limit(10).all():
                events.append({'type': 'Report', 'subtype': 'Hazard',
                    'title': (h.generic_hazard or h.description or 'Hazard Report')[:80],
                    'detail': h.id or '', 'status': h.status or '',
                    'date': h.date or (h.created_at.isoformat()[:10] if h.created_at else ''),
                    'ts': h.created_at})

            for a in ASRReport.query.filter(or_(
                    ASRReport.captain == emp_name,
                    ASRReport.copilot == emp_name)).order_by(ASRReport.created_at.desc()).limit(10).all():
                events.append({'type': 'Report', 'subtype': 'ASR',
                    'title': (a.occurrence_type or f'ASR {a.flight_no or ""}').strip() or 'ASR Report',
                    'detail': a.id or '', 'status': a.status or '',
                    'date': a.date or (a.created_at.isoformat()[:10] if a.created_at else ''),
                    'ts': a.created_at})

            for v in VoluntaryReport.query.filter(
                    VoluntaryReport.reporter_name == emp_name).order_by(VoluntaryReport.created_at.desc()).limit(5).all():
                events.append({'type': 'Report', 'subtype': 'Voluntary',
                    'title': (v.report_type or 'Voluntary Report'),
                    'detail': v.ref_number or '', 'status': v.status or '',
                    'date': v.date or (v.created_at.isoformat()[:10] if v.created_at else ''),
                    'ts': v.created_at})

            for ac in Action.query.filter(Action.owner == emp_name).order_by(Action.created_at.desc()).limit(10).all():
                events.append({'type': 'Action', 'subtype': ac.status or 'Open',
                    'title': (ac.description or 'Corrective Action')[:60],
                    'detail': f'Due: {ac.due_date or "—"}', 'status': ac.status or 'Open',
                    'date': ac.created_at.isoformat()[:10] if ac.created_at else '', 'ts': ac.created_at})

            for t in Training.query.filter(Training.employee_name == emp_name).order_by(Training.created_at.desc()).limit(5).all():
                events.append({'type': 'Training', 'subtype': t.status or 'Completed',
                    'title': t.training_program or t.training_type or 'Training',
                    'detail': f'Expires: {t.expiry_date or "—"}', 'status': t.status or 'Completed',
                    'date': t.training_date or '', 'ts': t.created_at})

        if uid_str:
            # Fetch more than needed so we can filter stale rows and still get 5 valid ones
            read_rows = SafetyPromoRead.query.filter_by(user_id=uid_str).order_by(
                SafetyPromoRead.read_at.desc()).limit(30).all()
            for r in read_rows:
                if not _content_exists(r.content_type, r.content_id):
                    continue  # source deleted — skip silently
                events.append({'type': 'Content', 'subtype': r.content_type,
                    'title': f'Read {r.content_type.capitalize()}',
                    'detail': r.content_id, 'status': 'Read',
                    'date': r.read_at.isoformat()[:10] if r.read_at else '', 'ts': r.read_at})

        events.sort(key=lambda x: (x.get('ts') or ''), reverse=True)
        result = [{'type': e['type'], 'subtype': e.get('subtype', ''), 'title': e['title'],
                   'detail': e['detail'], 'status': e['status'], 'date': e['date']}
                  for e in events[:30]]
        return api_ok(result, 'Timeline loaded')
    except Exception as e:
        return api_err(str(e)[:120], 500)


@app.route('/api/mobile/profile/notifications', methods=['GET'])
@csrf.exempt
def api_mobile_profile_notifications():
    """Employee Portal: In-app notifications from EmployeeNotificationLog."""
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    data = _verify_token(token)
    if not data:
        return api_err('Unauthorized', 401)
    try:
        uid_str, _, _, _ = _resolve_employee(data)
        records = []
        logs = EmployeeNotificationLog.query.filter_by(
            employee_user_id=uid_str
        ).order_by(EmployeeNotificationLog.sent_at.desc()).limit(50).all()
        for n in logs:
            records.append({
                'id':                n.id,
                'title':             n.title,
                'body':              n.body or '',
                'notification_type': n.notification_type or '',
                'content_type':      n.content_type or '',
                'content_id':        n.content_id or '',
                'is_read':           n.is_read,
                'sent_at':           n.sent_at.isoformat()[:16] if n.sent_at else '',
            })
        return api_ok(records, 'Notifications loaded')
    except Exception as e:
        return api_ok([], 'No notifications')


@app.route('/api/mobile/profile/notifications/mark-read', methods=['POST'])
@csrf.exempt
def api_mobile_profile_notifications_mark_read():
    """Mark one or all notifications as read."""
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    data = _verify_token(token)
    if not data:
        return api_err('Unauthorized', 401)
    try:
        uid_str = str(data['user_id'])
        body = request.get_json(silent=True) or {}
        notif_id = body.get('id')
        if notif_id:
            n = EmployeeNotificationLog.query.filter_by(id=notif_id, employee_user_id=uid_str).first()
            if n:
                n.is_read = True
        else:
            EmployeeNotificationLog.query.filter_by(employee_user_id=uid_str, is_read=False).update({'is_read': True})
        db.session.commit()
        return api_ok({}, 'Marked read')
    except Exception as e:
        db.session.rollback()
        return api_err(str(e)[:120], 500)


@app.route('/api/mobile/profile/feedback', methods=['GET'])
@csrf.exempt
def api_mobile_profile_feedback():
    """Employee Portal: Reporter feedback on submitted reports."""
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    data = _verify_token(token)
    if not data:
        return api_err('Unauthorized', 401)
    try:
        uid_str, _, _, _ = _resolve_employee(data)
        records = []
        feedbacks = ReportFeedback.query.filter_by(
            reporter_user_id=uid_str
        ).order_by(ReportFeedback.submitted_at.desc()).limit(20).all()
        for f in feedbacks:
            records.append({
                'ref':     f.report_ref or '—',
                'type':    f.report_type or 'Report',
                'stage':   f.stage_num or 1,
                'status':  f.stage_label or 'Submitted',
                'message': f.outcome_summary or '',
                'outcome': f.outcome_actions_taken or '',
                'date':    f.submitted_at.isoformat()[:10] if f.submitted_at else '',
            })
        return api_ok(records, 'Feedback loaded')
    except Exception as e:
        return api_ok([], 'No feedback found')


@app.route('/api/mobile/profile/preferences', methods=['PATCH'])
@csrf.exempt
def api_mobile_profile_preferences():
    """Flutter: Update user-owned preferences only.
    Admin-managed fields (position, base_station, department, etc.) are read-only
    from the mobile app — edit them through the Web Admin Employee Management module.
    """
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    data = _verify_token(token)
    if not data:
        return api_err('Unauthorized', 401)
    try:
        import json as _json
        _, emp, _, _ = _resolve_employee(data)
        if not emp:
            return api_err('Employee record not found', 404)
        body = request.get_json(silent=True) or {}
        # Only these fields may be changed from the mobile app
        if 'language_preference' in body:
            lang = (body['language_preference'] or 'en').strip()[:10]
            emp.language_preference = lang if lang in ('en', 'ar') else 'en'
        if 'dark_mode' in body:
            emp.dark_mode = bool(body['dark_mode'])
        if 'notification_prefs' in body and isinstance(body['notification_prefs'], dict):
            emp.notification_prefs = _json.dumps(body['notification_prefs'])
        if 'privacy_settings' in body and isinstance(body['privacy_settings'], dict):
            emp.privacy_settings = _json.dumps(body['privacy_settings'])
        db.session.commit()
        return api_ok({}, 'Preferences saved')
    except Exception as e:
        db.session.rollback()
        return api_err(str(e)[:120], 500)


@app.route('/api/mobile/profile/change-password', methods=['POST'])
@csrf.exempt
def api_mobile_profile_change_password():
    """Flutter: Change own password. Requires current password for verification."""
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    data = _verify_token(token)
    if not data:
        return api_err('Unauthorized', 401)
    try:
        from werkzeug.security import check_password_hash, generate_password_hash
        _, emp, _, _ = _resolve_employee(data)
        if not emp:
            return api_err('Employee record not found', 404)
        body = request.get_json(silent=True) or {}
        current_pw = body.get('current_password', '')
        new_pw     = body.get('new_password', '')
        if not current_pw or not new_pw:
            return api_err('current_password and new_password are required', 400)
        if len(new_pw) < 8:
            return api_err('New password must be at least 8 characters', 400)
        if not check_password_hash(emp.password_hash, current_pw):
            return api_err('Current password is incorrect', 403)
        emp.password_hash = generate_password_hash(new_pw)
        emp.password_changed_at = datetime.utcnow()
        db.session.commit()
        return api_ok({}, 'Password changed successfully')
    except Exception as e:
        db.session.rollback()
        return api_err(str(e)[:120], 500)


@app.route('/api/mobile/profile/upload_image', methods=['POST'])
@csrf.exempt
def api_mobile_profile_upload_image():
    """DISABLED: Profile photos are managed exclusively by administrators via the web portal."""
    return api_err('Profile photo changes must be made by an administrator via the web portal.', 403)


# ── Cloudinary helper ────────────────────────────────────────────────────────

def _cloudinary_configured():
    return all([
        os.environ.get('CLOUDINARY_CLOUD_NAME'),
        os.environ.get('CLOUDINARY_API_KEY'),
        os.environ.get('CLOUDINARY_API_SECRET'),
    ])

def _cloudinary_upload(file_stream, public_id):
    """Upload file to Cloudinary. Returns secure_url or raises."""
    import cloudinary
    import cloudinary.uploader
    cloudinary.config(
        cloud_name=os.environ['CLOUDINARY_CLOUD_NAME'],
        api_key=os.environ['CLOUDINARY_API_KEY'],
        api_secret=os.environ['CLOUDINARY_API_SECRET'],
    )
    result = cloudinary.uploader.upload(
        file_stream,
        public_id=public_id,
        overwrite=True,
        folder='avias/profile_images',
        transformation=[{'width': 400, 'height': 400, 'crop': 'fill', 'gravity': 'face'}],
    )
    return result['secure_url']

def _cloudinary_delete(secure_url):
    """Delete image from Cloudinary by its secure_url."""
    try:
        import cloudinary
        import cloudinary.uploader
        cloudinary.config(
            cloud_name=os.environ['CLOUDINARY_CLOUD_NAME'],
            api_key=os.environ['CLOUDINARY_API_KEY'],
            api_secret=os.environ['CLOUDINARY_API_SECRET'],
        )
        # Extract public_id: everything between /upload/v{ver}/ and the extension
        import re
        m = re.search(r'/upload/(?:v\d+/)?(.+)\.[a-z]+$', secure_url)
        if m:
            cloudinary.uploader.destroy(m.group(1))
    except Exception as e:
        app.logger.warning(f'Cloudinary delete failed: {e}')

def _profile_image_url(emp):
    """Return the correct profile image URL regardless of storage backend."""
    # Prefer the dedicated Cloudinary column (persistent across deploys)
    if getattr(emp, 'profile_photo_url', None):
        return emp.profile_photo_url
    if not emp.profile_image:
        return None
    v = emp.profile_image
    if v.startswith('https://') or v.startswith('http://'):
        return v
    # Legacy: filename saved on local disk
    return f'/static/profile_images/{v}'


# ── Admin: upload employee profile photo ────────────────────────────────────
@app.route('/api/admin/employees/<int:emp_id>/photo', methods=['POST'])
@csrf.exempt
@require_login
def api_admin_employee_photo_upload(emp_id):
    """Upload or replace employee profile photo.
    Uses Cloudinary when configured (persistent across deploys).
    Falls back to local static/ when env vars are absent (dev only).
    """
    import uuid, os
    emp = Employee.query.get_or_404(emp_id)
    if 'photo' not in request.files:
        return api_err('No photo file provided', 400)
    file = request.files['photo']
    if not file.filename:
        return api_err('Empty filename', 400)
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in {'.jpg', '.jpeg', '.png', '.webp'}:
        return api_err('Unsupported format. Use JPG, PNG or WEBP.', 400)

    if _cloudinary_configured():
        # Delete old Cloudinary image
        old_url = getattr(emp, 'profile_photo_url', None) or (
            emp.profile_image if emp.profile_image and emp.profile_image.startswith('https://') else None
        )
        if old_url:
            _cloudinary_delete(old_url)
        public_id = f'{emp.employee_id}_{uuid.uuid4().hex[:8]}'
        try:
            url = _cloudinary_upload(file.stream, public_id)
        except Exception as e:
            app.logger.error(f'Cloudinary upload error: {e}')
            return api_err('Image upload failed. Try again.', 500)
        emp.profile_photo_url = url  # store in dedicated column
        db.session.commit()
        return api_ok({'profile_image_url': url}, 'Photo updated')
    else:
        # Local fallback (dev / no Cloudinary configured)
        save_dir = os.path.join(app.static_folder, 'profile_images')
        os.makedirs(save_dir, exist_ok=True)
        if emp.profile_image and not emp.profile_image.startswith('http'):
            old_path = os.path.join(save_dir, emp.profile_image)
            if os.path.exists(old_path):
                os.remove(old_path)
        filename = f'{emp.employee_id}_{uuid.uuid4().hex[:8]}{ext}'
        file.save(os.path.join(save_dir, filename))
        emp.profile_image = filename
        db.session.commit()
        url = f'/static/profile_images/{filename}'
        return api_ok({'profile_image_url': url}, 'Photo updated')


# ── Admin: delete employee profile photo ────────────────────────────────────
@app.route('/api/admin/employees/<int:emp_id>/photo', methods=['DELETE'])
@csrf.exempt
@require_login
def api_admin_employee_photo_delete(emp_id):
    """Remove employee profile photo from Cloudinary or local storage."""
    import os
    emp = Employee.query.get_or_404(emp_id)
    cloudinary_url = getattr(emp, 'profile_photo_url', None) or (
        emp.profile_image if emp.profile_image and emp.profile_image.startswith('https://') else None
    )
    if not cloudinary_url and not emp.profile_image:
        return api_err('No photo to delete', 400)
    if cloudinary_url:
        _cloudinary_delete(cloudinary_url)
    elif emp.profile_image:
        save_dir = os.path.join(app.static_folder, 'profile_images')
        old_path = os.path.join(save_dir, emp.profile_image)
        if os.path.exists(old_path):
            os.remove(old_path)
    emp.profile_photo_url = None
    emp.profile_image = None
    db.session.commit()
    return api_ok({}, 'Photo removed')


# ═══════════════════════════════════════════════════════════════════════════════
#  WEB ADMIN — EMPLOYEE MANAGEMENT MODULE
#  All admin routes require a logged-in web session (require_login).
#  Flutter app has NO access to these routes.
# ═══════════════════════════════════════════════════════════════════════════════

@app.route('/employee-management')
@require_login
def employee_management():
    """Web Admin: Employee Management landing page."""
    departments = Department.query.order_by(Department.name).all()
    employees   = Employee.query.order_by(Employee.full_name).all()
    return render_template('employee_management.html',
                           employees=employees, departments=departments,
                           page_title='Employee Management')


@app.route('/employee-management/<int:emp_id>', methods=['GET'])
@require_login
def employee_detail(emp_id):
    """Web Admin: Employee detail/edit page."""
    from sqlalchemy import or_
    emp         = Employee.query.get_or_404(emp_id)
    departments = Department.query.order_by(Department.name).all()
    uid_str     = f'emp_{emp.id}'
    emp_name    = emp.full_name

    training_records     = Training.query.filter_by(employee_name=emp_name).order_by(Training.created_at.desc()).all()
    all_actions          = Action.query.filter(Action.owner == emp_name).all()
    total_reports        = (HazardReport.query.filter(or_(
                                HazardReport.reporter == emp_name,
                                HazardReport.reporter_user_id == uid_str)).count() +
                            ASRReport.query.filter(or_(
                                ASRReport.captain == emp_name,
                                ASRReport.copilot == emp_name)).count() +
                            VoluntaryReport.query.filter(
                                VoluntaryReport.reporter_name == emp_name).count())
    reads_count          = SafetyPromoRead.query.filter_by(user_id=uid_str).count()
    completed_trainings  = sum(1 for t in training_records if (t.status or '') == 'Completed')
    closed_actions       = sum(1 for a in all_actions if (a.status or '') == 'Closed')

    report_score    = min(30, total_reports * 5)
    training_score  = min(30, completed_trainings * 10)
    action_score    = min(25, closed_actions * 8)
    engagement_score = min(15, reads_count * 3)
    total_score     = report_score + training_score + action_score + engagement_score

    stats = {
        'total_reports':       total_reports,
        'completed_trainings': completed_trainings,
        'closed_actions':      closed_actions,
        'promos_read':         reads_count,
        'report_score':        report_score,
        'training_score':      training_score,
        'action_score':        action_score,
        'engagement_score':    engagement_score,
        'total_score':         total_score,
    }

    return render_template('employee_detail.html',
                           emp=emp, departments=departments,
                           training_records=training_records,
                           stats=stats,
                           emp_photo_url=_profile_image_url(emp),
                           page_title=f'Employee — {emp.full_name}')


@app.route('/api/admin/employees', methods=['GET'])
@require_login
def api_admin_employees_list():
    """Web Admin API: List employees with optional search."""
    q    = request.args.get('q', '').strip()
    dept = request.args.get('dept', '')
    page = int(request.args.get('page', 1))
    per  = min(int(request.args.get('per', 50)), 200)

    query = Employee.query
    if q:
        like = f'%{q}%'
        query = query.filter(
            db.or_(Employee.full_name.ilike(like),
                   Employee.employee_id.ilike(like),
                   Employee.email.ilike(like),
                   Employee.username.ilike(like))
        )
    if dept:
        query = query.filter(Employee.department_id == int(dept))

    total = query.count()
    emps  = query.order_by(Employee.full_name).offset((page - 1) * per).limit(per).all()

    return api_ok({
        'total': total, 'page': page, 'per': per,
        'employees': [{
            'id':                e.id,
            'employee_id':       e.employee_id,
            'full_name':         e.full_name,
            'username':          e.username,
            'email':             e.email or '',
            'mobile':            e.mobile or '',
            'role':              e.role or 'employee',
            'department_id':     e.department_id,
            'position':          e.position or '',
            'base_station':      e.base_station or '',
            'employment_status': e.employment_status or 'Active',
            'is_active':         e.is_active,
            'last_login':        e.last_login.isoformat() if e.last_login else '',
        } for e in emps]
    }, 'Employees loaded')


@app.route('/api/admin/employees', methods=['POST'])
@require_login
@csrf.exempt
def api_admin_employees_create():
    """Web Admin API: Create a new employee account."""
    from werkzeug.security import generate_password_hash
    body = request.get_json(silent=True) or {}
    required = ['employee_id', 'username', 'full_name', 'password']
    for field in required:
        if not body.get(field, '').strip():
            return api_err(f'{field} is required', 400)

    if Employee.query.filter_by(username=body['username'].strip()).first():
        return api_err('Username already exists', 409)
    if Employee.query.filter_by(employee_id=body['employee_id'].strip()).first():
        return api_err('Employee ID already exists', 409)

    try:
        emp = Employee(
            employee_id       = body['employee_id'].strip(),
            username          = body['username'].strip().lower(),
            password_hash     = generate_password_hash(body['password']),
            full_name         = body['full_name'].strip(),
            email             = body.get('email', '').strip(),
            mobile            = body.get('mobile', '').strip(),
            department_id     = body.get('department_id') or None,
            role              = body.get('role', 'employee').strip(),
            position          = body.get('position', '').strip(),
            base_station      = body.get('base_station', 'AMM').strip().upper(),
            join_date         = body.get('join_date', '').strip(),
            employment_status = body.get('employment_status', 'Active').strip(),
            is_active         = True,
        )
        db.session.add(emp)
        db.session.commit()
        return api_ok({'id': emp.id, 'employee_id': emp.employee_id}, 'Employee created', 201)
    except Exception as e:
        db.session.rollback()
        return api_err(str(e)[:120], 500)


@app.route('/api/admin/employees/<int:emp_id>', methods=['GET'])
@require_login
def api_admin_employees_get(emp_id):
    """Web Admin API: Get full employee detail."""
    import json as _json
    emp = Employee.query.get_or_404(emp_id)
    dept_name = emp.department.name if emp.department else ''
    try:
        notif_prefs   = _json.loads(emp.notification_prefs or '{}')
        privacy_prefs = _json.loads(emp.privacy_settings or '{}')
    except Exception:
        notif_prefs = privacy_prefs = {}
    return api_ok({
        'id':                 emp.id,
        'employee_id':        emp.employee_id,
        'username':           emp.username,
        'full_name':          emp.full_name,
        'email':              emp.email or '',
        'mobile':             emp.mobile or '',
        'role':               emp.role or 'employee',
        'department_id':      emp.department_id,
        'department':         dept_name,
        'position':           emp.position or '',
        'base_station':       emp.base_station or 'AMM',
        'join_date':          emp.join_date or '',
        'employment_status':  emp.employment_status or 'Active',
        'is_active':          emp.is_active,
        'profile_image_url':  _profile_image_url(emp),
        'created_at':         emp.created_at.isoformat() if emp.created_at else '',
        'last_login':         emp.last_login.isoformat() if emp.last_login else '',
        'language_preference': emp.language_preference or 'en',
        'dark_mode':           emp.dark_mode or False,
        'notification_prefs':  notif_prefs,
        'privacy_settings':    privacy_prefs,
        'password_changed_at': emp.password_changed_at.isoformat() if emp.password_changed_at else '',
    }, 'Employee loaded')


@app.route('/api/admin/employees/<int:emp_id>', methods=['DELETE'])
@require_login
@csrf.exempt
def api_admin_employees_delete(emp_id):
    """Web Admin API: Delete an employee account."""
    emp = Employee.query.get_or_404(emp_id)
    try:
        db.session.delete(emp)
        db.session.commit()
        return api_ok({}, 'Employee deleted')
    except Exception as e:
        db.session.rollback()
        return api_err(str(e)[:120], 500)


@app.route('/api/admin/employees/<int:emp_id>', methods=['PUT'])
@require_login
@csrf.exempt
def api_admin_employees_update(emp_id):
    """Web Admin API: Update admin-managed employee fields."""
    emp  = Employee.query.get_or_404(emp_id)
    body = request.get_json(silent=True) or {}
    admin_fields = {
        'full_name', 'email', 'mobile', 'role',
        'department_id', 'position', 'base_station',
        'join_date', 'employment_status',
    }
    try:
        for field in admin_fields:
            if field in body:
                val = body[field]
                if isinstance(val, str):
                    val = val.strip() or None
                setattr(emp, field, val)
        if 'employee_id' in body:
            new_eid = body['employee_id'].strip()
            if new_eid and new_eid != emp.employee_id:
                if Employee.query.filter(Employee.employee_id == new_eid, Employee.id != emp_id).first():
                    return api_err('Employee ID already taken', 409)
                emp.employee_id = new_eid
        db.session.commit()
        return api_ok({'id': emp.id}, 'Employee updated')
    except Exception as e:
        db.session.rollback()
        return api_err(str(e)[:120], 500)


@app.route('/api/admin/employees/<int:emp_id>/status', methods=['PATCH'])
@require_login
@csrf.exempt
def api_admin_employees_status(emp_id):
    """Web Admin API: Activate or deactivate an employee account."""
    emp  = Employee.query.get_or_404(emp_id)
    body = request.get_json(silent=True) or {}
    if 'is_active' not in body:
        return api_err('is_active field required', 400)
    try:
        emp.is_active         = bool(body['is_active'])
        emp.employment_status = 'Active' if emp.is_active else 'Inactive'
        db.session.commit()
        return api_ok({'is_active': emp.is_active}, 'Status updated')
    except Exception as e:
        db.session.rollback()
        return api_err(str(e)[:120], 500)


@app.route('/api/admin/employees/<int:emp_id>/reset-password', methods=['POST'])
@require_login
@csrf.exempt
def api_admin_employees_reset_password(emp_id):
    """Web Admin API: Reset employee password (no current password needed)."""
    from werkzeug.security import generate_password_hash
    emp  = Employee.query.get_or_404(emp_id)
    body = request.get_json(silent=True) or {}
    new_pw = (body.get('password') or body.get('new_password') or '').strip()
    if len(new_pw) < 8:
        return api_err('Password must be at least 8 characters', 400)
    try:
        emp.password_hash       = generate_password_hash(new_pw)
        emp.password_changed_at = datetime.utcnow()
        db.session.commit()
        return api_ok({}, 'Password reset successfully')
    except Exception as e:
        db.session.rollback()
        return api_err(str(e)[:120], 500)


@app.route('/api/admin/employees/<int:emp_id>/training', methods=['GET'])
@require_login
@csrf.exempt
def api_admin_employees_training_list(emp_id):
    """Web Admin API: Get training records for an employee."""
    emp      = Employee.query.get_or_404(emp_id)
    records  = Training.query.filter_by(employee_name=emp.full_name).order_by(Training.created_at.desc()).all()
    return api_ok([{
        'id':            t.id,
        'training_name': t.training_program or t.training_type or '',
        'category':      t.training_type or '',
        'status':        t.status or 'Completed',
        'training_date': t.training_date or '',
        'expiry_date':   t.expiry_date or '',
        'scheduled_date': t.scheduled_date or '',
        'instructor':    t.instructor or '',
    } for t in records], 'Training records loaded')


@app.route('/api/admin/employees/<int:emp_id>/training', methods=['POST'])
@require_login
@csrf.exempt
def api_admin_employees_training_create(emp_id):
    """Web Admin API: Add a training record for an employee."""
    emp  = Employee.query.get_or_404(emp_id)
    body = request.get_json(silent=True) or {}
    if not body.get('training_program', '').strip():
        return api_err('training_program is required', 400)
    try:
        t = Training(
            employee_name    = emp.full_name,
            training_program = body.get('training_program', '').strip(),
            training_type    = body.get('training_type', '').strip(),
            status           = body.get('status', 'Completed').strip(),
            training_date    = body.get('training_date', '').strip(),
            expiry_date      = body.get('expiry_date', '').strip(),
            scheduled_date   = body.get('scheduled_date', '').strip(),
            instructor       = body.get('instructor', '').strip(),
            department_id    = emp.department_id,
        )
        db.session.add(t)
        db.session.commit()
        push_notify_user(
            f'emp_{emp.id}',
            f'📚 Training Assigned: {t.training_program}',
            f'Scheduled for {t.scheduled_date or t.training_date or "TBD"}. Please review your training plan.',
            'action_assigned', 'training', t.id)
        return api_ok({'id': t.id}, 'Training record added', 201)
    except Exception as e:
        db.session.rollback()
        return api_err(str(e)[:120], 500)


@app.route('/api/admin/employees/training/<int:tid>', methods=['PUT'])
@require_login
@csrf.exempt
def api_admin_employees_training_update(tid):
    """Web Admin API: Update a training record."""
    t    = Training.query.get_or_404(tid)
    body = request.get_json(silent=True) or {}
    editable = {
        'training_program': 'training_program',
        'training_type':    'training_type',
        'status':           'status',
        'training_date':    'training_date',
        'expiry_date':      'expiry_date',
        'scheduled_date':   'scheduled_date',
        'instructor':       'instructor',
    }
    try:
        for key, attr in editable.items():
            if key in body:
                setattr(t, attr, (body[key] or '').strip())
        db.session.commit()
        return api_ok({'id': t.id}, 'Training record updated')
    except Exception as e:
        db.session.rollback()
        return api_err(str(e)[:120], 500)


@app.route('/api/admin/employees/training/<int:tid>', methods=['DELETE'])
@require_login
@csrf.exempt
def api_admin_employees_training_delete(tid):
    """Web Admin API: Delete a training record."""
    t = Training.query.get_or_404(tid)
    try:
        db.session.delete(t)
        db.session.commit()
        return api_ok({}, 'Training record deleted')
    except Exception as e:
        db.session.rollback()
        return api_err(str(e)[:120], 500)


@app.route('/api/logout', methods=['POST', 'OPTIONS'])
@csrf.exempt
def api_logout():
    """Flutter: Invalidate Bearer token — deletes from persistent token store.

    Previously decorated with @require_login (web session guard) which caused
    Flutter requests (Bearer token, no session cookie) to receive a 302 redirect
    instead of JSON, leaving the server-side token permanently active.

    Now: token-based validation. Logout is idempotent — returns 200 even if the
    token is already expired or invalid, so the client always clears its session.
    """
    if request.method == 'OPTIONS':
        return api_ok()
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    if token:
        try:
            ApiToken.query.filter_by(token=token).delete(synchronize_session=False)
            db.session.commit()
        except Exception:
            db.session.rollback()
    return api_ok({}, 'Logged out successfully')


# /api/setup-check removed — was a diagnostic endpoint exposing employee list and DB schema


# /api/debug-login removed — was a diagnostic endpoint exposing auth internals


@app.route('/api/my_reports', methods=['GET'])
def api_my_reports():
    """Flutter: Get reports submitted by the logged-in employee."""
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    data = _verify_token(token)
    if not data:
        return api_err('Unauthorized', 401)
    try:
        uid = data['user_id']
        # Resolve reporter name from employee or user
        uname = ''
        try:
            if str(uid).startswith('emp_'):
                emp = Employee.query.get(int(str(uid).replace('emp_','')))
                uname = emp.full_name if emp else ''
            else:
                user = User.query.get(int(str(uid).replace('usr_','')))
                uname = (user.full_name or user.username) if user else ''
        except Exception: pass

        records = []
        # All HazardReports (includes ASR, Voluntary, Confidential linked ones)
        q = HazardReport.query
        if uname:
            q = q.filter(HazardReport.reporter.ilike(f'%{uname}%'))
        for r in q.order_by(HazardReport.created_at.desc()).limit(50).all():
            try:
                wf_status, wf_color, wf_stage, timeline, wf_guidance, wf_responsible, wf_next = resolve_report_status(
                    hazard_id=r.hazard_id, hr_status=r.status)
                records.append({
                    'id':          r.id,
                    'type':        r.report_type or 'Hazard Report',
                    'title':       r.generic_hazard or r.description[:60] if r.description else r.id,
                    'description': (r.description or '')[:100],
                    'status':      wf_status,
                    'status_color':wf_color,
                    'stage':       wf_stage,
                    'guidance':    wf_guidance,
                    'responsible': wf_responsible,
                    'next_step':   wf_next,
                    'date':        r.date or '',
                    'location':    r.location or '',
                    'severity':    r.reporter_severity or 'Medium',
                    'hazard_id':   r.hazard_id or '',
                    'created_at':  r.created_at.isoformat() if r.created_at else '',
                    'timeline':    timeline[:8],
                })
            except Exception: pass

        # ASR by this user (not already in HazardReport)
        existing_ids = {r['hazard_id'] for r in records if r.get('hazard_id')}
        asr_q = ASRReport.query
        if uname:
            asr_q = asr_q.filter(ASRReport.captain.ilike(f'%{uname}%'))
        for r in asr_q.order_by(ASRReport.created_at.desc()).limit(20).all():
            try:
                if r.hazard_id and r.hazard_id in existing_ids:
                    continue  # Already included via HazardReport
                wf_status, wf_color, wf_stage, timeline, wf_guidance, wf_responsible, wf_next = resolve_report_status(
                    hazard_id=r.hazard_id, hr_status='Submitted')
                records.append({
                    'id':          r.id,
                    'type':        'ASR',
                    'title':       r.occurrence_type or 'Air Safety Report',
                    'description': f"Flight {r.flight_no or '—'}: {(r.event_description or '')[:80]}",
                    'status':      wf_status,
                    'status_color':wf_color,
                    'stage':       wf_stage,
                    'date':        r.date or '',
                    'location':    f'{r.route_from or ""}→{r.route_to or ""}',
                    'severity':    r.severity or 'C',
                    'hazard_id':   r.hazard_id or '',
                    'created_at':  r.created_at.isoformat() if r.created_at else '',
                    'timeline':    timeline[:5],
                })
            except Exception: pass

        records.sort(key=lambda x: x.get('created_at',''), reverse=True)
        return api_ok({'reports': records, 'total': len(records)}, 'My reports loaded')
    except Exception as e:
        return api_err(str(e)[:120], 500)


@app.route('/api/mobile/history', methods=['GET', 'OPTIONS'])
@csrf.exempt
def api_mobile_history():
    if request.method == 'OPTIONS':
        return api_ok({}, 'ok')
    """Flutter: Get reports for the authenticated employee only."""
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    identity = _get_identity(token)
    if not identity:
        return api_err('Unauthorized — please login', 401)

    try:
        limit = int(request.args.get('limit', 20))
        records = []
        reporter_name = identity['name']
        is_admin      = identity['is_admin']

        # ── 1. HazardReports (includes ASR-linked HazardReports) ──────────────
        q = HazardReport.query
        if not is_admin and reporter_name:
            q = q.filter(HazardReport.reporter.ilike(f'%{reporter_name}%'))
        for r in q.order_by(HazardReport.created_at.desc()).limit(limit * 3).all():
            try:
                wf_status, wf_color, wf_stage, timeline, wf_guidance, wf_responsible, wf_next = \
                    resolve_report_status(hazard_id=r.hazard_id, hr_status=r.status)
                # Phase 2: enrich with reporter feedback outcome if available
                fb_outcome = ''
                fb_actions = ''
                fb_risk    = ''
                if _ENFORCEMENT_MODELS and ReportFeedback:
                    try:
                        fb = ReportFeedback.query.filter_by(report_ref=r.id).first()
                        if fb and fb.outcome_shared:
                            fb_outcome = fb.outcome_summary or ''
                            fb_actions = fb.outcome_actions_taken or ''
                            fb_risk    = fb.outcome_risk_level or ''
                    except Exception:
                        pass
                records.append({
                    'id':                  r.id,
                    'type':                r.report_type or 'Hazard Report',
                    'title':               r.generic_hazard or (r.description or '')[:60],
                    'description':         (r.description or '')[:80],
                    'status':              wf_status,
                    'status_color':        wf_color,
                    'stage':               wf_stage,
                    'guidance':            wf_guidance,
                    'responsible':         wf_responsible,
                    'next_step':           wf_next,
                    'date':                r.date or '',
                    'location':            r.location or '',
                    'severity':            r.reporter_severity or 'Medium',
                    'created_at':          r.created_at.isoformat() if r.created_at else '',
                    'timeline':            timeline[:6],
                    'outcome_summary':     fb_outcome,
                    'outcome_actions_taken': fb_actions,
                    'outcome_risk_level':  fb_risk,
                })
            except Exception:
                pass

        # ── 2. Voluntary Reports ───────────────────────────────────────────────
        vq = VoluntaryReport.query
        if not is_admin and reporter_name:
            vq = vq.filter(VoluntaryReport.reporter_name.ilike(f'%{reporter_name}%'))
        for v in vq.order_by(VoluntaryReport.created_at.desc()).limit(limit).all():
            try:
                records.append({
                    'id':          v.ref_number or str(v.id),
                    'type':        'Voluntary',
                    'title':       (v.description or '')[:60],
                    'description': (v.description or '')[:80],
                    'status':      v.status or 'Submitted',
                    'status_color':'#22c55e',
                    'stage':       2 if (v.status or '') == 'Under Review' else (7 if (v.status or '') == 'Closed' else 1),
                    'guidance':    'Your voluntary report has been received and is under safety review.',
                    'responsible': 'Safety Manager',
                    'next_step':   'Await safety team assessment',
                    'date':        str(v.date) if v.date else '',
                    'location':    v.location or '',
                    'severity':    'Medium',
                    'created_at':  v.created_at.isoformat() if v.created_at else '',
                    'timeline':    [{'icon': '📝', 'event': 'Voluntary report submitted', 'date': str(v.date) if v.date else ''}],
                })
            except Exception:
                pass

        # ── 3. Confidential Reports ────────────────────────────────────────────
        if is_admin:
            # Only admins can see confidential reports in history
            for c in ConfidentialReport.query.order_by(ConfidentialReport.created_at.desc()).limit(limit).all():
                try:
                    records.append({
                        'id':          c.ref_number or str(c.id),
                        'type':        'Confidential',
                        'title':       'Confidential Safety Report',
                        'description': (c.description or '')[:80],
                        'status':      c.status or 'Submitted',
                        'status_color':'#7c3aed',
                        'stage':       2 if (c.status or '') == 'Under Review' else (7 if (c.status or '') == 'Closed' else 1),
                        'guidance':    'Confidential report received. Identity protected per SMS policy.',
                        'responsible': 'Accountable Manager',
                        'next_step':   'Confidential safety review',
                        'date':        str(c.date) if c.date else '',
                        'location':    c.location or '',
                        'severity':    'Medium',
                        'created_at':  c.created_at.isoformat() if c.created_at else '',
                        'timeline':    [{'icon': '🔒', 'event': 'Confidential report received', 'date': str(c.date) if c.date else ''}],
                    })
                except Exception:
                    pass

        records.sort(key=lambda x: x.get('created_at', ''), reverse=True)
        return api_ok({'reports': records[:limit], 'total': len(records)}, 'History loaded')
    except Exception as e:
        return api_err(str(e)[:120], 500)


@app.route('/api/mobile/stats', methods=['GET', 'OPTIONS'])
@csrf.exempt
def api_mobile_stats():
    if request.method == 'OPTIONS':
        return api_ok({}, 'ok')
    """Flutter: Dashboard stats — filtered by employee or full for admins."""
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    identity = _get_identity(token)
    if not identity:
        return api_err('Unauthorized', 401)

    try:
        is_admin      = identity['is_admin']
        reporter_name = identity['name']

        if is_admin:
            # Admins see full system stats
            return api_ok({
                'hazards_open':  Hazard.query.filter_by(status='Open').count(),
                'asr_total':     ASRReport.query.count(),
                'actions_open':  Action.query.filter(Action.status.notin_(['Closed'])).count(),
                'hr_this_month': HazardReport.query.count(),
            }, 'Stats loaded')
        else:
            # Employees see only their own counts
            my_hr = HazardReport.query.filter(
                HazardReport.reporter.ilike(f'%{reporter_name}%')).count() if reporter_name else 0
            my_asr = ASRReport.query.filter(
                ASRReport.captain.ilike(f'%{reporter_name}%')).count() if reporter_name else 0
            my_open = HazardReport.query.filter(
                HazardReport.reporter.ilike(f'%{reporter_name}%'),
                HazardReport.status.notin_(['Closed', 'Rejected'])).count() if reporter_name else 0
            my_closed = HazardReport.query.filter(
                HazardReport.reporter.ilike(f'%{reporter_name}%'),
                HazardReport.status == 'Closed').count() if reporter_name else 0
            # Real open action count for this employee (FL-010 fix)
            my_actions = Action.query.filter(
                Action.owner.ilike(f'%{reporter_name}%'),
                Action.status.notin_(['Closed', 'Completed', 'Cancelled'])
            ).count() if reporter_name else 0
            return api_ok({
                'hazards_open':  my_open,
                'asr_total':     my_asr,
                'actions_open':  my_actions,
                'hr_this_month': my_hr,
                'my_closed':     my_closed,
            }, 'Stats loaded')
    except Exception as e:
        return api_err(str(e)[:120], 500)


# ─────────────────────────────────────────────────────────────────────────────
#  SAFETY PROMOTION — Mobile API  (Phase 5)
# ─────────────────────────────────────────────────────────────────────────────

def _sp_user_read_set(user_id):
    """Return a set of (content_type, content_id) already read by this user."""
    try:
        rows = SafetyPromoRead.query.filter_by(user_id=str(user_id)).all()
        return {(r.content_type, r.content_id) for r in rows}
    except Exception:
        db.session.rollback()
        return set()


def _sp_user_ack_set(user_id):
    """Return a set of (content_type, content_id) already acked by this user."""
    try:
        rows = SafetyPromoAck.query.filter_by(user_id=str(user_id)).all()
        return {(r.content_type, r.content_id) for r in rows}
    except Exception:
        db.session.rollback()
        return set()


def _sp_item(obj, ctype, cid_attr, title_attr, date_attr, dept_attr=None,
             summary_attr=None, severity_attr=None, priority_attr=None,
             mandatory_attr='is_mandatory', read_set=None, ack_set=None,
             content_attr=None, author_attr=None, ref_attr=None):
    """Serialize a safety-promo model instance into a feed dict."""
    cid = str(getattr(obj, cid_attr, ''))
    is_read = (ctype, cid) in read_set if read_set is not None else False
    is_acked = (ctype, cid) in ack_set if ack_set is not None else False
    return {
        'type':        ctype,
        'id':          cid,
        'title':       getattr(obj, title_attr, '') or '',
        'summary':     getattr(obj, summary_attr, '') or '' if summary_attr else '',
        'content':     getattr(obj, content_attr, '') or '' if content_attr else '',
        'author':      getattr(obj, author_attr, '') or '' if author_attr else '',
        'ref_number':  getattr(obj, ref_attr, '') or '' if ref_attr else '',
        'date':        str(getattr(obj, date_attr, '') or ''),
        'dept_id':     getattr(obj, dept_attr, None) if dept_attr else None,
        'severity':    getattr(obj, severity_attr, '') or '' if severity_attr else '',
        'priority':    getattr(obj, priority_attr, 'Normal') or 'Normal' if priority_attr else 'Normal',
        'is_mandatory': bool(getattr(obj, mandatory_attr, False)),
        'is_read':     is_read,
        'is_acked':    is_acked,
        'has_attachment': bool(getattr(obj, 'attachment', None)),
    }


@app.route('/api/mobile/safety/ping', methods=['GET'])
@csrf.exempt
def api_mobile_safety_ping():
    """Diagnostic: confirm new code is deployed and show survey counts (no auth needed)."""
    try:
        surveys     = SafetySurvey.query.filter_by(status='Active').count()
        bulletins   = SafetyBulletin.query.filter(SafetyBulletin.status == 'Active').count()
        newsletters = SafetyNewsletter.query.filter_by(status='Published').count()
        lessons     = LessonLearned.query.filter_by(status='Published').count()
        try:
            reads = SafetyPromoRead.query.count()
            table_ok = True
        except Exception:
            reads = -1
            table_ok = False
        # Try to fetch first active survey and its questions
        survey_error = None
        survey_sample = None
        try:
            s = SafetySurvey.query.filter_by(status='Active').first()
            if s:
                survey_sample = {
                    'id': s.id,
                    'title': s.title,
                    'start_date': str(s.start_date or ''),
                    'has_questions': bool(s.questions),
                }
                try:
                    _ = SurveyResponse.query.filter_by(survey_id=s.id).count()
                except Exception as se:
                    survey_error = str(se)[:200]
        except Exception as se2:
            survey_error = str(se2)[:200]
        return api_ok({
            'version':      'phase5-v3',
            'surveys_active': surveys,
            'bulletins_active': bulletins,
            'newsletters_published': newsletters,
            'lessons_published': lessons,
            'promo_reads_table': table_ok,
            'promo_reads_count': reads,
            'survey_sample': survey_sample,
            'survey_error': survey_error,
        }, 'pong')
    except Exception as e:
        return api_err(str(e)[:300], 500)


@app.route('/api/mobile/safety/feed', methods=['GET', 'OPTIONS'])
@csrf.exempt
def api_mobile_safety_feed():
    """Flutter: unified chronological safety promotion feed.

    Query params:
      type   — alert | bulletin | newsletter | survey | lesson  (omit = all)
      limit  — default 50
      offset — default 0
    """
    if request.method == 'OPTIONS':
        return api_ok({}, 'ok')
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    identity = _get_identity(token)
    if not identity:
        return api_err('Unauthorized', 401)

    ftype  = request.args.get('type', '').lower()
    limit  = min(int(request.args.get('limit', 50)), 200)
    offset = int(request.args.get('offset', 0))
    uid    = str(identity.get('uid') or identity.get('user_id') or identity.get('id', ''))
    dept   = request.headers.get('X-Dept-Id')

    try:
        read_set = _sp_user_read_set(uid)
        ack_set  = _sp_user_ack_set(uid)
        items = []

        # Safety Bulletins (Alerts)
        if not ftype or ftype in ('alert', 'bulletin'):
            q = SafetyBulletin.query.filter(SafetyBulletin.status == 'Active')
            btype = 'alert' if ftype == 'alert' else None
            rows = q.all()
            for b in rows:
                ct = 'alert' if (b.bulletin_type or '').lower() in ('safety alert', 'alert') else 'bulletin'
                if ftype and ftype != ct:
                    continue
                items.append(_sp_item(b, ct, 'id', 'title', 'issue_date',
                                      dept_attr='department_id',
                                      summary_attr='content',
                                      severity_attr='severity',
                                      priority_attr='priority_level',
                                      read_set=read_set, ack_set=ack_set))

        # Newsletters
        if not ftype or ftype == 'newsletter':
            rows = SafetyNewsletter.query.filter(SafetyNewsletter.status == 'Published').all()
            for n in rows:
                items.append(_sp_item(n, 'newsletter', 'id', 'title', 'issue_date',
                                      dept_attr='department_id',
                                      summary_attr='summary',
                                      content_attr='content',
                                      author_attr='author',
                                      ref_attr='ref_number',
                                      priority_attr='priority_level',
                                      read_set=read_set, ack_set=ack_set))

        # Surveys
        if not ftype or ftype == 'survey':
            rows = SafetySurvey.query.filter(SafetySurvey.status == 'Active').all()
            for s in rows:
                item = _sp_item(s, 'survey', 'id', 'title', 'start_date',
                                dept_attr='department_id',
                                summary_attr='description',
                                priority_attr='priority_level',
                                read_set=read_set, ack_set=ack_set)
                # Include questions JSON so Flutter can render the form
                try:
                    item['questions'] = json.loads(s.questions) if s.questions else []
                except Exception:
                    item['questions'] = []
                item['end_date'] = str(s.end_date or '')
                # Check if this user already responded
                try:
                    already = SurveyResponse.query.filter_by(
                        survey_id=s.id).filter(
                        SurveyResponse.respondent_email == uid
                    ).first() if uid else None
                except Exception:
                    db.session.rollback()
                    already = None
                item['already_responded'] = bool(already)
                items.append(item)

        # Lessons Learned
        if not ftype or ftype == 'lesson':
            rows = LessonLearned.query.filter(LessonLearned.status == 'Published').all()
            for l in rows:
                items.append(_sp_item(l, 'lesson', 'id', 'title', 'date',
                                      dept_attr='department_id',
                                      summary_attr='lesson',
                                      priority_attr='priority_level',
                                      read_set=read_set, ack_set=ack_set))

        # Sort by date descending (ISO strings compare fine)
        items.sort(key=lambda x: x['date'] or '', reverse=True)

        total = len(items)
        page  = items[offset:offset + limit]
        return api_ok({'items': page, 'total': total, 'offset': offset, 'limit': limit},
                      'Feed loaded')
    except Exception as e:
        return api_err(str(e)[:200], 500)


@app.route('/api/mobile/safety/unread_count', methods=['GET', 'OPTIONS'])
@csrf.exempt
def api_mobile_safety_unread_count():
    """Flutter: per-type unread counts."""
    if request.method == 'OPTIONS':
        return api_ok({}, 'ok')
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    identity = _get_identity(token)
    if not identity:
        return api_err('Unauthorized', 401)

    uid = str(identity.get('uid') or identity.get('user_id') or identity.get('id', ''))
    try:
        read_set = _sp_user_read_set(uid)

        def _unread(qs, ctype, id_attr='id'):
            return sum(1 for r in qs if (ctype, str(getattr(r, id_attr, ''))) not in read_set)

        alerts      = SafetyBulletin.query.filter(
            SafetyBulletin.status == 'Active',
            SafetyBulletin.bulletin_type.ilike('%alert%')).all()
        bulletins   = SafetyBulletin.query.filter(
            SafetyBulletin.status == 'Active',
            ~SafetyBulletin.bulletin_type.ilike('%alert%')).all()
        newsletters = SafetyNewsletter.query.filter_by(status='Published').all()
        surveys     = SafetySurvey.query.filter_by(status='Active').all()
        lessons     = LessonLearned.query.filter_by(status='Published').all()

        counts = {
            'alert':      _unread(alerts,      'alert'),
            'bulletin':   _unread(bulletins,   'bulletin'),
            'newsletter': _unread(newsletters, 'newsletter'),
            'survey':     _unread(surveys,     'survey'),
            'lesson':     _unread(lessons,     'lesson'),
        }
        counts['total'] = sum(counts.values())
        return api_ok(counts, 'Counts loaded')
    except Exception as e:
        return api_err(str(e)[:200], 500)


@app.route('/api/mobile/safety/read', methods=['POST', 'OPTIONS'])
@csrf.exempt
def api_mobile_safety_read():
    """Flutter: mark an item as read. Body: {type, id}"""
    if request.method == 'OPTIONS':
        return api_ok({}, 'ok')
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    identity = _get_identity(token)
    if not identity:
        return api_err('Unauthorized', 401)

    data    = request.get_json(silent=True) or {}
    ctype   = data.get('type', '').strip()
    cid     = str(data.get('id', '')).strip()
    uid     = str(identity.get('uid') or identity.get('user_id') or identity.get('id', ''))

    if not ctype or not cid:
        return api_err('type and id required', 400)

    try:
        existing = SafetyPromoRead.query.filter_by(
            user_id=uid, content_type=ctype, content_id=cid).first()
        if not existing:
            db.session.add(SafetyPromoRead(
                user_id=uid, content_type=ctype, content_id=cid))
            db.session.commit()
        return api_ok({'marked': True}, 'Marked as read')
    except Exception as e:
        db.session.rollback()
        return api_err(str(e)[:200], 500)


@app.route('/api/mobile/safety/acknowledge', methods=['POST', 'OPTIONS'])
@csrf.exempt
def api_mobile_safety_acknowledge():
    """Flutter: record mandatory acknowledgment. Body: {type, id, full_name, device_info}"""
    if request.method == 'OPTIONS':
        return api_ok({}, 'ok')
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    identity = _get_identity(token)
    if not identity:
        return api_err('Unauthorized', 401)

    data      = request.get_json(silent=True) or {}
    ctype     = data.get('type', '').strip()
    cid       = str(data.get('id', '')).strip()
    uid       = str(identity.get('uid') or identity.get('user_id') or identity.get('id', ''))
    full_name = data.get('full_name', identity.get('name', ''))
    device    = data.get('device_info', '')

    if not ctype or not cid:
        return api_err('type and id required', 400)

    try:
        existing = SafetyPromoAck.query.filter_by(
            user_id=uid, content_type=ctype, content_id=cid).first()
        if not existing:
            db.session.add(SafetyPromoAck(
                user_id=uid, full_name=full_name,
                content_type=ctype, content_id=cid,
                device_info=device[:200] if device else None))
            db.session.commit()
        # Also mark as read
        if not SafetyPromoRead.query.filter_by(
                user_id=uid, content_type=ctype, content_id=cid).first():
            db.session.add(SafetyPromoRead(
                user_id=uid, content_type=ctype, content_id=cid))
            db.session.commit()
        return api_ok({'acknowledged': True}, 'Acknowledgment recorded')
    except Exception as e:
        db.session.rollback()
        return api_err(str(e)[:200], 500)


@app.route('/api/mobile/safety/search', methods=['GET', 'OPTIONS'])
@csrf.exempt
def api_mobile_safety_search():
    """Flutter: search across all safety promotion content.

    Query params:
      q    — search term (min 2 chars)
      type — optional filter
    """
    if request.method == 'OPTIONS':
        return api_ok({}, 'ok')
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    identity = _get_identity(token)
    if not identity:
        return api_err('Unauthorized', 401)

    q     = request.args.get('q', '').strip()
    ftype = request.args.get('type', '').lower()
    uid   = str(identity.get('uid') or identity.get('user_id') or identity.get('id', ''))

    if len(q) < 2:
        return api_err('Search term must be at least 2 characters', 400)

    pat = f'%{q}%'
    try:
        read_set = _sp_user_read_set(uid)
        ack_set  = _sp_user_ack_set(uid)
        items = []

        if not ftype or ftype in ('alert', 'bulletin'):
            rows = SafetyBulletin.query.filter(
                SafetyBulletin.status == 'Active',
                db.or_(SafetyBulletin.title.ilike(pat),
                       SafetyBulletin.content.ilike(pat),
                       SafetyBulletin.ref_number.ilike(pat))
            ).all()
            for b in rows:
                ct = 'alert' if (b.bulletin_type or '').lower() in ('safety alert', 'alert') else 'bulletin'
                if ftype and ftype != ct:
                    continue
                items.append(_sp_item(b, ct, 'id', 'title', 'issue_date',
                                      summary_attr='content', severity_attr='severity',
                                      read_set=read_set, ack_set=ack_set))

        if not ftype or ftype == 'newsletter':
            rows = SafetyNewsletter.query.filter(
                SafetyNewsletter.status == 'Published',
                db.or_(SafetyNewsletter.title.ilike(pat),
                       SafetyNewsletter.summary.ilike(pat),
                       SafetyNewsletter.content.ilike(pat))
            ).all()
            for n in rows:
                items.append(_sp_item(n, 'newsletter', 'id', 'title', 'issue_date',
                                      summary_attr='summary',
                                      content_attr='content',
                                      author_attr='author',
                                      ref_attr='ref_number',
                                      read_set=read_set, ack_set=ack_set))

        if not ftype or ftype == 'survey':
            rows = SafetySurvey.query.filter(
                SafetySurvey.status == 'Active',
                db.or_(SafetySurvey.title.ilike(pat),
                       SafetySurvey.description.ilike(pat))
            ).all()
            for s in rows:
                items.append(_sp_item(s, 'survey', 'id', 'title', 'start_date',
                                      summary_attr='description', read_set=read_set, ack_set=ack_set))

        if not ftype or ftype == 'lesson':
            rows = LessonLearned.query.filter(
                LessonLearned.status == 'Published',
                db.or_(LessonLearned.title.ilike(pat),
                       LessonLearned.lesson.ilike(pat),
                       LessonLearned.description.ilike(pat))
            ).all()
            for l in rows:
                items.append(_sp_item(l, 'lesson', 'id', 'title', 'date',
                                      summary_attr='lesson', read_set=read_set, ack_set=ack_set))

        items.sort(key=lambda x: x['date'] or '', reverse=True)
        return api_ok({'items': items, 'total': len(items)}, f'{len(items)} result(s)')
    except Exception as e:
        return api_err(str(e)[:200], 500)


@app.route('/api/mobile/safety/survey/respond', methods=['POST', 'OPTIONS'])
@csrf.exempt
def api_mobile_safety_survey_respond():
    """Flutter: submit survey answers.

    Body: {survey_id, answers: [{question_index, answer}], is_anonymous}
    """
    if request.method == 'OPTIONS':
        return api_ok({}, 'ok')
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    identity = _get_identity(token)
    if not identity:
        return api_err('Unauthorized', 401)

    data        = request.get_json(silent=True) or {}
    survey_id   = data.get('survey_id')
    answers     = data.get('answers', [])
    is_anon     = bool(data.get('is_anonymous', False))
    uid         = str(identity.get('uid') or identity.get('user_id') or identity.get('id', ''))
    name        = identity.get('name', '') or identity.get('username', uid)
    dept_id     = request.headers.get('X-Dept-Id')

    if not survey_id:
        return api_err('survey_id required', 400)

    try:
        survey = SafetySurvey.query.get(survey_id)
        if not survey:
            return api_err('Survey not found', 404)
        if survey.status != 'Active':
            return api_err('Survey is not active', 400)

        # Idempotency: one response per user per survey
        # Use respondent_email to store uid for dedup; respondent_name holds real name
        existing = SurveyResponse.query.filter_by(
            survey_id=survey_id, respondent_email=uid).first()
        if existing:
            return api_ok({'already_responded': True}, 'Already responded')

        resp = SurveyResponse(
            survey_id       = survey_id,
            respondent_name = 'Anonymous' if is_anon else (name or uid),
            respondent_email= '' if is_anon else uid,
            department_id   = int(dept_id) if dept_id and dept_id.isdigit() else None,
            is_anonymous    = is_anon,
            answers         = json.dumps(answers),
        )
        db.session.add(resp)
        # Increment response count
        if survey.response_count is not None:
            survey.response_count = (survey.response_count or 0) + 1
        # Also mark as read
        if not SafetyPromoRead.query.filter_by(
                user_id=uid, content_type='survey', content_id=str(survey_id)).first():
            db.session.add(SafetyPromoRead(
                user_id=uid, content_type='survey', content_id=str(survey_id)))
        db.session.commit()
        return api_ok({'submitted': True}, 'Response recorded — thank you!')
    except Exception as e:
        db.session.rollback()
        return api_err(str(e)[:200], 500)


# ═══════════════════════════════════════════════════════════════════════════════
#  PUSH NOTIFICATION ENGINE  (Firebase Admin SDK v1 API)
# ═══════════════════════════════════════════════════════════════════════════════

def _get_fcm_app():
    """Return initialized firebase_admin App, or None if not configured."""
    try:
        import firebase_admin
        from firebase_admin import credentials
        if not firebase_admin._apps:
            sa_json = os.environ.get('FIREBASE_SERVICE_ACCOUNT', '')
            if not sa_json:
                return None
            import json, tempfile
            sa_dict = json.loads(sa_json)
            cred = credentials.Certificate(sa_dict)
            firebase_admin.initialize_app(cred)
        return firebase_admin.get_app()
    except Exception:
        return None


def _fcm_send_multicast(tokens, title, body, data):
    """Send FCM v1 multicast message. Returns (success_count, failure_count)."""
    if not tokens:
        return 0, 0
    fcm_app = _get_fcm_app()
    if fcm_app is None:
        app.logger.warning(
            'FCM not configured: FIREBASE_SERVICE_ACCOUNT env var is missing or invalid. '
            'Push notification NOT sent. In-app log will still be written.')
        return 0, len(tokens)
    try:
        from firebase_admin import messaging
        msgs = [
            messaging.Message(
                notification=messaging.Notification(title=title, body=body),
                data={k: str(v) for k, v in data.items()},
                android=messaging.AndroidConfig(priority='high',
                    notification=messaging.AndroidNotification(
                        sound='default', channel_id='avias_alerts')),
                apns=messaging.APNSConfig(payload=messaging.APNSPayload(
                    aps=messaging.Aps(sound='default', badge=1))),
                token=t,
            ) for t in tokens
        ]
        ok = fail = 0
        for i in range(0, len(msgs), 500):
            resp = messaging.send_each(msgs[i:i+500])
            ok   += resp.success_count
            fail += resp.failure_count
            # Log individual failures so stale tokens are visible in logs
            for idx, r in enumerate(resp.responses):
                if not r.success:
                    app.logger.warning(
                        f'FCM token failure [{i+idx}]: {r.exception}')
        app.logger.info(f'FCM sent: {ok} success, {fail} failure, title="{title[:60]}"')
        return ok, fail
    except Exception as e:
        app.logger.error(f'FCM send_each exception: {e}')
        return 0, len(tokens)


def _log_notification(user_id, title, body, ntype, ctype, cid):
    """Persist an EmployeeNotificationLog row for in-app notification centre."""
    try:
        log = EmployeeNotificationLog(
            employee_user_id=user_id,
            title=title, body=body,
            notification_type=ntype,
            content_type=ctype,
            content_id=str(cid),
        )
        db.session.add(log)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        app.logger.warning(f'Notification log error: {e}')


def push_notify_all(title, body, ntype, ctype, cid):
    """
    Broadcast push + in-app log to ALL registered employees.
    Used for: bulletin, newsletter, survey.
    """
    try:
        tokens_rows = DeviceToken.query.all()
        app.logger.info(
            f'push_notify_all: title="{title[:60]}" '
            f'tokens={len(tokens_rows)} ctype={ctype} cid={cid}')
        if tokens_rows:
            data = {'type': ctype, 'id': str(cid), 'ntype': ntype}
            _fcm_send_multicast([r.fcm_token for r in tokens_rows], title, body, data)
        # Log for every active employee — always, even when no device tokens registered
        employees = Employee.query.filter_by(is_active=True).all()
        for emp in employees:
            _log_notification(f'emp_{emp.id}', title, body, ntype, ctype, str(cid))
        app.logger.info(f'push_notify_all: in-app log written for {len(employees)} employees')
    except Exception as e:
        app.logger.warning(f'push_notify_all error: {e}')


def push_notify_user(user_id, title, body, ntype, ctype, cid):
    """
    Push + in-app log to a SINGLE employee user.
    Used for: action assign, training assign, investigation update, audit finding.
    user_id format: 'emp_5'
    """
    try:
        tokens_rows = DeviceToken.query.filter_by(user_id=user_id).all()
        app.logger.info(
            f'push_notify_user: user={user_id} title="{title[:60]}" '
            f'tokens={len(tokens_rows)} ctype={ctype} cid={cid}')
        if tokens_rows:
            data = {'type': ctype, 'id': str(cid), 'ntype': ntype}
            _fcm_send_multicast([r.fcm_token for r in tokens_rows], title, body, data)
        _log_notification(user_id, title, body, ntype, ctype, cid)
    except Exception as e:
        app.logger.warning(f'push_notify_user error: {e}')


def push_notify_by_name(employee_name, title, body, ntype, ctype, cid):
    """
    Resolve employee full_name → user_id, then push_notify_user.
    Falls back to case-insensitive / stripped comparison if exact match fails.
    """
    if not employee_name:
        return
    try:
        name = employee_name.strip()
        emp = Employee.query.filter_by(full_name=name).first()
        if not emp:
            # Case-insensitive fallback
            emp = Employee.query.filter(
                db.func.lower(Employee.full_name) == name.lower()
            ).first()
        if emp:
            push_notify_user(f'emp_{emp.id}', title, body, ntype, ctype, cid)
        else:
            app.logger.warning(
                f'push_notify_by_name: no Employee found for name="{name}" — '
                f'in-app log NOT written, push NOT sent.')
    except Exception as e:
        app.logger.warning(f'push_notify_by_name error: {e}')


# Keep legacy name as alias (not used anywhere but kept for safety)
def _sp_push_notify(title, body, ctype, cid, dept_id=None):
    push_notify_all(title, body, 'safety_promo', ctype, cid)


# ── Admin: Notification monitoring ────────────────────────────────────────────
@app.route('/admin/notifications')
@require_login
def admin_notifications():
    page    = request.args.get('page', 1, type=int)
    ntype_f = request.args.get('ntype', '')
    q = EmployeeNotificationLog.query.order_by(EmployeeNotificationLog.sent_at.desc())
    if ntype_f:
        q = q.filter(EmployeeNotificationLog.notification_type == ntype_f)
    total   = q.count()
    logs    = q.offset((page-1)*50).limit(50).all()
    unread  = EmployeeNotificationLog.query.filter_by(is_read=False).count()
    by_type = db.session.query(
        EmployeeNotificationLog.notification_type,
        db.func.count(EmployeeNotificationLog.id)
    ).group_by(EmployeeNotificationLog.notification_type).all()
    return render_template('admin/notification_monitor.html',
        logs=logs, total=total, unread=unread, page=page,
        ntype_f=ntype_f, by_type=by_type,
        page_title='Notification Monitor')


# ── Alerts: content existence check ───────────────────────────────────────────

def _content_exists(ctype, cid):
    """Return True if the source record still exists in the DB."""
    try:
        if ctype == 'training':
            return db.session.get(Training, int(cid)) is not None
        if ctype == 'action':
            return db.session.get(Action, str(cid)) is not None
        if ctype == 'audit_finding':
            return db.session.get(AuditFinding, str(cid)) is not None
        if ctype == 'hazard':
            return db.session.get(HazardReport, str(cid)) is not None
        if ctype == 'investigation':
            return db.session.get(Investigation, str(cid)) is not None
        if ctype == 'moc':
            return db.session.get(MOC, str(cid)) is not None
        if ctype == 'bulletin':
            return db.session.get(SafetyBulletin, str(cid)) is not None
        if ctype == 'newsletter':
            return db.session.get(SafetyNewsletter, int(cid)) is not None
        if ctype == 'survey':
            return db.session.get(SafetySurvey, int(cid)) is not None
        return True
    except Exception:
        return True  # safe fallback — unknown types stay visible


# ── API: actionable alerts (Alerts tab) ───────────────────────────────────────
_ALERT_TYPES = {'action_assigned', 'system'}

@app.route('/api/mobile/alerts', methods=['GET'])
@csrf.exempt
def api_mobile_alerts():
    """
    Returns only actionable notifications for the Alerts tab.
    Stale entries (source deleted) are pruned automatically.
    """
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    data  = _verify_token(token)
    if not data:
        return api_err('Unauthorized', 401)
    try:
        uid_str, _, _, _ = _resolve_employee(data)
        logs = EmployeeNotificationLog.query.filter(
            EmployeeNotificationLog.employee_user_id == uid_str,
            EmployeeNotificationLog.notification_type.in_(_ALERT_TYPES),
        ).order_by(EmployeeNotificationLog.sent_at.desc()).limit(100).all()

        records = []
        stale_ids = []
        for n in logs:
            ctype = n.content_type or ''
            cid   = n.content_id   or ''
            # Prune stale entries whose source record was deleted
            if ctype and cid and not _content_exists(ctype, cid):
                stale_ids.append(n.id)
                continue
            records.append({
                'id':                n.id,
                'title':             n.title,
                'body':              n.body or '',
                'notification_type': n.notification_type or '',
                'content_type':      ctype,
                'content_id':        cid,
                'is_read':           n.is_read,
                'sent_at':           n.sent_at.isoformat()[:16] if n.sent_at else '',
            })

        if stale_ids:
            EmployeeNotificationLog.query.filter(
                EmployeeNotificationLog.id.in_(stale_ids)).delete(synchronize_session=False)
            db.session.commit()

        return api_ok(records, 'Alerts loaded')
    except Exception as e:
        app.logger.exception('api_mobile_alerts error')
        return api_ok([], 'No alerts')


# ── API: mark single alert read ────────────────────────────────────────────────
@app.route('/api/mobile/alerts/<int:alert_id>/read', methods=['POST'])
@csrf.exempt
def api_mobile_alert_read(alert_id):
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    data  = _verify_token(token)
    if not data:
        return api_err('Unauthorized', 401)
    try:
        uid_str = f"emp_{data['user_id']}"
        n = EmployeeNotificationLog.query.filter_by(
            id=alert_id, employee_user_id=uid_str).first()
        if n and not n.is_read:
            n.is_read = True
            db.session.commit()
        return api_ok({}, 'Marked read')
    except Exception as e:
        db.session.rollback()
        return api_err(str(e)[:120], 500)


# ── API: unread count for Flutter badge ───────────────────────────────────────
@app.route('/api/mobile/notifications/unread-count')
@csrf.exempt
def api_mobile_notif_unread_count():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    data  = _verify_token(token)
    if not data:
        return api_err('Unauthorized', 401)
    uid, _, _, _ = _resolve_employee(data)
    if not uid:
        return api_err('Employee not found', 404)
    user_id = f'emp_{uid}'
    # Count only actionable (Alerts tab) unread — not safety_promo informational
    count = EmployeeNotificationLog.query.filter(
        EmployeeNotificationLog.employee_user_id == user_id,
        EmployeeNotificationLog.is_read == False,
        EmployeeNotificationLog.notification_type.in_(_ALERT_TYPES),
    ).count()
    return api_ok({'unread_count': count})


@app.route('/api/admin/notifications/diagnose', methods=['GET'])
@csrf.exempt
@require_login
def api_notifications_diagnose():
    """
    Diagnostic endpoint — verifies the full notification pipeline without sending real events.
    Returns JSON with status of each component: FCM config, device tokens, DB log table.
    Optionally sends a test push if ?send_test=1&user_id=emp_5
    """
    report = {}

    # 1. FCM app initialisation
    fcm_app = _get_fcm_app()
    report['fcm_configured'] = fcm_app is not None
    report['firebase_service_account_set'] = bool(os.environ.get('FIREBASE_SERVICE_ACCOUNT', ''))

    # 2. Device tokens in DB
    try:
        all_tokens = DeviceToken.query.all()
        report['device_tokens_total'] = len(all_tokens)
        report['device_tokens'] = [
            {'user_id': dt.user_id,
             'token_preview': dt.fcm_token[:20] + '…',
             'updated_at': str(dt.updated_at)}
            for dt in all_tokens[:50]
        ]
    except Exception as e:
        report['device_tokens_error'] = str(e)

    # 3. EmployeeNotificationLog recent rows
    try:
        recent = EmployeeNotificationLog.query.order_by(
            EmployeeNotificationLog.sent_at.desc()).limit(5).all()
        report['notification_log_recent'] = [
            {'id': n.id, 'user': n.employee_user_id, 'title': n.title,
             'type': n.notification_type, 'is_read': n.is_read, 'sent_at': str(n.sent_at)}
            for n in recent
        ]
    except Exception as e:
        report['notification_log_error'] = str(e)

    # 4. Optional test push
    test_user = request.args.get('user_id', '')
    send_test = request.args.get('send_test', '') == '1'
    if send_test and test_user:
        try:
            push_notify_user(
                test_user,
                '🧪 AviaS Test Notification',
                'This is a diagnostic test from the AviaS notification system.',
                'system', 'system', '0')
            report['test_push_sent_to'] = test_user
            report['test_push_status']  = 'called — check device and logs'
        except Exception as e:
            report['test_push_error'] = str(e)

    return api_ok(report)


@app.route('/api/admin/notifications/send-overdue-reminders', methods=['POST'])
@csrf.exempt
@require_login
def api_send_overdue_reminders():
    """Send push reminders to owners of overdue actions. Call manually or via scheduler."""
    try:
        overdue = Action.query.filter_by(status='Overdue').all()
        sent = 0
        for a in overdue:
            if a.owner:
                push_notify_by_name(
                    a.owner,
                    f'⏰ Overdue Action: {a.title or a.action_id}',
                    f'This action was due {a.due_date}. Please update its status immediately.',
                    'action_assigned', 'action', str(a.id))
                sent += 1
        return api_ok({'reminders_sent': sent})
    except Exception as e:
        app.logger.error(f'Overdue reminder error: {e}')
        return api_err(str(e), 500)


@app.route('/admin/login', methods=['GET', 'POST'])
@csrf.exempt
def admin_login():
    # If already authenticated, go straight to dashboard
    if is_logged_in():
        return redirect(url_for('dashboard'))
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        user = User.query.filter_by(username=username, is_active=True).first()
        if user and check_pw(password, user.password_hash):
            # Silently upgrade legacy SHA-256 hash to pbkdf2 on first web login
            if _is_legacy_hash(user.password_hash):
                user.password_hash = hash_pw(password)
            session['admin_logged_in'] = True
            session['admin_user']      = user.username
            session['admin_role']      = user.role
            session['admin_name']      = user.full_name or user.username
            session.permanent          = True
            user.last_login            = datetime.utcnow()
            db.session.commit()
            # Sanitise next= to prevent open redirect: only allow same-origin paths
            raw_next = request.args.get('next', '')
            from urllib.parse import urlparse
            parsed = urlparse(raw_next)
            # Accept only relative paths that don't loop back to login
            if raw_next and not parsed.netloc and not parsed.scheme \
                    and not raw_next.startswith('/admin/login'):
                next_url = raw_next
            else:
                next_url = url_for('dashboard')
            return redirect(next_url)
        else:
            error = 'Invalid username or password.'
    return render_template('portal/login.html', error=error)


@app.route('/admin/logout')
@csrf.exempt
def admin_logout():
    session.clear()
    flash('You have been logged out.', 'info')
    return redirect(url_for('public_portal'))


# ── EMPLOYEE MANAGEMENT ────────────────────────────────────────────────────────

@app.route('/admin/employees')
@require_login
def admin_employees():
    """Employee management — mobile app users."""
    if session.get('admin_role') != 'admin':
        flash('Admin access required.', 'error')
        return redirect(url_for('dashboard'))
    employees = Employee.query.order_by(Employee.created_at.desc()).all()
    departments = Department.query.all()
    return render_template('portal/admin_employees.html',
                           employees=employees, departments=departments,
                           total=len(employees),
                           active=sum(1 for e in employees if e.is_active))


@app.route('/admin/employees/new', methods=['POST'])
@require_login
def admin_employee_new():
    if session.get('admin_role') != 'admin':
        return redirect(url_for('dashboard'))
    f = request.form
    if Employee.query.filter_by(username=f['username']).first():
        flash('Username already exists.', 'error')
        return redirect(url_for('admin_employees'))
    if Employee.query.filter_by(employee_id=f['employee_id']).first():
        flash('Employee ID already exists.', 'error')
        return redirect(url_for('admin_employees'))
    emp = Employee(
        employee_id   = f['employee_id'].strip(),
        username      = f['username'].strip(),
        password_hash = hash_pw(f['password']),
        full_name     = f.get('full_name','').strip(),
        email         = f.get('email','').strip(),
        mobile        = f.get('mobile','').strip(),
        role          = f.get('role','employee'),
        department_id = int(f['department_id']) if f.get('department_id') else None,
        is_active     = True,
    )
    db.session.add(emp)
    db.session.commit()
    flash(f'✓ Employee {emp.full_name} ({emp.employee_id}) created successfully.', 'success')
    return redirect(url_for('admin_employees'))


@app.route('/admin/employees/<int:eid>/toggle', methods=['POST'])
@require_login
def admin_employee_toggle(eid):
    if session.get('admin_role') != 'admin':
        return redirect(url_for('dashboard'))
    emp = Employee.query.get_or_404(eid)
    emp.is_active = not emp.is_active
    db.session.commit()
    flash(f'✓ {emp.full_name} {"activated" if emp.is_active else "deactivated"}.', 'success')
    return redirect(url_for('admin_employees'))


@app.route('/admin/employees/<int:eid>/reset-password', methods=['POST'])
@require_login
def admin_employee_reset_pw(eid):
    if session.get('admin_role') != 'admin':
        return redirect(url_for('dashboard'))
    emp = Employee.query.get_or_404(eid)
    new_pw = request.form.get('new_password','')
    if not new_pw or len(new_pw) < 6:
        flash('Password must be at least 6 characters.', 'error')
        return redirect(url_for('admin_employees'))
    emp.password_hash = hash_pw(new_pw)
    db.session.commit()
    flash(f'✓ Password reset for {emp.full_name}.', 'success')
    return redirect(url_for('admin_employees'))


@app.route('/admin/employees/<int:eid>/delete', methods=['POST'])
@require_login
def admin_employee_delete(eid):
    if session.get('admin_role') != 'admin':
        return redirect(url_for('dashboard'))
    emp = Employee.query.get_or_404(eid)
    name = emp.full_name
    db.session.delete(emp)
    db.session.commit()
    flash(f'✓ Employee {name} deleted.', 'success')
    return redirect(url_for('admin_employees'))


@app.route('/admin/users')
@require_login
def admin_users():
    """User management — admin only."""
    if session.get('admin_role') != 'admin':
        flash('⚠ Admin access required.', 'error')
        return redirect(url_for('dashboard'))
    users = User.query.order_by(User.created_at).all()
    return render_template('portal/admin_users.html', users=users)


@app.route('/admin/users/new', methods=['POST'])
@require_login
def admin_user_new():
    if session.get('admin_role') != 'admin':
        return redirect(url_for('dashboard'))
    f = request.form
    if User.query.filter_by(username=f['username']).first():
        flash('⚠ Username already exists.', 'error')
        return redirect(url_for('admin_users'))
    u = User(username=f['username'], password_hash=hash_pw(f['password']),
             full_name=f.get('full_name',''), role=f.get('role','safety_officer'),
             sag_role=f.get('sag_role',''),
             department_id=int(f['department_id']) if f.get('department_id') else None,
             is_active=True)
    db.session.add(u); db.session.commit()
    flash(f'✓ User {u.username} created.', 'success')
    return redirect(url_for('admin_users'))


@app.route('/admin/users/<int:uid>/edit', methods=['GET', 'POST'])
@require_login
def admin_user_edit(uid):
    if session.get('admin_role') != 'admin':
        flash('⚠ Admin access required.', 'error')
        return redirect(url_for('dashboard'))
    u = User.query.get_or_404(uid)
    if request.method == 'POST':
        f = request.form
        new_username = f.get('username', '').strip()
        if new_username and new_username != u.username:
            if User.query.filter_by(username=new_username).first():
                flash('⚠ Username already taken.', 'error')
                return redirect(url_for('admin_users'))
            u.username = new_username
        if f.get('full_name') is not None:
            u.full_name = f['full_name'].strip()
        new_pw = f.get('password', '').strip()
        if new_pw:
            u.password_hash = hash_pw(new_pw)
        u.role = f.get('role', u.role)
        u.sag_role = f.get('sag_role', u.sag_role)
        u.department_id = int(f['department_id']) if f.get('department_id') else None
        u.is_active = f.get('is_active') == '1'
        db.session.commit()
        flash(f'✓ User {u.username} updated.', 'success')
        return redirect(url_for('admin_users'))
    users = User.query.order_by(User.created_at).all()
    return render_template('portal/admin_users.html', users=users, edit_user=u)


@app.route('/admin/users/<int:uid>/toggle', methods=['POST'])
@require_login
def admin_user_toggle(uid):
    if session.get('admin_role') != 'admin':
        return redirect(url_for('dashboard'))
    u = User.query.get_or_404(uid)
    u.is_active = not u.is_active
    db.session.commit()
    flash(f'✓ User {u.username} {"activated" if u.is_active else "deactivated"}.', 'success')
    return redirect(url_for('admin_users'))


@app.route('/admin/reports')
@require_login
def admin_reports_inbox():
    """Safety Admin inbox — all submitted public reports."""
    vol   = VoluntaryReport.query.order_by(VoluntaryReport.created_at.desc()).all()
    conf  = ConfidentialReport.query.order_by(ConfidentialReport.created_at.desc()).all()
    haz   = HazardReport.query.order_by(HazardReport.created_at.desc()).limit(20).all()
    asr   = ASRReport.query.order_by(ASRReport.created_at.desc()).limit(20).all()
    total_new = (VoluntaryReport.query.filter_by(status='Submitted').count() +
                 ConfidentialReport.query.filter_by(status='Submitted').count())
    return render_template('portal/admin_reports.html',
                           vol=vol, conf=conf, haz=haz, asr=asr,
                           total_new=total_new)


@app.route('/admin/reports/voluntary/<int:rid>/review', methods=['POST'])
@require_login
def vol_report_review(rid):
    r = VoluntaryReport.query.get_or_404(rid)
    r.status = 'Under Review'
    db.session.commit()
    flash('✓ Marked as Under Review.', 'success')
    return redirect(url_for('admin_reports_inbox'))


@app.route('/admin/reports/confidential/<int:rid>/review', methods=['POST'])
@require_login
def conf_report_review(rid):
    r = ConfidentialReport.query.get_or_404(rid)
    r.status = 'Under Review'
    db.session.commit()
    flash('✓ Marked as Under Review.', 'success')
    return redirect(url_for('admin_reports_inbox'))


@app.route('/admin/reports/voluntary/<int:rid>/delete', methods=['POST'])
@require_login
def vol_report_delete(rid):
    r = VoluntaryReport.query.get_or_404(rid)
    db.session.delete(r)
    db.session.commit()
    flash('Voluntary report deleted.', 'success')
    return redirect(url_for('admin_reports_inbox'))


@app.route('/admin/reports/confidential/<int:rid>/delete', methods=['POST'])
@require_login
def conf_report_delete(rid):
    r = ConfidentialReport.query.get_or_404(rid)
    db.session.delete(r)
    db.session.commit()
    flash('Confidential report deleted.', 'success')
    return redirect(url_for('admin_reports_inbox'))


@app.route('/admin/reports/hazard/<rid>/delete', methods=['POST'])
@require_login
def hazard_report_delete(rid):
    r = HazardReport.query.get_or_404(rid)
    db.session.delete(r)
    db.session.commit()
    flash('Hazard report deleted.', 'success')
    return redirect(url_for('admin_reports_inbox'))


@app.route('/admin/reports/asr/<rid>/delete', methods=['POST'])
@require_login
def asr_report_delete(rid):
    r = ASRReport.query.get_or_404(rid)
    db.session.delete(r)
    db.session.commit()
    flash('ASR report deleted.', 'success')
    return redirect(url_for('admin_reports_inbox'))


# ═══════════════════════════════════════════════════════════════════════════════
#  INTERNAL ADMIN DASHBOARD  — login required
# ═══════════════════════════════════════════════════════════════════════════════

@app.route('/dashboard')
@require_login
def dashboard():
    check_overdue_actions()
    from sqlalchemy import func as sqf
    db.session.rollback()  # ensure clean PostgreSQL transaction state

    # ── HAZARDS & REPORTS ─────────────────────────────────────────────────────
    try:
        total_haz = Hazard.query.count()
        open_haz  = Hazard.query.filter_by(status='Open').count()
    except Exception:
        db.session.rollback()
        total_haz = open_haz = 0
    try:
        intol = Risk.query.filter_by(initial_tolerance='INTOLERABLE').count()
    except Exception:
        db.session.rollback()
        intol = 0
    try:
        asr_cnt = ASRReport.query.count()
    except Exception:
        db.session.rollback()
        asr_cnt = 0
    try:
        inv_cnt = Investigation.query.count()
    except Exception:
        db.session.rollback()
        inv_cnt = 0
    try:
        hr_cnt = HazardReport.query.count()
    except Exception:
        db.session.rollback()
        hr_cnt = 0

    # ── ACTIONS ───────────────────────────────────────────────────────────────
    try:
        open_act       = Action.query.filter(
                             Action.status.in_(['Open','In Progress','Overdue'])).count()
        overdue_act    = Action.query.filter_by(status='Overdue').count()
        pending_review = Action.query.filter(
                             Action.status.in_(['Mitigation Implemented',
                                               'Under Safety Review'])).count()
        closed_act     = Action.query.filter_by(status='Closed').count()
        total_act      = Action.query.count()
        closure_rate   = round(closed_act / total_act * 100) if total_act > 0 else 0
        unassigned_act = Action.query.filter(
                             (Action.sag_member == None) | (Action.sag_member == ''),
                             Action.status.notin_(['Closed'])).count()
    except Exception:
        db.session.rollback()
        open_act = overdue_act = pending_review = closed_act = 0
        total_act = closure_rate = unassigned_act = 0

    # ── AUDITS ────────────────────────────────────────────────────────────────
    try:
        audit_cnt       = AuditSchedule.query.count()
        active_audits   = AuditSchedule.query.filter_by(status='In Progress').count()
        open_findings   = AuditFinding.query.filter(
                              AuditFinding.status.notin_(['Closed','Accepted'])).count()
        closed_findings = AuditFinding.query.filter_by(status='Closed').count()
        total_findings  = AuditFinding.query.count()
        finding_close_rate = round(closed_findings / total_findings * 100) \
                             if total_findings > 0 else 100
        major_findings    = AuditFinding.query.filter_by(severity='Major').filter(
                                AuditFinding.status.notin_(['Closed'])).count()
        critical_findings = AuditFinding.query.filter_by(severity='Critical').filter(
                                AuditFinding.status.notin_(['Closed'])).count()
    except Exception:
        db.session.rollback()
        audit_cnt = active_audits = open_findings = closed_findings = 0
        total_findings = major_findings = critical_findings = 0
        finding_close_rate = 100

    # ── SPI ───────────────────────────────────────────────────────────────────
    # ICAO-correct: use calculated value field + statistical thresholds
    spi_alerts = spi_l2 = spi_l3 = 0
    try:
        for ind in SPIIndicator.query.filter_by(active=True).all():
            history = _spi_history(ind)
            if not history:
                continue
            all_hist_vals = [v for _, _, v in history]
            latest_val = all_hist_vals[-1]
            l1, l2, l3, _mean, _sd, _is_stat = _spi_thresholds(ind, all_hist_vals)
            is_pct = (ind.calc_type == 'PERCENT')
            if is_pct:
                if l3 and latest_val <= l3: spi_l3 += 1; spi_alerts += 1
                elif l2 and latest_val <= l2: spi_l2 += 1; spi_alerts += 1
                elif l1 and latest_val <= l1: spi_alerts += 1
            else:
                if l3 and latest_val >= l3: spi_l3 += 1; spi_alerts += 1
                elif l2 and latest_val >= l2: spi_l2 += 1; spi_alerts += 1
                elif l1 and latest_val >= l1: spi_alerts += 1
    except Exception:
        db.session.rollback()
        spi_alerts = spi_l2 = spi_l3 = 0

    # ── SAFETY PROMOTION ─────────────────────────────────────────────────────
    active_surveys = active_bulletins = active_campaigns = 0
    try:
        active_surveys = SafetySurvey.query.filter_by(status='Active').count()
    except Exception:
        db.session.rollback()
    try:
        active_bulletins = SafetyBulletin.query.filter_by(status='Active').count()
    except Exception:
        db.session.rollback()
    try:
        active_campaigns = SafetyCampaign.query.filter_by(status='Active').count()
    except Exception:
        db.session.rollback()

    # ── OTHER ─────────────────────────────────────────────────────────────────
    moc_cnt = doc_cnt = ra_open = 0
    try:
        moc_cnt = MOC.query.count()
        doc_cnt = SMSDocument.query.filter_by(status='Approved').count()
        ra_open = RiskAssessment.query.filter(
                      RiskAssessment.status.notin_(['Closed','Approved'])).count()
    except Exception:
        db.session.rollback()
        pass

    try:
        recent_haz = Hazard.query.order_by(Hazard.created_at.desc()).limit(5).all()
    except Exception:
        db.session.rollback()
        recent_haz = []
    try:
        recent_act = Action.query.filter(Action.status != 'Closed') \
                        .order_by(Action.created_at.desc()).limit(6).all()
    except Exception:
        db.session.rollback()
        recent_act = []
    try:
        overdue_actions = Action.query.filter_by(status='Overdue') \
                            .order_by(Action.due_date).limit(8).all()
    except Exception:
        db.session.rollback()
        overdue_actions = []
    try:
        pending_review_list = Action.query.filter(
            Action.status.in_(['Mitigation Implemented','Under Safety Review'])) \
            .order_by(Action.due_date).limit(6).all()
    except Exception:
        db.session.rollback()
        pending_review_list = []
    try:
        critical_act = Action.query.filter_by(priority='High') \
                        .filter(Action.status.notin_(['Closed'])).limit(6).all()
    except Exception:
        db.session.rollback()
        critical_act = []
    try:
        recent_findings = AuditFinding.query.filter(
            AuditFinding.status.notin_(['Closed'])) \
            .order_by(AuditFinding.created_at.desc()).limit(5).all()
    except Exception:
        db.session.rollback()
        recent_findings = []

    # ── DEPT ANALYTICS ────────────────────────────────────────────────────────
    dept_perf = []
    try:
        for dept in Department.query.all():
            d_total   = Action.query.filter_by(department_id=dept.id).count()
            d_open    = Action.query.filter_by(department_id=dept.id) \
                            .filter(Action.status.notin_(['Closed'])).count()
            d_overdue = Action.query.filter_by(department_id=dept.id,
                                               status='Overdue').count()
            d_closed  = Action.query.filter_by(department_id=dept.id,
                                               status='Closed').count()
            if d_total > 0:
                dept_perf.append({
                    'dept': dept, 'total': d_total, 'open': d_open,
                    'overdue': d_overdue, 'closed': d_closed,
                    'rate': round(d_closed / d_total * 100)
                })
        dept_perf.sort(key=lambda x: x['overdue'], reverse=True)
    except Exception:
        db.session.rollback()
        dept_perf = []

    # ── SEVERITY BREAKDOWN ────────────────────────────────────────────────────
    try:
        sev_data = {
            'Critical': AuditFinding.query.filter_by(severity='Critical').count(),
            'Major':    AuditFinding.query.filter_by(severity='Major').count(),
            'Minor':    AuditFinding.query.filter_by(severity='Minor').count(),
            'Obs':      AuditFinding.query.filter_by(severity='Observation').count(),
        }
    except Exception:
        db.session.rollback()
        sev_data = {'Critical': 0, 'Major': 0, 'Minor': 0, 'Obs': 0}

    # ── AUDIT PLAN ALERTS ─────────────────────────────────────────────────────
    now_month = datetime.now().month
    now_year  = datetime.now().year
    try:
        plan_this_month = AuditPlan.query.filter_by(year=now_year, month=now_month) \
                            .filter(AuditPlan.status != 'Completed').all()
        plan_overdue = AuditPlan.query.filter(
            AuditPlan.year == now_year, AuditPlan.month < now_month,
            AuditPlan.month != None, AuditPlan.status != 'Completed').all()
    except Exception:
        db.session.rollback()
        plan_this_month = []
        plan_overdue = []

    # ── VOLUNTARY & CONFIDENTIAL ─────────────────────────────────────────────
    vol_new  = 0; conf_new = 0; vol_total = 0; conf_total = 0
    try:
        vol_new   = VoluntaryReport.query.filter_by(status='Submitted').count()
        vol_total = VoluntaryReport.query.count()
        conf_new  = ConfidentialReport.query.filter_by(status='Submitted').count()
        conf_total= ConfidentialReport.query.count()
    except Exception:
        db.session.rollback()

    return render_template('dashboard/dashboard.html',
        # Hazards
        total_haz=total_haz, open_haz=open_haz, intol=intol,
        asr_cnt=asr_cnt, hr_cnt=hr_cnt,
        vol_new=vol_new, vol_total=vol_total,
        conf_new=conf_new, conf_total=conf_total,
        # Actions
        open_act=open_act, overdue_act=overdue_act, pending_review=pending_review,
        closed_act=closed_act, total_act=total_act, closure_rate=closure_rate,
        unassigned_act=unassigned_act,
        # Audits
        audit_cnt=audit_cnt, active_audits=active_audits,
        open_findings=open_findings, total_findings=total_findings,
        finding_close_rate=finding_close_rate,
        major_findings=major_findings, critical_findings=critical_findings,
        # SPI
        spi_alerts=spi_alerts, spi_l2=spi_l2, spi_l3=spi_l3,
        # Safety Promotion
        active_surveys=active_surveys, active_bulletins=active_bulletins,
        active_campaigns=active_campaigns,
        # Other
        moc_cnt=moc_cnt, doc_cnt=doc_cnt, inv_cnt=inv_cnt, ra_open=ra_open,
        # Lists
        recent_haz=recent_haz, recent_act=recent_act,
        overdue_actions=overdue_actions, pending_review_list=pending_review_list,
        critical_act=critical_act, recent_findings=recent_findings,
        dept_perf=dept_perf, sev_data=sev_data,
        plan_this_month=plan_this_month, plan_overdue=plan_overdue)

# ─── Hazard Report ────────────────────────────────────────────────────────────
@app.route('/hazard-report', methods=['GET','POST'])
@require_login
def hazard_report():
    if request.method == 'POST':
        f   = request.form
        rid = new_id('HR')
        hid = new_id('HAZ')
        dept_id = int(f['department_id'])

        # Resolve department name BEFORE any session operations
        dept = Department.query.get(dept_id)
        dept_name = dept.name if dept else ''

        # 1. Create and commit Hazard FIRST (HazardReport has FK to hazards.id)
        h = Hazard(
            id=hid,
            source='Hazard Report',
            linked_report_id=rid,
            department_id=dept_id,
            classification=f.get('classification','Operational'),
            type_of_activity=dept_name,
            generic_hazard=f.get('generic_hazard') or f.get('hazard_description','')[:100],
            specific_components=f.get('hazard_description',''),
            consequences=f.get('consequences','To Be Assessed'),
            status='Open',
            owner=None
        )
        db.session.add(h)
        db.session.commit()   # ← Hazard row exists in DB before report references it

        # 2. Now create HazardReport safely (hazard_id FK is satisfied)
        rep = HazardReport(
            id=rid,
            department_id=dept_id,
            location=f.get('location',''),
            date=f.get('date', date.today().isoformat()),
            description=f.get('hazard_description',''),
            classification=f.get('classification','Operational'),
            generic_hazard=f.get('generic_hazard',''),
            consequences=f.get('consequences',''),
            immediate_action=f.get('immediate_action',''),
            suggested_mitigation=f.get('suggested_mitigation',''),
            reporter_severity=f.get('reporter_severity',''),
            reporter=f.get('reporter','Anonymous') or 'Anonymous',
            report_type=f.get('report_type','Hazard Report'),
            status='Submitted',
            hazard_id=hid
        )
        db.session.add(rep)

        # 3. Update statuses and commit together
        h.status = 'Under Assessment'
        db.session.commit()

        # Auto-update matching SPI indicators (with deduplication)
        spi_auto_update(
            source_type   = 'hazard_report',
            department_id = dept_id,
            category      = f.get('classification',''),
            year          = datetime.now().year,
            month         = datetime.now().month,
            report_id     = rid
        )
        try:
            _spi_link_event(
                event_type    = 'hazard_report',
                event_id      = rid,
                event_title   = f.get('generic_hazard', f.get('hazard_description',''))[:120],
                department_id = dept_id,
                category      = f.get('classification',''),
                extra_text    = f.get('hazard_description','') + ' ' + f.get('consequences',''),
                event_date    = f.get('date',''),
            )
        except Exception:
            pass
        flash(f'✓ Hazard Report {rid} submitted successfully. Hazard {hid} created for assessment.', 'success')
        return redirect(url_for('hazard_report_detail', rid=rid))
    return render_template('reporting/hazard_report.html')


@app.route('/hazard-report/<rid>/safety-closure', methods=['POST'])
@require_login
def safety_closure_approve(rid):
    """
    Safety Manager Final Closure — ONLY authorized roles may close.
    This is the ONLY legitimate path to Closed status.
    """
    # Role check — only safety roles may close
    allowed_roles = ('admin', 'safety_manager', 'safety_officer', 'safety_admin')
    user_role = session.get('admin_role', '')
    if user_role not in allowed_roles:
        flash('⚠ Only Safety Management personnel may approve final closure.', 'error')
        return redirect(request.referrer or url_for('hazard_report_list'))

    hr = HazardReport.query.get_or_404(rid)
    notes = request.form.get('closure_notes', '')

    try:
        # Close the HazardReport
        hr.status = 'Closed'

        # Close the linked Hazard
        if hr.hazard_id:
            haz = Hazard.query.filter_by(id=hr.hazard_id).first()
            if haz:
                haz.status = 'Closed'
                # Also close any remaining open actions on this hazard
                for act in Action.query.filter_by(hazard_id=hr.hazard_id).all():
                    if act.status not in ('Closed',):
                        old_st = act.status
                        act.status = 'Closed'
                        log_action_history(act.id,
                            session.get('admin_username', 'Safety Manager'),
                            old_st, 'Closed',
                            f'Safety closure approved: {notes[:100]}', 'closure')

        db.session.commit()
        sync_report_status(hr.hazard_id)
        db.session.commit()
        # ── AVI Hook: Hazard report closure → verify hazard controls ─────────
        try:
            _avi_generate(
                source_module='hazard', source_record_id=hr.id,
                source_description=f'Hazard report closed: {(hr.generic_hazard or hr.description or "")[:200]}',
                linked_report_id=hr.id,
                linked_hazard_id=hr.hazard_id,
                operational_risk='High' if (hr.severity or '') in ('A','B') else 'Medium',
                override_objective=f'Verify that hazard controls implemented for report "{rid}" remain operationally effective and the hazard has not recurred.',
            )
            db.session.commit()
        except Exception:
            pass
        # Phase 2: advance reporter feedback to Closed with outcome summary
        try:
            outcome_summary = notes or 'Report reviewed and closed by Safety Management.'
            actor = session.get('admin_name', session.get('admin_user', 'System'))
            advance_report_feedback(
                rid, 'Closed', actor,
                outcome_summary=outcome_summary,
                actions_taken=f'Corrective actions implemented. See actions register for details.',
                risk_level=hr.severity or 'Medium',
            )
        except Exception:
            pass
        flash(f'✓ Occurrence {rid} officially closed by Safety Management. Closure approved.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'⚠ Error: {str(e)[:80]}', 'error')

    return redirect(request.referrer or url_for('hazard_report_list'))



@app.route('/hazard-reports')
@require_login
def hazard_report_list():
    """Aviation Operational Reporting Center — all report types."""
    dept_f  = request.args.get('dept', '')
    cat_f   = request.args.get('category', '')
    stat_f  = request.args.get('status', '')
    type_f  = request.args.get('type', '')
    q_f     = request.args.get('q', '').strip()

    # Check if report_type column exists (safe for old DBs)
    has_type_col = True
    try:
        HazardReport.query.filter_by(report_type='test').first()
    except Exception:
        has_type_col = False
        db.session.rollback()

    q = HazardReport.query
    if dept_f:  q = q.filter_by(department_id=int(dept_f))
    if cat_f:   q = q.filter_by(classification=cat_f)
    if stat_f:  q = q.filter_by(status=stat_f)
    if type_f and has_type_col:
        q = q.filter_by(report_type=type_f)
    if q_f:
        q = q.filter(
            HazardReport.description.ilike(f'%{q_f}%') |
            HazardReport.location.ilike(f'%{q_f}%') |
            HazardReport.generic_hazard.ilike(f'%{q_f}%') |
            HazardReport.reporter.ilike(f'%{q_f}%')
        )

    reports = q.order_by(HazardReport.created_at.desc()).all()

    # Sync status for all visible reports from live workflow data
    synced = False
    for rpt in reports:
        if rpt.hazard_id:
            try:
                sync_report_status(rpt.hazard_id)
                synced = True
            except Exception:
                pass
    if synced:
        try: db.session.commit()
        except Exception: db.session.rollback()

    total            = HazardReport.query.count()
    submitted        = HazardReport.query.filter_by(status='Submitted').count()
    under_assessment = HazardReport.query.filter_by(status='Under Assessment').count()
    actioned         = HazardReport.query.filter_by(status='Actioned').count()
    closed           = HazardReport.query.filter_by(status='Closed').count()

    # By report type — safe fallback if column missing
    cnt_hazard = cnt_asr = cnt_vol = cnt_conf = cnt_tech = 0
    if has_type_col:
        try:
            cnt_hazard = HazardReport.query.filter_by(report_type='Hazard Report').count()
            cnt_asr    = HazardReport.query.filter_by(report_type='ASR').count()
            cnt_vol    = HazardReport.query.filter_by(report_type='Voluntary').count()
            cnt_conf   = HazardReport.query.filter_by(report_type='Confidential').count()
            cnt_tech   = HazardReport.query.filter_by(report_type='Technical Log').count()
        except Exception:
            db.session.rollback()

    return render_template('reporting/hazard_report_list.html',
        reports=reports, total=total,
        submitted=submitted, under_assessment=under_assessment,
        actioned=actioned, closed=closed,
        cnt_hazard=cnt_hazard, cnt_asr=cnt_asr, cnt_vol=cnt_vol,
        cnt_conf=cnt_conf, cnt_tech=cnt_tech,
        type_f=type_f,
        dept_f=dept_f, cat_f=cat_f, stat_f=stat_f, q_f=q_f)

@app.route('/hazard-reports/<rid>')
@require_login
def hazard_report_detail(rid):
    """Full detail view for a single hazard report."""
    rep    = HazardReport.query.get_or_404(rid)
    hazard = Hazard.query.get(rep.hazard_id) if rep.hazard_id else None
    ra     = RiskAssessment.query.filter_by(hazard_id=rep.hazard_id).first() if rep.hazard_id else None
    actions = Action.query.filter_by(hazard_id=rep.hazard_id).all() if rep.hazard_id else []
    # SPI link context
    spi_indicators, spi_links = [], []
    try:
        spi_indicators = SPIIndicator.query.filter_by(active=True).order_by(SPIIndicator.code).all()
        if SPIEventLink is not None:
            spi_links = SPIEventLink.query.filter_by(
                event_type='hazard_report', event_id=str(rid)
            ).all()
    except Exception:
        db.session.rollback()
    return render_template('reporting/hazard_report_detail.html',
        rep=rep, hazard=hazard, ra=ra, actions=actions,
        spi_active_indicators=spi_indicators,
        spi_existing_links=spi_links,
        spi_link_event_type='hazard_report',
        spi_link_event_id=str(rid),
        spi_link_event_title=rep.generic_hazard or str(rid),
        spi_link_event_date=str(rep.date or ''),
        spi_link_severity=rep.reporter_severity or '',
        spi_link_dept_id=rep.department_id or '',
        spi_link_category=rep.classification or '',
        spi_return_url=f'/hazard-reports/{rid}',
    )

@app.route('/hazard-reports/<rid>/update-status', methods=['POST'])
@require_login
def hazard_report_update_status(rid):
    rep = HazardReport.query.get_or_404(rid)
    new_st = request.form.get('status', rep.status)
    rep.status = new_st
    db.session.commit()
    # Phase 2: advance reporter feedback lifecycle (ICAO Annex 19 §3.1.2)
    try:
        actor = session.get('admin_name', session.get('admin_user', 'System'))
        advance_report_feedback(rid, new_st, actor)
    except Exception:
        pass
    # Push notification to reporter
    if rep.reporter_user_id:
        status_msgs = {
            'Under Review': 'Your report is now under safety review.',
            'Action Assigned': 'A corrective action has been assigned based on your report.',
            'Closed': 'Your safety report has been reviewed and closed. Thank you.',
        }
        push_notify_user(
            rep.reporter_user_id,
            f'📄 Report Update: {rid}',
            status_msgs.get(new_st, f'Your report status has been updated to: {new_st}'),
            'system', 'hazard', rid)
    flash(f'✓ Report {rid} status updated to {rep.status}.', 'success')
    return redirect(url_for('hazard_report_detail', rid=rid))

# ─── ASR ─────────────────────────────────────────────────────────────────────
@app.route('/asr', methods=['GET','POST'])
@require_login
def asr():
    if request.method == 'POST':
        f   = request.form
        li  = int(f.get('likelihood', 3))
        se  = f.get('severity', 'C')
        ri  = f'{li}{se}'
        occ = f.get('occurrence_type', 'Flight Occurrence')
        hid = new_id('HAZ')
        aid = new_id('ASR')

        # Build full event description including dynamic category extras
        ev_desc = f.get('event_description', '')
        extras = [(k.replace('extra_','').replace('_',' ').title(), v)
                  for k,v in f.items() if k.startswith('extra_') and v]
        if extras:
            ev_desc += chr(10)*2 + '--- Occurrence Details ---' + chr(10)
            ev_desc += chr(10).join(f'{k}: {v}' for k,v in extras)

        # Resolve submitter's department from session (web admin user) or default to FO (1)
        _asr_dept_id = session.get('admin_dept_id') or 1
        if not _asr_dept_id:
            try:
                _web_user = User.query.filter_by(
                    username=session.get('admin_user'), is_active=True).first()
                _asr_dept_id = _web_user.department_id if _web_user else 1
            except Exception:
                _asr_dept_id = 1
        _asr_dept_id = _asr_dept_id or 1  # guard against None

        # 1. Create Hazard record (feeds Hazard Log + Dashboard)
        h = Hazard(id=hid, source='ASR', linked_report_id=aid,
                   department_id=_asr_dept_id, classification='Operational',
                   type_of_activity='Flight Operations',
                   generic_hazard=occ,
                   specific_components=ev_desc,
                   consequences=f.get('operational_impact', 'To Be Assessed'),
                   status='Open', owner='Flight Operations Manager')
        db.session.add(h)
        db.session.flush()

        # 2. Create Risk record
        r = Risk(id=new_id('RSK'), hazard_id=hid,
                 description=ev_desc,
                 initial_likelihood=li, initial_severity=se,
                 initial_risk_index=ri, initial_tolerance=get_tolerance(ri))
        db.session.add(r)

        # 3. Create ASR record with all fields
        asr_rec = ASRReport(id=aid,
            report_type      = f.get('report_type', 'Voluntary'),
            occurrence_type  = occ,
            captain          = f.get('captain', ''),
            captain_staff_no = f.get('captain_staff_no', ''),
            copilot          = f.get('copilot', ''),
            copilot_staff_no = f.get('copilot_staff_no', ''),
            date             = f.get('date', ''),
            time_local       = f.get('time_local', ''),
            time_utc         = f.get('time_utc', ''),
            flight_no        = f.get('flight_no', ''),
            route_from       = f.get('route_from', ''),
            route_to         = f.get('route_to', ''),
            diverted_to      = f.get('diverted_to', ''),
            squawk           = f.get('squawk', ''),
            aircraft_type    = f.get('aircraft_type', ''),
            registration     = f.get('registration', ''),
            pax              = int(f.get('pax') or 0),
            crew             = int(f.get('crew') or 0),
            altitude_ft      = int(f.get('altitude_ft') or 0),
            flight_phase     = f.get('flight_phase', ''),
            weather_wind     = f.get('weather_wind', ''),
            weather_vis_rvr  = f.get('weather_vis_rvr', ''),
            weather_clouds   = f.get('weather_clouds', ''),
            weather_temp_c   = int(f.get('weather_temp_c') or 0),
            weather_qnh      = int(f.get('weather_qnh') or 0),
            runway           = f.get('runway', ''),
            runway_state     = f.get('runway_state', ''),
            event_description= ev_desc,
            action_taken     = f.get('action_taken', ''),
            severity=se, likelihood=li, risk_index=ri, hazard_id=hid)
        db.session.add(asr_rec)
        db.session.flush()

        # 4. Create HazardReport record so ASR appears in Hazard Reports list
        try:
            hr = HazardReport(
                id           = new_id('HR'),
                hazard_id    = hid,
                department_id= _asr_dept_id,
                date         = f.get('date', ''),
                location     = f'{f.get("route_from","")}-{f.get("route_to","")}',
                description  = ev_desc,
                classification= 'Operational',
                generic_hazard= occ,
                consequences = f.get('operational_impact', 'To Be Assessed'),
                immediate_action = f.get('action_taken', ''),
                reporter     = f.get('captain', '') or 'Flight Crew',
                reporter_severity = se,
                report_type  = 'ASR',
                status       = 'Submitted',
            )
            db.session.add(hr)
            db.session.flush()
        except Exception:
            pass  # HazardReport creation is supplementary — never block ASR submission

        # 5. Auto-create Action for high-severity occurrences
        if se in ('D', 'E') or li >= 4:
            act = Action(
                id          = new_id('ACT'),
                source      = 'ASR',
                description = f'ASR Action: {occ} — Flight {f.get("flight_no","N/A")} on {f.get("date","")}',
                owner       = f.get('captain', 'Flight Operations Manager'),
                due_date    = (datetime.now() + __import__("datetime").timedelta(days=14)).strftime('%Y-%m-%d'),
                priority    = 'High' if se == 'E' else 'Medium',
                status      = 'Open',
                hazard_id   = hid,
            )
            db.session.add(act)

        db.session.commit()
        h.status = 'Under Assessment'
        db.session.commit()

        # 5. SPI auto-update — use submitter's actual department, not hardcoded 1
        try:
            spi_auto_update(
                source_type   = 'asr',
                department_id = _asr_dept_id,
                category      = 'ASR',
                year          = datetime.now().year,
                month         = datetime.now().month,
                report_id     = aid
            )
        except Exception:
            pass
        try:
            _spi_link_event(
                event_type    = 'asr',
                event_id      = aid,
                event_title   = f.get('occurrence_description', f.get('occurrence',''))[:120],
                department_id = _asr_dept_id,
                category      = f.get('occurrence_category', f.get('occurrence','')),
                severity      = 'High' if f.get('safety_effect','') in ('D','E') else 'Medium',
                extra_text    = f.get('occurrence_description','') + ' ' + f.get('immediate_action',''),
                event_date    = f.get('date',''),
            )
        except Exception:
            pass

        # ── AVI Hook: ASR submission → verify operational safety concern resolved ─
        try:
            asr_obj = ASRReport.query.get(aid)
            dept_id = asr_obj.department_id if asr_obj else None
            _avi_generate(
                source_module='asr', source_record_id=aid,
                source_description=f'ASR submitted: {f.get("occurrence_description","")[:200]}',
                department_id=dept_id,
                linked_report_id=aid,
                linked_hazard_id=hid,
                operational_risk='High' if f.get('safety_effect','') in ('D','E') else 'Medium',
            )
            db.session.commit()
        except Exception:
            pass
        flash(f'✓ ASR {aid} submitted successfully. Hazard {hid} created. Proceeding to Risk Assessment.', 'success')
        return redirect(url_for('ra_wizard_start', hid=hid))
    # GET - show list if ?list=1, otherwise show form
    all_asrs = ASRReport.query.order_by(ASRReport.created_at.desc()).all()
    return render_template('reporting/asr_report.html', all_asrs=all_asrs)


@app.route('/asr/list')
@require_login
def asr_list():
    """All ASR reports with delete capability."""
    page = request.args.get('page', 1, type=int)
    pg   = ASRReport.query.order_by(ASRReport.created_at.desc()).paginate(page=page, per_page=50, error_out=False)
    asrs = pg.items
    return render_template('reporting/asr_list.html', asrs=asrs, pagination=pg)

# ─── Hazard Log ───────────────────────────────────────────────────────────────
@app.route('/hazard-log')
@require_login
def hazard_log():
    dept_f = request.args.get('dept','')
    stat_f = request.args.get('status','')
    cls_f  = request.args.get('classification','')
    page   = request.args.get('page', 1, type=int)
    from sqlalchemy.orm import subqueryload as _sql
    hazards = []
    pg = None
    try:
        db.session.rollback()          # ensure clean state before every query
        q = Hazard.query.options(
            _sql(Hazard.actions),      # eagerly load actions — prevents lazy loads
            _sql(Hazard.risks),        # eagerly load risks   — inside template
            _sql(Hazard.department),
        )
        if dept_f: q = q.filter(Hazard.department_id == int(dept_f))
        if stat_f: q = q.filter(Hazard.status == stat_f)
        if cls_f:  q = q.filter(Hazard.classification == cls_f)
        pg      = q.order_by(Hazard.created_at.desc()).paginate(page=page, per_page=50, error_out=False)
        hazards = pg.items
    except Exception as e:
        db.session.rollback()
        print(f'hazard_log query error: {e}')
        hazards = []
        pg = None
    try:
        all_departments = Department.query.order_by(Department.name).all()
    except Exception as e:
        db.session.rollback()
        print(f'hazard_log dept query error: {e}')
        all_departments = []
    # Build a minimal pagination-compatible object when query fails
    if pg is None:
        class _NullPage:
            items = []; total = 0; pages = 1; page = 1
            has_prev = False; has_next = False
            prev_num = None; next_num = None
            def iter_pages(self, **kw): return []
        pg = _NullPage()
    return render_template('hazard/hazard_log.html', hazards=hazards,
        dept_f=dept_f, stat_f=stat_f, cls_f=cls_f, pagination=pg,
        all_departments=all_departments)

@app.route('/hazard-log/<hid>')
@require_login
def hazard_detail(hid):
    h = Hazard.query.get_or_404(hid)
    # SPI link context
    spi_link_indicators, spi_link_existing = [], []
    try:
        spi_link_indicators = SPIIndicator.query.filter_by(active=True).order_by(SPIIndicator.code).all()
        if SPIEventLink is not None:
            spi_link_existing = SPIEventLink.query.filter_by(
                event_type='hazard_report', event_id=str(hid)
            ).all()
    except Exception:
        db.session.rollback()
    return render_template('hazard/hazard_detail.html', h=h,
        spi_active_indicators=spi_link_indicators,
        spi_existing_links=spi_link_existing,
        spi_link_event_type='hazard_report',
        spi_link_event_id=str(hid),
        spi_link_event_title=(h.generic_hazard or str(hid))[:100],
        spi_link_event_date=str(h.created_at.strftime('%Y-%m-%d') if h.created_at else ''),
        spi_link_severity=h.status or '',
        spi_link_dept_id=h.department_id or '',
        spi_link_category=h.classification or '',
        spi_return_url=f'/hazard-log/{hid}',
    )

@app.route('/hazard-log/<hid>/update', methods=['POST'])
@require_login
def hazard_update(hid):
    h = Hazard.query.get_or_404(hid)
    f = request.form
    h.status = f.get('status', h.status)
    h.owner  = f.get('owner', h.owner)
    h.generic_hazard   = f.get('generic_hazard', h.generic_hazard)
    h.classification   = f.get('classification', h.classification)
    h.consequences     = f.get('consequences', h.consequences)
    db.session.commit()
    flash('✓ Hazard updated.', 'success')
    return redirect(url_for('hazard_detail', hid=hid))

# ─── Risk Register ────────────────────────────────────────────────────────────
@app.route('/hazard-log/<hid>/add-risk', methods=['POST'])
@require_login
def add_risk(hid):
    f  = request.form
    li = int(f['likelihood'])
    se = f['severity']
    ri = f'{li}{se}'
    r = Risk(id=new_id('RSK'), hazard_id=hid,
             description=f['description'],
             initial_likelihood=li, initial_severity=se,
             initial_risk_index=ri, initial_tolerance=get_tolerance(ri),
             residual_likelihood=int(f['res_likelihood']) if f.get('res_likelihood') else None,
             residual_severity=f.get('res_severity') or None)
    if r.residual_likelihood and r.residual_severity:
        rri = f"{r.residual_likelihood}{r.residual_severity}"
        r.residual_risk_index   = rri
        r.residual_tolerance    = get_tolerance(rri)
    db.session.add(r)
    db.session.commit()
    # Hook 4 — SPI Intelligence linkage for new risk assessment
    try:
        haz = Hazard.query.get(hid)
        _spi_link_event(
            event_type    = 'risk_assessment',
            event_id      = r.id,
            event_title   = (haz.generic_hazard[:120] if haz and haz.generic_hazard else f'Risk {r.id}'),
            department_id = haz.department_id if haz else None,
            category      = haz.classification if haz else '',
            severity      = 'High' if r.initial_tolerance in ('Unacceptable','High') else 'Medium',
            extra_text    = f.get('description', '') + ' ' + (haz.consequences if haz else ''),
            event_date    = '',
        )
    except Exception:
        pass
    flash('✓ Risk added.', 'success')
    return redirect(url_for('hazard_detail', hid=hid))

@app.route('/risk/<rid>/add-control', methods=['POST'])
@require_login
def add_control(rid):
    risk = Risk.query.get_or_404(rid)
    f = request.form
    c = Control(id=new_id('CTL'), risk_id=rid,
                control_type=f['control_type'],
                description=f['description'],
                owner=f.get('owner',''),
                effectiveness=f.get('effectiveness',''),
                review_date=f.get('review_date',''))
    db.session.add(c)
    db.session.commit()
    flash('✓ Control measure added.', 'success')
    return redirect(url_for('hazard_detail', hid=risk.hazard_id))

# ─── Actions ──────────────────────────────────────────────────────────────────
@app.route('/actions')
@require_login
def actions():
    """
    Centralized Action Management Dashboard.
    Shows ALL actions from ALL SMS modules in one place.
    """
    check_overdue_actions()
    stat_f = request.args.get('status', '')
    pri_f  = request.args.get('priority', '')
    src_f  = request.args.get('source', '')
    dept_f = request.args.get('dept', '')
    q_f    = request.args.get('q', '').strip()

    q = Action.query
    if stat_f:  q = q.filter_by(status=stat_f)
    if pri_f:   q = q.filter_by(priority=pri_f)
    if src_f:   q = q.filter_by(source=src_f)
    if q_f:
        q = q.filter(
            Action.description.ilike(f'%{q_f}%') |
            Action.owner.ilike(f'%{q_f}%') |
            Action.id.ilike(f'%{q_f}%')
        )
    page        = request.args.get('page', 1, type=int)
    _pg         = q.order_by(Action.created_at.desc()).paginate(page=page, per_page=50, error_out=False)
    all_actions = _pg.items

    # Status counts
    counts = {
        'open':        Action.query.filter_by(status='Open').count(),
        'prog':        Action.query.filter_by(status='In Progress').count(),
        'closed':      Action.query.filter_by(status='Closed').count(),
        'overdue':     Action.query.filter_by(status='Overdue').count(),
        'total':       Action.query.count(),
        'pending_rev': Action.query.filter(
                           Action.effectiveness.is_(None),
                           Action.status=='Closed'
                       ).count(),
        'high':        Action.query.filter_by(priority='High',
                           status='Open').count(),
    }

    # Source breakdown
    from sqlalchemy import func
    sources = db.session.query(
        Action.source, func.count(Action.id)
    ).filter(Action.status != 'Closed').group_by(Action.source).all()
    source_counts = {s: n for s, n in sources if s}

    # All distinct sources for filter
    all_sources = [r[0] for r in
                   db.session.query(Action.source).distinct().all() if r[0]]

    # Department breakdown
    from sqlalchemy import func as sqlfunc
    dept_actions = db.session.query(Action.source, sqlfunc.count(Action.id))                   .filter(Action.status.in_(['Open','In Progress','Overdue']))                   .group_by(Action.source).all()

    # Pending effectiveness review
    pending_eff = Action.query.filter(
        Action.status=='Closed', Action.effectiveness.is_(None)
    ).all()

    return render_template('action/action_list.html',
        actions=all_actions, counts=counts,
        source_counts=source_counts, all_sources=sorted(all_sources),
        pending_eff=pending_eff, pagination=_pg,
        stat_f=stat_f, pri_f=pri_f, src_f=src_f, dept_f=dept_f, q_f=q_f)


@app.route('/actions/new', methods=['GET', 'POST'])
@require_login
def new_action():
    """Create a new action. Can be called from any page with pre-filled fields."""
    if request.method == 'POST':
        f = request.form
        # Handle evidence file upload
        evidence_filename = None
        if 'evidence_file' in request.files:
            ef = request.files['evidence_file']
            if ef and ef.filename and allowed_file(ef.filename):
                from werkzeug.utils import secure_filename
                evidence_filename = f"{new_id('EV')}_{secure_filename(ef.filename)}"
                ef.save(os.path.join(app.config['UPLOAD_FOLDER'], evidence_filename))

        alert_month = int(f['spi_alert_month']) if f.get('spi_alert_month') else datetime.now().month
        alert_year  = int(f['spi_alert_year'])  if f.get('spi_alert_year')  else datetime.now().year
        spi_id_val  = int(f['spi_id']) if f.get('spi_id') else None

        # Link to the persisted escalation record for accurate trigger data
        esc_id = None
        if spi_id_val and f.get('spi_trigger_rule'):
            esc = SPIEscalation.query.filter_by(
                spi_id=spi_id_val,
                trigger_month=alert_month,
                trigger_rule=f.get('spi_trigger_rule', '')
            ).first()
            if esc:
                esc_id = esc.id
                esc.status = 'Actioned'

        a = Action(
            id=new_id('ACT'),
            source=f.get('source', 'Manual'),
            hazard_id=f.get('hazard_id') or None,
            linked_ref_id=f.get('linked_ref_id', ''),
            description=f['description'],
            owner=f['owner'],
            due_date=f['due_date'],
            priority=f.get('priority', 'Medium'),
            status='Open',
            # SPI lifecycle fields
            spi_id               = spi_id_val,
            spi_alert_level      = f.get('spi_alert_level', ''),
            spi_trigger_rule     = f.get('spi_trigger_rule', ''),
            spi_alert_month      = alert_month,
            spi_alert_year       = alert_year,
            spi_escalation_id    = esc_id,
            mitigation_description = f.get('mitigation_description', ''),
            corrective_description = f.get('corrective_description', ''),
            safety_notes           = f.get('safety_notes', ''),
            assigned_by            = f.get('assigned_by', ''),
            evidence               = f.get('evidence', ''),
            evidence_filename      = evidence_filename,
            mitigation_status      = 'Pending'
        )
        db.session.add(a)
        db.session.commit()
        sync_report_status(a.hazard_id)
        db.session.commit()
        # Hook 7 — SPI Intelligence linkage for new corrective action
        try:
            _spi_link_event(
                event_type    = 'action',
                event_id      = a.id,
                event_title   = a.description[:120] if a.description else '',
                department_id = None,
                category      = a.source or '',
                severity      = a.priority or 'Medium',
                extra_text    = (a.mitigation_description or '') + ' ' + (a.corrective_description or '') + ' ' + (a.safety_notes or ''),
                event_date    = a.due_date or '',
            )
        except Exception:
            pass
        if a.owner:
            push_notify_by_name(
                a.owner,
                f'⚡ Action Assigned: {a.id}',
                f'{a.description[:80] if a.description else "A corrective action has been assigned to you."}  Due: {a.due_date or "TBD"}',
                'action_assigned', 'action', a.id)
        flash(f'✓ Action {a.id} created successfully.', 'success')
        # Return to wherever the user came from
        return_url = f.get('return_url', url_for('actions'))
        return redirect(return_url)

    # GET — show the form, pre-fill from URL params if provided
    hazards = Hazard.query.filter_by(status='Open').order_by(Hazard.created_at.desc()).all()
    pre = {
        'source':           request.args.get('source', ''),
        'hazard_id':        request.args.get('hazard_id', ''),
        'linked_ref_id':    request.args.get('linked_ref_id', ''),
        'return_url':       request.args.get('return_url', url_for('actions')),
        'spi_id':           request.args.get('spi_id', ''),
        'spi_alert_level':  request.args.get('spi_alert_level', ''),
        'spi_trigger_rule': request.args.get('spi_trigger_rule', ''),
        'description':      request.args.get('description', ''),
        'priority':         request.args.get('priority', 'Medium'),
    }
    return render_template('action/action_form.html', hazards=hazards, pre=pre)


@app.route('/actions/<aid>/update', methods=['POST'])
@require_login
def update_action(aid):
    """
    Update an action status.
    Simple rules:
    - Closing requires effectiveness to be selected.
    - If Ineffective → action is re-opened automatically.
    """
    a   = Action.query.get_or_404(aid)
    f   = request.form
    old_status    = a.status   # capture before changes for history
    new_status    = f.get('status', a.status)
    # Map friendly status to canonical
    # Store the actual workflow status (not mapped) — allows richer lifecycle
    # Only map unknown values to avoid corruption
    VALID_STATUSES = {'Open','Assigned','In Progress','Mitigation Implemented',
                      'Under Safety Review','Effectiveness Verification',
                      'Returned','Closed','Overdue'}
    if new_status not in VALID_STATUSES:
        new_status = 'In Progress'
    effectiveness = f.get('effectiveness', '')
    return_url    = f.get('return_url', url_for('actions'))

    actor = session.get('admin_name', session.get('admin_user', 'System'))

    # SoD enforcement: action owner / assigner cannot close their own action
    if new_status == 'Closed':
        sod_result = enforce_sod('Action', aid, a, 'close', actor)
        if not sod_result['allowed']:
            flash(f'⛔ Segregation of Duties violation blocked: {sod_result["reason"]}', 'error')
            return redirect(return_url)

    # Full effectiveness gate — all fields required before closure
    if new_status == 'Closed':
        gate_errors = validate_action_closure(a, f)
        if gate_errors:
            for err in gate_errors:
                flash(f'⚠ {err}', 'error')
            return redirect(return_url)

    # If ineffective → auto-reopen with full audit trail
    if new_status == 'Closed' and effectiveness == 'Ineffective':
        a.effectiveness        = effectiveness
        a.effectiveness_review = f.get('effectiveness_review', '')
        a.verified_by          = f.get('verified_by', '')
        a.verified_date        = f.get('verified_date', '')
        result = process_ineffective_action(a, actor, f.get('effectiveness_review', ''))
        db.session.commit()
        sync_report_status(a.hazard_id)
        db.session.commit()
        flash(f'⚠ {result["message"]}', 'warning')
        if result.get('srb_escalation'):
            flash('🔴 This action has been reopened 2+ times — escalation to SRB is recommended.',
                  'error')
        return redirect(return_url)

    # Normal update
    a.status                 = new_status
    a.evidence               = f.get('evidence', a.evidence)
    a.mitigation_description = f.get('mitigation_description', a.mitigation_description)
    a.corrective_description = f.get('corrective_description', a.corrective_description)
    a.safety_notes           = f.get('safety_notes', a.safety_notes)
    a.follow_up_notes        = f.get('follow_up_notes', a.follow_up_notes)
    a.mitigation_status      = f.get('mitigation_status', a.mitigation_status)
    a.assigned_by            = f.get('assigned_by', a.assigned_by)
    a.closure_by             = f.get('closure_by', a.closure_by)
    a.verified_by            = f.get('verified_by', a.verified_by)
    a.verified_date          = f.get('verified_date', a.verified_date)
    # Implementation date
    if f.get('implementation_date'):
        a.implementation_date = f.get('implementation_date')
    # Safety review fields
    if f.get('safety_review_notes'):
        a.safety_review_notes = f.get('safety_review_notes')
    if f.get('safety_reviewer'):
        a.safety_reviewer = f.get('safety_reviewer')
    if new_status == 'Under Safety Review' and not a.safety_review_date:
        a.safety_review_date = datetime.now().strftime('%Y-%m-%d')
    if new_status == 'Closed' and not a.closed_date:
        a.closed_date = datetime.now().strftime('%Y-%m-%d')
        if not a.closure_by:
            a.closure_by = f.get('closure_by', '')
    # Handle new evidence file upload on update
    if 'evidence_file' in request.files:
        ef = request.files['evidence_file']
        if ef and ef.filename and allowed_file(ef.filename):
            from werkzeug.utils import secure_filename
            fn = f"{a.id}_{secure_filename(ef.filename)}"
            ef.save(os.path.join(app.config['UPLOAD_FOLDER'], fn))
            a.evidence_filename = fn
    a.owner        = f.get('owner', a.owner)
    a.due_date     = f.get('due_date', a.due_date)
    a.priority     = f.get('priority', a.priority)
    a.description  = f.get('description', a.description)

    if new_status == 'Closed':
        a.effectiveness        = effectiveness
        a.effectiveness_review = f.get('effectiveness_review', '')
        a.closed_date          = date.today().isoformat()

    # Save SAG governance fields if provided
    if f.get('sag_member'):
        a.sag_member = f.get('sag_member')
    if f.get('department_id'):
        try: a.department_id = int(f.get('department_id'))
        except: pass
    if f.get('root_cause'):
        a.root_cause = f.get('root_cause')
    if new_status == 'Returned':
        a.rejection_notes = f.get('rejection_notes', '')
        a.reopen_count = (a.reopen_count or 0) + 1

    try:
        db.session.flush()
        log_action_history(
            a.id,
            session.get('admin_name', session.get('sag_user', 'System')),
            old_status, new_status,
            f.get('safety_review_notes') or f.get('rejection_notes') or '',
            'status'
        )
        db.session.commit()
        # When all actions closed → set hazard to Awaiting Safety Approval
        # (Safety Manager must perform final closure — NOT automatic)
        if a.hazard_id and new_status == 'Closed':
            try:
                all_acts = Action.query.filter_by(hazard_id=a.hazard_id).all()
                if all_acts and all(x.status == 'Closed' for x in all_acts):
                    haz = Hazard.query.filter_by(id=a.hazard_id).first()
                    if haz and haz.status not in ('Closed', 'Awaiting Safety Approval'):
                        haz.status = 'Awaiting Safety Approval'
                        db.session.commit()
                        flash('✓ All actions closed — Report is now Awaiting Safety Approval.', 'success')
            except Exception:
                pass
        sync_report_status(a.hazard_id)
        db.session.commit()
        # ── Issue 5: Auto-close RAMitigation when its linked action closes ─────
        if new_status == 'Closed' and a.source == 'Risk Assessment':
            try:
                mit = RAMitigation.query.filter_by(action_id=a.id).first()
                if mit and mit.status != 'Closed':
                    mit.status = 'Closed'
                    db.session.commit()
                    # Auto-close RA if ALL its mitigations are closed
                    ra = RiskAssessment.query.get(mit.assessment_id)
                    if ra:
                        all_mits = RAMitigation.query.filter_by(assessment_id=ra.id).all()
                        if all_mits and all(m.status == 'Closed' for m in all_mits):
                            ra.status = 'Closed'
                            db.session.commit()
            except Exception:
                pass

        # ── AVI Hook: Action closure → verify corrective action effectiveness ─
        if new_status == 'Closed':
            try:
                _avi_generate(
                    source_module='action', source_record_id=a.id,
                    source_description=f'Action closed: {(a.description or "")[:200]}',
                    linked_action_id=a.id,
                    linked_hazard_id=a.hazard_id,
                    operational_risk='High' if (a.priority or '') == 'High' else 'Medium',
                    override_objective=f'Verify that corrective action "{(a.description or "")[:100]}" has produced lasting operational improvement and has not reverted.',
                )
                db.session.commit()
            except Exception:
                pass
        flash('✓ Action updated.', 'success')
    except Exception as e:
        db.session.rollback()
        err_str = str(e)
        if 'column' in err_str.lower() and 'does not exist' in err_str.lower():
            flash('⚠ Database schema update required.', 'error')
        else:
            flash(f'⚠ Could not save: {err_str[:120]}', 'error')
    return redirect(return_url)


@app.route('/actions/<aid>/report')
@require_login
def action_report(aid):
    """Print-ready action report — full traceability."""
    a = Action.query.get_or_404(aid)
    # Resolve source reference details
    source_ref = None
    if a.linked_ref_id:
        from models import AuditFinding
        source_ref = AuditFinding.query.get(a.linked_ref_id)
    hazard_ref = None
    if a.hazard_id:
        from models import Hazard
        hazard_ref = Hazard.query.get(a.hazard_id)
    MONTHS = ['January','February','March','April','May','June',
              'July','August','September','October','November','December']
    return render_template('action/action_report.html',
                           a=a, source_ref=source_ref, hazard_ref=hazard_ref,
                           MONTHS=MONTHS, now=datetime.utcnow())


@app.route('/actions/dashboard')
@require_login
def action_dashboard():
    """Aviation Action Management Dashboard — centralized view of all action streams."""
    check_overdue_actions()
    from sqlalchemy import func as sqlfunc
    from datetime import date

    # Core stats
    total    = Action.query.count()
    open_c   = Action.query.filter_by(status='Open').count()
    prog_c   = Action.query.filter_by(status='In Progress').count()
    overdue_c= Action.query.filter_by(status='Overdue').count()
    closed_c = Action.query.filter_by(status='Closed').count()
    high_c   = Action.query.filter_by(status='Open', priority='High').count()
    pend_eff = Action.query.filter(Action.status=='Closed',
                                    Action.effectiveness.is_(None)).count()

    # Source breakdown (open only)
    src_rows = db.session.query(Action.source, sqlfunc.count(Action.id))               .filter(Action.status.in_(['Open','In Progress','Overdue']))               .group_by(Action.source).all()
    source_data = sorted(src_rows, key=lambda x: x[1], reverse=True)

    # Critical overdue actions
    overdue_list = Action.query.filter_by(status='Overdue')                   .order_by(Action.due_date).limit(10).all()

    # Recently closed (last 30 days)
    from datetime import timedelta
    cutoff = (date.today() - timedelta(days=30)).isoformat()
    recently_closed = Action.query.filter(
        Action.status=='Closed', Action.closed_date >= cutoff
    ).order_by(Action.closed_date.desc()).limit(8).all()

    # High priority open
    high_priority = Action.query.filter_by(status='Open', priority='High')                    .order_by(Action.due_date).limit(8).all()

    # Pending effectiveness review
    pending_review = Action.query.filter(
        Action.status=='Closed', Action.effectiveness.is_(None)
    ).order_by(Action.closed_date.desc()).limit(8).all()

    return render_template('action/action_dashboard.html',
                           total=total, open_c=open_c, prog_c=prog_c,
                           overdue_c=overdue_c, closed_c=closed_c,
                           high_c=high_c, pend_eff=pend_eff,
                           source_data=source_data,
                           overdue_list=overdue_list,
                           recently_closed=recently_closed,
                           high_priority=high_priority,
                           pending_review=pending_review,
                           now=datetime.utcnow())


@app.route('/actions/<aid>')
@require_login
def action_detail(aid):
    """Single action detail page — full source records + SAG assignment."""
    a = Action.query.get_or_404(aid)

    # Load source records based on action source + linked refs
    hazard      = Hazard.query.get(a.hazard_id) if a.hazard_id else None
    hazard_rep  = HazardReport.query.filter_by(hazard_id=a.hazard_id).first() if a.hazard_id else None
    finding     = AuditFinding.query.get(a.linked_ref_id) if a.linked_ref_id and a.source == 'Audit Finding' else None
    investigation = Investigation.query.get(a.linked_ref_id) if a.linked_ref_id and a.source == 'Investigation' else None
    spi_ind     = SPIIndicator.query.get(a.spi_id) if a.spi_id else None
    ra          = RiskAssessment.query.get(a.linked_ref_id) if a.linked_ref_id and a.source in ('Risk Assessment','RA') else None

    # SAG members list — filtered by action's department if set
    sag_q = User.query.filter(User.sag_role != None, User.sag_role != '', User.is_active == True)
    if a.department_id:
        sag_members = sag_q.filter_by(department_id=a.department_id).all()
        if not sag_members:
            sag_members = sag_q.all()  # fallback: show all if no match
    else:
        sag_members = sag_q.all()

    # Action history
    history = ActionHistory.query.filter_by(action_id=aid)                  .order_by(ActionHistory.changed_at.desc()).limit(20).all()

    # SPI link context
    spi_link_indicators, spi_link_existing = [], []
    try:
        spi_link_indicators = SPIIndicator.query.filter_by(active=True).order_by(SPIIndicator.code).all()
        if SPIEventLink is not None:
            spi_link_existing = SPIEventLink.query.filter_by(
                event_type='action', event_id=str(aid)
            ).all()
    except Exception:
        db.session.rollback()
    return render_template('action/action_detail.html',
        a=a, hazard=hazard, hazard_rep=hazard_rep,
        finding=finding, investigation=investigation,
        spi_ind=spi_ind, ra=ra,
        sag_members=sag_members, history=history,
        spi_active_indicators=spi_link_indicators,
        spi_existing_links=spi_link_existing,
        spi_link_event_type='action',
        spi_link_event_id=str(aid),
        spi_link_event_title=a.description[:100] if a.description else str(aid),
        spi_link_event_date=str(a.due_date or ''),
        spi_link_severity=a.priority or '',
        spi_link_dept_id=a.department_id or '',
        spi_link_category=a.source or '',
        spi_return_url=f'/actions/{aid}',
    )



# ─── Audits ───────────────────────────────────────────────────────────────────
# ─── Legacy /audits/* routes — redirected to new audit system ────────────────
@app.route('/audits')
@require_login
def audits():
    return redirect(url_for('audit_schedule'))

@app.route('/audits/new')
@require_login
def new_audit():
    return redirect(url_for('new_audit_schedule'))

@app.route('/audits/<aid>')
@require_login
def audit_detail(aid):
    return redirect(url_for('audit_schedule'))

@app.route('/audits/<aid>/add-finding', methods=['GET','POST'])
@require_login
def add_finding(aid):
    return redirect(url_for('audit_schedule'))

@app.route('/audits/<aid>/update', methods=['GET','POST'])
@require_login
def update_audit(aid):
    return redirect(url_for('audit_schedule'))

# ─── Investigations ───────────────────────────────────────────────────────────
@app.route('/investigations')
@require_login
def investigations():
    q_f    = request.args.get('q', '').strip()
    stat_f = request.args.get('status', '')
    cls_f  = request.args.get('classification', '')
    dept_f = request.args.get('dept', '')
    page   = request.args.get('page', 1, type=int)

    pg = all_inv = None
    try:
        db.session.rollback()
        qry = Investigation.query
        if q_f:
            qry = qry.filter(db.or_(
                Investigation.title.ilike(f'%{q_f}%'),
                Investigation.id.ilike(f'%{q_f}%'),
                Investigation.investigator.ilike(f'%{q_f}%'),
            ))
        if stat_f:
            qry = qry.filter(Investigation.status == stat_f)
        if cls_f:
            qry = qry.filter(Investigation.classification == cls_f)
        if dept_f:
            qry = qry.filter(Investigation.department_id == int(dept_f))
        pg      = qry.order_by(Investigation.created_at.desc()).paginate(page=page, per_page=50, error_out=False)
        all_inv = pg.items
    except Exception:
        db.session.rollback()
        pg = None
        all_inv = []
    try:
        depts = Department.query.order_by(Department.name).all()
    except Exception:
        db.session.rollback()
        depts = []
    return render_template('investigation/investigation_list.html',
        investigations=all_inv, pagination=pg,
        q_f=q_f, stat_f=stat_f, cls_f=cls_f, dept_f=dept_f,
        all_departments=depts)

# ICAO occurrence categories (ECCAIRS taxonomy subset)
OCCURRENCE_CATEGORIES = [
    'ARC – Abnormal Runway Contact',
    'CFIT – Controlled Flight Into Terrain',
    'EVAC – Evacuation',
    'F-NI – Fire/Smoke Not Related to Ignition',
    'F-POST – Post Impact Fire',
    'FUEL – Fuel Related',
    'GCOL – Ground Collision',
    'GTOW – Glider Towing Related',
    'ICE – Icing',
    'LOC-G – Loss of Control on Ground',
    'LOC-I – Loss of Control In-Flight',
    'LOLI – Loss of Lifting Conditions',
    'MAC – Midair Collision',
    'MED – Medical',
    'RAMP – Ground Handling',
    'RE – Runway Excursion',
    'RI-VAP – Runway Incursion – Vehicle, Aircraft or Person',
    'SCF-NP – System/Component Failure – Non-Powerplant',
    'SCF-PP – System/Component Failure – Powerplant',
    'TURB – Turbulence Encounter',
    'UIMC – Unintended Flight in IMC',
    'UNK – Unknown / Undetermined',
    'USOS – Undershoot/Overshoot',
    'WS – Wind Shear / Microburst',
    'Other',
]

LIFECYCLE_STAGES = [
    'Notified',
    'Initial Assessment',
    'Under Investigation',
    'Root Cause Analysis',
    'Recommendations',
    'Pending Closure',
    'Closed',
]

def _next_lifecycle_stage(current):
    try:
        idx = LIFECYCLE_STAGES.index(current)
        return LIFECYCLE_STAGES[idx + 1] if idx + 1 < len(LIFECYCLE_STAGES) else current
    except ValueError:
        return 'Initial Assessment'

@app.route('/investigations/new', methods=['GET','POST'])
@require_login
def new_investigation():
    if request.method == 'POST':
        f = request.form
        classification = f.get('classification', 'Incident')
        inv = Investigation(id=new_id('INV'),
            title=f['title'],
            linked_report_id=f.get('linked_report_id',''),
            hazard_id=f.get('hazard_id') or None,
            department_id=int(f['department_id']) if f.get('department_id') else None,
            date_of_occurrence=f.get('date_of_occurrence',''),
            investigator=f.get('investigator',''),
            description=f.get('description',''),
            classification=classification,
            occurrence_category=f.get('occurrence_category',''),
            phase_of_flight=f.get('phase_of_flight',''),
            aircraft_type=f.get('aircraft_type',''),
            aircraft_reg=f.get('aircraft_reg',''),
            location=f.get('location',''),
            authority_notified=bool(f.get('authority_notified')),
            notification_date=f.get('notification_date',''),
            notification_ref=f.get('notification_ref',''),
            lifecycle_stage='Notified',
            target_close_date=f.get('target_close_date',''),
            why1=f.get('why1',''), why2=f.get('why2',''),
            why3=f.get('why3',''), why4=f.get('why4',''),
            why5=f.get('why5',''),
            root_cause=f.get('root_cause',''),
            human_factors=f.get('human_factors',''),
            technical_factors=f.get('technical_factors',''),
            organizational_factors=f.get('organizational_factors',''),
            environmental_factors=f.get('environmental_factors',''),
            recommendations=f.get('recommendations',''),
            status='Open')
        db.session.add(inv)
        db.session.flush()
        # Record initial timeline event
        from models import InvestigationEvent
        evt = InvestigationEvent(
            investigation_id=inv.id,
            event_type='stage_advance',
            from_stage=None,
            to_stage='Notified',
            note='Investigation opened.',
            performed_by=session.get('username','system'))
        db.session.add(evt)
        # Auto-create action from recommendations
        if f.get('recommendations'):
            act = Action(id=new_id('ACT'), source='Investigation',
                         hazard_id=f.get('hazard_id') or None,
                         linked_ref_id=inv.id,
                         description=f['recommendations'],
                         owner=f.get('investigator',''),
                         due_date=f.get('target_close_date',''),
                         priority='High' if classification in ('Accident','Serious Incident') else 'Medium',
                         status='Open')
            db.session.add(act)
        db.session.commit()
        if inv.hazard_id:
            sync_report_status(inv.hazard_id)
            db.session.commit()
        # Hook 3 — SPI Intelligence linkage for new investigation
        try:
            _spi_link_event(
                event_type    = 'investigation',
                event_id      = inv.id,
                event_title   = inv.title[:120] if inv.title else '',
                department_id = inv.department_id,
                category      = inv.classification or '',
                severity      = 'Critical' if inv.classification in ('Accident','Serious Incident') else 'High',
                extra_text    = (inv.description or '') + ' ' + (inv.occurrence_category or ''),
                event_date    = inv.occurrence_date or '',
            )
        except Exception:
            pass
        # ICAO Annex 13 timeline enforcement — set statutory deadlines
        try:
            tl = initialize_investigation_timeline(
                inv, session.get('admin_name', session.get('admin_user', 'System')))
            db.session.add(tl)
            db.session.commit()
            if inv.classification in ('Accident', 'Serious Incident'):
                flash('⚠ Statutory timelines set: CAA verbal notification required within 72 hours. '
                      'Preliminary report due within 30 days (ICAO Annex 13).', 'warning')
        except Exception:
            pass

        flash(f'✓ Investigation {inv.id} opened.', 'success')
        return redirect(url_for('investigation_detail', iid=inv.id))
    hazards = Hazard.query.order_by(Hazard.created_at.desc()).all()
    departments = Department.query.order_by(Department.name).all()
    return render_template('investigation/investigation_form.html',
        hazards=hazards, departments=departments,
        occurrence_categories=OCCURRENCE_CATEGORIES,
        lifecycle_stages=LIFECYCLE_STAGES)

@app.route('/investigations/<iid>')
@require_login
def investigation_detail(iid):
    from models import InvestigationEvent
    inv = Investigation.query.get_or_404(iid)
    actions = Action.query.filter_by(linked_ref_id=iid).all()
    timeline = InvestigationEvent.query.filter_by(investigation_id=iid)\
        .order_by(InvestigationEvent.created_at).all()
    # SPI link context
    spi_indicators, spi_links = [], []
    try:
        spi_indicators = SPIIndicator.query.filter_by(active=True).order_by(SPIIndicator.code).all()
        if SPIEventLink is not None:
            spi_links = SPIEventLink.query.filter_by(
                event_type='investigation', event_id=str(iid)
            ).all()
    except Exception:
        db.session.rollback()
    return render_template('investigation/investigation_detail.html',
        inv=inv, actions=actions, timeline=timeline,
        lifecycle_stages=LIFECYCLE_STAGES,
        occurrence_categories=OCCURRENCE_CATEGORIES,
        departments=Department.query.order_by(Department.name).all(),
        spi_active_indicators=spi_indicators,
        spi_existing_links=spi_links,
        spi_link_event_type='investigation',
        spi_link_event_id=str(iid),
        spi_link_event_title=inv.title or str(iid),
        spi_link_event_date=str(inv.date_of_occurrence or ''),
        spi_link_severity='High' if inv.classification in ('Accident','Serious Incident') else 'Medium',
        spi_link_dept_id=inv.department_id or '',
        spi_link_category=inv.occurrence_category or '',
        spi_return_url=f'/investigations/{iid}',
    )

@app.route('/investigations/<iid>/edit', methods=['GET','POST'])
@require_login
def edit_investigation(iid):
    from models import InvestigationEvent
    inv = Investigation.query.get_or_404(iid)
    if request.method == 'POST':
        f = request.form
        inv.title                   = f.get('title', inv.title)
        inv.date_of_occurrence      = f.get('date_of_occurrence', inv.date_of_occurrence)
        inv.investigator            = f.get('investigator', inv.investigator)
        inv.description             = f.get('description', inv.description)
        inv.classification          = f.get('classification', inv.classification)
        inv.occurrence_category     = f.get('occurrence_category', inv.occurrence_category)
        inv.phase_of_flight         = f.get('phase_of_flight', inv.phase_of_flight)
        inv.aircraft_type           = f.get('aircraft_type', inv.aircraft_type)
        inv.aircraft_reg            = f.get('aircraft_reg', inv.aircraft_reg)
        inv.location                = f.get('location', inv.location)
        inv.authority_notified      = bool(f.get('authority_notified'))
        inv.notification_date       = f.get('notification_date', inv.notification_date)
        inv.notification_ref        = f.get('notification_ref', inv.notification_ref)
        inv.target_close_date       = f.get('target_close_date', inv.target_close_date)
        inv.why1                    = f.get('why1', inv.why1)
        inv.why2                    = f.get('why2', inv.why2)
        inv.why3                    = f.get('why3', inv.why3)
        inv.why4                    = f.get('why4', inv.why4)
        inv.why5                    = f.get('why5', inv.why5)
        inv.root_cause              = f.get('root_cause', inv.root_cause)
        inv.human_factors           = f.get('human_factors', inv.human_factors)
        inv.technical_factors       = f.get('technical_factors', inv.technical_factors)
        inv.organizational_factors  = f.get('organizational_factors', inv.organizational_factors)
        inv.environmental_factors   = f.get('environmental_factors', inv.environmental_factors)
        inv.recommendations         = f.get('recommendations', inv.recommendations)
        if f.get('department_id'):
            inv.department_id = int(f['department_id'])
        db.session.commit()
        flash(f'✓ Investigation {inv.id} updated.', 'success')
        return redirect(url_for('investigation_detail', iid=iid))
    hazards = Hazard.query.order_by(Hazard.created_at.desc()).all()
    departments = Department.query.order_by(Department.name).all()
    return render_template('investigation/investigation_form.html',
        inv=inv, hazards=hazards, departments=departments,
        occurrence_categories=OCCURRENCE_CATEGORIES,
        lifecycle_stages=LIFECYCLE_STAGES, edit=True)

@app.route('/investigations/<iid>/advance', methods=['POST'])
@require_login
def advance_investigation(iid):
    from models import InvestigationEvent
    inv  = Investigation.query.get_or_404(iid)
    note = request.form.get('note', '')
    prev = inv.lifecycle_stage or 'Notified'
    nxt  = _next_lifecycle_stage(prev)
    if nxt == prev:
        flash('Investigation is already at the final stage.', 'info')
        return redirect(url_for('investigation_detail', iid=iid))
    inv.lifecycle_stage = nxt
    if nxt == 'Under Investigation':
        inv.status = 'In Progress'
    elif nxt == 'Closed':
        inv.status = 'Closed'
    # Notify investigation lead if set
    if inv.lead_investigator:
        push_notify_by_name(
            inv.lead_investigator,
            f'🔍 Investigation Update: {inv.ref_number or inv.id}',
            f'Stage advanced to "{nxt}". Please review the investigation.',
            'system', 'investigation', iid)
    evt = InvestigationEvent(
        investigation_id=iid,
        event_type='stage_advance',
        from_stage=prev,
        to_stage=nxt,
        note=note,
        performed_by=session.get('username','system'))
    db.session.add(evt)
    db.session.commit()
    flash(f'✓ Investigation advanced to "{nxt}".', 'success')
    return redirect(url_for('investigation_detail', iid=iid))

@app.route('/investigations/<iid>/close', methods=['POST'])
@require_login
def close_investigation(iid):
    from models import InvestigationEvent
    inv = Investigation.query.get_or_404(iid)
    if inv.status == 'Closed':
        flash('Investigation is already closed.', 'info')
        return redirect(url_for('investigation_detail', iid=iid))
    final_findings = request.form.get('final_findings', '')
    from datetime import date
    inv.status          = 'Closed'
    inv.lifecycle_stage = 'Closed'
    inv.final_findings  = final_findings
    inv.closed_date     = date.today().isoformat()
    inv.closed_by       = session.get('username', 'unknown')
    evt = InvestigationEvent(
        investigation_id=iid,
        event_type='closure',
        from_stage=inv.lifecycle_stage,
        to_stage='Closed',
        note=final_findings or 'Investigation closed.',
        performed_by=session.get('username','system'))
    db.session.add(evt)
    db.session.commit()
    # ── AVI Hook: Investigation closure → verify recommendation effectiveness ──
    try:
        recs = (inv.recommendations or '')[:300]
        _avi_generate(
            source_module='investigation', source_record_id=inv.id,
            source_description=f'Investigation closed: {inv.title}. Recommendations: {recs}',
            department_id=inv.department_id,
            linked_investigation_id=inv.id,
            linked_hazard_id=inv.hazard_id,
            operational_risk='Critical' if (inv.classification or '') in ('Accident','Serious Incident') else 'High',
            override_objective=f'Verify that recommendations from investigation "{inv.title}" have been fully implemented and are producing measurable operational improvement.',
        )
        db.session.commit()
    except Exception:
        pass
    flash(f'✓ Investigation {inv.id} closed.', 'success')
    return redirect(url_for('investigation_detail', iid=iid))

@app.route('/investigations/<iid>/notify', methods=['POST'])
@require_login
def notify_investigation(iid):
    from models import InvestigationEvent
    from datetime import date
    inv = Investigation.query.get_or_404(iid)
    inv.authority_notified = True
    inv.notification_date  = request.form.get('notification_date', date.today().isoformat())
    inv.notification_ref   = request.form.get('notification_ref', '')
    evt = InvestigationEvent(
        investigation_id=iid,
        event_type='notification',
        from_stage=inv.lifecycle_stage,
        to_stage=inv.lifecycle_stage,
        note=f'Authority notified. Ref: {inv.notification_ref}',
        performed_by=session.get('username','system'))
    db.session.add(evt)
    db.session.commit()
    flash('✓ Regulatory notification recorded.', 'success')
    return redirect(url_for('investigation_detail', iid=iid))

# ─── MOC ─────────────────────────────────────────────────────────────────────
@app.route('/moc')
@require_login
def moc_list():
    all_moc = MOC.query.order_by(MOC.created_at.desc()).all()
    return render_template('investigation/moc_list.html', mocs=all_moc)

def _moc_auto_generate_ra(m):
    """
    Auto-generate a Hazard + RiskAssessment when an MOC requires a pre-change RA.
    Creates:
      - Hazard (source='MOC', linked_report_id=m.id) in the Hazard Log
      - RiskAssessment (status='Draft') in the Risk Register, linked to that Hazard
      - MOCHazard junction record
    Back-links m.hazard_id, m.linked_ra_id, sets m.ra_status = 'In Progress'.
    Caller must call db.session.commit() after this function returns.
    Returns (hazard, ra) on success, (None, None) on failure.
    """
    try:
        # ── 1. Hazard ────────────────────────────────────────────────────────────
        haz_id = new_id('HAZ')
        hazard = Hazard(
            id               = haz_id,
            source           = 'MOC',
            linked_report_id = m.id,
            department_id    = m.department_id,
            classification   = 'Organizational',
            generic_hazard   = f'[MOC] {(m.title or "")[:180]}',
            consequences     = (m.proposed_change or m.description or '')[:500],
            status           = 'Open',
        )
        db.session.add(hazard)

        # ── 2. RiskAssessment ────────────────────────────────────────────────────
        dept      = Department.query.get(m.department_id)
        dept_code = dept.code if dept else 'SMS'
        ctrl_num  = gen_control_number(dept_code)
        ra_id     = new_id('RA')
        today     = datetime.utcnow().strftime('%Y-%m-%d')
        ra = RiskAssessment(
            id                  = ra_id,
            control_number      = ctrl_num,
            title               = f'MOC Risk Assessment — {(m.title or "")[:150]}',
            hazard_id           = haz_id,
            department_id       = m.department_id,
            general_description = (m.current_situation or m.proposed_change or '')[:1000],
            reasons             = f'Required by Management of Change {m.moc_number}. {(m.reason_for_change or "")}',
            status              = 'Draft',
            assessment_date     = today,
        )
        db.session.add(ra)

        # ── 3. MOCHazard junction record ─────────────────────────────────────────
        mh = MOCHazard(
            moc_id               = m.id,
            hazard_description   = f'[MOC] {(m.title or "")[:200]}',
            potential_consequence= (m.proposed_change or '')[:500],
            acceptance_status    = 'Pending',
            linked_hazard_id     = haz_id,
        )
        db.session.add(mh)

        # ── 4. Back-link MOC ─────────────────────────────────────────────────────
        m.hazard_id    = haz_id
        m.linked_ra_id = ra_id        # stores RA primary key so detail page can look it up
        m.ra_status    = 'In Progress'

        # ── 5. Audit trail ───────────────────────────────────────────────────────
        db.session.add(MOCUpdate(
            moc_id      = m.id,
            update_text = (f'Auto-generated: Hazard [{haz_id}] and Risk Assessment '
                           f'[{ctrl_num} / {ra_id}] created and linked to this MOC.'),
            update_by   = 'System',
            update_type = 'Progress',
        ))
        return hazard, ra
    except Exception as e:
        db.session.rollback()
        print(f'_moc_auto_generate_ra error: {e}')
        return None, None


@app.route('/moc/new', methods=['GET','POST'])
@require_login
def new_moc():
    if request.method == 'POST':
        f = request.form
        dept_id = int(f.get('department_id', 1))
        title   = f.get('title', '').strip()
        cat     = f.get('change_category', f.get('change_type', 'Operational'))
        mid_new = new_id('MOC')
        m = MOC(
            id                  = mid_new,
            moc_number          = _moc_number(),
            title               = title,
            description         = f.get('proposed_change', f.get('description', '')),
            proposed_change     = f.get('proposed_change', ''),
            current_situation   = f.get('current_situation', ''),
            reason_for_change   = f.get('reason_for_change', ''),
            expected_benefits   = f.get('expected_benefits', ''),
            department_id       = dept_id,
            change_category     = cat,
            change_type         = cat,
            initiator           = f.get('initiator', ''),
            date_raised         = f.get('date_raised', datetime.utcnow().strftime('%Y-%m-%d')),
            planned_date        = f.get('target_completion_date', f.get('planned_date', '')),
            target_completion_date = f.get('target_completion_date', ''),
            implementation_start_date = f.get('implementation_start_date', ''),
            pre_change_risk     = f.get('pre_change_risk', ''),
            safety_impact_level = f.get('safety_impact_level', 'Low'),
            risk_assessment_required = f.get('risk_assessment_required', 'no') == 'yes',
            ra_status           = f.get('ra_status', 'Not Started'),
            ae_approval_required = f.get('safety_impact_level','Low') in ('High','Critical'),
            status              = 'Draft',
            approval_status     = 'Pending',
            implementation_status = 'Not Started',
            # Impact assessment
            impact_aircraft_ops     = 'impact_aircraft_ops' in f,
            impact_flight_crew      = 'impact_flight_crew' in f,
            impact_cabin_crew       = 'impact_cabin_crew' in f,
            impact_ground_ops       = 'impact_ground_ops' in f,
            impact_maintenance      = 'impact_maintenance' in f,
            impact_occ              = 'impact_occ' in f,
            impact_training         = 'impact_training' in f,
            impact_safety_reporting = 'impact_safety_reporting' in f,
            impact_erp              = 'impact_erp' in f,
            impact_security         = 'impact_security' in f,
            impact_regulatory       = 'impact_regulatory' in f,
            impact_contractor       = 'impact_contractor' in f,
            # Regulatory
            icao_impact                  = 'icao_impact' in f,
            iosa_impact                  = 'iosa_impact' in f,
            easa_impact                  = 'easa_impact' in f,
            national_authority_impact    = 'national_authority_impact' in f,
            company_manual_impact        = 'company_manual_impact' in f,
            regulatory_approval_required = 'regulatory_approval_required' in f,
            regulatory_approval_ref      = f.get('regulatory_approval_ref', ''),
            # Implementation needs
            training_required              = 'training_required' in f,
            documentation_update_required  = 'documentation_update_required' in f,
            sop_revision_required          = 'sop_revision_required' in f,
            erp_update_required            = 'erp_update_required' in f,
            stakeholder_summary            = f.get('stakeholder_summary', ''),
        )
        db.session.add(m)
        # Log creation
        u = MOCUpdate(moc_id=mid_new, update_text=f'MOC {m.moc_number} created.',
                      update_by=session.get('username','System'), update_type='Progress')
        db.session.add(u)
        db.session.commit()
        # Auto-generate Hazard + Risk Assessment if RA is required
        if m.risk_assessment_required:
            _moc_auto_generate_ra(m)
            try:
                db.session.commit()
                flash(f'MOC {m.moc_number} created. Hazard and Risk Assessment auto-generated and linked.', 'success')
            except Exception as _e:
                db.session.rollback()
                flash(f'MOC {m.moc_number} created (RA auto-generation failed: {_e}).', 'warning')
        else:
            flash(f'MOC {m.moc_number} created successfully.', 'success')
        return redirect(url_for('moc_detail', mid=mid_new))
    all_departments = Department.query.order_by(Department.name).all()
    return render_template('investigation/moc_form.html', m=None, edit=False, all_departments=all_departments)

@app.route('/moc/<mid>/update', methods=['POST'])
@require_login
def update_moc(mid):
    m = MOC.query.get_or_404(mid)
    f = request.form
    new_impl_status   = f.get('implementation_status', m.implementation_status)
    new_approval      = f.get('approval_status', m.approval_status)
    new_approved_by   = f.get('approved_by', m.approved_by or '').strip()
    new_pcr           = f.get('post_change_review', m.post_change_review or '').strip()

    # ── Gate: Completed requires post-change review evidence (ICAO Doc 9859 §8.3)
    if new_impl_status == 'Completed' and not new_pcr:
        flash('⚠ Cannot mark MOC as Completed without a Post-Change Review. '
              'Document the review findings before closing.', 'error')
        return redirect(url_for('moc_list'))

    # ── Gate: Approved requires approver name
    if new_approval == 'Approved' and not new_approved_by:
        flash('⚠ Approval requires an Approved By name. '
              'Record the approving authority before submitting.', 'error')
        return redirect(url_for('moc_list'))

    m.approval_status       = new_approval
    m.approved_by           = new_approved_by
    m.implementation_status = new_impl_status
    m.post_change_review    = new_pcr
    db.session.commit()
    # ── AVI Hook: MOC completed → verify change didn't introduce new hazards ──
    if new_impl_status == 'Completed':
        try:
            _avi_generate(
                source_module='moc', source_record_id=m.id,
                source_description=f'MOC completed: {(m.title or "")[:200]}. Change type: {m.change_type}',
                department_id=m.department_id,
                linked_hazard_id=m.hazard_id,
                operational_risk='High',
                override_objective=f'Verify that MOC "{(m.title or "")[:100]}" has not introduced unacceptable safety risks and post-change review findings are resolved.',
            )
            db.session.commit()
        except Exception:
            pass
    flash('✓ MOC updated.', 'success')
    return redirect(url_for('moc_list'))

# ─────────────────────────────────────────────────────────────────────────────
#  MOC AIRLINE-GRADE MODULE  (ICAO Annex 19 / Doc 9859 §7 / IOSA ISM / EASA)
# ─────────────────────────────────────────────────────────────────────────────


def _moc_number():
    """Generate JAV/MOC/YYYY/NNN reference number."""
    yr = datetime.utcnow().year
    count = MOC.query.filter(MOC.moc_number.like(f'JAV/MOC/{yr}/%')).count()
    return f'JAV/MOC/{yr}/{count+1:03d}'

def _moc_status_color(status):
    return {
        'Draft': '#6b7280', 'Under Review': '#d97706', 'Approved': '#15803d',
        'Implementing': '#1d4ed8', 'Implemented': '#0e7490',
        'Post-Implementation Review': '#7c3aed', 'Closed': '#374151',
    }.get(status, '#6b7280')

def _moc_impact_color(level):
    return {'Low':'#15803d','Medium':'#d97706','High':'#dc2626','Critical':'#7c3aed'}.get(level,'#6b7280')

@app.route('/moc/<mid>/update-ra-status', methods=['POST'])
@require_login
def moc_update_ra_status(mid):
    """Update RA Required flag and RA Status from the detail page.
       If RA is now required and no Hazard/RA are linked yet, auto-generates them.
    """
    m = MOC.query.get_or_404(mid)
    f = request.form
    ra_required_new = f.get('risk_assessment_required', 'no') == 'yes'
    m.risk_assessment_required = ra_required_new
    m.ra_status    = f.get('ra_status', m.ra_status or 'Not Started')
    manual_ra_ref  = f.get('linked_ra_id', '').strip()
    if manual_ra_ref:
        m.linked_ra_id = manual_ra_ref   # user manually entered a reference
    db.session.commit()
    # Auto-generate Hazard + RA if RA is required and nothing linked yet
    if ra_required_new and not m.linked_ra_id:
        _moc_auto_generate_ra(m)
        try:
            db.session.commit()
            flash('Risk Assessment required — Hazard and Risk Assessment auto-generated and linked.', 'success')
            return redirect(url_for('moc_detail', mid=mid) + '#linked-records')
        except Exception as _e:
            db.session.rollback()
            flash(f'RA auto-generation failed: {_e}', 'warning')
    else:
        db.session.add(MOCUpdate(
            moc_id=mid,
            update_text=f'RA Status updated to: {m.ra_status}. RA Required: {"Yes" if m.risk_assessment_required else "No"}.',
            update_by=session.get('username', 'System'),
            update_type='Progress',
        ))
        db.session.commit()
        flash('Risk Assessment status updated.', 'success')
    return redirect(url_for('moc_detail', mid=mid) + '#ra-section')

@app.route('/moc/<mid>/detail')
@require_login
def moc_detail(mid):
    m = MOC.query.get_or_404(mid)
    all_departments = Department.query.order_by(Department.name).all()
    # Load actions with a fallback so a missing table never crashes the page
    try:
        actions = Action.query.filter_by(linked_ref_id=mid).all()
    except Exception:
        db.session.rollback()
        actions = []
    # Load linked hazards, risks
    try:
        moc_hazards     = MOCHazard.query.filter_by(moc_id=mid).order_by(MOCHazard.created_at).all()
        moc_milestones  = MOCMilestone.query.filter_by(moc_id=mid).order_by(MOCMilestone.target_date).all()
        moc_updates     = MOCUpdate.query.filter_by(moc_id=mid).order_by(MOCUpdate.created_at.desc()).all()
        moc_stakeholders= MOCStakeholder.query.filter_by(moc_id=mid).order_by(MOCStakeholder.created_at).all()
    except Exception:
        db.session.rollback()
        moc_hazards = moc_milestones = moc_updates = moc_stakeholders = []
    # Load linked investigations
    try:
        linked_investigations = Investigation.query.filter_by(linked_ref_id=mid).all()
    except Exception:
        db.session.rollback()
        linked_investigations = []
    # Load linked AVIs (Audit Verification Items)
    try:
        linked_avis = AuditVerificationItem.query.filter_by(source_record_id=mid).all()
    except Exception:
        db.session.rollback()
        linked_avis = []
    # Build real linked hazards list (from moc_hazards linked_hazard_id)
    linked_hazard_ids = [mh.linked_hazard_id for mh in moc_hazards if mh.linked_hazard_id]
    # Also include m.hazard_id if set and not already in the list
    if m.hazard_id and m.hazard_id not in linked_hazard_ids:
        linked_hazard_ids.append(m.hazard_id)
    try:
        linked_hazards = Hazard.query.filter(Hazard.id.in_(linked_hazard_ids)).all() if linked_hazard_ids else []
    except Exception:
        db.session.rollback()
        linked_hazards = []
    # Load the actual RiskAssessment object for the linked_ra_id
    linked_ra = None
    if m.linked_ra_id:
        try:
            linked_ra = RiskAssessment.query.get(m.linked_ra_id)
            # Fall back: maybe linked_ra_id stores control_number instead of id
            if linked_ra is None:
                linked_ra = RiskAssessment.query.filter_by(control_number=m.linked_ra_id).first()
        except Exception:
            db.session.rollback()
            linked_ra = None
    return render_template('investigation/moc_detail.html',
                           m=m, all_departments=all_departments, actions=actions,
                           moc_hazards=moc_hazards, moc_milestones=moc_milestones,
                           moc_updates=moc_updates, moc_stakeholders=moc_stakeholders,
                           linked_hazards=linked_hazards, linked_ra=linked_ra,
                           linked_investigations=linked_investigations,
                           linked_avis=linked_avis,
                           status_color=_moc_status_color(m.status or 'Draft'),
                           impact_color=_moc_impact_color(m.safety_impact_level or 'Low'))

@app.route('/moc/<mid>/edit', methods=['GET','POST'])
@require_login
def moc_edit(mid):
    m = MOC.query.get_or_404(mid)
    if m.status not in (None, 'Draft'):
        flash('Only Draft MOCs can be edited.', 'error')
        return redirect(url_for('moc_detail', mid=mid))
    if request.method == 'POST':
        f = request.form
        m.title = f.get('title', m.title)
        m.change_category = f.get('change_category', '')
        m.change_type = f.get('change_category', m.change_type)
        m.department_id = int(f.get('department_id', m.department_id or 1))
        m.initiator = f.get('initiator', m.initiator or '')
        m.date_raised = f.get('date_raised', '')
        m.planned_date = f.get('target_completion_date', m.planned_date or '')
        m.current_situation = f.get('current_situation', '')
        m.proposed_change = f.get('proposed_change', '')
        m.description = f.get('proposed_change', m.description or '')
        m.reason_for_change = f.get('reason_for_change', '')
        m.expected_benefits = f.get('expected_benefits', '')
        m.pre_change_risk = f.get('pre_change_risk', m.pre_change_risk or '')
        for fld in ['impact_aircraft_ops','impact_flight_crew','impact_cabin_crew',
                    'impact_ground_ops','impact_maintenance','impact_occ',
                    'impact_training','impact_safety_reporting','impact_erp',
                    'impact_security','impact_regulatory','impact_contractor']:
            setattr(m, fld, fld in f)
        m.safety_impact_level = f.get('safety_impact_level', 'Low')
        m.risk_assessment_required = f.get('risk_assessment_required', 'no') == 'yes'
        m.ra_status = f.get('ra_status', m.ra_status or 'Not Started')
        for fld in ['icao_impact','iosa_impact','easa_impact',
                    'national_authority_impact','company_manual_impact',
                    'regulatory_approval_required']:
            setattr(m, fld, fld in f)
        m.regulatory_approval_ref = f.get('regulatory_approval_ref', '')
        m.regulatory_approval_date = f.get('regulatory_approval_date', '')
        m.regulatory_evidence = f.get('regulatory_evidence', '')
        m.implementation_start_date = f.get('implementation_start_date', '')
        m.target_completion_date = f.get('target_completion_date', '')
        for fld in ['training_required','documentation_update_required',
                    'sop_revision_required','erp_update_required']:
            setattr(m, fld, fld in f)
        m.stakeholder_summary = f.get('stakeholder_summary', '')
        m.ae_approval_required = 'ae_approval_required' in f
        db.session.commit()
        flash('MOC updated.', 'success')
        return redirect(url_for('moc_detail', mid=mid))
    all_departments = Department.query.order_by(Department.name).all()
    return render_template('investigation/moc_form.html', m=m, edit=True, all_departments=all_departments)

@app.route('/moc/<mid>/submit', methods=['POST'])
@require_login
def moc_submit(mid):
    m = MOC.query.get_or_404(mid)
    if not m.title or not m.initiator:
        flash('MOC must have a title and initiator before submitting.', 'error')
        return redirect(url_for('moc_detail', mid=mid))
    m.status = 'Under Review'
    m.submitted_date = datetime.utcnow().strftime('%Y-%m-%d')
    m.approval_status = 'Pending'
    db.session.commit()
    u = MOCUpdate(moc_id=mid, update_text='MOC submitted for review.',
                  update_by=session.get('username','System'), update_type='Progress')
    db.session.add(u); db.session.commit()
    flash('MOC submitted for approval review.', 'success')
    return redirect(url_for('moc_detail', mid=mid))

@app.route('/moc/<mid>/approve-dept', methods=['POST'])
@require_login
def moc_approve_dept(mid):
    m = MOC.query.get_or_404(mid)
    f = request.form
    m.dept_manager_status = f.get('decision', 'Approved')
    m.dept_manager_name = f.get('approver_name', session.get('username',''))
    m.dept_manager_date = datetime.utcnow().strftime('%Y-%m-%d')
    m.dept_manager_comments = f.get('comments', '')
    db.session.commit()
    u = MOCUpdate(moc_id=mid, update_text=f"Dept Manager: {m.dept_manager_status} — {m.dept_manager_comments or 'No comments'}",
                  update_by=m.dept_manager_name, update_type='Progress')
    db.session.add(u); db.session.commit()
    flash(f'Department Manager decision: {m.dept_manager_status}', 'success')
    return redirect(url_for('moc_detail', mid=mid))

@app.route('/moc/<mid>/approve-safety', methods=['POST'])
@require_login
def moc_approve_safety(mid):
    m = MOC.query.get_or_404(mid)
    f = request.form
    m.safety_review_status = f.get('decision', 'Approved')
    m.safety_reviewer_name = f.get('approver_name', session.get('username',''))
    m.safety_review_date = datetime.utcnow().strftime('%Y-%m-%d')
    m.safety_review_comments = f.get('comments', '')
    db.session.commit()
    u = MOCUpdate(moc_id=mid, update_text=f"Safety Review: {m.safety_review_status} — {m.safety_review_comments or 'No comments'}",
                  update_by=m.safety_reviewer_name, update_type='Progress')
    db.session.add(u); db.session.commit()
    flash(f'Safety Review decision: {m.safety_review_status}', 'success')
    return redirect(url_for('moc_detail', mid=mid))

@app.route('/moc/<mid>/approve-sm', methods=['POST'])
@require_login
def moc_approve_sm(mid):
    m = MOC.query.get_or_404(mid)
    f = request.form
    m.sm_approval_status = f.get('decision', 'Approved')
    m.sm_name = f.get('approver_name', session.get('username',''))
    m.sm_date = datetime.utcnow().strftime('%Y-%m-%d')
    m.sm_comments = f.get('comments', '')
    # ── RA Gate: RA must be Approved before final sign-off ────────────────────
    if m.risk_assessment_required and (m.ra_status or 'Not Started') != 'Approved':
        flash('⚠ Risk Assessment is required for this MOC and must reach "Approved" status '
              'before the Safety Manager can approve. Update the RA status first.', 'error')
        return redirect(url_for('moc_detail', mid=mid))
    dept_ok = m.dept_manager_status == 'Approved'
    safety_ok = m.safety_review_status == 'Approved'
    sm_ok = m.sm_approval_status == 'Approved'
    ae_ok = (not m.ae_approval_required) or (m.ae_approval_status == 'Approved')
    if dept_ok and safety_ok and sm_ok and ae_ok:
        m.status = 'Approved'; m.approval_status = 'Approved'
        m.approved_by = m.sm_name; m.approved_date = datetime.utcnow().strftime('%Y-%m-%d')
        flash('Safety Manager approved. MOC is now APPROVED.', 'success')
    else:
        flash(f'Safety Manager decision: {m.sm_approval_status}', 'success')
    db.session.commit()
    u = MOCUpdate(moc_id=mid, update_text=f"Safety Manager: {m.sm_approval_status} — {m.sm_comments or 'No comments'}",
                  update_by=m.sm_name, update_type='Progress')
    db.session.add(u); db.session.commit()
    return redirect(url_for('moc_detail', mid=mid))

@app.route('/moc/<mid>/approve-ae', methods=['POST'])
@require_login
def moc_approve_ae(mid):
    m = MOC.query.get_or_404(mid)
    f = request.form
    m.ae_approval_status = f.get('decision', 'Approved')
    m.ae_name = f.get('approver_name', session.get('username',''))
    m.ae_date = datetime.utcnow().strftime('%Y-%m-%d')
    m.ae_comments = f.get('comments', '')
    dept_ok = m.dept_manager_status == 'Approved'
    safety_ok = m.safety_review_status == 'Approved'
    sm_ok = m.sm_approval_status == 'Approved'
    ae_ok = m.ae_approval_status == 'Approved'
    if dept_ok and safety_ok and sm_ok and ae_ok:
        m.status = 'Approved'; m.approval_status = 'Approved'
        m.approved_by = m.ae_name; m.approved_date = datetime.utcnow().strftime('%Y-%m-%d')
        flash('Accountable Executive approved. MOC is now APPROVED.', 'success')
    else:
        flash(f'AE decision: {m.ae_approval_status}', 'success')
    db.session.commit()
    u = MOCUpdate(moc_id=mid, update_text=f"Accountable Executive: {m.ae_approval_status} — {m.ae_comments or 'No comments'}",
                  update_by=m.ae_name, update_type='Progress')
    db.session.add(u); db.session.commit()
    return redirect(url_for('moc_detail', mid=mid))

@app.route('/moc/<mid>/start-implementation', methods=['POST'])
@require_login
def moc_start_implementation(mid):
    m = MOC.query.get_or_404(mid)
    if m.status != 'Approved':
        flash('MOC must be Approved before implementation can begin.', 'error')
        return redirect(url_for('moc_detail', mid=mid))
    m.status = 'Implementing'; m.implementation_status = 'In Progress'
    db.session.commit()
    u = MOCUpdate(moc_id=mid, update_text='Implementation started.',
                  update_by=session.get('username','System'), update_type='Progress')
    db.session.add(u); db.session.commit()
    flash('MOC status set to Implementing.', 'success')
    return redirect(url_for('moc_detail', mid=mid))

@app.route('/moc/<mid>/mark-implemented', methods=['POST'])
@require_login
def moc_mark_implemented(mid):
    m = MOC.query.get_or_404(mid)
    m.status = 'Implemented'; m.implementation_status = 'Completed'
    m.implemented_date = datetime.utcnow().strftime('%Y-%m-%d')
    db.session.commit()
    try:
        _avi_generate(source_module='moc', source_record_id=m.id,
                      source_description=f'MOC implemented: {(m.title or "")[:200]}',
                      department_id=m.department_id, linked_hazard_id=m.hazard_id,
                      operational_risk=m.safety_impact_level or 'Medium',
                      override_objective=f'Verify change "{(m.title or "")[:100]}" fully implemented without introducing new safety risks.')
        db.session.commit()
    except Exception: pass
    u = MOCUpdate(moc_id=mid, update_text='Change marked Implemented. Post-Implementation Review required.',
                  update_by=session.get('username','System'), update_type='Progress')
    db.session.add(u); db.session.commit()
    flash('MOC marked as Implemented. Post-Implementation Review must now be completed.', 'success')
    return redirect(url_for('moc_detail', mid=mid))

@app.route('/moc/<mid>/pir', methods=['POST'])
@require_login
def moc_pir(mid):
    m = MOC.query.get_or_404(mid)
    f = request.form
    if not f.get('pir_actual_outcome'):
        flash('Actual Outcome is required for Post-Implementation Review.', 'error')
        return redirect(url_for('moc_detail', mid=mid))
    m.pir_date = f.get('pir_date', datetime.utcnow().strftime('%Y-%m-%d'))
    m.pir_reviewer = f.get('pir_reviewer', session.get('username',''))
    m.pir_actual_outcome = f.get('pir_actual_outcome', '')
    m.pir_new_hazards = f.get('pir_new_hazards', '')
    m.pir_effectiveness = f.get('pir_effectiveness', '')
    m.pir_additional_actions = f.get('pir_additional_actions', '')
    m.pir_lessons_learned = f.get('pir_lessons_learned', '')
    m.post_change_review = f.get('pir_actual_outcome', '')
    m.status = 'Post-Implementation Review'
    db.session.commit()
    u = MOCUpdate(moc_id=mid, update_text=f'PIR completed. Effectiveness: {m.pir_effectiveness or "Pending"}.',
                  update_by=m.pir_reviewer, update_type='Progress')
    db.session.add(u); db.session.commit()
    flash('Post-Implementation Review recorded.', 'success')
    return redirect(url_for('moc_detail', mid=mid))

@app.route('/moc/<mid>/close', methods=['POST'])
@require_login
def moc_close(mid):
    m = MOC.query.get_or_404(mid)
    if not m.pir_actual_outcome:
        flash('Post-Implementation Review must be completed before closing this MOC.', 'error')
        return redirect(url_for('moc_detail', mid=mid))
    m.status = 'Closed'; m.closed_date = datetime.utcnow().strftime('%Y-%m-%d')
    m.implementation_status = 'Completed'
    db.session.commit()
    u = MOCUpdate(moc_id=mid, update_text='MOC closed.',
                  update_by=session.get('username','System'), update_type='Progress')
    db.session.add(u); db.session.commit()
    flash('MOC closed successfully.', 'success')
    return redirect(url_for('moc_detail', mid=mid))

@app.route('/moc/<mid>/add-hazard', methods=['POST'])
@require_login
def moc_add_hazard(mid):
    m = MOC.query.get_or_404(mid)
    f = request.form
    hazard_desc   = f.get('hazard_description', '')
    consequence   = f.get('potential_consequence', '')
    acceptance    = f.get('acceptance_status', 'Pending')
    authority     = f.get('acceptance_authority', '')
    # RA fields belong in the Risk Assessment module — use defaults here
    # The Risk record created below is a placeholder; full RA is done separately
    il, isev, iidx, itol = (2, 'B', '2B', 'Tolerable')  # default pending full RA
    rl, rsev, ridx, rtol = (1, 'A', '1A', 'Tolerable')

    # ── 1. Create real Hazard record in the Hazard Log ───────────────────────
    haz_id = new_id('HAZ')
    real_haz = Hazard(
        id                  = haz_id,
        source              = 'MOC',
        linked_report_id    = mid,
        department_id       = m.department_id,
        classification      = 'Organizational',
        type_of_activity    = 'Management of Change',
        generic_hazard      = f'MOC: {m.title or mid}',
        specific_components = hazard_desc,
        consequences        = consequence,
        status              = 'Open',
    )
    db.session.add(real_haz)
    db.session.flush()   # get haz_id into DB before Risk FK

    # ── 2. Create Risk record in the Risk Register ───────────────────────────
    risk_id = new_id('RSK')
    real_risk = Risk(
        id                  = risk_id,
        hazard_id           = haz_id,
        description         = hazard_desc,
        initial_likelihood  = il,
        initial_severity    = isev,
        initial_risk_index  = iidx,
        initial_tolerance   = itol,
        residual_likelihood = rl,
        residual_severity   = rsev,
        residual_risk_index = ridx,
        residual_tolerance  = rtol,
    )
    db.session.add(real_risk)
    db.session.flush()

    # ── 3. Create Controls (existing + proposed) ─────────────────────────────
    existing_ctrl = f.get('existing_controls', '').strip()
    proposed_ctrl = f.get('proposed_controls', '').strip()
    if existing_ctrl:
        db.session.add(Control(
            id           = new_id('CTL'),
            risk_id      = risk_id,
            control_type = 'Preventive',
            description  = existing_ctrl,
        ))
    if proposed_ctrl:
        db.session.add(Control(
            id           = new_id('CTL'),
            risk_id      = risk_id,
            control_type = 'Preventive',
            description  = proposed_ctrl,
        ))

    # ── 4. Create MOCHazard register entry linked to the real Hazard ─────────
    mh = MOCHazard(
        moc_id               = mid,
        hazard_description   = hazard_desc,
        potential_consequence= consequence,
        # RA-specific fields (existing/proposed controls, risk matrix) belong
        # in the Risk Assessment module — left blank here for governance separation
        acceptance_status    = acceptance,
        acceptance_authority = authority,
        linked_hazard_id     = haz_id,
    )
    db.session.add(mh)

    # ── 6. Log the event ─────────────────────────────────────────────────────
    db.session.add(MOCUpdate(
        moc_id      = mid,
        update_text = f'Hazard "{hazard_desc[:60]}" added to register. '
                      f'Hazard Log: {haz_id} | Risk Register: {risk_id}.',
        update_by   = session.get('username', 'System'),
        update_type = 'Hazard',
    ))
    db.session.commit()
    flash(f'Hazard added. Entries created in Hazard Log ({haz_id}) and Risk Register ({risk_id}).', 'success')
    return redirect(url_for('moc_detail', mid=mid) + '#hazards')

def _safe_delete_hazard(hid):
    """Shared helper: safely delete a Hazard and all its children.
    Handles NOT NULL FK constraints on risk_occurrences and risk_actions.
    Called from moc_delete_hazard and delete_moc to avoid code duplication.
    """
    nullable_tables = ['hazard_reports', 'asr_reports', 'actions', 'investigations',
                       'audit_findings', 'audit_actions', 'risk_actions', 'risk_assessments']
    for tbl in nullable_tables:
        try:
            db.session.execute(
                db.text(f"UPDATE {tbl} SET hazard_id = NULL WHERE hazard_id = :hid"),
                {'hid': hid}
            )
        except Exception:
            pass
    # DELETE occurrences (NOT NULL — cannot nullify)
    RiskOccurrence.query.filter_by(hazard_id=hid).delete(synchronize_session=False)
    # Clean RiskAction per risk (NOT NULL risk_id) then delete risks
    for r in Risk.query.filter_by(hazard_id=hid).all():
        RiskAction.query.filter_by(risk_id=r.id).delete(synchronize_session=False)
        try:
            db.session.execute(
                db.text("UPDATE ra_rows SET risk_id = NULL WHERE risk_id = :rid"),
                {'rid': r.id}
            )
        except Exception:
            pass
        Control.query.filter_by(risk_id=r.id).delete(synchronize_session=False)
        db.session.delete(r)
    db.session.flush()
    haz = Hazard.query.get(hid)
    if haz:
        _avi_purge(linked_hazard_id=hid)
        db.session.delete(haz)


@app.route('/moc/hazard/<int:hid>/delete', methods=['POST'])
@require_login
def moc_delete_hazard(hid):
    mh = MOCHazard.query.get_or_404(hid)
    mid = mh.moc_id
    try:
        if mh.linked_hazard_id:
            _safe_delete_hazard(mh.linked_hazard_id)
        db.session.delete(mh)
        db.session.commit()
        flash('Hazard removed from register and Hazard Log.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'⚠ Could not remove hazard: {str(e)[:120]}', 'error')
    return redirect(url_for('moc_detail', mid=mid) + '#hazards')

@app.route('/moc/<mid>/add-milestone', methods=['POST'])
@require_login
def moc_add_milestone(mid):
    MOC.query.get_or_404(mid)
    f = request.form
    ms = MOCMilestone(moc_id=mid, description=f.get('description',''),
                      responsible_person=f.get('responsible_person',''),
                      target_date=f.get('target_date',''), status=f.get('status','Pending'),
                      notes=f.get('notes',''))
    db.session.add(ms); db.session.commit()
    flash('Milestone added.', 'success')
    return redirect(url_for('moc_detail', mid=mid) + '#milestones')

@app.route('/moc/milestone/<int:msid>/update', methods=['POST'])
@require_login
def moc_update_milestone(msid):
    ms = MOCMilestone.query.get_or_404(msid)
    mid = ms.moc_id
    ms.status             = request.form.get('status', ms.status)
    ms.description        = request.form.get('description', ms.description or '')
    ms.responsible_person = request.form.get('responsible_person', ms.responsible_person or '')
    ms.target_date        = request.form.get('target_date', ms.target_date or '')
    ms.notes              = request.form.get('notes', ms.notes or '')
    if ms.status == 'Complete' and not ms.completed_date:
        ms.completed_date = datetime.utcnow().strftime('%Y-%m-%d')
    db.session.commit()
    flash('Milestone updated.', 'success')
    return redirect(url_for('moc_detail', mid=mid) + '#milestones')

@app.route('/moc/<mid>/add-stakeholder', methods=['POST'])
@require_login
def moc_add_stakeholder(mid):
    MOC.query.get_or_404(mid)
    f = request.form
    sk = MOCStakeholder(moc_id=mid, department_name=f.get('department_name',''),
                        contact_name=f.get('contact_name',''),
                        consultation_date=f.get('consultation_date',''),
                        comments=f.get('comments',''), reviewed='reviewed' in f)
    db.session.add(sk); db.session.commit()
    flash('Stakeholder consultation recorded.', 'success')
    return redirect(url_for('moc_detail', mid=mid) + '#stakeholders')

@app.route('/moc/<mid>/add-update', methods=['POST'])
@require_login
def moc_add_update(mid):
    MOC.query.get_or_404(mid)
    f = request.form
    m = MOC.query.get_or_404(mid)
    u = MOCUpdate(moc_id=mid, update_text=f.get('update_text',''),
                  update_by=f.get('update_by', session.get('username','')),
                  update_type=f.get('update_type','Progress'))
    db.session.add(u); db.session.commit()
    if m.initiator:
        push_notify_by_name(
            m.initiator,
            f'🔄 MoC Update: {m.change_title or mid}',
            f'{u.update_type}: {(u.update_text or "")[:100]}',
            'system', 'moc', str(mid))
    flash('Update logged.', 'success')
    return redirect(url_for('moc_detail', mid=mid) + '#updates')

# ─── SPI ─────────────────────────────────────────────────────────────────────
# ═══════════════════════════════════════════════════════════════════════════════
#  SPI — Safety Performance Indicators  (ICAO Annex 19 / Doc 9859)
# ═══════════════════════════════════════════════════════════════════════════════

# ═══════════════════════════════════════════════════════════════════════════════
#  SPI ENGINE — ICAO Annex 19 §6.3 / Doc 9859 Chapter 7
#  Statistical monitoring: Mean + Standard Deviation alert thresholds
#  Baseline collection mode → automatic transition to stat mode
#  Three ICAO calculation methods: COUNT / RATE / PERCENT
# ═══════════════════════════════════════════════════════════════════════════════

import math

def _spi_calc(ind, events, exposure, total_events):
    """
    ICAO three calculation methods:
      COUNT   : SPI = events
      RATE    : SPI = (events / exposure) × 1000
      PERCENT : SPI = (events / total_events) × 100
    """
    try:
        if ind.calc_type == 'COUNT':
            return float(events)
        elif ind.calc_type == 'PERCENT':
            return round(events / total_events * 100, 2) if total_events else 0.0
        else:  # RATE (default)
            return round(events / exposure * 1000, 4) if exposure else 0.0
    except Exception:
        return 0.0


def _spi_history(ind):
    """Return list of (year, month, value) sorted chronologically for all data."""
    rows = SPIData.query.filter_by(spi_id=ind.id).filter(
        SPIData.value.isnot(None)).order_by(
        SPIData.year, SPIData.month).all()
    return [(r.year, r.month, r.value) for r in rows]


def _spi_statistics(values):
    """Calculate mean and population standard deviation."""
    if not values:
        return 0.0, 0.0
    n    = len(values)
    mean = sum(values) / n
    if n < 2:
        return mean, 0.0
    variance = sum((v - mean) ** 2 for v in values) / n
    sd = math.sqrt(variance)
    return round(mean, 4), round(sd, 4)


def _spi_thresholds(ind, all_values=None):
    """
    ICAO Statistical Monitoring thresholds — per CAAS/ICAO SPM methodology
    (Doc 9859 Ch.4, APRAST-6 Workshop):

      Alert levels for the CURRENT monitoring period are set from the
      PRECEDING period's (preceding year's) Average and Standard Deviation:
        L1 = Preceding Year Mean + 1 SD
        L2 = Preceding Year Mean + 2 SD
        L3 = Preceding Year Mean + 3 SD

      For PERCENT-type indicators (higher = better), direction is reversed:
        L1 = Preceding Year Mean - 1 SD  (falling below = worse)
        L2 = Preceding Year Mean - 2 SD
        L3 = Preceding Year Mean - 3 SD

    Period logic:
      1. Try to use the full preceding calendar year (cur_year - 1).
      2. If insufficient (<baseline_months), fall back to the oldest rolling
         12-month window available in all history.
      3. If still insufficient data, use SPT-based fixed thresholds.

    Returns: (l1, l2, l3, mean, sd, is_stat_mode)
    """
    is_pct = (ind.calc_type == 'PERCENT')
    spt    = ind.spt_target or 0
    baseline_needed = ind.baseline_months or 3

    # ── Fetch full history with year/month info ───────────────────────────────
    full_history = SPIData.query.filter_by(spi_id=ind.id).filter(
        SPIData.value.isnot(None)).order_by(SPIData.year, SPIData.month).all()

    if not full_history:
        # Zero data — pure SPT fallback
        if is_pct:
            return round(spt*0.90,2), round(spt*0.80,2), round(spt*0.70,2), 0.0, 0.0, False
        return round(spt*1.20,2), round(spt*1.40,2), round(spt*1.60,2), 0.0, 0.0, False

    # Current monitoring period = year of the latest data point
    cur_year = full_history[-1].year

    # ── Strategy 1: preceding calendar year ──────────────────────────────────
    prec_vals = [r.value for r in full_history if r.year == cur_year - 1]

    # ── Strategy 2: rolling lookback (if no preceding year or too few months) ─
    if len(prec_vals) < baseline_needed:
        # Take all data BEFORE the current year as the baseline window
        prec_vals = [r.value for r in full_history if r.year < cur_year]
        if len(prec_vals) < baseline_needed:
            # No preceding data at all — use all collected data so far
            prec_vals = [r.value for r in full_history]

    # ── Strategy 3: SPT fallback if still not enough ─────────────────────────
    if len(prec_vals) < baseline_needed:
        mean_now = sum(prec_vals)/len(prec_vals) if prec_vals else 0.0
        sd_now   = _spi_statistics(prec_vals)[1] if len(prec_vals) >= 2 else 0.0
        if is_pct:
            return round(spt*0.90,2), round(spt*0.80,2), round(spt*0.70,2), mean_now, sd_now, False
        return round(spt*1.20,2), round(spt*1.40,2), round(spt*1.60,2), mean_now, sd_now, False

    # ── Statistical mode — ICAO Mean ± SD from preceding period ──────────────
    mean, sd = _spi_statistics(prec_vals)
    if is_pct:
        l1 = max(0, round(mean - sd,     2))
        l2 = max(0, round(mean - 2 * sd, 2))
        l3 = max(0, round(mean - 3 * sd, 2))
    else:
        l1 = round(mean + sd,     2)
        l2 = round(mean + 2 * sd, 2)
        l3 = round(mean + 3 * sd, 2)
    return l1, l2, l3, mean, sd, True


def _spi_target(ind, all_values=None):
    """
    SPT = Safety Performance Target.
    ALWAYS returns ind.spt_target unchanged.
    SPT is fixed — defined by Safety Management, never auto-modified.
    
    improvement_target (shown separately on dashboard) =
      Avg × (1 − improvement_pct/100) for COUNT/RATE
      Avg × (1 + improvement_pct/100) for PERCENT
    """
    return ind.spt_target or 0.0


def _spi_improvement_target(ind, all_values=None):
    """
    ICAO improvement target — per CAAS/ICAO SPM PDF (APRAST-6):
      Target = Preceding Year Average × (1 − improvement%)
      For PERCENT type: Target = Preceding Year Average × (1 + improvement%)

    "A planned 5% reduction of the current period's average over the
     preceding period's average rate." — ICAO SPM Presentation p.9

    Preceding year = calendar year before the most recent data point.
    Falls back to SPT if insufficient preceding year data.
    """
    improvement = (ind.improvement_pct or 5.0) / 100.0

    # Fetch preceding year data
    full_history = SPIData.query.filter_by(spi_id=ind.id).filter(
        SPIData.value.isnot(None)).order_by(SPIData.year, SPIData.month).all()

    if not full_history:
        return ind.spt_target or 0.0

    cur_year  = full_history[-1].year
    prec_vals = [r.value for r in full_history if r.year == cur_year - 1]

    # Fall back to all data before current year if not a full year
    if len(prec_vals) < 3:
        prec_vals = [r.value for r in full_history if r.year < cur_year]
    if len(prec_vals) < 3:
        return ind.spt_target or 0.0

    prec_avg = sum(prec_vals) / len(prec_vals)
    if ind.calc_type == 'PERCENT':
        return round(prec_avg * (1 + improvement), 2)
    return round(prec_avg * (1 - improvement), 2)


def _spi_status(value, ind, all_values=None):
    """
    ICAO alert levels based on statistical thresholds.
    Returns (label, color, level 0-3)
    """
    l1, l2, l3, mean, sd, is_stat = _spi_thresholds(ind, all_values)
    spt    = ind.spt_target or 0
    is_pct = ind.calc_type == 'PERCENT'

    if is_pct:
        if value <= l3:   return ('🔴 CRITICAL',   '#dc2626', 3)
        elif value <= l2: return ('🟠 WARNING L2', '#ea580c', 2)
        elif value <= l1: return ('🟡 WATCH L1',   '#d97706', 1)
        elif value < spt: return ('🔵 BELOW SPT',  '#1d4ed8', 0)
        else:             return ('🟢 OK',          '#15803d', 0)
    else:
        if value >= l3:   return ('🔴 CRITICAL',   '#dc2626', 3)
        elif value >= l2: return ('🟠 WARNING L2', '#ea580c', 2)
        elif value >= l1: return ('🟡 WATCH L1',   '#d97706', 1)
        elif value > spt: return ('🔵 EXCEEDS SPT','#1d4ed8', 0)
        else:             return ('🟢 OK',          '#15803d', 0)


def _spi_trend(values_list):
    """
    Multi-point trend analysis using linear regression slope.
    Uses up to last 6 data points for a statistically meaningful direction.

    Returns:
      '↑ Worsening'  / '↓ Improving'  for COUNT/RATE (up = worse)
      '↑ Improving'  / '↓ Worsening'  labelled neutrally as ↑/↓ Increasing/Decreasing
      '→ Stable' when slope is within ±5% of mean per period

    We label direction factually (Increasing/Decreasing) so the dashboard
    can apply colour based on calc_type context.
    """
    vals = [v for v in values_list if v is not None]
    if len(vals) < 2:
        return '— Insufficient data'
    if len(vals) == 2:
        pct_change = (vals[-1] - vals[-2]) / vals[-2] if vals[-2] != 0 else 0
        if pct_change > 0.05:
            return '↑ Increasing'
        elif pct_change < -0.05:
            return '↓ Decreasing'
        return '→ Stable'

    # Linear regression slope (least squares) over the available points
    n = len(vals)
    xs = list(range(n))
    x_mean = sum(xs) / n
    y_mean = sum(vals) / n
    numerator   = sum((xs[i] - x_mean) * (vals[i] - y_mean) for i in range(n))
    denominator = sum((xs[i] - x_mean) ** 2 for i in range(n))
    slope = numerator / denominator if denominator != 0 else 0

    # Normalise slope relative to mean to get percentage-per-period
    if y_mean != 0:
        slope_pct = slope / abs(y_mean)
    else:
        slope_pct = 0

    if slope_pct > 0.03:       # > 3% increase per period
        return '↑ Increasing'
    elif slope_pct < -0.03:    # > 3% decrease per period
        return '↓ Decreasing'
    return '→ Stable'


def _spi_trigger_check(values_chronological, l1, l2, l3, is_pct=False):
    """
    ICAO alert trigger rules — detects abnormal safety trends.
    Returns triggered rule or None.
    (Thin wrapper — use _spi_trigger_detail for full month/value info.)
    """
    detail = _spi_trigger_detail(values_chronological, l1, l2, l3, is_pct)
    return detail['rule'] if detail else None


def _spi_trigger_detail(values_with_months, l1_current, l2_current, l3_current,
                        is_pct=False, spt=None):
    """
    ICAO alert trigger rules with EXACT MONTH identification.

    CORRECT ICAO LOGIC: For each month i, compute thresholds from
    months [0..i-1] (prior data only), then test month i against those
    prior thresholds. This prevents the spike itself from inflating the
    mean/SD and hiding the trigger.

    Requires a minimum of 2 prior points to compute meaningful statistics.
    Falls back to the passed-in l1/l2/l3 thresholds for the full dataset
    when insufficient prior data exists.

    values_with_months : list of (month_number, value) tuples, sorted asc.
    l1/l2/l3_current   : pre-calculated thresholds from full dataset
                         (used as fallback for the latest point).

    Returns dict or None:
      { 'rule', 'trigger_month', 'trigger_value', 'level', 'description' }
    Priority: A > B > C. Returns MOST RECENT trigger.
    """
    if not values_with_months:
        return None

    # Normalise
    if isinstance(values_with_months[0], (list, tuple)):
        pairs = [(int(x[0]), float(x[1]))
                 for x in values_with_months if x[1] is not None]
    else:
        pairs = [(i + 1, float(v))
                 for i, v in enumerate(values_with_months) if v is not None]

    if not pairs:
        return None

    months = [p[0] for p in pairs]
    vals   = [p[1] for p in pairs]
    n      = len(pairs)

    def above(v, thr):
        if thr is None:
            return False
        return (v <= thr) if is_pct else (v >= thr)

    def thresholds_for_prior(prior_vals, spt_fallback=None):
        """
        Compute L1/L2/L3 from prior values.
        - Needs >= 2 values with non-zero SD for statistical mode
        - If SD=0 or insufficient data, use SPT-based fallback:
            L1=SPT×1.2, L2=SPT×1.4, L3=SPT×1.6
        Returns (l1, l2, l3) — never None (always returns something usable)
        """
        if len(prior_vals) >= 2:
            m, s = _spi_statistics(prior_vals)
            if s > 0:
                if is_pct:
                    return (round(m - s, 4), round(m - 2*s, 4), round(m - 3*s, 4))
                return (round(m + s, 4), round(m + 2*s, 4), round(m + 3*s, 4))
        # Fall back to SPT-based thresholds
        if spt_fallback and spt_fallback > 0:
            if is_pct:
                return (spt_fallback*0.90, spt_fallback*0.80, spt_fallback*0.70)
            return (spt_fallback*1.20, spt_fallback*1.40, spt_fallback*1.60)
        # Last resort: use passed-in full-dataset thresholds
        return l1_current, l2_current, l3_current

    # For each month i, build its effective thresholds from prior data
    # Store per-month (l1, l2, l3)
    effective = []
    for i in range(n):
        prior = vals[:i]                    # everything BEFORE this month
        tl1, tl2, tl3 = thresholds_for_prior(prior, spt_fallback=spt)
        effective.append((tl1, tl2, tl3))

    # ── Rule A: any single point >= its own L3 ────────────────────────────
    rule_a_month = rule_a_val = None
    for i in range(n - 1, -1, -1):
        tl1, tl2, tl3 = effective[i]
        if tl3 is not None and above(vals[i], tl3):
            rule_a_month, rule_a_val = months[i], vals[i]
            break

    # ── Rule B: 2 consecutive >= their respective L2 ─────────────────────
    rule_b_month = rule_b_val = None
    for i in range(n - 1, 0, -1):
        _, tl2_i,  _ = effective[i]
        _, tl2_i1, _ = effective[i-1]
        if (tl2_i is not None and tl2_i1 is not None and
                above(vals[i], tl2_i) and above(vals[i-1], tl2_i1)):
            rule_b_month = months[i-1]
            rule_b_val   = vals[i-1]
            break

    # ── Rule C: 3 consecutive >= their respective L1 ─────────────────────
    rule_c_month = rule_c_val = None
    for i in range(n - 1, 1, -1):
        tl1_i,  _, _ = effective[i]
        tl1_i1, _, _ = effective[i-1]
        tl1_i2, _, _ = effective[i-2]
        if (tl1_i is not None and tl1_i1 is not None and tl1_i2 is not None and
                above(vals[i], tl1_i) and
                above(vals[i-1], tl1_i1) and
                above(vals[i-2], tl1_i2)):
            rule_c_month = months[i-2]
            rule_c_val   = vals[i-2]
            break

    # Return highest-priority result
    if rule_a_month is not None:
        _, _, tl3 = effective[months.index(rule_a_month)]
        return {
            'rule':          'A',
            'trigger_month': rule_a_month,
            'trigger_value': rule_a_val,
            'level':         'L3',
            'description':   (f'Rule A: One point ({rule_a_val:.3f}) exceeded '
                               f'L3 (Mean+3SD = {tl3:.3f})')
        }
    if rule_b_month is not None:
        idx = months.index(rule_b_month)
        _, tl2, _ = effective[idx]
        return {
            'rule':          'B',
            'trigger_month': rule_b_month,
            'trigger_value': rule_b_val,
            'level':         'L2',
            'description':   (f'Rule B: Two consecutive points exceeded '
                               f'L2 (Mean+2SD = {tl2:.3f})')
        }
    if rule_c_month is not None:
        idx = months.index(rule_c_month)
        tl1, _, _ = effective[idx]
        return {
            'rule':          'C',
            'trigger_month': rule_c_month,
            'trigger_value': rule_c_val,
            'level':         'L1',
            'description':   (f'Rule C: Three consecutive points exceeded '
                               f'L1 (Mean+1SD = {tl1:.3f})')
        }
    return None


def _spi_build_table(indicators, cur_year):
    """Build complete SPI table data for dashboard."""
    MONTHS = ['Jan','Feb','Mar','Apr','May','Jun',
              'Jul','Aug','Sep','Oct','Nov','Dec']
    table = []
    for ind in indicators:
        # Current year display data (for table cells)
        month_vals = {}
        for d in SPIData.query.filter_by(spi_id=ind.id, year=cur_year).all():
            month_vals[d.month] = d.value if d.value is not None else (d.rate or 0.0)

        # All historical values (full history — for context)
        all_history = _spi_history(ind)
        all_values  = [v for _, _, v in all_history]
        baseline_needed = ind.baseline_months or 3

        # Statistics — uses preceding year per ICAO methodology
        l1, l2, l3, mean, sd, is_stat = _spi_thresholds(ind, all_values)
        target   = _spi_target(ind, all_values)

        # Auto-update stat_mode flag in DB when statistical mode activates
        if is_stat and not ind.stat_mode:
            try:
                ind.stat_mode = True
                db.session.add(ind)
            except Exception:
                pass

        # YTD + 3M avg (current year only)
        sorted_months = sorted(month_vals)
        vals_yr  = [month_vals[m] for m in sorted_months]
        ytd      = round(sum(vals_yr) / len(vals_yr), 2) if vals_yr else 0.0
        recent_3 = vals_yr[-3:]
        avg3     = round(sum(recent_3) / len(recent_3), 2) if recent_3 else 0.0
        trend    = _spi_trend(all_values[-6:])   # use last 6 points for better trend
        latest   = vals_yr[-1] if vals_yr else (all_values[-1] if all_values else 0.0)

        # SPT is FIXED — never auto-modified
        spt_fixed = ind.spt_target or 0.0
        # Improvement target — preceding year average × (1 - improvement%)
        impr_target = _spi_improvement_target(ind)

        # Status based on latest value vs statistical thresholds
        status = _spi_status(latest, ind, all_values)

        # ── ICAO Trigger check — CROSS-YEAR AWARE ────────────────────────────
        # Build (sequence_number, value) from LAST 24 MONTHS of all history
        # so Rule B and C triggers across year boundaries are detected.
        last_24 = all_history[-24:]   # (year, month, value) tuples, chronological
        trigger_pairs = [(i + 1, v) for i, (_, _, v) in enumerate(last_24)]
        # Map seq index → real month number for escalation reporting
        seq_to_month  = {i + 1: m for i, (_, m, _) in enumerate(last_24)}
        seq_to_year   = {i + 1: y for i, (y, _, _) in enumerate(last_24)}

        trigger_detail = _spi_trigger_detail(
            trigger_pairs, l1, l2, l3, ind.calc_type == 'PERCENT',
            spt=ind.spt_target)
        trigger = trigger_detail['rule'] if trigger_detail else None

        # Resolve real calendar month/year from sequence number
        if trigger_detail:
            seq = trigger_detail['trigger_month']
            trigger_detail['trigger_month'] = seq_to_month.get(seq, seq)
            trigger_detail['trigger_year']  = seq_to_year.get(seq, cur_year)

        # Extract exact escalation info AND persist a record
        if trigger_detail:
            escalation_month  = trigger_detail['trigger_month']
            escalation_year   = cur_year
            escalation_value  = trigger_detail['trigger_value']
            escalation_level  = trigger_detail['level']
            escalation_desc   = trigger_detail['description']
            # Compute threshold that was crossed
            esc_thr_map       = {'L3': l3, 'L2': l2, 'L1': l1}
            escalation_threshold = esc_thr_map.get(escalation_level, l1)
            escalation_diff   = (round(escalation_value - escalation_threshold, 4)
                                 if escalation_threshold else None)
            # Persist escalation record (idempotent)
            try:
                _spi_record_escalation(ind, trigger_detail, l1, l2, l3, mean, sd)
                db.session.commit()
            except Exception:
                db.session.rollback()
        else:
            escalation_month     = None
            escalation_year      = cur_year
            escalation_value     = None
            escalation_level     = None
            escalation_desc      = None
            escalation_threshold = None
            escalation_diff      = None

        # Baseline progress
        months_collected = len(all_values)
        baseline_pct     = min(100, int(months_collected / baseline_needed * 100))

        # Dept info
        dept_codes, dept_names = [], []
        for did in (ind.department_ids or '').split(','):
            try:
                d = Department.query.get(int(did.strip()))
                if d:
                    dept_codes.append(d.code)
                    dept_names.append(d.name)
            except Exception:
                pass

        last_month = max(month_vals.keys()) if month_vals else None

        # Linked actions for this SPI indicator
        spi_actions = Action.query.filter_by(spi_id=ind.id).order_by(
            Action.created_at.desc()).all()

        table.append(dict(
            spi_actions=spi_actions,
            ind=ind,
            month_vals=month_vals,
            ytd=ytd, avg3=avg3, trend=trend,
            status=status, trigger=trigger,
            trigger_detail=trigger_detail,
            # Exact escalation — the REAL month the threshold was crossed
            escalation_month=escalation_month,
            escalation_year=escalation_year,
            escalation_value=escalation_value,
            escalation_level=escalation_level,
            escalation_desc=escalation_desc,
            escalation_threshold=escalation_threshold,
            escalation_diff=escalation_diff,
            latest_val=latest,
            l1=l1, l2=l2, l3=l3,
            mean=mean, sd=round(sd, 4),
            spt=spt_fixed,
            impr_target=impr_target,
            target=spt_fixed,
            is_stat=is_stat,
            all_values=all_values,
            months_collected=months_collected,
            baseline_needed=baseline_needed,
            baseline_pct=baseline_pct,
            depts=', '.join(dept_codes),
            dept_names=', '.join(dept_names),
            last_month=last_month,
        ))
    return table, MONTHS


def _spi_record_escalation(ind, trigger_detail, l1, l2, l3, mean, sd):
    """
    Persist an SPIEscalation record when a trigger is detected.
    Idempotent — checks if this trigger month/rule/indicator already recorded.
    Returns the SPIEscalation instance (new or existing).
    """
    if not trigger_detail:
        return None

    rule  = trigger_detail['rule']
    month = trigger_detail['trigger_month']
    level = trigger_detail['level']
    value = trigger_detail['trigger_value']
    desc  = trigger_detail['description']

    # Which threshold was exceeded?
    thr_map = {'L3': l3, 'L2': l2, 'L1': l1}
    threshold = thr_map.get(level, l1)

    # Idempotency: don't duplicate for same indicator+month+rule
    existing = SPIEscalation.query.filter_by(
        spi_id=ind.id,
        trigger_month=month,
        trigger_rule=rule,
        alert_level=level
    ).first()
    if existing:
        return existing

    esc = SPIEscalation(
        spi_id          = ind.id,
        trigger_month   = month,
        trigger_year    = datetime.now().year,
        trigger_rule    = rule,
        alert_level     = level,
        spi_value       = round(value, 4),
        threshold_value = round(threshold, 4) if threshold else None,
        mean_value      = round(mean, 4),
        sd_value        = round(sd, 4),
        description     = desc,
        status          = 'Open',
    )
    db.session.add(esc)
    db.session.flush()   # get ID without committing
    # ── AVI Hook: SPI exceedance → verify operational corrective actions ───────
    try:
        risk_map = {'L3': 'Critical', 'L2': 'High', 'L1': 'Medium'}
        dept_ids = [int(x) for x in (ind.department_ids or '').split(',') if x.strip().isdigit()]
        dept_id  = dept_ids[0] if dept_ids else None
        _avi_generate(
            source_module='spi', source_record_id=f'{ind.id}-{month}-{rule}',
            source_description=f'SPI {ind.code} ({ind.name}) exceeded {level} threshold. Rule {rule} triggered in month {month}. Value: {round(value,4)}',
            department_id=dept_id,
            linked_spi_id=ind.id,
            operational_risk=risk_map.get(level, 'High'),
            override_objective=f'Verify that operational actions taken after {ind.code} exceeded {level} have restored the Safety Performance Indicator to within the SPT target.',
        )
    except Exception:
        pass
    return esc



# ============================================================================
#  SPI INTELLIGENCE LINKAGE ENGINE
#  ICAO Annex 19 s.6.3 / Doc 9859 Chapter 7 -- Safety Performance Monitoring
#  Additive architecture -- NEVER modifies existing SPI calculations.
# ============================================================================

# Aviation occurrence taxonomy: topic -> keyword phrases
_SPI_TAXONOMY = {
    'runway':     ['runway excursion','runway incursion','overrun','undershoot',
                   'overshoot','usos','re ','ri-vap','landing roll','rejected takeoff'],
    'approach':   ['approach','unstable approach','cfit','loc-i','controlled flight',
                   'terrain','go-around','missed approach','glide','ils','gpws'],
    'ground_ops': ['fod','foreign object','ground collision','gcol','pushback','towing',
                   'ground handling','ramp','taxiway','apron','wing strike','ground damage'],
    'technical':  ['engine failure','engine','scf-pp','scf-np','component failure',
                   'hydraulic','electrical','avionics','pressurisation','apu','bird strike'],
    'turbulence': ['turbulence','turb ','wake turbulence','wind shear','ws ','microburst',
                   'icing','ice accumulation','uimc'],
    'fatigue':    ['fatigue','rest period','duty','fdt','fdp','hours flown','sleep','roster'],
    'fire_smoke': ['fire','smoke','f-ni','f-post','fumes','odour','odor','vapor','fume event'],
    'fuel':       ['fuel','fuel leak','fuel exhaustion','fuel contamination','defuelling','misfuel'],
    'airspace':   ['airspace','atc','separation','tcas','mac ','midair','collision avoidance','loss of separation'],
    'maintenance':['maintenance','mx ','airworthiness','service bulletin','inspection','defect'],
    'training':   ['training','simulator','recurrent','line check','proficiency','currency','competency'],
    'security':   ['security','dangerous goods','dg ','hazmat','lithium','unauthorized'],
    'medical':    ['medical','incapacitation','illness','injury','first aid'],
    'sms_general':['safety report','hazard report','near miss','occurrence','safety event'],
}

_SEVERITY_MAP = {
    'asr':            'High',
    'investigation':  'Critical',
    'audit_finding':  'Medium',
    'risk_assessment':'High',
    'erp_activation': 'Critical',
    'hazard_report':  'Medium',
    'action':         'Low',
    'safety_promo':   'Low',
}


def _spi_detect_topics(text):
    """Return set of taxonomy topic keys matching free text."""
    text_lower = (text or '').lower()
    hits = set()
    for topic, keywords in _SPI_TAXONOMY.items():
        for kw in keywords:
            if kw in text_lower:
                hits.add(topic)
                break
    return hits


def _spi_link_event(event_type, event_id, event_title,
                    department_id, category='', severity='',
                    extra_text='', event_date=None):
    """
    Intelligent SPI linkage engine.
    Matches an operational event to every relevant active SPI indicator
    and creates SPIEventLink records. Purely additive -- never touches
    SPIData, SPIIndicator values, or any existing calculation.
    Always safe to call from try/except.
    """
    now = datetime.utcnow()
    event_date = event_date or now.strftime('%Y-%m-%d')
    severity = severity or _SEVERITY_MAP.get(event_type, 'Medium')

    corpus = ' '.join([str(category or ''), str(event_title or ''), str(extra_text or '')])
    detected_topics = _spi_detect_topics(corpus)
    cat_lower = (category or '').lower()
    linked_spi_ids = set()

    try:
        indicators = SPIIndicator.query.filter_by(active=True).all()
    except Exception:
        return []

    for ind in indicators:
        reasons = []

        # 1. Direct source-type + auto_category match
        if ind.auto_source and ind.auto_source == event_type:
            if not ind.auto_category or ind.auto_category.lower() in cat_lower:
                reasons.append('source:' + event_type)

        # 2. Taxonomy topic overlap between event and indicator
        ind_corpus = ' '.join([(ind.name or ''), (ind.description or ''),
                                (ind.category or '')]).lower()
        ind_topics = _spi_detect_topics(ind_corpus)
        overlap = detected_topics & ind_topics
        if overlap:
            reasons.append('topic:' + ','.join(sorted(overlap)))

        # 3. Direct category string match
        if ind.category and ind.category.lower() in cat_lower:
            reasons.append('category:' + ind.category)
        if ind.auto_category and ind.auto_category.lower() in cat_lower:
            reasons.append('auto_cat:' + ind.auto_category)

        if not reasons:
            continue

        # 4. Department filter
        dept_ids = [x.strip() for x in (ind.department_ids or '').split(',') if x.strip()]
        if dept_ids and str(department_id) not in dept_ids:
            continue

        if ind.id in linked_spi_ids:
            continue

        # Deduplication
        try:
            existing = SPIEventLink.query.filter_by(
                spi_id=ind.id, event_type=event_type, event_id=str(event_id)
            ).first()
        except Exception:
            existing = None
        if existing:
            continue

        link = SPIEventLink(
            spi_id=ind.id,
            event_type=event_type,
            event_id=str(event_id),
            event_title=str(event_title or '')[:200],
            event_date=event_date,
            department_id=department_id,
            category=str(category or '')[:100],
            severity=severity,
            match_reason=('; '.join(reasons))[:300],
        )
        db.session.add(link)
        linked_spi_ids.add(ind.id)

    try:
        if linked_spi_ids:
            db.session.commit()
    except Exception:
        db.session.rollback()

    return list(linked_spi_ids)


def _spi_recurrence_analysis(spi_id, lookback_months=6):
    """Detect recurring patterns for a SPI indicator. Read-only analytics."""
    from collections import Counter
    from datetime import date, timedelta
    try:
        cutoff = (date.today() - timedelta(days=lookback_months * 30)).isoformat()
        links  = SPIEventLink.query.filter(
            SPIEventLink.spi_id == spi_id,
            SPIEventLink.event_date >= cutoff).all()
    except Exception:
        return {'total': 0, 'category_counts': [], 'type_counts': [],
                'dept_counts': [], 'severity_counts': {},
                'recurring_categories': [], 'has_recurrence': False}

    cat_counts  = Counter(l.category for l in links if l.category)
    type_counts = Counter(l.event_type for l in links)
    dept_counts = Counter(l.department_id for l in links if l.department_id)
    sev_counts  = Counter(l.severity for l in links if l.severity)
    recurring_cats = [(c, n) for c, n in cat_counts.most_common(5) if n >= 2]
    return {
        'total':                len(links),
        'category_counts':      cat_counts.most_common(8),
        'type_counts':          type_counts.most_common(8),
        'dept_counts':          dept_counts.most_common(5),
        'severity_counts':      dict(sev_counts),
        'recurring_categories': recurring_cats,
        'has_recurrence':       len(recurring_cats) > 0,
    }


def _spi_intelligence_summary():
    """System-wide intelligence snapshot. Read-only."""
    from collections import Counter, defaultdict
    from datetime import date, timedelta
    try:
        cutoff_90 = (date.today() - timedelta(days=90)).isoformat()
        cutoff_30 = (date.today() - timedelta(days=30)).isoformat()
        all_links = SPIEventLink.query.filter(
            SPIEventLink.event_date >= cutoff_90).all()
    except Exception:
        return {'total_links_90d': 0, 'type_counts': [], 'dept_counts': [],
                'critical_30d': 0, 'srb_items': [], 'top_spis': []}

    spi_counts   = Counter(l.spi_id for l in all_links)
    type_counts  = Counter(l.event_type for l in all_links)
    dept_counts  = Counter(l.department_id for l in all_links if l.department_id)
    cat_by_spi   = defaultdict(Counter)
    for l in all_links:
        if l.category:
            cat_by_spi[l.spi_id][l.category] += 1
    critical_30  = sum(1 for l in all_links
                       if l.severity in ('Critical', 'High') and l.event_date >= cutoff_30)

    # SRB feed: SPIs currently exceeding SPT
    srb_items = []
    try:
        for ind in SPIIndicator.query.filter_by(active=True).all():
            recent = SPIData.query.filter_by(spi_id=ind.id).order_by(
                SPIData.year.desc(), SPIData.month.desc()).first()
            # Use .value (calculated field); .rate is legacy compat only
            recent_val = recent.value if recent and recent.value is not None else None
            if recent_val is not None and ind.spt_target:
                if ind.calc_type == 'PERCENT':
                    exceeded = recent_val < ind.spt_target
                else:
                    exceeded = recent_val > ind.spt_target
                if exceeded:
                    rec = _spi_recurrence_analysis(ind.id, lookback_months=3)
                    srb_items.append({
                        'ind':         ind,
                        'value':       recent_val,
                        'spt':         ind.spt_target,
                        'event_count': spi_counts.get(ind.id, 0),
                        'recurring':   rec['has_recurrence'],
                        'top_cats':    rec['recurring_categories'],
                    })
    except Exception:
        pass

    # Top SPIs by linked event count (90d)
    top_spis = []
    for spi_id, cnt in spi_counts.most_common(10):
        try:
            ind = SPIIndicator.query.get(spi_id)
            if ind:
                top_spis.append({'ind': ind, 'link_count': cnt,
                                 'cats': cat_by_spi[spi_id].most_common(3)})
        except Exception:
            pass

    return {
        'total_links_90d': len(all_links),
        'type_counts':     type_counts.most_common(),
        'dept_counts':     dept_counts.most_common(8),
        'critical_30d':    critical_30,
        'srb_items':       srb_items,
        'top_spis':        top_spis,
    }

def spi_auto_update(source_type, department_id, category, year, month,
                    report_id=None):
    """
    Auto-match a submitted report to SPIs and increment event count.
    
    DEDUPLICATION: Uses report_id to track which reports have already
    been counted. If report_id is provided and already in entry.notes
    (as a comma-separated log), the report is NOT counted again.
    
    source_type : 'hazard_report' | 'asr' | 'audit'
    report_id   : unique ID of the source record (e.g. 'HR-SMS-01')
    """
    matched = SPIIndicator.query.filter_by(
        auto_source=source_type, active=True).all()

    for ind in matched:
        # Category match (empty auto_category matches all)
        if ind.auto_category:
            if ind.auto_category.lower() not in (category or '').lower():
                continue
        # Department match
        dept_ids = [x.strip() for x in (ind.department_ids or '').split(',')]
        if str(department_id) not in dept_ids and '' not in dept_ids:
            continue

        entry = SPIData.query.filter_by(
            spi_id=ind.id, year=year, month=month).first()

        if entry:
            # DEDUPLICATION: check if this report was already counted
            already_counted = (
                report_id and
                entry.notes and
                report_id in entry.notes.split(',')
            )
            if already_counted:
                continue  # skip — already counted this report
            entry.events = (entry.events or 0) + 1
            # Log report_id to prevent future duplicate
            if report_id:
                existing = entry.notes or ''
                ids = [x for x in existing.split(',') if x]
                ids.append(report_id)
                entry.notes = ','.join(ids[-50:])  # keep last 50 IDs max
        else:
            entry = SPIData(
                spi_id=ind.id, year=year, month=month,
                events=1, exposure=1, total_events=1,
                source='auto',
                notes=report_id or ''
            )
            db.session.add(entry)

        # Recalculate SPI value
        entry.value = _spi_calc(
            ind, entry.events,
            entry.exposure or 1,
            entry.total_events or entry.events
        )
        entry.rate = entry.value  # legacy compat

        # Auto-transition to stat mode when baseline complete
        all_vals = [r.value for r in SPIData.query.filter_by(spi_id=ind.id).all()
                    if r.value is not None]
        if len(all_vals) >= (ind.baseline_months or 3) and not ind.stat_mode:
            ind.stat_mode = True

    db.session.commit()


@app.route('/spi', methods=['GET','POST'])
@require_login
def spi():
    """SPI Dashboard — ICAO Annex 19 §6.3 / Doc 9859 Chapter 7."""
    cur_year = datetime.now().year
    dept_f   = request.args.get('dept', '')
    # Year filter — default to current year, but user can select any year with data
    try:
        sel_year = int(request.args.get('year', cur_year))
    except (ValueError, TypeError):
        sel_year = cur_year

    if request.method == 'POST':
        f        = request.form
        ind      = SPIIndicator.query.get_or_404(int(f['spi_id']))
        year     = int(f.get('year', sel_year))
        month    = int(f['month'])
        events   = int(f.get('events', 0))
        exposure = float(f.get('exposure', 1) or 1)
        total_ev = int(f.get('total_events', events) or events)

        entry = SPIData.query.filter_by(
            spi_id=ind.id, year=year, month=month).first()
        if entry:
            entry.events      = events
            entry.exposure    = exposure
            entry.total_events = total_ev
        else:
            entry = SPIData(spi_id=ind.id, year=year, month=month,
                            events=events, exposure=exposure,
                            total_events=total_ev, source='manual')
            db.session.add(entry)

        entry.value   = _spi_calc(ind, events, exposure, total_ev)
        entry.rate    = entry.value
        entry.flights = int(exposure)
        db.session.flush()

        # Auto stat-mode transition
        all_vals = [r.value for r in SPIData.query.filter_by(spi_id=ind.id).all()
                    if r.value is not None]
        if len(all_vals) >= (ind.baseline_months or 3) and not ind.stat_mode:
            ind.stat_mode = True

        db.session.commit()
        flash(f'✓ {ind.code} logged: {month}/{year} = {entry.value:.3f}', 'success')
        return redirect(url_for('spi', dept=dept_f, year=sel_year))

    indicators = SPIIndicator.query.filter_by(active=True).all()
    if dept_f:
        indicators = [i for i in indicators
                      if dept_f in (i.department_ids or '').split(',')]

    # Build table for the SELECTED year (for monthly display columns)
    table, MONTHS = _spi_build_table(indicators, sel_year)

    # Gather all years that have SPI data recorded, then always include the
    # last 3 calendar years so the selector is navigable even when a year
    # has no entries yet (user can switch there to log historical data).
    try:
        year_rows = db.session.execute(
            text("SELECT DISTINCT year FROM spi_data ORDER BY year DESC")
        ).fetchall()
        available_years = [r[0] for r in year_rows if r[0]]
    except Exception:
        available_years = []
    for y in [cur_year, cur_year - 1, cur_year - 2]:
        if y not in available_years:
            available_years.append(y)
    available_years = sorted(set(available_years), reverse=True)

    critical = sum(1 for r in table if r['status'][2] >= 3)
    warning  = sum(1 for r in table if r['status'][2] == 2)
    watch    = sum(1 for r in table if r['status'][2] == 1)
    ok_count = sum(1 for r in table
                   if r['status'][2] == 0 and r['latest_val'] <= (r['ind'].spt_target or 999))

    # Any triggered ICAO rules?
    triggered = [r for r in table if r['trigger']]

    return render_template('spi/spi_dashboard.html',
        table=table, MONTHS=MONTHS,
        indicators=SPIIndicator.query.filter_by(active=True).all(),
        dept_f=dept_f, cur_year=cur_year, sel_year=sel_year,
        available_years=available_years,
        critical=critical, warning=warning, watch=watch, ok_count=ok_count,
        triggered=triggered,
        enumerate=enumerate)


@app.route('/spi/actions')
@require_login
def spi_actions_list():
    """List of all SPI-linked actions with full escalation traceability."""
    status_f = request.args.get('status', '')
    dept_f   = request.args.get('dept', '')
    level_f  = request.args.get('level', '')

    q = Action.query.filter(Action.spi_id.isnot(None))
    if status_f: q = q.filter_by(status=status_f)
    if level_f:  q = q.filter_by(spi_alert_level=level_f)
    actions = q.order_by(Action.spi_alert_year.desc(),
                         Action.spi_alert_month.desc()).all()

    # For each action, load or recompute escalation data
    action_data = []
    for a in actions:
        esc = None
        if a.spi_escalation_id:
            esc = SPIEscalation.query.get(a.spi_escalation_id)
        if not esc and a.spi_id and a.spi_alert_month:
            esc = SPIEscalation.query.filter_by(
                spi_id=a.spi_id,
                trigger_month=a.spi_alert_month
            ).first()
        action_data.append({'action': a, 'esc': esc})

    return render_template('spi/spi_actions_list.html',
                           action_data=action_data,
                           actions=actions,
                           status_f=status_f,
                           dept_f=dept_f, level_f=level_f)

@app.route('/spi/escalation/<int:esc_id>')
@require_login
def spi_escalation_detail(esc_id):
    """Escalation event detail — the source record for SPI actions."""
    esc = SPIEscalation.query.get_or_404(esc_id)
    ind = SPIIndicator.query.get(esc.spi_id) if esc.spi_id else None
    actions = Action.query.filter_by(spi_id=esc.spi_id,
                                     spi_alert_month=esc.trigger_month,
                                     spi_alert_year=esc.trigger_year).all()
    MONTHS = ['January','February','March','April','May','June',
              'July','August','September','October','November','December']
    return render_template('spi/spi_escalation_detail.html',
                           esc=esc, ind=ind, actions=actions,
                           MONTHS=MONTHS, now=datetime.utcnow())

@app.route('/spi/action-report/<action_id>')
@require_login
def spi_action_report(action_id):
    """
    Print-ready SPI Alert Mitigation Report.

    Escalation data priority:
      1. Linked SPIEscalation record (most accurate — created when dashboard loaded)
      2. Recompute from SPI history using stored spi_alert_month (reliable fallback)
      3. Action's own spi_alert_* fields (last resort)

    This ensures the report ALWAYS shows the real trigger month,
    not the action creation date.
    """
    a   = Action.query.filter_by(id=action_id).first_or_404()
    ind = SPIIndicator.query.get(a.spi_id) if a.spi_id else None

    MONTHS = ['January','February','March','April','May','June',
              'July','August','September','October','November','December']

    # ── Step 1: try linked escalation record ─────────────────────────────
    esc = None
    if a.spi_escalation_id:
        esc = SPIEscalation.query.get(a.spi_escalation_id)

    # ── Step 2: search existing escalation records ────────────────────────
    if not esc and a.spi_id and a.spi_alert_month:
        esc = SPIEscalation.query.filter_by(
            spi_id       = a.spi_id,
            trigger_month= a.spi_alert_month,
            trigger_rule = a.spi_trigger_rule or ''
        ).first()
        if not esc:
            # also try without rule filter (broader match)
            esc = SPIEscalation.query.filter_by(
                spi_id        = a.spi_id,
                trigger_month = a.spi_alert_month
            ).first()

    # ── Step 3: recompute from live SPI data ──────────────────────────────
    # This is the guaranteed fallback — uses the stored trigger month and
    # recomputes what the SPI value and thresholds were at that point.
    esc_computed = None
    if not esc and ind and a.spi_alert_month:
        try:
            history = _spi_history(ind)
            all_vals = [v for _, _, v in history]
            l1, l2, l3, mean, sd, is_stat = _spi_thresholds(ind, all_vals)

            # Find the SPI value for the stored trigger month
            from models import SPIData
            dp = SPIData.query.filter_by(
                spi_id = ind.id,
                year   = a.spi_alert_year or datetime.now().year,
                month  = a.spi_alert_month
            ).first()
            spi_val_at_trigger = dp.value if dp and dp.value else None

            # Which threshold was triggered
            level = a.spi_alert_level or 'L1'
            thr_map  = {'L3': l3, 'L2': l2, 'L1': l1}
            thr_val  = thr_map.get(level)
            diff_val = round(spi_val_at_trigger - thr_val, 4)                        if spi_val_at_trigger and thr_val else None
            rule_desc = {
                'A': f'Rule A: One point exceeded L3 (Mean+3SD = {l3:.4f})',
                'B': f'Rule B: Two consecutive exceeded L2 (Mean+2SD = {l2:.4f})',
                'C': f'Rule C: Three consecutive exceeded L1 (Mean+1SD = {l1:.4f})',
            }.get(a.spi_trigger_rule or '', '')

            esc_computed = {
                'trigger_month'  : a.spi_alert_month,
                'trigger_year'   : a.spi_alert_year or datetime.now().year,
                'trigger_rule'   : a.spi_trigger_rule,
                'alert_level'    : level,
                'spi_value'      : spi_val_at_trigger,
                'threshold_value': round(thr_val, 4) if thr_val else None,
                'mean_value'     : round(mean, 4),
                'sd_value'       : round(sd, 4),
                'description'    : rule_desc,
                'diff_value'     : diff_val,
            }
        except Exception as ex:
            esc_computed = None

    return render_template('spi/spi_action_report.html',
                           a=a, ind=ind, esc=esc,
                           esc_computed=esc_computed,
                           now=datetime.utcnow(), MONTHS=MONTHS)

@app.route('/spi/indicators', methods=['GET','POST'])
@require_login
def spi_indicators():
    """Manage SPI indicator definitions."""
    if request.method == 'POST':
        f        = request.form
        dept_ids = ','.join(request.form.getlist('department_ids'))
        ind = SPIIndicator(
            code            = f['code'].upper().strip(),
            name            = f['name'],
            department_ids  = dept_ids,
            category        = f.get('category',''),
            description     = f.get('description',''),
            calc_type       = f.get('calc_type','RATE'),
            exposure_type   = f.get('exposure_type','Flights'),
            unit            = f.get('unit',''),
            frequency       = f.get('frequency','Monthly'),
            spt_target      = float(f['spt_target']) if f.get('spt_target') else None,
            improvement_pct = float(f['improvement_pct']) if f.get('improvement_pct') else 5.0,
            baseline_months = int(f['baseline_months']) if f.get('baseline_months') else 3,
            alert_l1        = float(f['alert_l1']) if f.get('alert_l1') else None,
            alert_l2        = float(f['alert_l2']) if f.get('alert_l2') else None,
            alert_l3        = float(f['alert_l3']) if f.get('alert_l3') else None,
            auto_source     = f.get('auto_source','manual'),
            auto_category   = f.get('auto_category',''),
            active          = True,
            stat_mode       = False
        )
        db.session.add(ind)
        db.session.commit()
        flash(f'✓ SPI Indicator {ind.code} created. Collecting baseline ({ind.baseline_months} months needed).', 'success')
        return redirect(url_for('spi_indicators'))

    indicators = SPIIndicator.query.order_by(SPIIndicator.code).all()
    return render_template('spi/spi_indicators.html', indicators=indicators)


@app.route('/spi/indicators/<int:iid>/delete', methods=['POST'])
@require_login
def spi_delete_indicator(iid):
    ind = SPIIndicator.query.get_or_404(iid)
    _avi_purge(spi_indicator_id=iid)
    db.session.delete(ind)
    db.session.commit()
    flash(f'✓ Indicator {ind.code} deleted.', 'success')
    return redirect(url_for('spi_indicators'))


@app.route('/spi/evidence/<filename>')
@require_login
def spi_evidence_file(filename):
    """Serve uploaded evidence files."""
    from flask import send_from_directory
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

@app.route('/spi/indicators/<int:iid>/toggle', methods=['POST'])
@require_login
def spi_toggle_indicator(iid):
    ind = SPIIndicator.query.get_or_404(iid)
    ind.active = not ind.active
    db.session.commit()
    flash(f'✓ {ind.code} {"activated" if ind.active else "deactivated"}.', 'success')
    return redirect(url_for('spi_indicators'))


@app.route('/spi/indicators/<int:iid>/set-alerts', methods=['POST'])
@require_login
def spi_set_alerts(iid):
    """Update alert thresholds (L1/L2/L3) and SPT target for an existing indicator."""
    ind = SPIIndicator.query.get_or_404(iid)
    f   = request.form
    try:
        ind.spt_target = float(f['spt_target']) if f.get('spt_target') else ind.spt_target
        ind.alert_l1   = float(f['alert_l1'])   if f.get('alert_l1')   else None
        ind.alert_l2   = float(f['alert_l2'])   if f.get('alert_l2')   else None
        ind.alert_l3   = float(f['alert_l3'])   if f.get('alert_l3')   else None
        db.session.commit()
        flash(f'✓ Alert thresholds updated for {ind.code}.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {e}', 'danger')
    return_url = request.form.get('return_url') or f'/spi/indicators/{iid}/detail'
    return redirect(return_url)


# ─── SPI Intelligence: Indicator Detail ───────────────────────────────────────
@app.route('/spi/indicators/<int:iid>/detail')
@require_login
def spi_indicator_detail(iid):
    """Full intelligence view for a single SPI indicator — Task #65."""
    ind  = SPIIndicator.query.get_or_404(iid)
    data = SPIData.query.filter_by(spi_id=iid).order_by(SPIData.year, SPIData.month).all()

    # ── Linked events (all time) ──────────────────────────────────────────────
    def _safe(fn, default):
        try:
            return fn()
        except Exception:
            db.session.rollback()
            return default

    all_links = _safe(
        lambda: SPIEventLink.query.filter_by(spi_id=iid)
                    .order_by(SPIEventLink.created_at.desc()).all(),
        []
    )

    # ── Trend driver categories (frequency count) ─────────────────────────────
    from collections import Counter
    category_counts = Counter(lnk.category for lnk in all_links if lnk.category)
    top_categories  = category_counts.most_common(5)

    # ── Department exposure ───────────────────────────────────────────────────
    dept_counts = Counter(lnk.department_id for lnk in all_links if lnk.department_id)
    dept_names  = {}
    for did in dept_counts:
        d = _safe(lambda: Department.query.get(did), None)
        dept_names[did] = d.name if d else f'Dept {did}'

    # ── Recurrence analysis ───────────────────────────────────────────────────
    recurrence = _safe(lambda: _spi_recurrence_analysis(iid, lookback_months=12), {})

    # ── Severity breakdown ────────────────────────────────────────────────────
    sev_counts = Counter(lnk.severity for lnk in all_links if lnk.severity)

    # ── Recent 90-day links ───────────────────────────────────────────────────
    from datetime import timedelta
    cutoff_90 = (datetime.utcnow() - timedelta(days=90)).strftime('%Y-%m-%d')
    recent_links = [lnk for lnk in all_links if (lnk.event_date or '') >= cutoff_90]

    # ── Build value chart data (last 24 months) ───────────────────────────────
    chart_labels = [f"{r.month}/{r.year}" for r in data[-24:]]
    chart_values = [round(float(r.value), 4) if r.value is not None else None for r in data[-24:]]
    spt_target = float(ind.spt_target) if ind.spt_target else None

    # ── ICAO Statistical thresholds (auto-calculated from preceding year) ─────
    # These are ALWAYS computed regardless of whether manual overrides are saved.
    # Manual fields (ind.alert_l1/l2/l3) override auto if explicitly set.
    history_tuples = _spi_history(ind)
    all_hist_vals  = [v for _, _, v in history_tuples]
    stat_l1, stat_l2, stat_l3, stat_mean, stat_sd, is_stat = _spi_thresholds(ind, all_hist_vals)
    impr_target = _spi_improvement_target(ind)
    baseline_needed = ind.baseline_months or 3
    months_collected = len(all_hist_vals)

    # Chart lines: prefer statistical auto-values; allow manual override
    al1 = float(ind.alert_l1) if ind.alert_l1 else (round(stat_l1, 4) if stat_l1 else None)
    al2 = float(ind.alert_l2) if ind.alert_l2 else (round(stat_l2, 4) if stat_l2 else None)
    al3 = float(ind.alert_l3) if ind.alert_l3 else (round(stat_l3, 4) if stat_l3 else None)

    # Current status of the latest data point
    latest_val  = all_hist_vals[-1] if all_hist_vals else None
    curr_status = _spi_status(latest_val, ind, all_hist_vals) if latest_val is not None else ('— No data', '#9ca3af', 0)

    # Trigger check on full history
    last_24_hist  = history_tuples[-24:]
    trigger_pairs = [(i+1, v) for i, (_, _, v) in enumerate(last_24_hist)]
    trigger_detail = _spi_trigger_detail(trigger_pairs, stat_l1, stat_l2, stat_l3,
                                         ind.calc_type == 'PERCENT', spt=ind.spt_target)

    return render_template('spi/spi_indicator_detail.html',
        ind=ind,
        data=data,
        all_links=all_links,
        recent_links=recent_links,
        top_categories=top_categories,
        dept_counts=dept_counts,
        dept_names=dept_names,
        recurrence=recurrence,
        sev_counts=sev_counts,
        chart_labels=chart_labels,
        chart_values=chart_values,
        spt_target=spt_target,
        # Auto-calculated ICAO thresholds
        al1=al1, al2=al2, al3=al3,
        stat_l1=round(stat_l1,4) if stat_l1 else None,
        stat_l2=round(stat_l2,4) if stat_l2 else None,
        stat_l3=round(stat_l3,4) if stat_l3 else None,
        stat_mean=round(stat_mean,4) if stat_mean else None,
        stat_sd=round(stat_sd,4) if stat_sd else None,
        is_stat=is_stat,
        impr_target=impr_target,
        latest_val=latest_val,
        curr_status=curr_status,
        trigger_detail=trigger_detail,
        months_collected=months_collected,
        baseline_needed=baseline_needed,
        now=datetime.utcnow(),
    )


# ─── SPI Intelligence Hub ─────────────────────────────────────────────────────
@app.route('/spi/intelligence')
@require_login
def spi_intelligence():
    """System-wide SPI analytics hub + SRB auto-feed — Task #66."""

    def _safe(fn, default):
        try:
            return fn()
        except Exception:
            db.session.rollback()
            return default

    # ── System-wide intelligence summary ──────────────────────────────────────
    intel = _safe(lambda: _spi_intelligence_summary(), {})

    # ── All active indicators with link counts ────────────────────────────────
    indicators = _safe(lambda: SPIIndicator.query.filter_by(active=True).all(), [])

    ind_link_counts = {}
    for ind in indicators:
        ind_link_counts[ind.id] = _safe(
            lambda i=ind: SPIEventLink.query.filter_by(spi_id=i.id).count(), 0
        )

    # ── Top linked indicators (by event count) ────────────────────────────────
    top_indicators = sorted(indicators, key=lambda i: ind_link_counts.get(i.id, 0), reverse=True)[:8]

    # ── Recent high-severity links (last 30 days) ─────────────────────────────
    from datetime import timedelta
    cutoff_30 = (datetime.utcnow() - timedelta(days=30)).strftime('%Y-%m-%d')
    high_severity_links = _safe(
        lambda: SPIEventLink.query.filter(
            SPIEventLink.severity.in_(['Critical','High']),
            SPIEventLink.event_date >= cutoff_30
        ).order_by(SPIEventLink.created_at.desc()).limit(20).all(),
        []
    )

    # ── SRB feed: upcoming meetings that need SPI agenda items ───────────────
    srb_meetings = _safe(
        lambda: SRBMeeting.query.filter(SRBMeeting.status == 'Scheduled')
                    .order_by(SRBMeeting.meeting_date).limit(3).all(),
        []
    )

    # ── Cross-module event type breakdown ─────────────────────────────────────
    from collections import Counter
    all_links = _safe(
        lambda: SPIEventLink.query.all(), []
    )
    event_type_counts = Counter(lnk.event_type for lnk in all_links)
    dept_exposure = Counter(lnk.department_id for lnk in all_links if lnk.department_id)

    dept_names = {}
    for did in list(dept_exposure.keys())[:10]:
        d = _safe(lambda: Department.query.get(did), None)
        dept_names[did] = d.name if d else f'Dept {did}'

    # ── Recurring hazard patterns ─────────────────────────────────────────────
    recurrence_alerts = []
    for ind in top_indicators[:5]:
        rec = _safe(lambda i=ind: _spi_recurrence_analysis(i.id, lookback_months=6), {})
        if rec.get('recurring_categories'):
            recurrence_alerts.append({
                'indicator': ind,
                'recurrence': rec,
            })

    return render_template('spi/spi_intelligence.html',
        intel=intel,
        indicators=indicators,
        ind_link_counts=ind_link_counts,
        top_indicators=top_indicators,
        high_severity_links=high_severity_links,
        srb_meetings=srb_meetings,
        event_type_counts=dict(event_type_counts),
        dept_exposure=dict(dept_exposure),
        dept_names=dept_names,
        recurrence_alerts=recurrence_alerts,
        total_links=len(all_links),
        now=datetime.utcnow(),
    )


# ─── Manual SPI Link ──────────────────────────────────────────────────────────
@app.route('/spi/link', methods=['POST'])
@require_login
def spi_manual_link():
    """Create a manual SPIEventLink from any report detail page — Task #69."""
    if SPIEventLink is None:
        flash('SPI linkage not yet available — please redeploy models.', 'warning')
        return redirect(request.referrer or '/spi')

    spi_id      = request.form.get('spi_id', type=int)
    event_type  = request.form.get('event_type', '')
    event_id    = request.form.get('event_id', '')
    event_title = request.form.get('event_title', '')
    event_date  = request.form.get('event_date', '')
    severity    = request.form.get('severity', '')
    dept_id     = request.form.get('department_id', type=int)
    category    = request.form.get('category', '')
    return_url  = request.form.get('return_url') or request.referrer or '/spi'

    if not spi_id:
        flash('Please select an SPI indicator.', 'warning')
        return redirect(return_url)

    # Check for duplicate (same event already linked to same indicator)
    existing = None
    try:
        existing = SPIEventLink.query.filter_by(
            spi_id=spi_id, event_type=event_type, event_id=event_id
        ).first()
    except Exception:
        db.session.rollback()

    if existing:
        flash(f'This report is already linked to {existing.indicator.code if existing.indicator else "that indicator"}.', 'info')
        return redirect(return_url)

    try:
        link = SPIEventLink(
            spi_id=spi_id,
            event_type=event_type,
            event_id=event_id,
            event_title=event_title,
            event_date=event_date or datetime.utcnow().strftime('%Y-%m-%d'),
            department_id=dept_id,
            category=category,
            severity=severity,
            match_reason='Manual link by ' + (session.get('admin_user') or 'user'),
        )
        db.session.add(link)
        db.session.commit()
        ind = SPIIndicator.query.get(spi_id)
        flash(f'✓ Report linked to SPI indicator {ind.code if ind else spi_id}.', 'success')
    except Exception as e:
        db.session.rollback()
        app.logger.error('spi_manual_link error: %s', e)
        flash('Error creating SPI link. Please try again.', 'danger')

    return redirect(return_url)


@app.route('/spi/unlink/<int:link_id>', methods=['POST'])
@require_login
def spi_unlink(link_id):
    """Remove a manual SPIEventLink."""
    if SPIEventLink is None:
        flash('SPI linkage not available.', 'warning')
        return redirect(request.referrer or '/spi')
    return_url = request.form.get('return_url') or request.referrer or '/spi'
    try:
        lnk = SPIEventLink.query.get_or_404(link_id)
        code = lnk.indicator.code if lnk.indicator else str(link_id)
        db.session.delete(lnk)
        db.session.commit()
        flash(f'✓ SPI link to {code} removed.', 'success')
    except Exception as e:
        db.session.rollback()
        app.logger.error('spi_unlink error: %s', e)
        flash('Error removing SPI link.', 'danger')
    return redirect(return_url)


# ─── Safety Promotion ─────────────────────────────────────────────────────────
# ═══════════════════════════════════════════════════════════════════════════════
#  SAFETY PROMOTION MODULE
# ═══════════════════════════════════════════════════════════════════════════════

@app.route('/safety-promotion')
@require_login
def safety_promotion():
    def _safe(fn, default):
        try: return fn()
        except Exception as _ex:
            db.session.rollback()
            app.logger.warning('safety_promotion query failed: %s', _ex)
            return default
    bulletins   = _safe(lambda: SafetyBulletin.query.filter_by(status='Active').order_by(SafetyBulletin.created_at.desc()).limit(5).all(), [])
    newsletters = _safe(lambda: SafetyNewsletter.query.filter_by(status='Published').order_by(SafetyNewsletter.created_at.desc()).limit(4).all(), [])
    trainings   = _safe(lambda: Training.query.order_by(Training.created_at.desc()).limit(8).all(), [])
    campaigns   = _safe(lambda: SafetyCampaign.query.filter_by(status='Active').order_by(SafetyCampaign.created_at.desc()).limit(4).all(), [])
    lessons     = _safe(lambda: LessonLearned.query.order_by(LessonLearned.created_at.desc()).limit(4).all(), [])
    surveys     = _safe(lambda: SafetySurvey.query.filter_by(status='Active').all(), [])
    overdue_training = _safe(lambda: Training.query.filter_by(status='Expired').count(), 0)
    due_soon    = _safe(lambda: Training.query.filter_by(status='Due Soon').count(), 0)
    emails_sent = _safe(lambda: EmailLog.query.count(), 0)
    dist_count  = _safe(lambda: DistributionList.query.filter_by(is_active=True).count(), 0)
    avg_response_rate = 0
    try:
        surveyed = SafetySurvey.query.filter(SafetySurvey.target_count > 0).all()
        if surveyed:
            avg_response_rate = round(sum(
                (s.response_count or 0) / s.target_count * 100 for s in surveyed
            ) / len(surveyed), 1)
    except Exception:
        db.session.rollback()
    return render_template('spi/safety_promotion.html',
                           bulletins=bulletins, newsletters=newsletters,
                           trainings=trainings, campaigns=campaigns,
                           lessons=lessons, surveys=surveys,
                           overdue_training=overdue_training, due_soon=due_soon,
                           emails_sent=emails_sent, dist_count=dist_count,
                           avg_response_rate=avg_response_rate)

@app.route('/safety-promotion/bulletins')
@require_login
def sp_bulletins():
    status_f = request.args.get('status','')
    type_f   = request.args.get('type','')
    q = SafetyBulletin.query
    if status_f: q = q.filter_by(status=status_f)
    if type_f:   q = q.filter_by(bulletin_type=type_f)
    bulletins = q.order_by(SafetyBulletin.created_at.desc()).all()
    return render_template('spi/sp_bulletins.html', bulletins=bulletins, status_f=status_f, type_f=type_f)

@app.route('/safety-promotion/bulletin/new', methods=['GET','POST'])
@require_login
def new_bulletin():
    if request.method == 'POST':
        f = request.form
        bid = new_id('SB')
        ef = None
        if 'attachment' in request.files:
            att = request.files['attachment']
            if att and att.filename and allowed_file(att.filename):
                from werkzeug.utils import secure_filename
                ef = f'{bid}_{secure_filename(att.filename)}'
                att.save(os.path.join(app.config['UPLOAD_FOLDER'], ef))
        b = SafetyBulletin(id=bid, ref_number=f.get('ref_number',bid), title=f['title'],
            bulletin_type=f.get('bulletin_type','Bulletin'), severity=f.get('severity','Information'),
            department_id=int(f['department_id']) if f.get('department_id') else None,
            issue_date=f.get('issue_date',datetime.now().strftime('%Y-%m-%d')),
            content=f['content'], recommendations=f.get('recommendations',''),
            issued_by=f.get('issued_by','Safety Department'), status='Active',
            attachment=ef, linked_hazard_id=f.get('linked_hazard_id') or None)
        db.session.add(b); db.session.commit()
        # Hook 8 — SPI Intelligence linkage for new safety bulletin
        try:
            _spi_link_event(
                event_type    = 'safety_promo',
                event_id      = bid,
                event_title   = f['title'][:120],
                department_id = int(f['department_id']) if f.get('department_id') else None,
                category      = f.get('bulletin_type', 'Bulletin'),
                severity      = f.get('severity', 'Information'),
                extra_text    = f.get('content', '') + ' ' + f.get('recommendations', ''),
                event_date    = f.get('issue_date', ''),
            )
        except Exception:
            pass
        push_notify_all(
            f'🛡 Safety Bulletin: {f["title"]}',
            f.get("recommendations") or f.get("content", "")[:100] or 'New safety bulletin published.',
            'safety_promo', 'bulletin', bid)
        flash(f'✓ Bulletin {bid} published.', 'success')
        return redirect(url_for('sp_bulletins'))
    return render_template('spi/sp_bulletin_form.html',
                           now=datetime.utcnow())

@app.route('/safety-promotion/bulletin/<bid>')
@require_login
def sp_bulletin_detail(bid):
    b = SafetyBulletin.query.get_or_404(bid)
    return render_template('spi/sp_bulletin_detail.html', b=b, now=datetime.utcnow())

@app.route('/safety-promotion/bulletin/<bid>/archive', methods=['POST'])
@require_login
def sp_bulletin_archive(bid):
    b = SafetyBulletin.query.get_or_404(bid)
    b.status = 'Archived'; db.session.commit()
    flash('✓ Bulletin archived.', 'success')
    return redirect(url_for('sp_bulletins'))

@app.route('/safety-promotion/newsletters')
@require_login
def sp_newsletters():
    newsletters = SafetyNewsletter.query.order_by(SafetyNewsletter.created_at.desc()).all()
    return render_template('spi/sp_newsletters.html', newsletters=newsletters)

@app.route('/safety-promotion/newsletter/new', methods=['GET','POST'])
@require_login
def sp_newsletter_new():
    if request.method == 'POST':
        f = request.form
        nid_str = new_id('NL')
        ef = None
        if 'attachment' in request.files:
            att = request.files['attachment']
            if att and att.filename and allowed_file(att.filename):
                from werkzeug.utils import secure_filename
                ef = f'{nid_str}_{secure_filename(att.filename)}'
                att.save(os.path.join(app.config['UPLOAD_FOLDER'], ef))
        n = SafetyNewsletter(ref_number=f.get('ref_number',nid_str), title=f['title'],
            issue_number=f.get('issue_number',''),
            department_id=int(f['department_id']) if f.get('department_id') else None,
            issue_date=f.get('issue_date',datetime.now().strftime('%Y-%m-%d')),
            author=f.get('author','Safety Department'), summary=f.get('summary',''),
            content=f.get('content',''), status=f.get('status','Draft'), attachment=ef)
        db.session.add(n); db.session.commit()
        flash(f'✓ Newsletter saved.', 'success')
        return redirect(url_for('sp_newsletters'))
    return render_template('spi/sp_newsletter_form.html',
                           now=datetime.utcnow())

@app.route('/safety-promotion/newsletter/<int:nid>')
@require_login
def sp_newsletter_detail(nid):
    n = SafetyNewsletter.query.get_or_404(nid)
    return render_template('spi/sp_newsletter_detail.html', n=n, now=datetime.utcnow())

@app.route('/safety-promotion/newsletter/<int:nid>/publish', methods=['POST'])
@require_login
def sp_newsletter_publish(nid):
    n = SafetyNewsletter.query.get_or_404(nid)
    n.status = 'Published'; db.session.commit()
    push_notify_all(
        f'📰 New Newsletter: {n.title}',
        n.summary or 'A new safety newsletter has been published.',
        'safety_promo', 'newsletter', nid)
    flash('✓ Newsletter published.', 'success')
    return redirect(url_for('sp_newsletters'))

@app.route('/safety-promotion/training')
@require_login
def sp_training():
    """Training records list with auto status refresh."""
    dept_f   = request.args.get('dept', '')
    status_f = request.args.get('status', '')
    type_f   = request.args.get('type', '')

    # Auto-refresh expired / due soon statuses
    from datetime import date as dt_date, timedelta
    today = dt_date.today()
    due_threshold = (today + timedelta(days=60)).isoformat()
    today_str = today.isoformat()
    for t in Training.query.filter(
            Training.status.in_(['Scheduled','In Progress','Completed','Current'])).all():
        if t.expiry_date and t.status not in ('Cancelled',):
            if t.expiry_date < today_str:
                t.status = 'Expired'
            elif t.expiry_date < due_threshold and t.status == 'Completed':
                t.status = 'Due Soon'
    db.session.commit()

    q = Training.query
    if dept_f:   q = q.filter_by(department_id=int(dept_f))
    if status_f: q = q.filter_by(status=status_f)
    if type_f:   q = q.filter_by(training_type=type_f)
    trainings = q.order_by(Training.created_at.desc()).all()

    stats = {
        'total':     Training.query.count(),
        'scheduled': Training.query.filter_by(status='Scheduled').count(),
        'in_progress':Training.query.filter_by(status='In Progress').count(),
        'completed': Training.query.filter_by(status='Completed').count(),
        'due_soon':  Training.query.filter(Training.status.in_(['Due Soon','Current'])).count(),
        'expired':   Training.query.filter_by(status='Expired').count(),
        'overdue':   Training.query.filter_by(status='Overdue').count(),
    }
    return render_template('spi/sp_training.html', trainings=trainings, stats=stats,
                           dept_f=dept_f, status_f=status_f, type_f=type_f)


@app.route('/safety-promotion/training/new', methods=['GET', 'POST'])
@require_login
def new_training():
    if request.method == 'POST':
        f = request.form
        cert = _save_upload('certificate', f'CERT_')
        evid = _save_upload('evidence', f'EV_')
        t = Training(
            employee_name    = f.get('employee_name', ''),
            employee_id      = f.get('employee_id', ''),
            department_id    = int(f['department_id']) if f.get('department_id') else None,
            position         = f.get('position', ''),
            training_type    = f.get('training_type', 'SMS Training'),
            training_program = f.get('training_program', ''),
            course_code      = f.get('course_code', ''),
            instructor       = f.get('instructor', ''),
            location         = f.get('location', ''),
            scheduled_date   = f.get('scheduled_date', ''),
            training_date    = f.get('training_date', ''),
            completion_date  = f.get('completion_date', ''),
            expiry_date      = f.get('expiry_date', ''),
            duration_hours   = float(f['duration_hours']) if f.get('duration_hours') else None,
            status           = f.get('status', 'Scheduled'),
            certificate      = cert,
            evidence         = evid,
            is_recurrent     = 'is_recurrent' in f,
            recurrence_months= int(f['recurrence_months']) if f.get('recurrence_months') else None,
            notes            = f.get('notes', ''),
        )
        db.session.add(t)
        db.session.commit()
        # Notify the assigned employee
        if t.employee_name:
            push_notify_by_name(
                t.employee_name,
                f'📚 Training Assigned: {t.training_program}',
                f'Scheduled for {t.scheduled_date or t.training_date or "TBD"}. Please review your training plan.',
                'action_assigned', 'training', t.id)
        flash(f'✓ Training record saved for {t.employee_name}.', 'success')
        return redirect(url_for('sp_training'))
    return render_template('spi/sp_training_form.html', now=datetime.utcnow(), editing=False)


@app.route('/safety-promotion/training/<int:tid>', methods=['GET', 'POST'])
@require_login
def sp_training_detail(tid):
    """View and edit a single training record."""
    t = Training.query.get_or_404(tid)
    if request.method == 'POST':
        f = request.form
        action = f.get('action', 'update')

        if action == 'delete':
            db.session.delete(t)
            db.session.commit()
            flash('Training record deleted.', 'success')
            return redirect(url_for('sp_training'))

        # Handle file uploads
        new_cert = _save_upload('certificate', 'CERT_')
        new_evid = _save_upload('evidence', 'EV_')

        t.employee_name     = f.get('employee_name', t.employee_name)
        t.employee_id       = f.get('employee_id', t.employee_id)
        t.department_id     = int(f['department_id']) if f.get('department_id') else t.department_id
        t.position          = f.get('position', t.position)
        t.training_type     = f.get('training_type', t.training_type)
        t.training_program  = f.get('training_program', t.training_program)
        t.course_code       = f.get('course_code', t.course_code)
        t.instructor        = f.get('instructor', t.instructor)
        t.location          = f.get('location', t.location)
        t.scheduled_date    = f.get('scheduled_date', t.scheduled_date)
        t.training_date     = f.get('training_date', t.training_date)
        t.completion_date   = f.get('completion_date', t.completion_date)
        t.expiry_date       = f.get('expiry_date', t.expiry_date)
        t.duration_hours    = float(f['duration_hours']) if f.get('duration_hours') else t.duration_hours
        t.status            = f.get('status', t.status)
        t.is_recurrent      = 'is_recurrent' in f
        t.recurrence_months = int(f['recurrence_months']) if f.get('recurrence_months') else t.recurrence_months
        t.notes             = f.get('notes', t.notes)
        if new_cert: t.certificate = new_cert
        if new_evid: t.evidence    = new_evid
        t.updated_at = datetime.utcnow()
        db.session.commit()
        flash('✓ Training record updated.', 'success')
        return redirect(url_for('sp_training_detail', tid=tid))
    return render_template('spi/sp_training_detail.html', t=t, now=datetime.utcnow())


@app.route('/safety-promotion/training/export/xlsx')
@require_login
def sp_training_export_xlsx():
    """Export training records to Excel with formatting."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from flask import send_file

    dept_f   = request.args.get('dept', '')
    status_f = request.args.get('status', '')
    q = Training.query
    if dept_f:   q = q.filter_by(department_id=int(dept_f))
    if status_f: q = q.filter_by(status=status_f)
    records = q.order_by(Training.status, Training.expiry_date).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Training Records'

    # Header styling
    navy   = PatternFill('solid', start_color='0F1C3F')
    gold   = PatternFill('solid', start_color='C9A84C')
    red_f  = PatternFill('solid', start_color='FEE2E2')
    amber_f= PatternFill('solid', start_color='FEF9C3')
    green_f= PatternFill('solid', start_color='DCFCE7')
    blue_f = PatternFill('solid', start_color='DBEAFE')
    white_f= Font(color='FFFFFF', bold=True, name='Arial')
    bold_f = Font(bold=True, name='Arial')
    std_f  = Font(name='Arial')
    center = Alignment(horizontal='center', vertical='center')
    left   = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin   = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )

    # Title row
    ws.merge_cells('A1:L1')
    ws['A1'] = 'AviaS — SAFETY TRAINING RECORDS'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = navy
    ws['A1'].alignment = center

    # Subtitle
    ws.merge_cells('A2:L2')
    ws['A2'] = f'Generated: {datetime.utcnow().strftime("%d %b %Y")}   |   Total Records: {len(records)}'
    ws['A2'].font = Font(name='Arial', size=10, italic=True, color='C9A84C')
    ws['A2'].fill = PatternFill('solid', start_color='0F1C3F')
    ws['A2'].alignment = center

    # Column headers
    headers = [
        ('Employee Name', 22), ('Employee ID', 12), ('Department', 18),
        ('Position', 18), ('Training Type', 20), ('Program', 28),
        ('Instructor', 18), ('Scheduled', 12), ('Completion', 12),
        ('Expiry Date', 12), ('Status', 14), ('Certificate', 14)
    ]
    for col, (hdr, width) in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=hdr)
        cell.font = white_f
        cell.fill = PatternFill('solid', start_color='1D4ED8')
        cell.alignment = center
        cell.border = thin
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 20
    ws.freeze_panes = 'A4'

    # Data rows
    STATUS_FILLS = {
        'Expired':    red_f,
        'Overdue':    red_f,
        'Due Soon':   amber_f,
        'Completed':  green_f,
        'Scheduled':  blue_f,
        'In Progress':blue_f,
        'Cancelled':  PatternFill('solid', start_color='F3F4F6'),
    }
    for row_n, t in enumerate(records, 4):
        fill = STATUS_FILLS.get(t.status, PatternFill())
        row_data = [
            t.employee_name or '',
            t.employee_id or '',
            t.department.name if t.department else '',
            t.position or '',
            t.training_type or '',
            t.training_program or '',
            t.instructor or '',
            t.scheduled_date or t.training_date or '',
            t.completion_date or '',
            t.expiry_date or '',
            t.status or '',
            '✓ Uploaded' if t.certificate else '✗ Missing',
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_n, column=col, value=val)
            cell.font = std_f
            cell.border = thin
            cell.alignment = left
            if col in (11, 12):  # status + cert cols
                cell.fill = fill
                cell.alignment = center
        ws.row_dimensions[row_n].height = 16

    # Summary sheet
    ws2 = wb.create_sheet('Summary')
    ws2.merge_cells('A1:C1')
    ws2['A1'] = 'Training Status Summary'
    ws2['A1'].font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
    ws2['A1'].fill = navy
    ws2['A1'].alignment = center
    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 14
    ws2.column_dimensions['C'].width = 14

    statuses = ['Scheduled','In Progress','Completed','Due Soon','Expired','Overdue','Cancelled']
    ws2.cell(row=2, column=1, value='Status').font = bold_f
    ws2.cell(row=2, column=2, value='Count').font   = bold_f
    ws2.cell(row=2, column=3, value='% of Total').font = bold_f
    for i, st in enumerate(statuses, 3):
        cnt = sum(1 for r in records if r.status == st)
        ws2.cell(row=i, column=1, value=st)
        ws2.cell(row=i, column=2, value=cnt)
        ws2.cell(row=i, column=3, value=f'=B{i}/B{3+len(statuses)}' if cnt else 0)
        if st in STATUS_FILLS:
            for col in range(1, 4):
                ws2.cell(row=i, column=col).fill = STATUS_FILLS[st]
    total_row = 3 + len(statuses)
    ws2.cell(row=total_row, column=1, value='TOTAL').font = bold_f
    ws2.cell(row=total_row, column=2, value=f'=SUM(B3:B{total_row-1})').font = bold_f

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'Training_Records_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/safety-promotion/surveys')
@require_login
def sp_surveys():
    surveys = SafetySurvey.query.order_by(SafetySurvey.created_at.desc()).all()
    return render_template('spi/sp_surveys.html', surveys=surveys)

@app.route('/safety-promotion/survey/new', methods=['GET','POST'])
@require_login
def sp_survey_new():
    import json as _j
    depts = Department.query.order_by(Department.name).all()
    if request.method == 'POST':
        f = request.form
        # Accept rich JSON questions from builder, fall back to legacy plain text list
        raw_q = f.get('questions_json', '').strip()
        if raw_q:
            try:
                questions_data = _j.loads(raw_q)
            except Exception:
                questions_data = []
        else:
            questions_data = [{'id': f'q_{i}', 'text': q.strip(), 'type': 'text',
                                'description': '', 'required': True, 'order': i, 'options': []}
                               for i, q in enumerate(f.getlist('question')) if q.strip()]
        s = SafetySurvey(
            title          = f.get('title', '').strip(),
            survey_type    = f.get('survey_type', 'Safety Culture Survey'),
            department_id  = int(f['department_id']) if f.get('department_id') else None,
            start_date     = f.get('start_date', ''),
            end_date       = f.get('end_date', ''),
            description    = f.get('description', ''),
            questions      = _j.dumps(questions_data),
            status         = 'Draft',
            target_count   = int(f.get('target_count', 0) or 0),
        )
        # target_audience stored in description prefix if column not available yet
        target_aud = f.get('target_audience', 'all')
        try:
            s.target_audience = target_aud
        except Exception:
            pass
        db.session.add(s)
        db.session.commit()
        flash('✓ Survey created successfully.', 'success')
        return redirect(f'/safety-promotion/survey/{s.id}')
    return render_template('spi/sp_survey_form.html',
                           survey=None, departments=depts, now=datetime.utcnow())


@app.route('/safety-promotion/survey/<int:sid>/edit', methods=['GET','POST'])
@require_login
def sp_survey_edit(sid):
    import json as _j
    s     = SafetySurvey.query.get_or_404(sid)
    depts = Department.query.order_by(Department.name).all()
    if request.method == 'POST':
        if s.status == 'Active':
            flash('⚠ Cannot edit an active survey. Close it first.', 'warning')
            return redirect(f'/safety-promotion/survey/{sid}')
        f = request.form
        raw_q = f.get('questions_json', '').strip()
        if raw_q:
            try:
                questions_data = _j.loads(raw_q)
            except Exception:
                questions_data = _j.loads(s.questions or '[]')
        else:
            questions_data = _j.loads(s.questions or '[]')
        s.title         = f.get('title', s.title).strip()
        s.survey_type   = f.get('survey_type', s.survey_type)
        s.department_id = int(f['department_id']) if f.get('department_id') else None
        s.start_date    = f.get('start_date', s.start_date)
        s.end_date      = f.get('end_date', s.end_date)
        s.description   = f.get('description', s.description)
        s.questions     = _j.dumps(questions_data)
        s.target_count  = int(f.get('target_count', s.target_count or 0) or 0)
        try:
            s.target_audience = f.get('target_audience', 'all')
        except Exception:
            pass
        db.session.commit()
        flash('✓ Survey updated.', 'success')
        return redirect(f'/safety-promotion/survey/{sid}')
    try:
        questions_data = _j.loads(s.questions or '[]')
    except Exception:
        questions_data = []
    return render_template('spi/sp_survey_form.html',
                           survey=s, questions_json=_j.dumps(questions_data),
                           departments=depts, now=datetime.utcnow())

@app.route('/safety-promotion/survey/<int:sid>/activate', methods=['POST'])
@require_login
def sp_survey_activate(sid):
    s = SafetySurvey.query.get_or_404(sid)
    s.status = 'Active'; db.session.commit()
    push_notify_all(
        f'📋 Safety Survey: {s.title}',
        s.description or 'A new safety survey has been assigned to you. Please complete it.',
        'safety_promo', 'survey', sid)
    flash('✓ Survey activated.', 'success')
    return redirect('/safety-promotion/surveys')

@app.route('/safety-promotion/survey/<int:sid>/close', methods=['POST'])
@require_login
def sp_survey_close(sid):
    s = SafetySurvey.query.get_or_404(sid)
    s.status = 'Closed'; db.session.commit()
    flash('✓ Survey closed.', 'success')
    return redirect('/safety-promotion/surveys')

@app.route('/safety-promotion/survey/<int:sid>/respond', methods=['POST'])
@require_login
def sp_survey_respond(sid):
    s = SafetySurvey.query.get_or_404(sid)
    s.response_count = (s.response_count or 0) + 1; db.session.commit()
    flash('✓ Response recorded.', 'success')
    return redirect('/safety-promotion/surveys')

@app.route('/safety-promotion/campaigns')
@require_login
def sp_campaigns():
    campaigns = SafetyCampaign.query.order_by(SafetyCampaign.created_at.desc()).all()
    return render_template('spi/sp_campaigns.html', campaigns=campaigns)

@app.route('/safety-promotion/campaign/new', methods=['GET','POST'])
@require_login
def sp_campaign_new():
    if request.method == 'POST':
        f = request.form
        ef = None
        if 'attachment' in request.files:
            att = request.files['attachment']
            if att and att.filename and allowed_file(att.filename):
                from werkzeug.utils import secure_filename
                ef = f'CAM_{secure_filename(att.filename)}'
                att.save(os.path.join(app.config['UPLOAD_FOLDER'], ef))
        sc = SafetyCampaign(title=f['title'], campaign_type=f.get('campaign_type','Monthly'),
            department_id=int(f['department_id']) if f.get('department_id') else None,
            start_date=f.get('start_date',''), end_date=f.get('end_date',''),
            description=f.get('description',''), objectives=f.get('objectives',''),
            status='Active', attachment=ef)
        db.session.add(sc); db.session.commit()
        flash('✓ Campaign created.', 'success')
        return redirect(url_for('sp_campaigns'))
    return render_template('spi/sp_campaign_form.html',
                           now=datetime.utcnow())

@app.route('/safety-promotion/lessons')
@require_login
def sp_lessons():
    cat_f=request.args.get('category',''); q_f=request.args.get('q','').strip()
    q = LessonLearned.query
    if cat_f: q = q.filter_by(category=cat_f)
    if q_f:   q = q.filter(LessonLearned.title.ilike(f'%{q_f}%')|LessonLearned.description.ilike(f'%{q_f}%'))
    lessons = q.order_by(LessonLearned.created_at.desc()).all()
    return render_template('spi/sp_lessons.html', lessons=lessons, cat_f=cat_f, q_f=q_f)

@app.route('/safety-promotion/lesson/new', methods=['GET','POST'])
@require_login
def sp_lesson_new():
    if request.method == 'POST':
        f = request.form
        lid = new_id('LL')
        ef = None
        if 'attachment' in request.files:
            att = request.files['attachment']
            if att and att.filename and allowed_file(att.filename):
                from werkzeug.utils import secure_filename
                ef = f'{lid}_{secure_filename(att.filename)}'
                att.save(os.path.join(app.config['UPLOAD_FOLDER'], ef))
        ll = LessonLearned(ref_number=f.get('ref_number',lid), title=f['title'],
            category=f.get('category','Incident'),
            department_id=int(f['department_id']) if f.get('department_id') else None,
            date=f.get('date',datetime.now().strftime('%Y-%m-%d')),
            author=f.get('author','Safety Department'), description=f.get('description',''),
            lesson=f.get('lesson',''), recommendations=f.get('recommendations',''),
            status='Published', attachment=ef, linked_hazard_id=f.get('linked_hazard_id') or None)
        db.session.add(ll); db.session.commit()
        flash(f'✓ Lesson Learned {lid} published.', 'success')
        return redirect(url_for('sp_lessons'))
    return render_template('spi/sp_lesson_form.html', now=datetime.utcnow())

# ── Bulletin PDF print ────────────────────────────────────────────────────────


@app.route('/safety-promotion/distribution')
@require_login
def distribution_list():
    depts = Department.query.order_by(Department.name).all()
    dept_f = request.args.get('dept', '')
    q = DistributionList.query
    if dept_f:
        q = q.filter_by(department_id=int(dept_f))
    recipients = q.order_by(DistributionList.name).all()
    total = DistributionList.query.filter_by(is_active=True).count()
    return render_template('spi/sp_distribution.html',
                           recipients=recipients, total=total, dept_f=dept_f, depts=depts)

@app.route('/safety-promotion/distribution/add', methods=['POST'])
@require_login
def distribution_add():
    f = request.form
    email = f.get('email', '').strip()
    if not email:
        flash('Email required.', 'error')
        return redirect(url_for('distribution_list'))
    if DistributionList.query.filter_by(email=email).first():
        flash(email + ' already in list.', 'warning')
        return redirect(url_for('distribution_list'))
    db.session.add(DistributionList(
        name=f.get('name','').strip(), email=email,
        department_id=int(f['department_id']) if f.get('department_id') else None,
        position=f.get('position','').strip(), is_active=True))
    db.session.commit()
    flash('Added ' + email, 'success')
    return redirect(url_for('distribution_list'))

@app.route('/safety-promotion/distribution/<int:rid>/toggle', methods=['POST'])
@require_login
def distribution_toggle(rid):
    r = DistributionList.query.get_or_404(rid)
    r.is_active = not r.is_active
    db.session.commit()
    flash(r.email + ' updated.', 'success')
    return redirect(url_for('distribution_list'))

@app.route('/safety-promotion/distribution/<int:rid>/delete', methods=['POST'])
@require_login
def distribution_delete(rid):
    r = DistributionList.query.get_or_404(rid)
    db.session.delete(r)
    db.session.commit()
    flash('Removed.', 'success')
    return redirect(url_for('distribution_list'))

@app.route('/safety-promotion/distribution/import', methods=['POST'])
@require_login
def distribution_import():
    added = skipped = 0
    for line in request.form.get('csv_text','').strip().splitlines():
        parts = [p.strip() for p in line.split(',')]
        if len(parts) < 2: continue
        name, email = parts[0], parts[1]
        if not email or DistributionList.query.filter_by(email=email).first():
            skipped += 1; continue
        dept_id = int(parts[2]) if len(parts)>2 and parts[2].isdigit() else None
        db.session.add(DistributionList(name=name, email=email, department_id=dept_id,
                                        position=parts[3] if len(parts)>3 else ''))
        added += 1
    db.session.commit()
    flash('Imported ' + str(added) + ' (' + str(skipped) + ' skipped).', 'success')
    return redirect(url_for('distribution_list'))

@app.route('/safety-promotion/email-log')
@require_login
def email_log_list():
    logs = EmailLog.query.order_by(EmailLog.sent_at.desc()).limit(100).all()
    total_sent = EmailLog.query.filter_by(status='Sent').count()
    total_recv = db.session.query(db.func.sum(EmailLog.recipient_count)).scalar() or 0
    return render_template('spi/sp_email_log.html', logs=logs,
                           total_sent=total_sent, total_recipients=int(total_recv))

@app.route('/safety-promotion/bulletin/<bid>/send-email', methods=['POST'])
@require_login
def bulletin_send_email(bid):
    b = SafetyBulletin.query.get_or_404(bid)
    dept_ids = [int(d) for d in request.form.getlist('dept_ids') if d.isdigit()] or None
    to = get_recipients(dept_ids)
    if not to:
        flash('No recipients. Add to Distribution List first.', 'warning')
        return redirect(url_for('sp_bulletin_detail', bid=bid))
    sev = {'Critical':'#dc2626','High':'#ea580c','Information':'#1d4ed8'}.get(b.severity,'#374151')
    body = ('<div style="background:'+sev+';color:#fff;padding:6px 14px;border-radius:5px;font-size:12px;font-weight:700;display:inline-block;margin-bottom:12px">'
            +b.severity.upper()+' — '+(b.bulletin_type or '')+'</div>'
            +'<p>'+(b.content or '').replace('\n','<br>')+'</p>'
            +('<p style="margin-top:12px"><strong>Recommendations:</strong><br>'+(b.recommendations or '').replace('\n','<br>')+'</p>' if b.recommendations else ''))
    subj = '[Safety Bulletin] '+(b.ref_number or b.id)+' — '+b.title
    sent, err = send_email(to, subj, email_html(b.title,'Safety Bulletin — '+(b.ref_number or b.id), body,
                           b.ref_number or b.id, b.issue_date or b.created_at.strftime('%d %b %Y')))
    dept_lbl = 'All' if not dept_ids else ','.join(d.name for d in Department.query.filter(Department.id.in_(dept_ids)).all())
    db.session.add(EmailLog(subject=subj,content_type='Bulletin',content_ref=str(bid),
        sent_by=session.get('admin_name','System'),recipient_count=sent,
        dept_filter=dept_lbl, status='Sent' if not err else 'Failed',error_message=err))
    db.session.commit()
    flash('Emailed to '+str(sent)+' recipient(s).', 'success')
    return redirect(url_for('sp_bulletin_detail', bid=bid))

@app.route('/safety-promotion/newsletter/<int:nid>/send-email', methods=['POST'])
@require_login
def newsletter_send_email(nid):
    n = SafetyNewsletter.query.get_or_404(nid)
    dept_ids = [int(d) for d in request.form.getlist('dept_ids') if d.isdigit()] or None
    to = get_recipients(dept_ids)
    if not to:
        flash('No recipients.', 'warning')
        return redirect('/safety-promotion/newsletter/'+str(nid))
    subj = '[Safety Newsletter] '+n.title
    body = '<p>'+(n.summary or '').replace('\n','<br>')+'</p>'+('<div>'+(n.content or '').replace('\n','<br>')+'</div>' if n.content else '')
    sent, err = send_email(to, subj, email_html(n.title,'Safety Newsletter',body))
    dept_lbl = 'All' if not dept_ids else ','.join(d.name for d in Department.query.filter(Department.id.in_(dept_ids)).all())
    db.session.add(EmailLog(subject=subj,content_type='Newsletter',content_ref=str(nid),
        sent_by=session.get('admin_name','System'),recipient_count=sent,
        dept_filter=dept_lbl,status='Sent' if not err else 'Failed',error_message=err))
    db.session.commit()
    flash('Emailed to '+str(sent)+' recipient(s).', 'success')
    return redirect('/safety-promotion/newsletter/'+str(nid))

@app.route('/safety-promotion/lesson/<int:lid>/send-email', methods=['POST'])
@require_login
def lesson_send_email(lid):
    ll = LessonLearned.query.get_or_404(lid)
    recip = _get_dist_list()
    if not recip:
        flash('No recipients in distribution list.', 'warning')
        return redirect(f'/safety-promotion/lesson/{lid}')
    emails = [r.email for r in recip]
    body = (
        '<p><b>Category:</b> ' + (ll.category or '') +
        ' &nbsp;&middot;&nbsp; <b>Date:</b> ' + (ll.date or '') +
        '<br/><b>Author:</b> ' + (ll.author or 'Safety') + '</p>'
        '<hr style="border:none;border-top:1px solid #e5e7eb;margin:14px 0"/>'
        '<p>' + (ll.description or '').replace('\n','<br/>') + '</p>'
        + ('<p><b>Lesson:</b><br/>' + (ll.lesson or '').replace('\n','<br/>') + '</p>'
           if ll.lesson else '')
        + ('<p><b>Recommendations:</b><br/>' + (ll.recommendations or '').replace('\n','<br/>') + '</p>'
           if ll.recommendations else '')
    )
    subject = '[Lessons Learned] ' + ll.title
    html = _email_html(ll.title, 'Lessons Learned', body, '#15803d')
    sent, err = _do_send(emails, subject, html)
    _write_log('Lesson', lid, subject, sent, 'All',
               'Sent' if not err else 'Failed', err)
    flash(('⚠ SMTP not configured — email not sent.' if err == 'SMTP_NOT_CONFIGURED' else f'Lesson shared with {sent} recipients.'), 'warning' if err == 'SMTP_NOT_CONFIGURED' else 'success')
    return redirect(f'/safety-promotion/lesson/{lid}')
@app.route('/safety-promotion/bulletin/<bid>/print')
@require_login
def sp_bulletin_print(bid):
    b = SafetyBulletin.query.get_or_404(bid)
    return render_template('spi/sp_bulletin_print.html', b=b, now=datetime.utcnow())


# ── Newsletter edit ────────────────────────────────────────────────────────────

@app.route('/safety-promotion/newsletter/<int:nid>/edit', methods=['GET','POST'])
@require_login
def sp_newsletter_edit(nid):
    n = SafetyNewsletter.query.get_or_404(nid)
    if request.method == 'POST':
        f = request.form
        n.title        = f.get('title', n.title)
        n.issue_number = f.get('issue_number', n.issue_number)
        n.author       = f.get('author', n.author)
        n.issue_date   = f.get('issue_date', n.issue_date)
        n.summary      = f.get('summary', n.summary)
        n.content      = f.get('content', n.content)
        n.status       = f.get('status', n.status)
        if 'attachment' in request.files:
            att = request.files['attachment']
            if att and att.filename and allowed_file(att.filename):
                from werkzeug.utils import secure_filename
                fn = f'NL{n.id}_{secure_filename(att.filename)}'
                att.save(os.path.join(app.config['UPLOAD_FOLDER'], fn))
                n.attachment = fn
        db.session.commit()
        flash('✓ Newsletter updated.', 'success')
        return redirect(url_for('sp_newsletter_detail', nid=nid))
    return render_template('spi/sp_newsletter_form.html', n=n, editing=True,
                           now=datetime.utcnow())


@app.route('/safety-promotion/newsletter/<int:nid>/archive', methods=['POST'])
@require_login
def sp_newsletter_archive(nid):
    n = SafetyNewsletter.query.get_or_404(nid)
    n.status = 'Archived'
    db.session.commit()
    flash('✓ Newsletter archived.', 'success')
    return redirect(url_for('sp_newsletters'))


@app.route('/safety-promotion/newsletter/<int:nid>/print')
@require_login
def sp_newsletter_print(nid):
    n = SafetyNewsletter.query.get_or_404(nid)
    return render_template('spi/sp_newsletter_print.html', n=n, now=datetime.utcnow())


# ── Survey results dashboard ──────────────────────────────────────────────────

@app.route('/safety-promotion/survey/<int:sid>')
@require_login
def sp_survey_detail(sid):
    import json as _j
    s = SafetySurvey.query.get_or_404(sid)
    questions = []
    try:
        raw = _j.loads(s.questions or '[]')
        for q in raw:
            if isinstance(q, str):
                questions.append({'text': q, 'type': 'text', 'required': True, 'options': []})
            else:
                questions.append(q)
    except Exception:
        pass
    pct = int((s.response_count or 0) / max(s.target_count or 1, 1) * 100)
    responses = SurveyResponse.query.filter_by(survey_id=sid).order_by(SurveyResponse.submitted_at.desc()).all()
    all_departments = Department.query.order_by(Department.name).all()
    return render_template('spi/sp_survey_detail.html', s=s,
                           questions=questions, pct=pct,
                           responses=responses, all_departments=all_departments,
                           now=datetime.utcnow())


@app.route('/safety-promotion/survey/<int:sid>/analytics')
@require_login
def sp_survey_analytics(sid):
    import json as _j, collections
    s = SafetySurvey.query.get_or_404(sid)
    questions = []
    try:
        raw = _j.loads(s.questions or '[]')
        for i, q in enumerate(raw):
            if isinstance(q, str):
                questions.append({'id': f'q_{i}', 'text': q, 'type': 'text', 'options': [], 'order': i})
            else:
                q.setdefault('id', f'q_{i}')
                q.setdefault('order', i)
                questions.append(q)
    except Exception:
        pass

    responses = SurveyResponse.query.filter_by(survey_id=sid).all()
    # Build per-question answer tallies
    analytics = []
    for idx, q in enumerate(questions):
        tally   = collections.Counter()
        texts   = []
        qtype   = q.get('type', 'text')
        key     = str(idx)
        for r in responses:
            try:
                ans_raw = _j.loads(r.answers or '{}')
                if isinstance(ans_raw, list):
                    ans_raw = {str(item.get('question_index', i)): item.get('answer', '')
                               for i, item in enumerate(ans_raw) if isinstance(item, dict)}
            except Exception:
                ans_raw = {}
            val = ans_raw.get(key, '')
            if not val and val != 0:
                continue
            if qtype in ('single', 'yes_no', 'likert3', 'likert5', 'dropdown', 'rating5', 'rating10'):
                tally[str(val)] += 1
            elif qtype == 'multiple':
                if isinstance(val, list):
                    for v in val:
                        tally[str(v)] += 1
                else:
                    tally[str(val)] += 1
            else:
                texts.append(str(val))
        total = sum(tally.values()) or len(texts)
        pcts  = {k: round(v / max(sum(tally.values()), 1) * 100, 1) for k, v in tally.items()}
        analytics.append({
            'question': q,
            'tally':    dict(tally),
            'pcts':     pcts,
            'texts':    texts,
            'total':    total,
        })

    dept_breakdown = {}
    for r in responses:
        dept_name = r.department.name if r.department else 'Unknown'
        dept_breakdown[dept_name] = dept_breakdown.get(dept_name, 0) + 1

    resp_rate = round((s.response_count or 0) / max(s.target_count or 1, 1) * 100, 1)
    return render_template('spi/sp_survey_analytics.html',
                           s=s, questions=questions, analytics=analytics,
                           responses=responses, dept_breakdown=dept_breakdown,
                           resp_rate=resp_rate, now=datetime.utcnow())


# ── Campaign detail & close ───────────────────────────────────────────────────

@app.route('/safety-promotion/campaign/<int:cid>')
@require_login
def sp_campaign_detail(cid):
    c = SafetyCampaign.query.get_or_404(cid)
    return render_template('spi/sp_campaign_detail.html', c=c, now=datetime.utcnow())


@app.route('/safety-promotion/campaign/<int:cid>/complete', methods=['POST'])
@require_login
def sp_campaign_complete(cid):
    c = SafetyCampaign.query.get_or_404(cid)
    c.status = 'Completed'
    db.session.commit()
    flash('✓ Campaign marked as Completed.', 'success')
    return redirect(url_for('sp_campaigns'))


# ── Training PDF report ───────────────────────────────────────────────────────

@app.route('/safety-promotion/training/<int:tid>/edit', methods=['GET', 'POST'])
@require_login
def sp_training_edit(tid):
    """Full edit form — alias for detail page with editing mode."""
    t = Training.query.get_or_404(tid)
    if request.method == 'POST':
        return sp_training_detail(tid)
    return render_template('spi/sp_training_form.html', t=t, editing=True,
                           now=datetime.utcnow())


@app.route('/safety-promotion/training/report')
@require_login
def sp_training_report():
    dept_f   = request.args.get('dept', '')
    status_f = request.args.get('status', '')
    q = Training.query
    if dept_f:   q = q.filter_by(department_id=int(dept_f))
    if status_f: q = q.filter_by(status=status_f)
    trainings = q.order_by(Training.expiry_date).all()
    stats = {
        'total':     Training.query.count(),
        'scheduled': Training.query.filter_by(status='Scheduled').count(),
        'completed': Training.query.filter_by(status='Completed').count(),
        'expired':   Training.query.filter_by(status='Expired').count(),
        'due_soon':  Training.query.filter(Training.status.in_(['Due Soon','Current'])).count(),
    }
    return render_template('spi/sp_training_report.html',
                           trainings=trainings, stats=stats, now=datetime.utcnow())


# ── Lessons Learned PDF ───────────────────────────────────────────────────────

@app.route('/safety-promotion/lesson/<int:lid>/print')
@require_login
def sp_lesson_print(lid):
    ll = LessonLearned.query.get_or_404(lid)
    return render_template('spi/sp_lesson_print.html', ll=ll, now=datetime.utcnow())


@app.route('/safety-promotion/lesson/<int:lid>')
@require_login

def sp_lesson_detail(lid):
    ll = LessonLearned.query.get_or_404(lid)
    return render_template('spi/sp_lesson_detail.html', ll=ll, now=datetime.utcnow())


# ═══════════════════════════════════════════════════════════════════════════════
#  REGULATORY COMPLIANCE REGISTER — ICAO Annex 19 §2 / Doc 9859 §4.1
# ═══════════════════════════════════════════════════════════════════════════════

_COMP_BODIES     = ['ICAO', 'JCAR', 'IOSA', 'EASA', 'FAA', 'JAA', 'Internal']
_COMP_STATUSES   = ['Compliant', 'Partially Compliant', 'Non-Compliant',
                    'Under Review', 'Not Applicable', 'Exempt']
_COMP_OBL_TYPES  = ['Ongoing', 'Periodic', 'One-Time', 'Conditional']
_COMP_FREQ       = ['Monthly', 'Quarterly', 'Semi-Annual', 'Annual', 'As Required']


@app.route('/compliance')
@require_login
def compliance_list():
    status_filter   = request.args.get('status', '')
    body_filter     = request.args.get('body', '')
    priority_filter = request.args.get('priority', '')

    q = ComplianceObligation.query
    if status_filter:
        q = q.filter_by(compliance_status=status_filter)
    if body_filter:
        q = q.filter_by(regulation_body=body_filter)
    if priority_filter:
        q = q.filter_by(priority=priority_filter)
    obligations = q.order_by(ComplianceObligation.regulation_body,
                              ComplianceObligation.ref_number).all()

    all_obs = ComplianceObligation.query.all()
    today_str = datetime.utcnow().strftime('%Y-%m-%d')

    return render_template('compliance/compliance_list.html',
        obligations     = obligations,
        total           = len(all_obs),
        compliant       = sum(1 for o in all_obs if o.compliance_status == 'Compliant'),
        partial         = sum(1 for o in all_obs if o.compliance_status == 'Partially Compliant'),
        non_compliant   = sum(1 for o in all_obs if o.compliance_status == 'Non-Compliant'),
        under_review    = sum(1 for o in all_obs if o.compliance_status == 'Under Review'),
        overdue         = sum(1 for o in all_obs
                              if o.next_review_due and o.next_review_due < today_str),
        statuses        = _COMP_STATUSES,
        bodies          = _COMP_BODIES,
        status_filter   = status_filter,
        body_filter     = body_filter,
        priority_filter = priority_filter,
    )


@app.route('/compliance/<int:oid>')
@require_login
def compliance_detail(oid):
    ob = ComplianceObligation.query.get_or_404(oid)
    return render_template('compliance/compliance_detail.html', ob=ob)


@app.route('/compliance/new', methods=['GET', 'POST'])
@require_login
def new_compliance():
    if request.method == 'POST':
        last = ComplianceObligation.query.order_by(
                   ComplianceObligation.id.desc()).first()
        seq  = (last.id + 1) if last else 1
        ref  = request.form.get('ref_number', '').strip() or f'COMP-{seq:04d}'
        dept_id = request.form.get('department_id') or None
        ob = ComplianceObligation(
            ref_number           = ref,
            regulation_body      = request.form.get('regulation_body', ''),
            standard_ref         = request.form.get('standard_ref', ''),
            requirement_title    = request.form.get('requirement_title', ''),
            requirement_text     = request.form.get('requirement_text', ''),
            applicability        = request.form.get('applicability', ''),
            obligation_type      = request.form.get('obligation_type', 'Ongoing'),
            compliance_status    = request.form.get('compliance_status', 'Under Review'),
            priority             = request.form.get('priority', 'Medium'),
            evidence_description = request.form.get('evidence_description', ''),
            evidence_location    = request.form.get('evidence_location', ''),
            finding_ref          = request.form.get('finding_ref', ''),
            linked_action_id     = request.form.get('linked_action_id', ''),
            responsible_person   = request.form.get('responsible_person', ''),
            responsible_dept     = request.form.get('responsible_person', ''),
            department_id        = int(dept_id) if dept_id else None,
            review_frequency     = request.form.get('review_frequency', 'Annual'),
            next_review_due      = request.form.get('next_review_due', ''),
            last_reviewed        = request.form.get('last_reviewed', ''),
            notes                = request.form.get('notes', ''),
            created_by           = session.get('user', {}).get('username', ''),
        )
        db.session.add(ob)
        db.session.commit()
        flash('Compliance obligation created.', 'success')
        return redirect(url_for('compliance_detail', oid=ob.id))

    last = ComplianceObligation.query.order_by(
               ComplianceObligation.id.desc()).first()
    seq  = (last.id + 1) if last else 1
    return render_template('compliance/compliance_form.html',
        edit              = False,
        ob                = ComplianceObligation(),
        suggested_ref     = f'COMP-{seq:04d}',
        bodies            = _COMP_BODIES,
        statuses          = _COMP_STATUSES,
        obligation_types  = _COMP_OBL_TYPES,
        review_frequencies= _COMP_FREQ,
        departments       = Department.query.order_by(Department.name).all(),
    )


@app.route('/compliance/<int:oid>/edit', methods=['GET', 'POST'])
@require_login
def edit_compliance(oid):
    ob = ComplianceObligation.query.get_or_404(oid)
    if request.method == 'POST':
        dept_id = request.form.get('department_id') or None
        ob.regulation_body      = request.form.get('regulation_body', ob.regulation_body)
        ob.standard_ref         = request.form.get('standard_ref', ob.standard_ref)
        ob.requirement_title    = request.form.get('requirement_title', ob.requirement_title)
        ob.requirement_text     = request.form.get('requirement_text', ob.requirement_text)
        ob.applicability        = request.form.get('applicability', ob.applicability)
        ob.obligation_type      = request.form.get('obligation_type', ob.obligation_type)
        ob.compliance_status    = request.form.get('compliance_status', ob.compliance_status)
        ob.priority             = request.form.get('priority', ob.priority)
        ob.evidence_description = request.form.get('evidence_description', ob.evidence_description)
        ob.evidence_location    = request.form.get('evidence_location', ob.evidence_location)
        ob.finding_ref          = request.form.get('finding_ref', ob.finding_ref)
        ob.linked_action_id     = request.form.get('linked_action_id', ob.linked_action_id)
        ob.responsible_person   = request.form.get('responsible_person', ob.responsible_person)
        ob.department_id        = int(dept_id) if dept_id else None
        ob.review_frequency     = request.form.get('review_frequency', ob.review_frequency)
        ob.next_review_due      = request.form.get('next_review_due', ob.next_review_due)
        ob.last_reviewed        = request.form.get('last_reviewed', ob.last_reviewed)
        ob.notes                = request.form.get('notes', ob.notes)
        if request.form.get('ref_number', '').strip():
            ob.ref_number = request.form['ref_number'].strip()
        db.session.commit()
        flash('Obligation updated.', 'success')
        return redirect(url_for('compliance_detail', oid=ob.id))

    return render_template('compliance/compliance_form.html',
        edit              = True,
        ob                = ob,
        suggested_ref     = ob.ref_number,
        bodies            = _COMP_BODIES,
        statuses          = _COMP_STATUSES,
        obligation_types  = _COMP_OBL_TYPES,
        review_frequencies= _COMP_FREQ,
        departments       = Department.query.order_by(Department.name).all(),
    )


@app.route('/compliance/<int:oid>/update-status', methods=['POST'])
@require_login
def update_compliance_status(oid):
    ob = ComplianceObligation.query.get_or_404(oid)
    ob.compliance_status    = request.form.get('compliance_status', ob.compliance_status)
    ob.last_reviewed        = request.form.get('last_reviewed', ob.last_reviewed)
    ob.next_review_due      = request.form.get('next_review_due', ob.next_review_due)
    ob.evidence_description = request.form.get('evidence_description', ob.evidence_description)
    ob.notes                = request.form.get('notes', ob.notes)
    db.session.commit()
    flash('Compliance status updated.', 'success')
    return redirect(url_for('compliance_detail', oid=ob.id))


# ═══════════════════════════════════════════════════════════════════════════════
#  TESTING-PHASE DELETE ROUTES — Safe cascade deletion for all major modules
# ═══════════════════════════════════════════════════════════════════════════════

@app.route('/delete/hazard-report/<rid>', methods=['POST'])
@require_login
def delete_hazard_report(rid):
    """Safe delete: nullify FKs first, then cascade in correct dependency order."""
    rep = HazardReport.query.get_or_404(rid)
    hid = rep.hazard_id
    try:
        # Step 1: Nullify hazard_id on the report (removes the FK reference)
        rep.hazard_id = None
        db.session.flush()

        # Step 2: Delete linked Actions from SAG Portal
        try:
            linked_acts = Action.query.filter_by(hazard_id=hid).all() if hid else []
            linked_acts += Action.query.filter_by(linked_ref_id=rid).all()
            for la in linked_acts:
                ActionHistory.query.filter_by(action_id=la.id).delete(synchronize_session=False)
                db.session.delete(la)
            db.session.flush()
        except Exception: pass

        # Step 3: Delete the report row + its AVIs
        _avi_purge(source_record_id=rid, linked_report_id=rid)
        db.session.delete(rep)
        db.session.flush()

        # Step 4: If a linked Hazard exists, use the safe helper
        if hid:
            # Clean Phase 2 reporter feedback
            try:
                if _ENFORCEMENT_MODELS and ReportFeedback:
                    ReportFeedback.query.filter_by(
                        report_ref=rid).delete(synchronize_session=False)
            except Exception:
                pass
            _safe_delete_hazard(hid)
        db.session.commit()
        flash(f'✓ Hazard Report {rid} deleted.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'⚠ Could not delete {rid}: {str(e)[:120]}', 'error')
    return redirect(request.form.get('return_url', url_for('hazard_report_list')))


@app.route('/delete/hazard/<hid>', methods=['POST'])
@require_login
def delete_hazard(hid):
    """Safe delete hazard: cascade-delete or nullify ALL FK references first."""
    h = Hazard.query.get_or_404(hid)
    try:
        # Step 1: Nullify NULLABLE hazard_id references (these columns allow NULL)
        nullable_tables = [
            'hazard_reports', 'asr_reports', 'actions', 'investigations',
            'audit_findings', 'audit_actions', 'risk_actions', 'risk_assessments',
        ]
        for tbl in nullable_tables:
            db.session.execute(
                db.text(f"UPDATE {tbl} SET hazard_id = NULL WHERE hazard_id = :hid"),
                {'hid': hid}
            )
        db.session.flush()

        # Step 2: DELETE risk_occurrences — hazard_id is NOT NULL so must delete, not nullify
        RiskOccurrence.query.filter_by(hazard_id=hid).delete(synchronize_session=False)
        db.session.flush()

        # Step 3: Delete risks under this hazard
        # — clean RiskAction (risk_id is NOT NULL FK) before deleting each Risk
        for r in Risk.query.filter_by(hazard_id=hid).all():
            RiskAction.query.filter_by(risk_id=r.id).delete(synchronize_session=False)
            db.session.execute(
                db.text("UPDATE ra_rows SET risk_id = NULL WHERE risk_id = :rid"),
                {'rid': r.id}
            )
            Control.query.filter_by(risk_id=r.id).delete(synchronize_session=False)
            db.session.delete(r)
        db.session.flush()

        _avi_purge(linked_hazard_id=hid)
        db.session.delete(h)
        db.session.commit()
        flash(f'✓ Hazard {hid} deleted.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'⚠ Could not delete {hid}: {str(e)[:120]}', 'error')
    return redirect(url_for('hazard_log'))


@app.route('/delete/asr/<aid>', methods=['POST'])
@require_login
def delete_asr(aid):
    rec = ASRReport.query.get_or_404(aid)
    try:
        if rec.hazard_id:
            hid = rec.hazard_id
            HazardReport.query.filter_by(hazard_id=hid).delete(synchronize_session=False)
            # Clean actions and their history
            acts = Action.query.filter_by(hazard_id=hid).all()
            for a in acts:
                ActionHistory.query.filter_by(action_id=a.id).delete(synchronize_session=False)
                db.session.delete(a)
            db.session.flush()
            # RiskOccurrence.hazard_id is NOT NULL — delete, not nullify
            RiskOccurrence.query.filter_by(hazard_id=hid).delete(synchronize_session=False)
            # Clean RiskAction per risk before deleting risks
            for r in Risk.query.filter_by(hazard_id=hid).all():
                RiskAction.query.filter_by(risk_id=r.id).delete(synchronize_session=False)
                Control.query.filter_by(risk_id=r.id).delete(synchronize_session=False)
                db.session.delete(r)
            db.session.flush()
            haz = Hazard.query.get(hid)
            if haz:
                db.session.delete(haz)
        _avi_purge(source_record_id=aid)
        db.session.delete(rec)
        db.session.commit()
        flash(f'✓ ASR {aid} deleted successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting ASR: {str(e)[:80]}', 'danger')
    return redirect('/asr/list')


@app.route('/delete/action/<aid>', methods=['POST'])
@require_login
def delete_action(aid):
    a = Action.query.get_or_404(aid)
    try:
        # Delete child records first (FK constraint)
        ActionHistory.query.filter_by(action_id=aid).delete(synchronize_session=False)
        db.session.flush()
        _avi_purge(source_record_id=aid, linked_action_id=aid)
        db.session.delete(a)
        db.session.commit()
        flash(f'✓ Action {aid} deleted.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'⚠ Could not delete: {str(e)[:120]}', 'error')
    return redirect(request.form.get('return_url', '/actions'))


@app.route('/delete/risk-assessment/<ra_id>', methods=['POST'])
@require_login
def delete_risk_assessment(ra_id):
    """Safe delete RA: RARow/RAMitigation use assessment_id FK."""
    ra = RiskAssessment.query.get_or_404(ra_id)
    try:
        # RAMitigation and RARow use assessment_id (not ra_id)
        RAMitigation.query.filter_by(assessment_id=ra_id).delete(synchronize_session=False)
        RAReview.query.filter_by(assessment_id=ra_id).delete(synchronize_session=False)
        RAChecklistItem.query.filter_by(assessment_id=ra_id).delete(synchronize_session=False)
        # RARows: nullify risk_id FK first, then delete
        for row in RARow.query.filter_by(assessment_id=ra_id).all():
            row.risk_id = None
        db.session.flush()
        RARow.query.filter_by(assessment_id=ra_id).delete(synchronize_session=False)
        db.session.flush()
        _avi_purge(source_record_id=ra_id)
        db.session.delete(ra)
        db.session.commit()
        flash(f'✓ Risk Assessment {ra_id} deleted.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'⚠ Could not delete {ra_id}: {str(e)[:120]}', 'error')
    return redirect(url_for('ra_list'))


@app.route('/delete/audit-schedule/<sid>', methods=['POST'])
@require_login
def delete_audit_schedule(sid):
    """Safe delete audit: AuditAction.finding_id and
    AuditChecklist.linked_finding_id must be nullified before finding delete."""
    s = AuditSchedule.query.get_or_404(sid)
    try:
        findings = AuditFinding.query.filter_by(schedule_id=sid).all()
        fids = [f.id for f in findings]

        # Nullify linked_finding_id on checklists that reference these findings
        if fids:
            AuditChecklist.query.filter(
                AuditChecklist.linked_finding_id.in_(fids)
            ).update({'linked_finding_id': None}, synchronize_session=False)
            db.session.flush()
            # Delete audit actions linked to findings
            AuditAction.query.filter(
                AuditAction.finding_id.in_(fids)
            ).delete(synchronize_session=False)
            db.session.flush()

        # Delete checklist items for the schedule
        AuditChecklist.query.filter_by(schedule_id=sid).delete(synchronize_session=False)
        db.session.flush()

        # Delete linked Actions from SAG Portal (linked to findings)
        try:
            if fids:
                acts = Action.query.filter(
                    Action.linked_ref_id.in_([str(fid) for fid in fids])
                ).all()
                for la in acts:
                    ActionHistory.query.filter_by(action_id=la.id).delete(synchronize_session=False)
                    db.session.delete(la)
                db.session.flush()
        except Exception: pass

        # Now delete findings (and their AVIs)
        for f in findings:
            _avi_purge(source_record_id=f.id, linked_finding_id=f.id)
            db.session.delete(f)
        db.session.flush()

        _avi_purge(linked_audit_id=sid)
        db.session.delete(s)
        db.session.commit()
        flash(f'✓ Audit Schedule {sid} deleted.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'⚠ Could not delete {sid}: {str(e)[:120]}', 'error')
    return redirect(url_for('audit_schedule'))


@app.route('/delete/audit-finding/<fid>', methods=['POST'])
@require_login
def delete_audit_finding(fid):
    """Safe delete finding: nullify linked_finding_id in checklists, delete actions first."""
    f = AuditFinding.query.get_or_404(fid)
    schedule_id = f.schedule_id
    try:
        # Nullify checklist linked_finding_id references
        AuditChecklist.query.filter_by(linked_finding_id=str(fid)).update(
            {'linked_finding_id': None}, synchronize_session=False)
        db.session.flush()
        AuditAction.query.filter_by(finding_id=str(fid)).delete(synchronize_session=False)
        db.session.flush()
        # Delete linked Actions from SAG Portal
        try:
            linked_acts = Action.query.filter_by(linked_ref_id=str(fid)).all()
            if f.linked_action_id:
                la = Action.query.get(f.linked_action_id)
                if la and la not in linked_acts:
                    linked_acts.append(la)
            for la in linked_acts:
                ActionHistory.query.filter_by(action_id=la.id).delete(synchronize_session=False)
                db.session.delete(la)
            db.session.flush()
        except Exception: pass
        _avi_purge(source_record_id=fid, linked_finding_id=fid)
        db.session.delete(f)
        db.session.commit()
        flash('✓ Audit Finding and linked SAG actions deleted.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'⚠ Could not delete finding: {str(e)[:120]}', 'error')
    return redirect(url_for('audit_execution', sid=schedule_id))


@app.route('/delete/investigation/<iid>', methods=['POST'])
@require_login
def delete_investigation(iid):
    inv = Investigation.query.get_or_404(iid)
    try:
        # Step 1: Clean up SAG actions linked to this investigation
        linked_acts = Action.query.filter_by(linked_ref_id=str(iid)).all()
        for la in linked_acts:
            ActionHistory.query.filter_by(action_id=la.id).delete(synchronize_session=False)
            db.session.delete(la)
        db.session.flush()

        # Step 2: Delete InvestigationEvent timeline records (NOT NULL FK — CRITICAL)
        InvestigationEvent.query.filter_by(
            investigation_id=iid).delete(synchronize_session=False)
        db.session.flush()

        # Step 3: Nullify ERPActivation.investigation_id (nullable FK)
        try:
            db.session.execute(
                db.text("UPDATE erp_activations SET investigation_id = NULL "
                        "WHERE investigation_id = :iid"),
                {'iid': iid}
            )
        except Exception:
            pass
        db.session.flush()

        # Step 4: Clean Phase 2 models
        try:
            if _ENFORCEMENT_MODELS and InvestigationTimeline:
                InvestigationTimeline.query.filter_by(
                    investigation_id=iid).delete(synchronize_session=False)
        except Exception:
            pass
        try:
            if _ENFORCEMENT_MODELS and SoDViolationBlock:
                SoDViolationBlock.query.filter_by(
                    entity_type='Investigation',
                    entity_id=str(iid)).delete(synchronize_session=False)
        except Exception:
            pass

        # Step 5: Delete AVIs and the investigation record itself
        _avi_purge(source_record_id=iid, linked_investigation_id=iid)
        db.session.delete(inv)
        db.session.commit()
        flash('✓ Investigation and all linked records deleted.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'⚠ Could not delete: {str(e)[:120]}', 'error')
    return redirect('/investigations')


@app.route('/delete/training/<int:tid>', methods=['POST'])
@require_login
def delete_training(tid):
    t = Training.query.get_or_404(tid)
    db.session.delete(t)
    db.session.commit()
    flash('✓ Training record deleted.', 'success')
    return redirect(url_for('sp_training'))


@app.route('/delete/bulletin/<bid>', methods=['POST'])
@require_login
def delete_bulletin(bid):
    b = SafetyBulletin.query.get_or_404(bid)
    db.session.delete(b)
    db.session.commit()
    flash('✓ Bulletin deleted.', 'success')
    return redirect(url_for('sp_bulletins'))


@app.route('/delete/newsletter/<int:nid>', methods=['POST'])
@require_login
def delete_newsletter(nid):
    n = SafetyNewsletter.query.get_or_404(nid)
    db.session.delete(n)
    db.session.commit()
    flash('✓ Newsletter deleted.', 'success')
    return redirect(url_for('sp_newsletters'))


@app.route('/delete/survey/<int:sid>', methods=['POST'])
@require_login
def delete_survey(sid):
    s = SafetySurvey.query.get_or_404(sid)
    try:
        # SurveyResponse has NOT NULL FK to safety_surveys.id — must delete first
        SurveyResponse.query.filter_by(survey_id=sid).delete(synchronize_session=False)
        db.session.flush()
        db.session.delete(s)
        db.session.commit()
        flash('✓ Survey and all responses deleted.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'⚠ Could not delete survey: {str(e)[:120]}', 'error')
    return redirect('/safety-promotion/surveys')


@app.route('/delete/campaign/<int:cid>', methods=['POST'])
@require_login
def delete_campaign(cid):
    sc = SafetyCampaign.query.get_or_404(cid)
    db.session.delete(sc)
    db.session.commit()
    flash('✓ Campaign deleted.', 'success')
    return redirect(url_for('sp_campaigns'))


@app.route('/delete/lesson/<int:lid>', methods=['POST'])
@require_login
def delete_lesson(lid):
    ll = LessonLearned.query.get_or_404(lid)
    db.session.delete(ll)
    db.session.commit()
    flash('✓ Lesson Learned deleted.', 'success')
    return redirect(url_for('sp_lessons'))


@app.route('/delete/spi-data/<int:did>', methods=['POST'])
@require_login
def delete_spi_data(did):
    d = SPIData.query.get_or_404(did)
    iid = d.spi_id
    db.session.delete(d)
    db.session.commit()
    flash('✓ SPI data point deleted.', 'success')
    return redirect(url_for('spi_indicator_detail', iid=iid))


@app.route('/delete/moc/<mid>', methods=['POST'])
@require_login
def delete_moc(mid):
    m = MOC.query.get_or_404(mid)
    try:
        if m.hazard_id:
            # Unlink actions first (don't delete — they may be referenced elsewhere)
            Action.query.filter_by(hazard_id=m.hazard_id).update(
                {'hazard_id': None}, synchronize_session=False)
            db.session.flush()
            # Use safe helper to cascade-delete hazard and all children
            _safe_delete_hazard(m.hazard_id)
        db.session.delete(m)
        db.session.commit()
        flash('✓ MOC record deleted.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'⚠ Could not delete MOC: {str(e)[:120]}', 'error')
    return redirect(url_for('moc_list'))


# ── Admin cleanup dashboard ───────────────────────────────────────────────────

@app.route('/admin/cleanup')
@admin_required
def admin_cleanup():
    counts = {
        'hazard_reports':   HazardReport.query.count(),
        'hazards':          Hazard.query.count(),
        'asr_reports':      ASRReport.query.count(),
        'actions':          Action.query.count(),
        'risk_assessments': RiskAssessment.query.count(),
        'audit_schedules':  AuditSchedule.query.count(),
        'investigations':   Investigation.query.count(),
        'training':         Training.query.count(),
        'bulletins':        SafetyBulletin.query.count(),
        'newsletters':      SafetyNewsletter.query.count(),
        'surveys':          SafetySurvey.query.count(),
        'campaigns':        SafetyCampaign.query.count(),
        'lessons':          LessonLearned.query.count(),
        'spi_data':         SPIData.query.count(),
    }
    return render_template('admin_cleanup.html', counts=counts, total=sum(counts.values()))


# ─── Risk Matrix Reference ────────────────────────────────────────────────────

# ═══════════════════════════════════════════════════════════════════════════════
#  SAFETY PROMOTION — EMAIL DISTRIBUTION & RESPONSE TRACKING
# ═══════════════════════════════════════════════════════════════════════════════

def _smtp_cfg():
    return {
        'host': os.environ.get('SMTP_HOST',''),
        'port': int(os.environ.get('SMTP_PORT','587')),
        'user': os.environ.get('SMTP_USER',''),
        'password': os.environ.get('SMTP_PASSWORD',''),
        'from_email': os.environ.get('SMTP_FROM','safety@aviation.jo'),
    }

def _get_dist_list(dept_id=None):
    q = DistributionList.query.filter_by(is_active=True)
    if dept_id:
        q = q.filter_by(department_id=int(dept_id))
    return q.all()

def _do_send(emails, subject, html):
    cfg = _smtp_cfg()
    if not cfg['host']:
        # SMTP not configured — return 0 sent and a clear warning
        app.logger.warning('SMTP not configured — email not sent: "%s" (%d recipients)', subject, len(emails))
        return 0, 'SMTP_NOT_CONFIGURED'
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    try:
        srv = smtplib.SMTP(cfg['host'], cfg['port'])
        srv.starttls()
        srv.login(cfg['user'], cfg['password'])
        sent = 0
        for addr in emails:
            m = MIMEMultipart('alternative')
            m['Subject'] = subject
            m['From']    = cfg['from_email']
            m['To']      = addr
            m.attach(MIMEText(html, 'html'))
            try:
                srv.sendmail(cfg['from_email'], [addr], m.as_string())
                sent += 1
            except Exception:
                pass
        srv.quit()
        return sent, None
    except Exception as e:
        return 0, str(e)

def _write_log(ctype, cref, subject, count, dept_label, status='Sent', err=None):
    db.session.add(EmailLog(
        content_type=ctype, content_ref=str(cref), subject=subject,
        recipient_count=count, dept_filter=dept_label or 'All',
        sent_by=session.get('admin_name','System'),
        status=status, error_message=err
    ))
    db.session.commit()

def _email_html(title, subtitle, body, color='#0f1c3f'):
    return (
        '<!DOCTYPE html><html><head><meta charset="UTF-8"/><style>'
        'body{font-family:Arial,sans-serif;background:#f0f2f8;margin:0;padding:0}'
        '.w{max-width:620px;margin:20px auto;background:#fff;border-radius:10px;overflow:hidden}'
        '.h{background:' + color + ';padding:22px 28px;text-align:center}'
        '.hl{color:#c9a84c;font-size:22px;font-weight:800}'
        '.hs{color:rgba(255,255,255,.6);font-size:11px;margin-top:3px}'
        '.tb{background:' + color + '22;border-bottom:3px solid ' + color + ';padding:14px 28px}'
        '.tb h1{font-size:17px;font-weight:800;color:' + color + ';margin:0}'
        '.tb p{font-size:12px;color:#6b7280;margin:3px 0 0}'
        '.b{padding:22px 28px;font-size:14px;color:#374151;line-height:1.7}'
        '.f{background:#f8f9fc;padding:12px 28px;font-size:11px;color:#9ca3af;text-align:center;border-top:1px solid #e5e7eb}'
        '</style></head><body><div class="w">'
        '<div class="h"><div class="hl">&#x2708; AviaS</div>'
        '<div class="hs">Safety Management System</div></div>'
        '<div class="tb"><h1>' + title + '</h1><p>' + subtitle + '</p></div>'
        '<div class="b">' + body + '</div>'
        '<div class="f">AviaS Safety Management System &middot; Official Safety Communication</div>'
        '</div></body></html>'
    )


# ── Distribution List ─────────────────────────────────────────────────────────

# NOTE: duplicate route removed — /safety-promotion/distribution is served by distribution_list (above)
def sp_distribution():
    recipients = DistributionList.query.order_by(DistributionList.department_id, DistributionList.name).all()
    total = DistributionList.query.filter_by(is_active=True).count()
    depts = Department.query.all()
    return render_template('spi/sp_distribution.html',
                           recipients=recipients, total=total, depts=depts)


# NOTE: duplicate route removed — served by distribution_add (above)
def sp_distribution_add():
    f = request.form
    if not f.get('email') or not f.get('name'):
        flash('Name and email required.', 'error')
        return redirect(url_for('distribution_list'))
    if DistributionList.query.filter_by(email=f['email'].strip()).first():
        flash(f'{f["email"]} already in distribution list.', 'warning')
        return redirect(url_for('distribution_list'))
    db.session.add(DistributionList(
        name=f['name'].strip(), email=f['email'].strip(),
        position=f.get('position',''),
        department_id=int(f['department_id']) if f.get('department_id') else None,
        is_active=True,
    ))
    db.session.commit()
    flash(f'+ {f["name"]} added to distribution list.', 'success')
    return redirect(url_for('distribution_list'))


# NOTE: duplicate route removed — served by distribution_toggle (above)
def sp_distribution_toggle(rid):
    r = DistributionList.query.get_or_404(rid)
    r.is_active = not r.is_active
    db.session.commit()
    flash(f'{"Activated" if r.is_active else "Deactivated"}: {r.name}', 'success')
    return redirect(url_for('distribution_list'))


# NOTE: duplicate route removed — served by distribution_delete (above)
def sp_distribution_delete(rid):
    r = DistributionList.query.get_or_404(rid)
    db.session.delete(r); db.session.commit()
    flash('Recipient removed.', 'success')
    return redirect(url_for('distribution_list'))


# ── Email Log ─────────────────────────────────────────────────────────────────

# NOTE: duplicate route removed — served by email_log_list (above)
def sp_email_log():
    logs             = EmailLog.query.order_by(EmailLog.sent_at.desc()).all()
    total_sent       = sum(1 for l in logs if 'Sent' in (l.status or ''))
    total_failed     = sum(1 for l in logs if l.status == 'Failed')
    total_recipients = sum(l.recipient_count or 0 for l in logs)
    return render_template('spi/sp_email_log.html',
                           logs=logs, total_sent=total_sent,
                           total_failed=total_failed,
                           total_recipients=total_recipients)


# ── Send Bulletin ─────────────────────────────────────────────────────────────

# NOTE: duplicate route removed — served by bulletin_send_email (above)
def sp_send_bulletin_email(bid):
    b    = SafetyBulletin.query.get_or_404(bid)
    dept_id = request.form.get('dept_id') or None
    recip   = _get_dist_list(dept_id)
    if not recip:
        flash('No active recipients. Add employees to Distribution List first.', 'warning')
        return redirect(url_for('sp_bulletin_detail', bid=bid))
    emails     = [r.email for r in recip]
    dept_label = Department.query.get(int(dept_id)).name if dept_id else 'All'
    clr        = {'Critical':'#dc2626','High':'#ea580c','Medium':'#d97706','Low':'#15803d'}.get(b.severity,'#0f1c3f')
    body       = ('<p><b>Ref:</b> ' + (b.ref_number or b.id) + ' &nbsp;&middot;&nbsp; '
                  '<b>Severity:</b> <span style="color:' + clr + '">' + (b.severity or '') + '</span><br/>'
                  '<b>Issued by:</b> ' + (b.issued_by or 'Safety') + ' &nbsp;&middot;&nbsp; '
                  '<b>Date:</b> ' + (b.issue_date or '') + '</p>'
                  '<hr style="border:none;border-top:1px solid #e5e7eb;margin:14px 0"/>'
                  '<p>' + (b.content or '').replace('\n','<br/>') + '</p>'
                  + ('<p><b>Recommendations:</b><br/>' + (b.recommendations or '').replace('\n','<br/>') + '</p>'
                     if b.recommendations else ''))
    subject    = '[Safety Bulletin] ' + b.title
    html       = _email_html(b.title, 'Safety Bulletin ' + (b.severity or ''), body, clr)
    sent, err  = _do_send(emails, subject, html)
    _write_log('Bulletin', bid, subject, sent, dept_label,
               'Sent' if not err else 'Failed', err)
    flash(('⚠ SMTP not configured — email not sent. Set SMTP_HOST in environment.' if err == 'SMTP_NOT_CONFIGURED' else f'Bulletin emailed to {sent} recipients ({dept_label}).'), 'warning' if err == 'SMTP_NOT_CONFIGURED' else 'success')
    return redirect(url_for('sp_bulletin_detail', bid=bid))


# ── Send Newsletter ───────────────────────────────────────────────────────────

# NOTE: duplicate route removed — served by newsletter_send_email (above)
def sp_send_newsletter_email(nid):
    n       = SafetyNewsletter.query.get_or_404(nid)
    dept_id = request.form.get('dept_id') or None
    recip   = _get_dist_list(dept_id)
    if not recip:
        flash('No active recipients in distribution list.', 'warning')
        return redirect(url_for('sp_newsletter_detail', nid=nid))
    emails     = [r.email for r in recip]
    dept_label = Department.query.get(int(dept_id)).name if dept_id else 'All'
    body       = ('<p><b>Issue:</b> ' + (n.issue_number or '') +
                  ' &nbsp;&middot;&nbsp; <b>Date:</b> ' + (n.issue_date or '') +
                  '<br/><b>Author:</b> ' + (n.author or 'Safety') + '</p>'
                  + ('<p><em>' + (n.summary or '') + '</em></p>' if n.summary else '')
                  + '<hr style="border:none;border-top:1px solid #e5e7eb;margin:14px 0"/>'
                  + '<p>' + (n.content or '').replace('\n','<br/>') + '</p>')
    subject    = '[Safety Newsletter] ' + n.title
    html       = _email_html(n.title, 'Safety Newsletter', body)
    sent, err  = _do_send(emails, subject, html)
    _write_log('Newsletter', nid, subject, sent, dept_label,
               'Sent' if not err else 'Failed', err)
    flash(('⚠ SMTP not configured — email not sent. Set SMTP_HOST in environment.' if err == 'SMTP_NOT_CONFIGURED' else f'Newsletter emailed to {sent} recipients.'), 'warning' if err == 'SMTP_NOT_CONFIGURED' else 'success')
    return redirect(url_for('sp_newsletter_detail', nid=nid))


# ── Send Survey ───────────────────────────────────────────────────────────────

@app.route('/safety-promotion/survey/<int:sid>/send-email', methods=['POST'])
@require_login
def sp_send_survey_email(sid):
    s       = SafetySurvey.query.get_or_404(sid)
    dept_id = request.form.get('dept_id') or None
    recip   = _get_dist_list(dept_id)
    if not recip:
        flash('⚠ No active recipients found. Please add employees to the Distribution List first at /safety-promotion/distribution', 'warning')
        return redirect(url_for('sp_survey_detail', sid=sid))
    emails     = [r.email for r in recip]
    dept_label = Department.query.get(int(dept_id)).name if dept_id else 'All'
    pub_url    = request.host_url.rstrip('/') + '/safety-promotion/survey/' + str(sid) + '/respond-public'
    body       = ('<p>The Safety Department invites you to participate in this safety survey.</p>'
                  '<p><b>' + s.title + '</b><br/>'
                  + (s.description or '') + '</p>'
                  + '<p><b>Deadline:</b> ' + (s.end_date or 'Open ended') + '</p>'
                  '<p style="text-align:center;margin-top:18px">'
                  '<a href="' + pub_url + '" style="background:#7c3aed;color:#fff;padding:11px 24px;'
                  'border-radius:7px;font-weight:700;font-size:13px;text-decoration:none">'
                  'Complete Survey &rarr;</a></p>')
    subject    = '[Safety Survey] ' + s.title
    html       = _email_html(s.title, 'Safety Survey — Your Participation Required', body, '#7c3aed')
    sent, err  = _do_send(emails, subject, html)
    s.target_count = (s.target_count or 0) + sent
    _write_log('Survey', sid, subject, sent, dept_label,
               'Sent' if not err else 'Failed', err)
    flash(('⚠ SMTP not configured — email not sent.' if err == 'SMTP_NOT_CONFIGURED' else f'Survey invitation sent to {sent} recipients.'), 'warning' if err == 'SMTP_NOT_CONFIGURED' else 'success')
    return redirect(url_for('sp_survey_detail', sid=sid))


# ── Send Campaign ─────────────────────────────────────────────────────────────

@app.route('/safety-promotion/campaign/<int:cid>/send-email', methods=['POST'])
@require_login
def sp_send_campaign_email(cid):
    camp    = SafetyCampaign.query.get_or_404(cid)
    dept_id = request.form.get('dept_id') or None
    recip   = _get_dist_list(dept_id)
    if not recip:
        flash('No active recipients in distribution list.', 'warning')
        return redirect(url_for('sp_campaign_detail', cid=cid))
    emails     = [r.email for r in recip]
    dept_label = Department.query.get(int(dept_id)).name if dept_id else 'All'
    body       = ('<p><b>Type:</b> ' + (camp.campaign_type or '') +
                  ' &nbsp;&middot;&nbsp; <b>Period:</b> ' + (camp.start_date or '') +
                  ' to ' + (camp.end_date or '') + '</p>'
                  '<p>' + (camp.description or '').replace('\n','<br/>') + '</p>'
                  + ('<p><b>Objectives:</b><br/>' + (camp.objectives or '').replace('\n','<br/>') + '</p>'
                     if camp.objectives else ''))
    subject    = '[Safety Campaign] ' + camp.title
    html       = _email_html(camp.title, 'Safety Campaign', body, '#d97706')
    sent, err  = _do_send(emails, subject, html)
    _write_log('Campaign', cid, subject, sent, dept_label,
               'Sent' if not err else 'Failed', err)
    flash(('⚠ SMTP not configured — email not sent.' if err == 'SMTP_NOT_CONFIGURED' else f'Campaign emailed to {sent} recipients.'), 'warning' if err == 'SMTP_NOT_CONFIGURED' else 'success')
    return redirect(url_for('sp_campaign_detail', cid=cid))


# ── Send Lesson ───────────────────────────────────────────────────────────────

# NOTE: duplicate route removed — served by lesson_send_email (above)
def sp_send_lesson_email(lid):
    ll   = LessonLearned.query.get_or_404(lid)
    recip = _get_dist_list()
    if not recip:
        flash('No active recipients in distribution list.', 'warning')
        return redirect(url_for('sp_lesson_detail', lid=lid))
    emails = [r.email for r in recip]
    body   = ('<p><b>Category:</b> ' + (ll.category or '') +
              ' &nbsp;&middot;&nbsp; <b>Date:</b> ' + (ll.date or '') +
              '<br/><b>Author:</b> ' + (ll.author or 'Safety') + '</p>'
              '<hr style="border:none;border-top:1px solid #e5e7eb;margin:14px 0"/>'
              '<p>' + (ll.description or '').replace('\n','<br/>') + '</p>'
              + ('<p><b>Lesson:</b><br/>' + (ll.lesson or '').replace('\n','<br/>') + '</p>'
                 if ll.lesson else '')
              + ('<p><b>Recommendations:</b><br/>' + (ll.recommendations or '').replace('\n','<br/>') + '</p>'
                 if ll.recommendations else ''))
    subject = '[Lessons Learned] ' + ll.title
    html    = _email_html(ll.title, 'Lessons Learned', body, '#15803d')
    sent, err = _do_send(emails, subject, html)
    _write_log('Lesson', lid, subject, sent, 'All',
               'Sent' if not err else 'Failed', err)
    flash(('⚠ SMTP not configured — email not sent.' if err == 'SMTP_NOT_CONFIGURED' else f'Lesson emailed to {sent} recipients.'), 'warning' if err == 'SMTP_NOT_CONFIGURED' else 'success')
    return redirect(url_for('sp_lesson_detail', lid=lid))


# ── Public Survey Response ────────────────────────────────────────────────────

@app.route('/safety-promotion/survey/<int:sid>/respond-public', methods=['GET','POST'])
def sp_survey_respond_public(sid):
    s = SafetySurvey.query.get_or_404(sid)
    if s.status != 'Active':
        return render_template('spi/sp_survey_closed.html', survey=s)

    import json as _j
    questions = []
    try:
        raw = _j.loads(s.questions or '[]')
        # Normalize: strings → dicts so template can call q.get('text')
        for q in raw:
            if isinstance(q, str):
                questions.append({'text': q, 'type': 'text'})
            elif isinstance(q, dict):
                questions.append(q)
    except Exception:
        pass

    if request.method == 'POST':
        f       = request.form
        is_anon = bool(f.get('is_anonymous'))
        answers = {str(i): f.get('q_' + str(i), '') for i in range(len(questions))}
        db.session.add(SurveyResponse(
            survey_id=sid,
            respondent_name='' if is_anon else (f.get('respondent_name','') or f.get('name','')),
            respondent_email='' if is_anon else (f.get('respondent_email','') or f.get('email','')),
            department_id=int(f['department_id']) if f.get('department_id') else None,
            is_anonymous=is_anon,
            answers=_j.dumps(answers),
            ip_address=request.remote_addr,
        ))
        s.response_count = (s.response_count or 0) + 1
        db.session.commit()
        return render_template('spi/sp_survey_thanks.html', survey=s)

    return render_template('spi/sp_survey_public.html', survey=s, questions=questions)


# ── Survey Response Tracking ──────────────────────────────────────────────────

@app.route('/safety-promotion/survey/<int:sid>/responses')
@require_login
def sp_survey_responses(sid):
    s         = SafetySurvey.query.get_or_404(sid)
    responses = SurveyResponse.query.filter_by(survey_id=sid)\
                    .order_by(SurveyResponse.submitted_at.desc()).all()
    total_sent    = s.target_count or 0
    resp_count    = len(responses)
    response_rate = round((resp_count / total_sent * 100) if total_sent > 0 else 0, 1)

    import json as _j
    from collections import Counter, defaultdict

    # Parse questions
    questions = []
    try:
        raw = _j.loads(s.questions or '[]')
        for q in raw:
            if isinstance(q, str):
                questions.append({'text': q, 'type': 'text'})
            elif isinstance(q, dict):
                questions.append(q)
    except Exception:
        pass

    # Pre-parse answers for each response so template can display them
    parsed_responses = []
    for resp in responses:
        try:
            ans_raw = _j.loads(resp.answers or '{}')
        except Exception:
            ans_raw = {}
        # Answers may be a dict {str(index): value} or a list [{question_index, answer}]
        if isinstance(ans_raw, list):
            ans = {str(item.get('question_index', i)): item.get('answer', '')
                   for i, item in enumerate(ans_raw) if isinstance(item, dict)}
        elif isinstance(ans_raw, dict):
            ans = ans_raw
        else:
            ans = {}
        parsed_responses.append({
            'obj':             resp,
            'answers':         ans,
            'answer_list':     [ans.get(str(i), '') for i in range(len(questions))],
        })

    # Department participation
    dept_counts = Counter(r.department_id for r in responses if r.department_id)
    all_depts   = Department.query.all()
    dept_names  = {d.id: d.name for d in all_depts}

    # Question analytics: for each question, tally all answers
    q_analytics = []
    for i, q in enumerate(questions):
        tally = Counter()
        for pr in parsed_responses:
            ans = pr['answer_list'][i]
            if ans:
                tally[ans] += 1
        q_analytics.append({'question': q, 'tally': dict(tally), 'total': len(tally)})

    # Timeline: responses by date
    from collections import OrderedDict
    timeline = Counter()
    for r in responses:
        if r.submitted_at:
            timeline[r.submitted_at.strftime('%d %b')] += 1
    timeline = dict(sorted(timeline.items()))

    anon_count    = sum(1 for r in responses if r.is_anonymous)
    tracked_count = len(responses) - anon_count

    return render_template('spi/sp_survey_responses.html',
                           survey=s,
                           responses=parsed_responses,
                           response_rate=response_rate,
                           questions=questions,
                           q_analytics=q_analytics,
                           dept_counts=dept_counts,
                           dept_names=dept_names,
                           all_depts=all_depts,
                           total_sent=total_sent,
                           anon_count=anon_count,
                           tracked_count=tracked_count,
                           timeline=timeline)



@app.route('/safety-promotion/survey/<int:sid>/response/<int:rid>')
@require_login
def sp_survey_response_detail(sid, rid):
    """Individual survey response detail view."""
    import json as _j
    s    = SafetySurvey.query.get_or_404(sid)
    resp = SurveyResponse.query.get_or_404(rid)
    questions = []
    try:
        raw = _j.loads(s.questions or '[]')
        for q in raw:
            questions.append({'text': q} if isinstance(q, str) else q)
    except Exception:
        pass
    try:
        answers = _j.loads(resp.answers or '{}')
    except Exception:
        answers = {}
    # Mobile submits answers as a list; web stores as a dict — normalise to dict
    if isinstance(answers, list):
        answers = {str(item.get('question_index', i)): item.get('answer', '')
                   for i, item in enumerate(answers) if isinstance(item, dict)}
    qa_pairs = [(questions[i], answers.get(str(i), '—'))
                for i in range(len(questions))]
    return render_template('spi/sp_survey_response_detail.html',
                           survey=s, resp=resp, qa_pairs=qa_pairs)



# ═══════════════════════════════════════════════════════════════════════════════
#  SAG PORTAL — Separate login, separate access, same PostgreSQL database
#  /sag-login  /sag/dashboard  /sag/action/<id>
#  SAG members see ONLY their own assigned actions — nothing else.
# ═══════════════════════════════════════════════════════════════════════════════

def is_sag_logged_in():
    return session.get('sag_logged_in') is True

def require_sag(f):
    import functools
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        if not is_sag_logged_in():
            return redirect(url_for('sag_login', next=request.path))
        return f(*args, **kwargs)
    return decorated


@app.route('/sag-login', methods=['GET', 'POST'])
@csrf.exempt
def sag_login():
    if is_sag_logged_in():
        return redirect(url_for('sag_dashboard'))
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        user = User.query.filter_by(username=username, is_active=True).first()
        if user and check_pw(password, user.password_hash):
            session['sag_logged_in'] = True
            session['sag_user']      = user.username
            session['sag_name']      = user.full_name or user.username
            session['sag_role']      = user.sag_role or user.role
            session['sag_dept_id']   = user.department_id
            session.permanent        = True
            user.last_login = datetime.utcnow()
            db.session.commit()
            return redirect(url_for('sag_dashboard'))
        error = 'Invalid username or password.'
    return render_template('portal/sag_login.html', error=error)


@app.route('/sag-logout')
def sag_logout():
    for k in ['sag_logged_in','sag_user','sag_name','sag_role','sag_dept_id']:
        session.pop(k, None)
    return redirect(url_for('sag_login'))


@app.route('/sag/dashboard')
@require_sag
def sag_dashboard():
    check_overdue_actions()
    username = session.get('sag_user', '')
    # Load only valid actions — filter orphaned ones (source deleted)
    all_my = Action.query.filter_by(sag_member=username).order_by(Action.due_date).all()
    my_actions = []
    for _a in all_my:
        try:
            if _a.source == 'Audit Finding' and _a.linked_ref_id:
                if not AuditFinding.query.get(_a.linked_ref_id):
                    continue  # source finding deleted — skip orphan
            if _a.source in ('Hazard Report','Hazard') and _a.hazard_id:
                if not Hazard.query.get(_a.hazard_id):
                    continue  # source hazard deleted — skip orphan
        except Exception:
            pass
        my_actions.append(_a)
    open_c     = sum(1 for a in my_actions if a.status == 'Open')
    prog_c     = sum(1 for a in my_actions if a.status == 'In Progress')
    overdue_c  = sum(1 for a in my_actions if a.status == 'Overdue')
    review_c   = sum(1 for a in my_actions
                     if a.status in ('Mitigation Implemented','Under Safety Review'))
    returned_c = sum(1 for a in my_actions if a.status == 'Returned')
    closed_c   = sum(1 for a in my_actions if a.status == 'Closed')
    active_actions = [a for a in my_actions if a.status != 'Closed']
    closed_actions = [a for a in my_actions if a.status == 'Closed'][-10:]
    return render_template('portal/sag_dashboard.html',
                           active_actions=active_actions,
                           closed_actions=closed_actions,
                           open_c=open_c, prog_c=prog_c,
                           overdue_c=overdue_c, review_c=review_c,
                           returned_c=returned_c, closed_c=closed_c)


@app.route('/sag/action/<aid>', methods=['GET', 'POST'])
@require_sag
def sag_action_detail(aid):
    a = Action.query.get_or_404(aid)
    if a.sag_member != session.get('sag_user', ''):
        flash('⚠ This action is not assigned to you.', 'error')
        return redirect(url_for('sag_dashboard'))

    if request.method == 'POST':
        f          = request.form
        action_btn = f.get('action_btn', '')
        old_status = a.status

        if action_btn in ('save_progress', 'submit_review'):
            # Core fields (all workflows)
            a.root_cause             = f.get('root_cause',             a.root_cause or '')
            a.corrective_description = f.get('corrective_description', a.corrective_description or '')
            a.mitigation_description = f.get('mitigation_description', a.mitigation_description or '')
            a.safety_notes           = f.get('safety_notes',           a.safety_notes or '')
            a.follow_up_notes        = f.get('follow_up_notes',        a.follow_up_notes or '')
            a.evidence               = f.get('evidence',               a.evidence or '')
            if f.get('implementation_date'):
                a.implementation_date = f.get('implementation_date')
            if f.get('mitigation_status'):
                a.mitigation_status = f.get('mitigation_status')
            # Workflow-specific fields (stored in existing columns)
            if f.get('contributing_factors'):
                a.rejection_notes = f.get('contributing_factors')   # reuse for extra RCA detail
            if f.get('residual_risk'):
                a.safety_notes = (a.safety_notes or '') + ' | Residual Risk: ' + f.get('residual_risk','')
            if f.get('cap_responsible'):
                a.owner = f.get('cap_responsible', a.owner)
            if f.get('cap_target_date'):
                a.due_date = f.get('cap_target_date')
            if f.get('recovery_timeline'):
                a.follow_up_notes = (a.follow_up_notes or '') + ' | Recovery: ' + f.get('recovery_timeline','')
            if f.get('mitigation_status'):
                a.mitigation_status = f.get('mitigation_status')
            # Handle evidence file upload
            ev_file = request.files.get('evidence_file')
            if ev_file and ev_file.filename:
                try:
                    from werkzeug.utils import secure_filename as sf
                    import os
                    ext = ev_file.filename.rsplit('.', 1)[-1].lower() if '.' in ev_file.filename else ''
                    ALLOWED = {'pdf','doc','docx','xls','xlsx','png','jpg','jpeg','gif','bmp','txt','zip'}
                    if ext in ALLOWED:
                        fname = sf(f'evidence_{aid}_{ev_file.filename}')
                        upload_dir = app.config.get('UPLOAD_FOLDER', 'uploads')
                        os.makedirs(upload_dir, exist_ok=True)
                        ev_file.save(os.path.join(upload_dir, fname))
                        a.evidence_filename = fname
                    else:
                        flash(f'⚠ File type .{ext} not allowed.', 'error')
                except Exception as _e:
                    flash(f'⚠ File upload failed: {str(_e)[:80]}', 'error')

        if action_btn == 'save_progress':
            if a.status in ('Open', 'Assigned'):
                a.status = 'In Progress'
            log_action_history(aid, session['sag_user'], old_status, a.status,
                               'Progress saved by SAG member', 'progress')
            db.session.commit()
            sync_report_status(a.hazard_id)
            db.session.commit()
            flash('✓ Progress saved successfully.', 'success')

        elif action_btn == 'submit_review':
            a.mitigation_status = 'Completed'
            a.status = 'Mitigation Implemented'
            log_action_history(aid, session['sag_user'], old_status,
                               'Mitigation Implemented',
                               'Submitted for Safety Review by ' + session.get('sag_name','SAG'),
                               'status')
            db.session.commit()
            # SAG completion → set to Awaiting Safety Approval (NOT auto-close)
            if a.hazard_id:
                try:
                    haz = Hazard.query.filter_by(id=a.hazard_id).first()
                    if haz and haz.status not in ('Closed', 'Awaiting Safety Approval'):
                        haz.status = 'Awaiting Safety Approval'
                        db.session.commit()
                except Exception:
                    pass
            sync_report_status(a.hazard_id)
            db.session.commit()
            flash('✓ Mitigation submitted for Safety Review. Awaiting Safety Manager approval before closure.', 'success')

        return redirect(url_for('sag_action_detail', aid=aid))

    # Load full source records for SAG member visibility
    hazard_rep    = HazardReport.query.filter_by(hazard_id=a.hazard_id).first() if a.hazard_id else None
    finding       = AuditFinding.query.get(a.linked_ref_id) if a.linked_ref_id and a.source == 'Audit Finding' else None
    investigation = Investigation.query.get(a.linked_ref_id) if a.linked_ref_id and a.source == 'Investigation' else None
    spi_ind       = SPIIndicator.query.get(a.spi_id) if a.spi_id else None
    ra            = RiskAssessment.query.get(a.linked_ref_id) if a.linked_ref_id and a.source in ('Risk Assessment','RA') else None

    history = ActionHistory.query.filter_by(action_id=aid)                  .order_by(ActionHistory.changed_at.desc()).limit(15).all()
    return render_template('portal/sag_action_detail.html',
        a=a, history=history,
        hazard_rep=hazard_rep, finding=finding,
        investigation=investigation, spi_ind=spi_ind, ra=ra)


# ── SAG Governance Dashboard (Safety Admin view inside main SMS) ───────────────

@app.route('/sag/governance')
@require_sag
def sag_governance():
    check_overdue_actions()
    total_open     = Action.query.filter(Action.status.notin_(['Closed'])).count()
    overdue        = Action.query.filter_by(status='Overdue').count()
    pending_review = Action.query.filter(
        Action.status.in_(['Mitigation Implemented','Under Safety Review'])).count()
    returned       = Action.query.filter_by(status='Returned').count()
    closed_month   = Action.query.filter(
        Action.status=='Closed',
        Action.closed_date >= date.today().replace(day=1).isoformat()
    ).count()
    high_risk = Action.query.filter(
        Action.priority=='High', Action.status.notin_(['Closed'])
    ).order_by(Action.due_date).limit(10).all()
    for_review = Action.query.filter(
        Action.status.in_(['Mitigation Implemented','Under Safety Review'])
    ).order_by(Action.due_date).limit(15).all()
    overdue_list = Action.query.filter_by(status='Overdue').order_by(Action.due_date).limit(15).all()
    returned_list = Action.query.filter_by(status='Returned').order_by(Action.due_date).limit(10).all()
    unassigned = Action.query.filter(
        (Action.sag_member == None) | (Action.sag_member == ''),
        Action.status.notin_(['Closed'])
    ).order_by(Action.due_date).limit(10).all()
    from sqlalchemy import func as sqf
    src_data = db.session.query(Action.source, sqf.count(Action.id)).filter(
        Action.status.notin_(['Closed'])
    ).group_by(Action.source).order_by(sqf.count(Action.id).desc()).all()
    dept_perf = []
    for dept in Department.query.all():
        d_total   = Action.query.filter_by(department_id=dept.id).count()
        d_open    = Action.query.filter_by(department_id=dept.id).filter(
                        Action.status.notin_(['Closed'])).count()
        d_overdue = Action.query.filter_by(department_id=dept.id, status='Overdue').count()
        d_closed  = Action.query.filter_by(department_id=dept.id, status='Closed').count()
        if d_total > 0:
            dept_perf.append({'dept': dept, 'total': d_total, 'open': d_open,
                              'overdue': d_overdue, 'closed': d_closed,
                              'rate': round(d_closed/d_total*100)})
    dept_perf.sort(key=lambda x: x['overdue'], reverse=True)
    sag_members = User.query.filter_by(is_active=True).all()
    return render_template('action/sag_governance.html',
                           total_open=total_open, overdue=overdue,
                           pending_review=pending_review, returned=returned,
                           closed_month=closed_month,
                           high_risk=high_risk, for_review=for_review,
                           overdue_list=overdue_list, returned_list=returned_list,
                           unassigned=unassigned, src_data=src_data,
                           dept_perf=dept_perf, sag_members=sag_members)


@app.route('/sag/assign/<aid>', methods=['POST'])
@require_sag
def sag_assign(aid):
    a = Action.query.get_or_404(aid)
    old_owner = a.sag_member or 'Unassigned'
    a.sag_member    = request.form.get('sag_member', '')
    a.department_id = int(request.form['department_id']) if request.form.get('department_id') else a.department_id
    a.priority      = request.form.get('priority', a.priority)
    a.due_date      = request.form.get('due_date', a.due_date)
    a.root_cause    = request.form.get('root_cause', a.root_cause)
    if a.status == 'Open':
        a.status = 'Assigned'
    log_action_history(aid, session.get('admin_name','Admin'),
                       old_owner, a.sag_member,
                       f'Assigned to {a.sag_member}', 'assignment')
    db.session.commit()
    sync_report_status(a.hazard_id)
    db.session.commit()
    if a.sag_member:
        push_notify_by_name(a.sag_member,
            f'Action Assigned — {aid}',
            f'You have been assigned a corrective action: {a.description[:80] if a.description else aid}.',
            'action_assigned', 'action', aid)
    flash(f'✓ Action {aid} assigned to {a.sag_member}.', 'success')
    return redirect(request.form.get('return_url', url_for('sag_governance')))


@app.route('/actions/<aid>/history')
@require_login
def action_history_view(aid):
    a       = Action.query.get_or_404(aid)
    history = ActionHistory.query.filter_by(action_id=aid)                  .order_by(ActionHistory.changed_at.desc()).all()
    return render_template('action/action_history.html', a=a, history=history)


@app.route('/risk-matrix')
@require_login
def risk_matrix():
    return render_template('risk/risk_matrix.html')


# ═══════════════════════════════════════════════════════════════════════════════
#  AUDIT MANAGEMENT MODULE ROUTES
#  ICAO Annex 19 / IOSA ISM compliant
#  Added as extension — existing routes unchanged
# ═══════════════════════════════════════════════════════════════════════════════

# ── Checklist templates per department ───────────────────────────────────────
CHECKLIST_TEMPLATES = {
    'default': [
        ('SOP Compliance',      'ISM 1.1.1',  'Are Standard Operating Procedures current, approved and accessible to all relevant personnel?'),
        ('SOP Compliance',      'ISM 1.1.2',  'Have SOPs been reviewed within the required timeframe?'),
        ('Training Records',    'ISM 2.1.1',  'Are training records complete, current and properly filed for all personnel?'),
        ('Training Records',    'ISM 2.1.2',  'Do all personnel hold valid certifications required for their role?'),
        ('Safety Reporting',    'ISM 3.1.1',  'Is the safety reporting system accessible and promoted to all staff?'),
        ('Safety Reporting',    'ISM 3.1.2',  'Are safety reports reviewed and actioned within defined timeframes?'),
        ('Risk Management',     'ISM 4.1.1',  'Are hazard identification processes implemented and records maintained?'),
        ('Risk Management',     'ISM 4.1.2',  'Are risk assessments reviewed and updated when changes occur?'),
        ('Operational Procedures', 'ISM 5.1.1', 'Are operational procedures aligned with current regulatory requirements?'),
        ('Operational Procedures', 'ISM 5.1.2', 'Are emergency/contingency procedures known and practised by personnel?'),
    ],
    'FO': [
        ('SOP Compliance',      'FO-1.1',  'Are Operations Manual revisions current and controlled?'),
        ('SOP Compliance',      'FO-1.2',  'Are MEL/CDL procedures understood and correctly applied?'),
        ('Training Records',    'FO-2.1',  'Are all flight crew recency requirements met (line checks, simulator)?'),
        ('Training Records',    'FO-2.2',  'Are CRM and UPRT training records current for all pilots?'),
        ('Safety Reporting',    'FO-3.1',  'Are ASR reports submitted within 72 hours of occurrence?'),
        ('Safety Reporting',    'FO-3.2',  'Are TCAS RA events reported and entered into the safety system?'),
        ('Risk Management',     'FO-4.1',  'Are NOTAM and weather briefing procedures followed for all flights?'),
        ('Risk Management',     'FO-4.2',  'Are fuel policy and alternates planned per company policy?'),
        ('Operational Procedures', 'FO-5.1', 'Are sterile cockpit procedures enforced during critical phases of flight?'),
        ('Operational Procedures', 'FO-5.2', 'Are fatigue risk management procedures followed and documented?'),
    ],
    'ME': [
        ('SOP Compliance',      'ME-1.1',  'Are maintenance procedures documented and consistent with approved data?'),
        ('SOP Compliance',      'ME-1.2',  'Are tooling calibration records maintained and current?'),
        ('Training Records',    'ME-2.1',  'Do all certifying engineers hold valid licences and type ratings?'),
        ('Training Records',    'ME-2.2',  'Are Human Factors in Maintenance training records current?'),
        ('Safety Reporting',    'ME-3.1',  'Are occurrence reports filed for all significant maintenance events?'),
        ('Safety Reporting',    'ME-3.2',  'Are defect reporting and follow-up processes properly implemented?'),
        ('Risk Management',     'ME-4.1',  'Are safety risk assessments conducted before non-routine tasks?'),
        ('Risk Management',     'ME-4.2',  'Are foreign object damage (FOD) prevention procedures in place?'),
        ('Operational Procedures', 'ME-5.1', 'Are shift handover procedures formally documented and followed?'),
        ('Operational Procedures', 'ME-5.2', 'Are critical maintenance tasks subject to independent inspection?'),
    ],
    'GO': [
        ('SOP Compliance',      'GO-1.1',  'Are ramp operations procedures current and followed by all ground staff?'),
        ('SOP Compliance',      'GO-1.2',  'Are aircraft loading instructions and mass & balance procedures complied with?'),
        ('Training Records',    'GO-2.1',  'Are all ramp agents trained and current on vehicle airside driving?'),
        ('Training Records',    'GO-2.2',  'Are dangerous goods handling training records maintained for all staff?'),
        ('Safety Reporting',    'GO-3.1',  'Are ramp incidents and near-misses reported to the safety system?'),
        ('Risk Management',     'GO-4.1',  'Are FOD inspection procedures conducted before aircraft movement?'),
        ('Operational Procedures', 'GO-5.1', 'Are pushback procedures followed including communication protocols?'),
        ('Operational Procedures', 'GO-5.2', 'Are fuelling safety procedures and bonding requirements enforced?'),
    ],
    'CC': [
        ('SOP Compliance',      'CC-1.1',  'Are cabin crew procedures consistent with approved cabin safety manual?'),
        ('Training Records',    'CC-2.1',  'Are all cabin crew recurrent safety training records current?'),
        ('Training Records',    'CC-2.2',  'Are SEP (Safety & Emergency Procedures) drills completed on schedule?'),
        ('Safety Reporting',    'CC-3.1',  'Are cabin safety incidents reported through the SMS system?'),
        ('Risk Management',     'CC-4.1',  'Are pre-flight safety checks documented and completed for all flights?'),
        ('Operational Procedures', 'CC-5.1', 'Are passenger safety briefings conducted per approved procedure?'),
        ('Operational Procedures', 'CC-5.2', 'Are turbulence/emergency protocols reviewed and practised by crew?'),
    ],
    'SD': [
        ('SOP Compliance',      'SD-1.1',  'Is the SMS manual current, approved and distributed to all departments?'),
        ('Training Records',    'SD-2.1',  'Have all staff completed SMS awareness training within the required period?'),
        ('Safety Reporting',    'SD-3.1',  'Are all safety reports triaged, investigated and actioned within KPI timelines?'),
        ('Safety Reporting',    'SD-3.2',  'Are safety statistics reported to management at defined intervals?'),
        ('Risk Management',     'SD-4.1',  'Is the hazard register reviewed and updated quarterly?'),
        ('Risk Management',     'SD-4.2',  'Are SPI/SPT targets reviewed by the Safety Review Board?'),
        ('Operational Procedures', 'SD-5.1', 'Are Safety Review Board meetings held per schedule with full attendance?'),
        ('Operational Procedures', 'SD-5.2', 'Is the audit programme implemented as planned for the current year?'),
    ],
}

def get_checklist_template(dept_code):
    return CHECKLIST_TEMPLATES.get(dept_code, CHECKLIST_TEMPLATES['default'])


# ─── AUDIT PLAN ───────────────────────────────────────────────────────────────
# ─── CHECKLIST TEMPLATE MANAGEMENT ──────────────────────────────────────────

@app.route('/audit-checklists')
@require_login
def checklist_templates():
    """List all department checklist templates."""
    try:
        templates = ChecklistTemplate.query.filter_by(is_active=True).all()
    except Exception:
        try:
            templates = ChecklistTemplate.query.all()
        except Exception:
            templates = []
    return render_template('audit/checklist_templates.html',
                           templates=templates)


@app.route('/audit-checklists/<int:dept_id>', methods=['GET','POST'])
@require_login
def checklist_template_dept(dept_id):
    """
    Editable checklist template for a specific department.
    GET  = show current template with edit UI
    POST = save updated template as new version → becomes active
    """
    dept = Department.query.get_or_404(dept_id)
    audit_type = request.args.get('audit_type', 'Internal')

    if request.method == 'POST':
        action = request.form.get('action', 'save')

        if action == 'save':
            # Deactivate old templates for this dept+type (safe for PostgreSQL)
            try:
                old_tmpls = ChecklistTemplate.query.filter_by(
                    department_id=dept_id, audit_type=audit_type
                ).all()
                for ot in old_tmpls:
                    ot.is_active = False
                db.session.flush()
            except Exception:
                pass

            # Get version number
            last = ChecklistTemplate.query.filter_by(
                department_id=dept_id, audit_type=audit_type
            ).order_by(ChecklistTemplate.version.desc()).first()
            ver = (last.version + 1) if last else 1

            tmpl = ChecklistTemplate(
                department_id=dept_id,
                audit_type=audit_type,
                name=f'{dept.name} — {audit_type} Audit Checklist v{ver}',
                version=ver,
                is_active=True,
            )
            db.session.add(tmpl)
            db.session.flush()

            # Save items from form
            refs     = request.form.getlist('item_ref')
            cats     = request.form.getlist('category')
            qs       = request.form.getlist('question')
            iosa_refs= request.form.getlist('iosa_ref')
            for i, (ref, cat, q, iosa) in enumerate(zip(refs, cats, qs, iosa_refs)):
                if q.strip():
                    db.session.add(ChecklistTemplateItem(
                        template_id=tmpl.id,
                        item_ref=ref.strip(),
                        category=cat.strip(),
                        question=q.strip(),
                        iosa_ref=iosa.strip(),
                        sequence=i,
                    ))

            db.session.commit()
            flash(f'✓ Checklist saved as version {ver}. Will be used for future {dept.name} audits.', 'success')

        elif action == 'add_item':
            # Quick add single item
            tmpl = ChecklistTemplate.query.filter_by(
                department_id=dept_id, audit_type=audit_type, is_active=True
            ).first()
            if tmpl:
                seq = len(tmpl.items)
                db.session.add(ChecklistTemplateItem(
                    template_id=tmpl.id,
                    item_ref=request.form.get('item_ref',''),
                    category=request.form.get('category','General'),
                    question=request.form.get('question',''),
                    iosa_ref=request.form.get('iosa_ref',''),
                    sequence=seq,
                ))
                db.session.commit()
                flash('✓ Item added.', 'success')

        return redirect(url_for('checklist_template_dept',
                                dept_id=dept_id, audit_type=audit_type))

    # GET — load active template (safe for PostgreSQL)
    try:
        tmpl = ChecklistTemplate.query.filter_by(
            department_id=dept_id, audit_type=audit_type, is_active=True
        ).first()
    except Exception:
        tmpl = None
    try:
        all_versions = ChecklistTemplate.query.filter_by(
            department_id=dept_id, audit_type=audit_type
        ).order_by(ChecklistTemplate.version.desc()).all()
    except Exception:
        all_versions = []
    # Group by category for display
    grouped = {}
    if tmpl:
        for item in tmpl.items:
            grouped.setdefault(item.category, []).append(item)

    audit_types = ['Internal', 'External', 'IOSA', 'Regulatory', 'Safety']
    return render_template('audit/checklist_template_edit.html',
                           dept=dept, tmpl=tmpl, grouped=grouped,
                           audit_type=audit_type, audit_types=audit_types,
                           all_versions=all_versions)


@app.route('/audit-findings-report')
@require_login
def audit_findings_report():
    """System-wide Audit Findings List Report with filters."""
    dept_f   = request.args.get('dept', '')
    sev_f    = request.args.get('severity', '')
    status_f = request.args.get('status', '')
    audit_f  = request.args.get('audit', '')
    q_f      = request.args.get('q', '').strip()

    query = AuditFinding.query
    if dept_f:
        # Filter via schedule → department
        query = query.join(AuditSchedule, AuditFinding.schedule_id == AuditSchedule.id)                     .filter(AuditSchedule.department_id == int(dept_f))
    if sev_f:
        query = query.filter(AuditFinding.severity == sev_f)
    if status_f:
        query = query.filter(AuditFinding.status == status_f)
    if audit_f:
        query = query.filter(AuditFinding.schedule_id == audit_f)
    if q_f:
        query = query.filter(
            AuditFinding.description.ilike(f'%{q_f}%') |
            AuditFinding.root_cause.ilike(f'%{q_f}%') |
            AuditFinding.finding_ref.ilike(f'%{q_f}%')
        )

    findings = query.order_by(AuditFinding.created_at.desc()).all()

    # Stats
    total    = AuditFinding.query.count()
    open_c   = AuditFinding.query.filter_by(status='Open').count()
    closed_c = AuditFinding.query.filter_by(status='Closed').count()
    overdue_c= AuditFinding.query.filter_by(status='Overdue').count()

    schedules = AuditSchedule.query.all()

    return render_template('audit/audit_findings_report.html',
                           findings=findings, total=total,
                           open_c=open_c, closed_c=closed_c, overdue_c=overdue_c,
                           dept_f=dept_f, sev_f=sev_f, status_f=status_f,
                           audit_f=audit_f, q_f=q_f, schedules=schedules)


@app.route('/audit-schedule-calendar')
@require_login
def audit_schedule_calendar():
    """Calendar view of all scheduled audits."""
    year  = int(request.args.get('year',  datetime.now().year))
    month = int(request.args.get('month', datetime.now().month))
    dept_f= request.args.get('dept', '')

    # Get all schedules for the selected month/year
    from calendar import monthrange
    _, days_in_month = monthrange(year, month)

    q = AuditSchedule.query
    if dept_f:
        q = q.filter_by(department_id=int(dept_f))
    all_schedules = q.all()

    # Group by day for calendar display
    cal_data = {}  # day → list of schedules
    for s in all_schedules:
        if s.scheduled_date:
            try:
                from datetime import date as dt_date
                sd = dt_date.fromisoformat(s.scheduled_date)
                if sd.year == year and sd.month == month:
                    cal_data.setdefault(sd.day, []).append(s)
            except Exception:
                pass

    # Prev/next month
    if month == 1:
        prev_year, prev_month = year - 1, 12
    else:
        prev_year, prev_month = year, month - 1
    if month == 12:
        next_year, next_month = year + 1, 1
    else:
        next_year, next_month = year, month + 1

    import calendar
    first_weekday = calendar.monthrange(year, month)[0]  # 0=Mon

    MONTH_NAMES = ['January','February','March','April','May','June',
                   'July','August','September','October','November','December']

    return render_template('audit/audit_schedule_calendar.html',
                           year=year, month=month,
                           month_name=MONTH_NAMES[month-1],
                           days_in_month=days_in_month,
                           first_weekday=first_weekday,
                           cal_data=cal_data,
                           prev_year=prev_year, prev_month=prev_month,
                           next_year=next_year, next_month=next_month,
                           dept_f=dept_f)


@app.route('/audit-plans')
@require_login
def audit_plans():
    year_f     = request.args.get('year', datetime.now().year, type=int)
    dept_f     = request.args.get('dept', '', type=str)
    month_f    = request.args.get('month', 0, type=int)
    all_plans  = AuditPlan.query.order_by(AuditPlan.year.desc(), AuditPlan.month).all()
    years      = list(range(datetime.now().year - 1, datetime.now().year + 3))
    this_month = datetime.now().month
    this_year  = datetime.now().year

    # Build monthly grid
    grid = {m: [] for m in range(1, 13)}
    year_plans = []
    for p in all_plans:
        if p.year == year_f:
            # Apply dept filter
            if dept_f and str(p.department_id) != dept_f:
                continue
            # Apply month filter
            if month_f and p.month != month_f:
                continue
            year_plans.append(p)
            if p.month:
                grid[p.month].append(p)

    total_p     = len(year_plans)
    completed_p = sum(1 for p in year_plans if p.status == 'Completed')
    active_p    = sum(1 for p in year_plans if p.status == 'Active')

    # Alert data — audits THIS month
    alerts_this_month = [
        p for p in all_plans
        if p.year == this_year and p.month == this_month and p.status != 'Completed'
    ]
    # Upcoming — next 2 months
    upcoming = [
        p for p in all_plans
        if p.year == this_year
        and p.month in [this_month + 1, this_month + 2]
        and p.status != 'Completed'
    ]
    # Overdue — past months this year, not completed
    overdue_plans = [
        p for p in all_plans
        if p.year == this_year
        and p.month is not None
        and p.month < this_month
        and p.status != 'Completed'
    ]

    return render_template('audit/audit_plan_list.html',
                           plans=year_plans, grid=grid,
                           year_f=year_f, years=years,
                           dept_f=dept_f, month_f=month_f,
                           total_p=total_p, completed_p=completed_p, active_p=active_p,
                           alerts_this_month=alerts_this_month,
                           upcoming=upcoming,
                           overdue_plans=overdue_plans,
                           this_month=this_month)

@app.route('/audit-plans/new', methods=['GET', 'POST'])
@require_login
def new_audit_plan():
    if request.method == 'POST':
        f   = request.form
        pid = new_id('PLAN')
        p   = AuditPlan(
            id=pid,
            year=int(f['year']),
            month=int(f['month']) if f.get('month') else None,
            department_id=int(f['department_id']),
            audit_type=f['audit_type'],
            frequency=f.get('frequency', 'Annual'),
            responsible_manager=f.get('responsible_manager', ''),
            auditor_name=f.get('auditor_name', ''),
            planned_week=int(f['planned_week']) if f.get('planned_week') else None,
            scope=f.get('scope', ''),
            objectives=f.get('objectives', ''),
            iosa_reference=f.get('iosa_reference', ''),
            status='Planned'
        )
        db.session.add(p)
        db.session.commit()
        flash(f'✓ Audit Plan {pid} saved — {p.audit_type} for {f["year"]}.', 'success')
        return redirect(url_for('audit_plans') + f'?year={f["year"]}')
    years = list(range(datetime.now().year, datetime.now().year + 3))
    return render_template('audit/audit_plan_form.html', years=years)


@app.route('/audit-plans/<pid>/edit', methods=['GET','POST'])
@require_login
def edit_audit_plan(pid):
    p = AuditPlan.query.get_or_404(pid)
    if request.method == 'POST':
        f = request.form
        p.year               = int(f['year'])
        p.month              = int(f['month']) if f.get('month') else None
        p.department_id      = int(f['department_id'])
        p.audit_type         = f['audit_type']
        p.frequency          = f.get('frequency', p.frequency)
        p.responsible_manager = f.get('responsible_manager', p.responsible_manager)
        p.auditor_name       = f.get('auditor_name', p.auditor_name)
        p.planned_week       = int(f['planned_week']) if f.get('planned_week') else None
        p.scope              = f.get('scope', p.scope)
        p.objectives         = f.get('objectives', p.objectives)
        p.iosa_reference     = f.get('iosa_reference', p.iosa_reference)
        p.status             = f.get('status', p.status)
        db.session.commit()
        flash(f'✓ Audit Plan {p.id} updated.', 'success')
        return redirect(url_for('audit_plans') + f'?year={p.year}')
    years = list(range(datetime.now().year, datetime.now().year + 3))
    return render_template('audit/audit_plan_form.html', years=years, edit=p)

@app.route('/audit-plans/<pid>/delete', methods=['POST'])
@require_login
def delete_audit_plan(pid):
    p = AuditPlan.query.get_or_404(pid)
    year = p.year
    try:
        # Safely cascade: unlink schedules plan_id first, then delete schedules
        schedules = AuditSchedule.query.filter_by(plan_id=pid).all() if hasattr(AuditSchedule, 'plan_id') else (p.schedules if hasattr(p, 'schedules') else [])
        for s in schedules:
            # Each schedule: nullify checklist linked_finding_id, delete actions/findings/checklists
            findings = AuditFinding.query.filter_by(schedule_id=s.id).all()
            fids = [f.id for f in findings]
            if fids:
                AuditChecklist.query.filter(
                    AuditChecklist.linked_finding_id.in_(fids)
                ).update({'linked_finding_id': None}, synchronize_session=False)
                AuditAction.query.filter(
                    AuditAction.finding_id.in_(fids)
                ).delete(synchronize_session=False)
            AuditChecklist.query.filter_by(schedule_id=s.id).delete(synchronize_session=False)
            for f in findings:
                db.session.delete(f)
            db.session.flush()
            db.session.delete(s)
        db.session.flush()
        db.session.delete(p)
        db.session.commit()
        flash(f'✓ Audit Plan {pid} deleted.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'⚠ Could not delete {pid}: {str(e)[:120]}', 'error')
    return redirect(url_for('audit_plans') + f'?year={year}')

@app.route('/audit-plans/<pid>/complete', methods=['POST'])
@require_login
def complete_audit_plan(pid):
    """Mark an audit plan entry as Completed."""
    p = AuditPlan.query.get_or_404(pid)
    p.status = 'Completed'
    db.session.commit()
    flash(f'✓ Audit plan {p.id} marked as Completed.', 'success')
    return redirect(url_for('audit_plans'))

@app.route('/audit-plans/<pid>/schedule', methods=['POST'])
@require_login
def schedule_from_plan(pid):
    """Convert an audit plan entry into a scheduled audit."""
    plan = AuditPlan.query.get_or_404(pid)
    f    = request.form
    sid  = new_id('AUD')
    s = AuditSchedule(
        id=sid,
        plan_id=pid,
        department_id=plan.department_id,
        audit_type=plan.audit_type,
        scheduled_date=f['scheduled_date'],
        lead_auditor=f['lead_auditor'],
        audit_team=f.get('audit_team', ''),
        scope=plan.scope,
        objectives=plan.objectives,
        status='Planned'
    )
    db.session.add(s)
    # Auto-populate checklist from template
    dept = Department.query.get(plan.department_id)
    template = get_checklist_template(dept.code if dept else 'default')
    for idx, (cat, ref, question) in enumerate(template):
        item = AuditChecklist(
            schedule_id=sid, category=cat,
            item_ref=ref, question=question, sequence=idx
        )
        db.session.add(item)
    db.session.commit()
    flash(f'✓ Audit {sid} scheduled. Checklist auto-populated ({len(template)} items).', 'success')
    return redirect(url_for('audit_schedule'))


# ─── AUDIT SCHEDULE ───────────────────────────────────────────────────────────
@app.route('/audit-schedule')
@require_login
def audit_schedule():
    dept_f   = request.args.get('dept', '')
    status_f = request.args.get('status', '')
    page     = request.args.get('page', 1, type=int)
    q = AuditSchedule.query
    if dept_f:   q = q.filter_by(department_id=int(dept_f))
    if status_f: q = q.filter_by(status=status_f)
    pg        = q.order_by(AuditSchedule.scheduled_date).paginate(page=page, per_page=50, error_out=False)
    schedules = pg.items

    # Check/update overdue audit actions
    today = date.today().isoformat()
    changed = False
    for s in AuditSchedule.query.all():
        for f2 in s.findings:
            for a in f2.actions:
                if a.status in ('Open', 'In Progress') and a.due_date and a.due_date < today:
                    a.status = 'Overdue'
                    changed = True
    if changed:
        db.session.commit()

    return render_template('audit/audit_schedule.html', schedules=schedules,
                           dept_f=dept_f, status_f=status_f, pagination=pg)

@app.route('/audit-schedule/new', methods=['GET', 'POST'])
@require_login
def new_audit_schedule():
    """Create a scheduled audit without an existing plan (ad-hoc)."""
    if request.method == 'POST':
        f   = request.form
        sid = new_id('AUD')
        dept = Department.query.get(int(f['department_id']))
        s = AuditSchedule(
            id=sid, plan_id=None,
            department_id=int(f['department_id']),
            audit_type=f['audit_type'],
            scheduled_date=f['scheduled_date'],
            lead_auditor=f['lead_auditor'],
            audit_team=f.get('audit_team', ''),
            scope=f.get('scope', ''),
            objectives=f.get('objectives', ''),
            status='Planned'
        )
        db.session.add(s)
        # Auto-populate checklist
        template = get_checklist_template(dept.code if dept else 'default')
        for idx, (cat, ref, question) in enumerate(template):
            item = AuditChecklist(
                schedule_id=sid, category=cat,
                item_ref=ref, question=question, sequence=idx
            )
            db.session.add(item)
        db.session.commit()
        flash(f'✓ Audit {sid} created. Checklist auto-populated.', 'success')
        return redirect(url_for('audit_schedule'))
    return render_template('audit/audit_schedule_form.html')


# ─── AUDIT EXECUTION ─────────────────────────────────────────────────────────
@app.route('/audit-schedule/<sid>')
@require_login
def audit_execution(sid):
    s = AuditSchedule.query.get_or_404(sid)
    # Group checklist by category
    checklist = {}
    for item in sorted(s.checklist_items, key=lambda x: x.sequence):
        checklist.setdefault(item.category, []).append(item)
    total = len(s.checklist_items)
    done  = sum(1 for i in s.checklist_items if i.response)
    nc    = sum(1 for i in s.checklist_items if i.response == 'No')

    # Checklist NO items without findings
    no_items_without_findings = [
        i for cat_items in checklist.values()
        for i in cat_items
        if i.response == 'No' and not i.linked_finding_id
    ]
    all_no_have_findings = len(no_items_without_findings) == 0

    # All findings closed (using the new lifecycle status OR old AuditAction check)
    all_findings_closed = all(
        f.status == 'Closed' for f in s.findings
    ) if s.findings else True

    # AuditAction-based checks (legacy — some findings have AuditActions)
    findings_with_audit_actions = [f for f in s.findings if f.actions]
    all_findings_actioned = True  # satisfied if all findings are Closed
    all_actions_closed    = True
    all_verified          = True

    if findings_with_audit_actions:
        all_actions_closed = all(
            all(a.status == 'Closed' for a in f.actions)
            for f in findings_with_audit_actions
        )
        # Effectiveness is optional — don't block closure on it
        all_verified = True

    # A finding is OK if Closed, or its linked SAG Action is Closed
    def finding_ok(f):
        if f.status == 'Closed': return True
        if f.linked_action_id:
            la = Action.query.get(f.linked_action_id)
            return la is not None and la.status == 'Closed'
        return False
    findings_ready = all(finding_ok(f) for f in s.findings) if s.findings else True

    # CAN CLOSE: checklist complete + all findings closed/resolved
    can_close = (
        all_no_have_findings and
        findings_ready and
        all_actions_closed
    )

    # ── Safety Assurance: AVIs linked to this audit schedule ─────────────────
    linked_avis = []
    avi_pending_count = 0
    try:
        if AuditVerificationItem:
            linked_avis = AuditVerificationItem.query.filter_by(
                scheduled_audit_id=sid
            ).order_by(AuditVerificationItem.operational_risk).all()
            avi_pending_count = sum(
                1 for a in linked_avis
                if a.status in ('Scheduled', 'In Verification', 'Pending')
            )
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
        linked_avis = []
        avi_pending_count = 0

    # can_close is also blocked if there are unverified AVIs linked to this audit
    if avi_pending_count > 0:
        can_close = False

    return render_template('audit/audit_execution.html',
        s=s, checklist=checklist, total=total, done=done, nc=nc,
        can_close=can_close,
        no_items_without_findings=no_items_without_findings,
        all_no_have_findings=all_no_have_findings,
        all_findings_closed=all_findings_closed,
        all_findings_actioned=all_findings_actioned,
        all_actions_closed=all_actions_closed,
        all_verified=all_verified,
        linked_avis=linked_avis,
        avi_pending_count=avi_pending_count,
        today_date=date.today().isoformat())

@app.route('/audit-schedule/<sid>/start', methods=['POST'])
@require_login
def start_audit(sid):
    s = AuditSchedule.query.get_or_404(sid)
    s.status       = 'In Progress'
    s.actual_date  = date.today().isoformat()
    s.opening_meeting = request.form.get('opening_meeting', date.today().isoformat())

    # ── Safety Assurance: move linked AVIs Scheduled → In Verification ────────
    try:
        if AuditVerificationItem:
            avis = AuditVerificationItem.query.filter_by(
                scheduled_audit_id=sid, status='Scheduled'
            ).all()
            for avi in avis:
                avi.status = 'In Verification'
                db.session.add(avi)
    except Exception:
        pass

    # Always load LATEST active checklist template when starting audit
    if s.department_id:
        # Load latest active template for this dept
        tmpl = None
        try:
            tmpl = ChecklistTemplate.query.filter_by(
                department_id=s.department_id,
                audit_type=s.audit_type or 'Internal',
                is_active=True
            ).order_by(ChecklistTemplate.version.desc()).first()
        except Exception:
            pass

        if tmpl and tmpl.items:
            # Clear existing checklist items (replace with latest template)
            try:
                AuditChecklist.query.filter_by(schedule_id=sid).delete(
                    synchronize_session=False)
                db.session.flush()
            except Exception:
                pass
            for ti in tmpl.items:
                db.session.add(AuditChecklist(
                    schedule_id=sid,
                    category=ti.category,
                    item_ref=ti.item_ref,
                    question=ti.question,
                    sequence=ti.sequence,
                ))
            db.session.flush()
            flash(f'✓ Audit started. Loaded {len(tmpl.items)} items from {tmpl.name}.', 'success')
        else:
            # Fall back to static template only if no saved template exists
            if not AuditChecklist.query.filter_by(schedule_id=sid).first():
                dept = Department.query.get(s.department_id)
                static_tmpl = get_checklist_template(dept.code if dept else 'default')
                for cat, items in static_tmpl.items():
                    for seq, (ref, q) in enumerate(items):
                        db.session.add(AuditChecklist(
                            schedule_id=sid,
                            category=cat,
                            item_ref=ref,
                            question=q,
                            sequence=seq,
                        ))
                flash('✓ Audit started. Checklist loaded from default template.', 'success')

    db.session.commit()
    return redirect(url_for('audit_execution', sid=sid))

@app.route('/audit-schedule/<sid>/checklist', methods=['POST'])
@require_login
def save_checklist(sid):
    """
    Save checklist responses.
    For any item answered 'No' without a linked finding:
    Auto-create a Finding inheriting checklist metadata.
    """
    s = AuditSchedule.query.get_or_404(sid)
    new_findings = []

    for item in s.checklist_items:
        prev_resp = item.response
        item.response = request.form.get(f'resp_{item.id}', '')
        item.comment  = request.form.get(f'comment_{item.id}', '')
        item.evidence = request.form.get(f'evidence_{item.id}', '')

        # Handle evidence file upload per checklist item
        ef_key = f'evidence_file_{item.id}'
        if ef_key in request.files:
            ef = request.files[ef_key]
            if ef and ef.filename and allowed_file(ef.filename):
                from werkzeug.utils import secure_filename
                fn = f'CL{item.id}_{secure_filename(ef.filename)}'
                ef.save(os.path.join(app.config['UPLOAD_FOLDER'], fn))
                item.evidence_filename = fn

        # Auto-create Finding when answer flips to 'No' and no finding yet
        if item.response == 'No' and not item.linked_finding_id:
            count       = AuditFinding.query.filter_by(schedule_id=sid).count() + len(new_findings) + 1
            finding_ref = f'F-{count:03d}'
            fid = new_id('FND')
            # Get auditor-entered finding details (if provided via expanded form)
            _ftitle    = request.form.get(f'finding_title_{item.id}', '').strip()
            _fdesc     = request.form.get(f'finding_desc_{item.id}', '').strip()
            _fevidence = request.form.get(f'finding_evidence_{item.id}', '').strip()
            _fseverity = request.form.get(f'finding_severity_{item.id}', 'Minor')
            _frequire  = request.form.get(f'finding_require_{item.id}', '').strip()
            finding = AuditFinding(
                id           = fid,
                schedule_id  = sid,
                finding_ref  = finding_ref,
                finding_title = _ftitle or f'Non-Conformance: {item.item_ref}',
                description  = _fdesc or item.question or '',
                category     = item.category or 'Operational',
                severity     = _fseverity,
                standard_ref = _frequire or item.item_ref or '',
                requirement  = item.question or '',
                evidence     = _fevidence or item.comment or item.evidence or '',
                assigned_to  = s.lead_auditor or '',
                status       = 'Open',
            )
            db.session.add(finding)
            db.session.flush()
            item.linked_finding_id = fid
            new_findings.append(finding_ref)

        # If answer changed FROM 'No' to something else → unlink auto-finding if still Open
        elif item.response != 'No' and prev_resp == 'No' and item.linked_finding_id:
            linked = AuditFinding.query.get(item.linked_finding_id)
            if linked and linked.status == 'Open':
                db.session.delete(linked)
                item.linked_finding_id = None

    db.session.commit()

    if new_findings:
        flash(f'✓ Checklist saved. Auto-created findings: {", ".join(new_findings)} for NO items.', 'success')
    else:
        flash('✓ Checklist saved.', 'success')
    return redirect(url_for('audit_execution', sid=sid))

@app.route('/audit-schedule/<sid>/close', methods=['POST'])
@require_login
def close_audit(sid):
    s = AuditSchedule.query.get_or_404(sid)
    # Validate: all NO checklist items must have linked findings
    no_without = [i for i in s.checklist_items
                  if i.response == 'No' and not i.linked_finding_id]
    if no_without:
        refs = ', '.join(i.item_ref or f'item {i.id}' for i in no_without[:3])
        flash(f'✗ Cannot close: {len(no_without)} checklist NO item(s) still need findings ({refs}).', 'error')
        return redirect(url_for('audit_execution', sid=sid))
    # Validate: all findings must be Closed OR their linked SAG Action is Closed
    if s.findings:
        for finding in s.findings:
            if finding.status == 'Closed':
                continue  # already closed — OK

            # Check linked main Action (SAG workflow)
            linked_act = None
            if finding.linked_action_id:
                linked_act = Action.query.get(finding.linked_action_id)

            if linked_act and linked_act.status == 'Closed':
                # SAG action is closed — auto-close the finding too
                finding.status = 'Closed'
                finding.closure_date = date.today().isoformat()
                finding.closure_notes = f'Auto-closed: linked action {linked_act.id} was closed.'
                continue

            # Check legacy AuditActions
            if finding.actions:
                open_act = [a for a in finding.actions if a.status != 'Closed']
                if open_act:
                    flash(f'✗ Cannot close: Finding {finding.finding_ref} — '
                          f'{len(open_act)} action(s) not yet closed.', 'error')
                    return redirect(url_for('audit_execution', sid=sid))
            else:
                flash(f'✗ Cannot close: Finding {finding.finding_ref} is not Closed '
                      f'(status: {finding.status}). Close the linked action in the SAG portal first.', 'error')
                return redirect(url_for('audit_execution', sid=sid))
    s.status          = 'Completed'
    s.closure_date    = date.today().isoformat()
    s.closed_by       = request.form.get('closed_by', 'Safety Manager')
    s.final_remarks   = request.form.get('final_remarks', '')
    s.closing_meeting = request.form.get('closing_meeting', date.today().isoformat())
    # Save additional review fields if columns exist
    try:
        s.audit_result      = request.form.get('audit_result', 'Satisfactory')
        s.followup_required = request.form.get('followup_required', 'No')
    except Exception:
        pass
    # ── Safety Assurance: handle AVIs linked to this completed audit ──────────
    try:
        if AuditVerificationItem:
            unverified = AuditVerificationItem.query.filter(
                AuditVerificationItem.scheduled_audit_id == sid,
                AuditVerificationItem.status.in_(['Scheduled', 'In Verification'])
            ).all()
            for avi in unverified:
                # Return to Pending — auditor must reschedule in a future cycle
                avi.status = 'Pending'
                avi.scheduled_audit_id = None
                note = f'[Returned to Pending: Audit {sid} closed without verification — must be rescheduled]'
                avi.effectiveness_notes = ((avi.effectiveness_notes or '') + '\n' + note).strip()
                db.session.add(avi)
            if unverified:
                flash(f'⚠ {len(unverified)} verification item(s) were not verified during this audit and have been returned to Pending for rescheduling.', 'warning')
    except Exception:
        pass
    db.session.commit()
    flash(f'✓ Audit {sid} closed by Safety Department. Final report available.', 'success')
    return redirect(url_for('audit_execution', sid=sid))


# ─── AUDIT FINDINGS ───────────────────────────────────────────────────────────
@app.route('/audit-schedule/<sid>/findings/new', methods=['POST'])
@require_login
def new_finding(sid):
    s   = AuditSchedule.query.get_or_404(sid)
    f   = request.form
    fid = new_id('FND')

    # Count findings for this audit to generate ref
    count      = len(s.findings) + 1
    finding_ref = f'F-{count:03d}'

    finding = AuditFinding(
        id=fid, schedule_id=sid,
        finding_ref=finding_ref,
        description=f['description'],
        category=f['category'],
        severity=f['severity'],
        standard_ref=f.get('standard_ref', ''),
        root_cause=f.get('root_cause', ''),
        evidence=f.get('evidence', ''),
        requirement=f.get('requirement', ''),
        status='Open'
    )
    db.session.add(finding)
    db.session.flush()

    # Auto-create Hazard in main SMS Hazard Log
    auto_hazard = f.get('auto_hazard') == 'yes'
    hid = None
    if auto_hazard:
        hid  = new_id('HAZ')
        sev  = f.get('risk_severity', 'C')
        lik  = int(f.get('risk_likelihood', 3))
        ri   = f'{lik}{sev}'
        tol  = get_tolerance(ri)
        h = Hazard(
            id=hid, source='Audit',
            linked_report_id=fid,
            department_id=s.department_id,
            classification='Organizational',
            type_of_activity='Audit Finding',
            generic_hazard=f['description'][:120],
            specific_components=f.get('root_cause', ''),
            consequences='To be assessed by Safety Department',
            status='Open',
            owner=f.get('action_owner', 'Safety Manager')
        )
        db.session.add(h)
        db.session.flush()
        risk = Risk(
            id=new_id('RSK'), hazard_id=hid,
            description=f['description'],
            initial_likelihood=lik, initial_severity=sev,
            initial_risk_index=ri, initial_tolerance=tol
        )
        db.session.add(risk)
        finding.hazard_id = hid

    # Auto-create in BOTH tables: unified Action (visible everywhere) + AuditAction (for verification)
    act_desc = f.get('action_description', f['description'])
    act_owner = f.get('action_owner', '')
    act_due   = f.get('due_date', '')
    act_pri   = 'High' if f['severity'] == 'Major' else 'Medium'

    # 1. Unified Action table — appears in main Actions dashboard
    unified_id = new_id('ACT')
    unified_action = Action(
        id=unified_id, source='Audit',
        hazard_id=hid, linked_ref_id=fid,
        description=act_desc, owner=act_owner,
        due_date=act_due, priority=act_pri, status='Open'
    )
    db.session.add(unified_action)

    # 2. AuditAction table — for audit-specific effectiveness/verification tracking
    audit_action = AuditAction(
        id=new_id('ACT'), finding_id=fid,
        hazard_id=hid, description=act_desc,
        action_type='Corrective', owner=act_owner,
        due_date=act_due, priority=act_pri, status='Open'
    )
    db.session.add(audit_action)
    finding.status = 'Actioned'
    db.session.commit()
    # Hook 5 — SPI Intelligence linkage for new audit finding
    try:
        sched = AuditSchedule.query.get(sid)
        _spi_link_event(
            event_type    = 'audit_finding',
            event_id      = fid,
            event_title   = f['description'][:120],
            department_id = sched.department_id if sched else None,
            category      = f.get('category', ''),
            severity      = f.get('severity', 'Minor'),
            extra_text    = f.get('root_cause', '') + ' ' + f.get('requirement', ''),
            event_date    = '',
        )
    except Exception:
        pass

    msg = f'✓ Finding {finding_ref} recorded. Action {unified_action.id} created.'
    if hid: msg += f' Hazard {hid} created in SMS Hazard Log.'
    flash(msg, 'success')
    return redirect(url_for('audit_execution', sid=sid))


# ─── FINDING DETAIL ───────────────────────────────────────────────────────────
@app.route('/audit-findings/<fid>', methods=['GET','POST'])
@require_login
def finding_detail(fid):
    """Finding detail with full lifecycle: auditee CAP submission + safety review."""
    finding = AuditFinding.query.get_or_404(fid)
    schedule = AuditSchedule.query.get(finding.schedule_id)

    if request.method == 'POST':
        action = request.form.get('form_action', '')
        f = request.form

        if action == 'assign':
            finding.assigned_to   = f.get('assigned_to', '')
            finding.assigned_dept = f.get('assigned_dept', '')
            finding.assigned_date = datetime.now().strftime('%Y-%m-%d')
            finding.status = 'Assigned'

            # Auto-create or update a linked Action assigned to the SAG member
            sag_user = f.get('sag_member', '')
            dept_id  = int(f['department_id']) if f.get('department_id') else None
            if sag_user:
                # Check if action already linked
                existing_act = Action.query.filter_by(
                    linked_ref_id=fid, source='Audit Finding').first()
                if existing_act:
                    existing_act.sag_member   = sag_user
                    existing_act.department_id = dept_id
                    existing_act.status = 'Assigned'
                else:
                    new_act = Action(
                        id=new_id('ACT'),
                        source='Audit Finding',
                        description=f'CAP: {(finding.description or "")[:100]}',
                        owner=finding.assigned_to or sag_user,
                        due_date=finding.cap_due_date or '',
                        priority='High' if finding.severity=='Major' else 'Medium',
                        status='Assigned',
                        linked_ref_id=fid,
                        sag_member=sag_user,
                        department_id=dept_id,
                        assigned_by=session.get('admin_name','Admin'),
                    )
                    db.session.add(new_act)
                    db.session.flush()
                    finding.linked_action_id = new_act.id
                    log_action_history(new_act.id, session.get('admin_name','Admin'),
                                       'New', 'Assigned',
                                       f'Created from Audit Finding {fid}', 'assignment')

            db.session.commit()
            if finding.assigned_to:
                push_notify_by_name(
                    finding.assigned_to,
                    f'🔎 Audit Finding Assigned: {fid}',
                    f'You have been assigned an audit finding requiring a Corrective Action Plan. Due: {finding.cap_due_date or "TBD"}',
                    'action_assigned', 'audit_finding', fid)
            flash(f'✓ Finding {fid} assigned to {finding.assigned_to}'
                  + (f' · Action routed to SAG member {sag_user}' if sag_user else '') + '.', 'success')

        elif action == 'submit_root_cause':
            finding.root_cause           = f.get('root_cause', '')
            finding.investigation_notes  = f.get('investigation_notes', '')
            finding.contributing_factors = f.get('contributing_factors', '')
            finding.immediate_action     = f.get('immediate_action', '')
            finding.longterm_action      = f.get('longterm_action', '')
            finding.root_cause_submitted_at = datetime.utcnow()
            finding.status = 'Root Cause Submitted'
            db.session.commit()
            flash('✓ Root cause and corrective actions submitted.', 'success')

        elif action == 'submit_cap':
            finding.cap_responsible    = f.get('cap_responsible', '')
            finding.cap_due_date       = f.get('cap_due_date', '')
            finding.cap_completion_pct = int(f.get('cap_completion_pct', 0))
            finding.cap_status         = 'In Progress'
            finding.cap_submitted_at   = datetime.utcnow()
            # Handle evidence file uploads
            files = request.files.getlist('evidence_files')
            new_files = []
            for ef in files:
                if ef and ef.filename and allowed_file(ef.filename):
                    from werkzeug.utils import secure_filename
                    fname = f"{fid}_{secure_filename(ef.filename)}"
                    ef.save(os.path.join(app.config['UPLOAD_FOLDER'], fname))
                    new_files.append(fname)
            if new_files:
                existing = [x for x in (finding.evidence_files or '').split(',') if x]
                finding.evidence_files = ','.join(existing + new_files)
            finding.status = 'CAP Submitted'
            # Auto-create Action in main Action module if not already done
            if not finding.linked_action_id and f.get('create_action') == 'yes':
                aid = new_id('ACT')
                a = Action(
                    id=aid, source='Audit',
                    linked_ref_id=fid,
                    description=f'[{finding.finding_ref}] {(finding.immediate_action or finding.description or "")[:200]}',
                    owner=finding.cap_responsible or finding.assigned_to or '',
                    due_date=finding.cap_due_date or '',
                    priority='High' if finding.severity == 'Major' else 'Medium',
                    status='Open'
                )
                db.session.add(a)
                finding.linked_action_id = aid
            db.session.commit()
            flash('✓ CAP submitted and sent for Safety review.', 'success')

        elif action == 'submit_for_review':
            finding.status = 'Under Safety Review'
            db.session.commit()
            flash('✓ Finding submitted for Safety review.', 'success')

        elif action == 'safety_accept':
            finding.reviewed_by   = f.get('reviewed_by', '')
            finding.review_date   = datetime.now().strftime('%Y-%m-%d')
            finding.review_notes  = f.get('review_notes', '')
            finding.status = 'Accepted'
            db.session.commit()
            flash('✓ Finding accepted by Safety.', 'success')

        elif action == 'safety_return':
            finding.revision_reason = f.get('revision_reason', '')
            finding.reviewed_by     = f.get('reviewed_by', '')
            finding.status = 'Returned for Revision'
            db.session.commit()
            flash('⚠ Finding returned to auditee for revision.', 'warning')

        elif action == 'close_finding':
            finding.closure_verified_by = f.get('closure_verified_by', '')
            finding.closure_date        = datetime.now().strftime('%Y-%m-%d')
            finding.closure_notes       = f.get('closure_notes', '')
            finding.sig_dept_manager    = f.get('sig_dept_manager', '')
            finding.sig_auditor         = f.get('sig_auditor', '')
            finding.sig_safety_manager  = f.get('sig_safety_manager', '')
            finding.sig_date            = datetime.now().strftime('%Y-%m-%d')
            finding.cap_status          = 'Completed'
            finding.cap_completion_pct  = 100
            finding.status = 'Closed'
            # Update linked Action to Closed
            if finding.linked_action_id:
                la = Action.query.get(finding.linked_action_id)
                if la:
                    la.status = 'Closed'
                    la.closed_date = finding.closure_date
            db.session.commit()
            # ── AVI Hook: AuditFinding closure → schedule next-cycle verification ─
            try:
                sched = AuditSchedule.query.get(finding.schedule_id)
                dept_id = sched.department_id if sched else None
                _avi_generate(
                    source_module='audit_finding', source_record_id=finding.id,
                    source_description=f'Finding {finding.finding_ref} closed: {(finding.finding_title or finding.description or "")[:200]}',
                    department_id=dept_id,
                    linked_finding_id=finding.id,
                    linked_audit_id=finding.schedule_id,
                    linked_action_id=finding.linked_action_id,
                    operational_risk='Critical' if finding.severity == 'Major' else ('High' if finding.severity == 'Minor' else 'Medium'),
                    override_objective=f'Verify that CAP for finding "{finding.finding_ref}: {(finding.finding_title or "")[:80]}" has corrected the non-conformity and will not recur in the next audit cycle.',
                )
                db.session.commit()
            except Exception:
                pass
            flash(f'✓ Finding {fid} closed and verified.', 'success')

        elif action == 'reopen':
            finding.status = 'Assigned'
            finding.closure_date = None
            db.session.commit()
            flash('⚠ Finding reopened.', 'warning')

        elif action == 'update_cap_progress':
            finding.cap_completion_pct = int(f.get('cap_completion_pct', 0))
            finding.cap_status = f.get('cap_status', finding.cap_status)
            # Check overdue
            if finding.cap_due_date:
                from datetime import date
                try:
                    due = date.fromisoformat(finding.cap_due_date)
                    if date.today() > due and finding.cap_status != 'Completed':
                        finding.status = 'Overdue'
                        finding.cap_status = 'Overdue'
                except Exception:
                    pass
            db.session.commit()
            flash('✓ CAP progress updated.', 'success')

        return redirect(url_for('finding_detail', fid=fid))

    # Check overdue status on GET
    if finding.cap_due_date and finding.status not in ('Closed', 'Accepted'):
        from datetime import date
        try:
            due = date.fromisoformat(finding.cap_due_date)
            if date.today() > due and finding.status not in ('Closed', 'Accepted'):
                finding.status = 'Overdue'
                finding.cap_status = 'Overdue'
                db.session.commit()
        except Exception:
            pass

    # Evidence files list
    evidence_file_list = [x for x in (finding.evidence_files or '').split(',') if x]

    # Linked Action
    linked_action = None
    if finding.linked_action_id:
        linked_action = Action.query.get(finding.linked_action_id)

    # SAG members for assignment panel
    sag_members = User.query.filter(
        User.sag_role != None, User.sag_role != '', User.is_active == True
    ).all()

    # SPI link context
    spi_link_indicators, spi_link_existing = [], []
    try:
        spi_link_indicators = SPIIndicator.query.filter_by(active=True).order_by(SPIIndicator.code).all()
        if SPIEventLink is not None:
            spi_link_existing = SPIEventLink.query.filter_by(
                event_type='audit_finding', event_id=str(fid)
            ).all()
    except Exception:
        db.session.rollback()
    return render_template('audit/finding_detail.html',
                           finding=finding, schedule=schedule,
                           evidence_file_list=evidence_file_list,
                           linked_action=linked_action,
                           sag_members=sag_members,
                           now=datetime.utcnow(),
                           spi_active_indicators=spi_link_indicators,
                           spi_existing_links=spi_link_existing,
                           spi_link_event_type='audit_finding',
                           spi_link_event_id=str(fid),
                           spi_link_event_title=(finding.description or str(fid))[:100],
                           spi_link_event_date=str(finding.created_at.strftime('%Y-%m-%d') if finding.created_at else ''),
                           spi_link_severity='High' if finding.severity == 'Major' else 'Medium',
                           spi_link_dept_id='',
                           spi_link_category=finding.category or '',
                           spi_return_url=f'/audit-findings/{fid}',
                           )


@app.route('/audit-findings/<fid>/report')
@require_login
def finding_report(fid):
    """Dynamic NCR Report — safe, fully defensive, full lifecycle."""
    try:
        finding  = AuditFinding.query.get_or_404(fid)
        schedule = AuditSchedule.query.get(finding.schedule_id) if finding.schedule_id else None
        evidence_file_list = [x.strip() for x in (finding.evidence_files or '').split(',') if x.strip()]
        checklist_item = None
        try:
            checklist_item = AuditChecklist.query.filter_by(linked_finding_id=str(fid)).first()
        except Exception: pass
        linked_action  = None
        action_history = []
        sag_user       = None
        action_evidence_files = []
        if finding.linked_action_id:
            try:
                linked_action = Action.query.get(finding.linked_action_id)
                if linked_action:
                    action_history = ActionHistory.query.filter_by(
                        action_id=linked_action.id).order_by(ActionHistory.changed_at).all()
                    if linked_action.sag_member:
                        sag_user = User.query.filter_by(username=linked_action.sag_member).first()
                    if linked_action.evidence_filename:
                        action_evidence_files = [linked_action.evidence_filename]
            except Exception: pass
        dept = None
        try:
            if schedule and schedule.department_id:
                dept = Department.query.get(schedule.department_id)
        except Exception: pass
        ncr_number = finding.finding_ref or finding.id
        MONTHS = ['January','February','March','April','May','June',
                  'July','August','September','October','November','December']
        return render_template('audit/finding_report.html',
                               finding=finding, schedule=schedule, dept=dept,
                               checklist_item=checklist_item,
                               linked_action=linked_action, action_history=action_history,
                               sag_user=sag_user, evidence_file_list=evidence_file_list,
                               action_evidence_files=action_evidence_files,
                               ncr_number=ncr_number, now=datetime.utcnow(), MONTHS=MONTHS)
    except Exception as _e:
        import traceback; traceback.print_exc()
        app.logger.error('NCR report error fid=%s: %s', fid, _e)
        flash('Could not generate the NCR report. Please try again.', 'error')
        return redirect(f'/audit-findings/{fid}')


@app.route('/audit-schedule/<sid>/final-report')
@require_login
def audit_final_report(sid):
    """Enterprise Aviation Final Audit & NCR/CAPA Report Package — safe rendering."""
    try:
        schedule = AuditSchedule.query.get_or_404(sid)
        plan     = AuditPlan.query.get(schedule.plan_id) if schedule.plan_id else None
        findings = AuditFinding.query.filter_by(schedule_id=sid)                       .order_by(AuditFinding.finding_ref).all()
        checklist_items = []
        try:
            checklist_items = AuditChecklist.query.filter_by(schedule_id=sid)                                  .order_by(AuditChecklist.category, AuditChecklist.sequence).all()
        except Exception: pass
        # Safe overdue check
        for f in findings:
            if f.cap_due_date and f.status not in ('Closed','Accepted'):
                try:
                    if date.today() > date.fromisoformat(f.cap_due_date):
                        f.status = 'Overdue'; f.cap_status = 'Overdue'
                except Exception: pass
        try: db.session.commit()
        except Exception: db.session.rollback()
        # Build finding_data safely — each item fully defensive
        finding_data = []
        for f in findings:
            la = None; hist = []; sag_user = None; cl_item = None
            ev_files = []; act_ev = []
            try:
                if f.linked_action_id:
                    la = Action.query.get(f.linked_action_id)
            except Exception: pass
            try:
                cl_item = AuditChecklist.query.filter_by(linked_finding_id=str(f.id)).first()
            except Exception: pass
            try:
                ev_files = [x.strip() for x in (f.evidence_files or '').split(',') if x.strip()]
            except Exception: pass
            if la:
                try:
                    hist = ActionHistory.query.filter_by(action_id=la.id)                               .order_by(ActionHistory.changed_at).all()
                except Exception: pass
                try:
                    if la.sag_member:
                        sag_user = User.query.filter_by(username=la.sag_member).first()
                except Exception: pass
                try:
                    if la.evidence_filename:
                        act_ev = [la.evidence_filename]
                except Exception: pass
            finding_data.append({'finding':f,'action':la,'history':hist,
                                  'sag_user':sag_user,'cl_item':cl_item,
                                  'ev_files':ev_files,'act_ev':act_ev})
        # Analytics — all safe
        total       = len(findings)
        closed      = sum(1 for f in findings if f.status=='Closed')
        open_f      = sum(1 for f in findings if f.status not in ('Closed','Accepted'))
        overdue     = sum(1 for f in findings if f.status=='Overdue')
        major       = sum(1 for f in findings if (f.severity or '')=='Major')
        minor       = sum(1 for f in findings if (f.severity or '')=='Minor')
        critical    = sum(1 for f in findings if (f.severity or '')=='Critical')
        observation = sum(1 for f in findings if (f.severity or '')=='Observation')
        cl_total    = len(checklist_items)
        cl_yes      = sum(1 for i in checklist_items if i.response=='Yes')
        cl_no       = sum(1 for i in checklist_items if i.response=='No')
        cl_na       = sum(1 for i in checklist_items if i.response=='N/A')
        compliance_pct = round((cl_yes/cl_total*100) if cl_total>0 else 0)
        dept = None
        try:
            if schedule.department_id:
                dept = Department.query.get(schedule.department_id)
        except Exception: pass
        all_closed = (closed==total) if total>0 else True
        return render_template('audit/audit_final_report.html',
                               schedule=schedule, plan=plan, dept=dept,
                               findings=findings, finding_data=finding_data,
                               checklist_items=checklist_items, all_closed=all_closed,
                               total=total, closed=closed, open_f=open_f, overdue=overdue,
                               major=major, minor=minor, critical=critical, observation=observation,
                               cl_total=cl_total, cl_yes=cl_yes, cl_no=cl_no, cl_na=cl_na,
                               compliance_pct=compliance_pct, now=datetime.utcnow())
    except Exception as _e:
        import traceback; traceback.print_exc()
        app.logger.error('Final report error sid=%s: %s', sid, _e)
        flash('Could not generate the final audit report. Please try again.', 'error')
        return redirect(f'/audit-schedule/{sid}')


# ─── AUDIT ACTIONS ────────────────────────────────────────────────────────────
@app.route('/audit-actions')
@require_login
def audit_actions():
    today    = date.today().isoformat()
    status_f = request.args.get('status', '')
    pri_f    = request.args.get('priority', '')
    q = AuditAction.query
    if status_f: q = q.filter_by(status=status_f)
    if pri_f:    q = q.filter_by(priority=pri_f)

    # Auto-mark overdue
    for a in AuditAction.query.filter(AuditAction.status.in_(['Open', 'In Progress'])).all():
        if a.due_date and a.due_date < today:
            a.status = 'Overdue'
    db.session.commit()

    actions  = q.order_by(AuditAction.created_at.desc()).all()
    open_c   = AuditAction.query.filter_by(status='Open').count()
    inprog   = AuditAction.query.filter_by(status='In Progress').count()
    overdue  = AuditAction.query.filter_by(status='Overdue').count()
    closed   = AuditAction.query.filter_by(status='Closed').count()
    return render_template('audit/audit_actions.html',
        actions=actions, open_c=open_c, inprog=inprog,
        overdue=overdue, closed=closed,
        status_f=status_f, pri_f=pri_f)

@app.route('/audit-actions/<aid>/update', methods=['POST'])
@require_login
def update_audit_action(aid):
    a = AuditAction.query.get_or_404(aid)
    f = request.form
    a.status               = f.get('status', a.status)
    a.owner                = f.get('owner', a.owner)
    a.due_date             = f.get('due_date', a.due_date)
    a.priority             = f.get('priority', a.priority)
    a.implementation_notes = f.get('implementation_notes', a.implementation_notes)
    a.effectiveness        = f.get('effectiveness', a.effectiveness)
    a.effectiveness_notes  = f.get('effectiveness_notes', a.effectiveness_notes)
    a.verified_by          = f.get('verified_by', a.verified_by)
    a.verification_date    = f.get('verification_date', a.verification_date)
    if a.status == 'Closed':
        a.closed_date = date.today().isoformat()
        # Update parent finding status
        if a.finding and all(x.status == 'Closed' for x in a.finding.actions):
            a.finding.status = 'Closed'
    # Reopen if ineffective
    if a.effectiveness == 'Ineffective' and a.reopen_reason is None:
        a.status         = 'Open'
        a.reopen_reason  = f.get('reopen_reason', 'Re-opened: action was ineffective')
        a.effectiveness  = None
        if a.finding:
            a.finding.status = 'Open'
        flash('⚠ Action re-opened: effectiveness was Ineffective.', 'error')
    else:
        flash('✓ Action updated.', 'success')
    db.session.commit()
    return redirect(url_for('audit_actions'))


# ═══════════════════════════════════════════════════════════════════════════════
#  ENTERPRISE CONTINUOUS COMPLIANCE & AUDIT VERIFICATION ENGINE
#  ICAO Annex 19 · Doc 9859 · IOSA ISM · EASA SMS Oversight
#  Phase 2: Auto-Generation Engine + Hooks
# ═══════════════════════════════════════════════════════════════════════════════

# ── AVI auto-generation templates per source module ───────────────────────────
_AVI_TEMPLATES = {
    'asr': {
        'area': 'Flight Operations',
        'objective': 'Verify that the reported air safety concern has been operationally resolved and recurrence prevention measures are effective.',
        'evidence': 'SOP revision records, training records, operational briefing logs, recurrence data.',
        'criteria': 'No recurrence of same event category within 6 months; SOPs updated; crews briefed.',
        'risk': 'High',
    },
    'hazard': {
        'area': 'Safety Assurance',
        'objective': 'Verify that identified hazard controls remain operationally effective and have not introduced new hazards.',
        'evidence': 'Risk register updates, control implementation records, inspection reports.',
        'criteria': 'Hazard risk rating maintained at acceptable level; controls verified in place.',
        'risk': 'High',
    },
    'investigation': {
        'area': 'Safety Assurance',
        'objective': 'Verify that investigation recommendations have been fully implemented and are producing measurable operational improvement.',
        'evidence': 'Implementation records, follow-up inspection reports, SPI trend data, training completion records.',
        'criteria': 'All recommendations closed; no recurrence of root-cause scenario; SPI improving.',
        'risk': 'Critical',
    },
    'risk': {
        'area': 'Risk Management',
        'objective': 'Verify that risk mitigation measures remain adequate and risk level has not been re-elevated by operational changes.',
        'evidence': 'Risk assessment review records, control audit evidence, SPI data.',
        'criteria': 'Risk rating at or below accepted level; mitigations still in place and effective.',
        'risk': 'High',
    },
    'action': {
        'area': 'Safety Assurance',
        'objective': 'Verify that the closed corrective/preventive action has produced lasting operational improvement and not reverted.',
        'evidence': 'Post-implementation inspection, operational data, follow-up audit checklist response.',
        'criteria': 'Issue not recurred; implementation sustained; no regression observed.',
        'risk': 'Medium',
    },
    'spi': {
        'area': 'Safety Performance Monitoring',
        'objective': 'Verify that operational actions taken in response to SPI exceedance have restored performance to within the Safety Performance Target.',
        'evidence': 'SPI trend data for 3 months post-action, operational briefing records, SOP compliance checks.',
        'criteria': 'SPI value below L1 threshold for 3 consecutive months; corrective actions closed.',
        'risk': 'Critical',
    },
    'audit_finding': {
        'area': 'Audit & Compliance',
        'objective': 'Verify that the CAP implemented for this finding has corrected the non-conformity and will prevent recurrence in the next audit cycle.',
        'evidence': 'CAP completion records, updated procedures, next-cycle checklist response, evidence files.',
        'criteria': 'Same finding does not recur; CAP verified complete; next audit result: Conforming.',
        'risk': 'High',
    },
    'moc': {
        'area': 'Management of Change',
        'objective': 'Verify that the implemented change has not introduced unacceptable safety risks and post-change review findings are resolved.',
        'evidence': 'Post-change inspection records, hazard reassessment, operational feedback reports.',
        'criteria': 'No new hazards identified; change integrated into operations without safety degradation.',
        'risk': 'Medium',
    },
    'erp': {
        'area': 'Emergency Response',
        'objective': 'Verify that ERP drill findings and corrective actions have improved emergency response capability.',
        'evidence': 'Drill follow-up records, revised ERP procedures, next drill performance metrics.',
        'criteria': 'All drill findings closed; ERP procedures updated; next drill performance improved.',
        'risk': 'High',
    },
    'safety_promo': {
        'area': 'Safety Promotion',
        'objective': 'Verify that safety promotion activities have reached target audience and produced measurable safety culture improvement.',
        'evidence': 'Attendance records, survey results, knowledge assessment scores, safety reporting trend.',
        'criteria': 'Target attendance achieved; measurable improvement in safety culture indicators.',
        'risk': 'Low',
    },
    'cap': {
        'area': 'Corrective Action',
        'objective': 'Verify that the Corrective Action Plan has been fully implemented and the root cause eliminated.',
        'evidence': 'CAP completion evidence, root cause elimination verification, recurrence check.',
        'criteria': 'All CAP tasks complete; root cause addressed; no recurrence within 90 days.',
        'risk': 'High',
    },
}


def _avi_generate(source_module, source_record_id, source_description,
                  department_id=None, linked_report_id=None, linked_hazard_id=None,
                  linked_investigation_id=None, linked_spi_id=None,
                  linked_action_id=None, linked_audit_id=None,
                  linked_finding_id=None, linked_risk_id=None,
                  override_objective=None, override_evidence=None,
                  override_criteria=None, operational_risk=None):
    """
    Core AVI auto-generation function.
    Creates an AuditVerificationItem from any source module.
    Safe to call from any route — wrapped in try/except internally.
    Never raises; returns the created AVI id or None on failure.
    """
    try:
        # Prevent duplicates: one pending AVI per source_module + source_record_id
        existing = AuditVerificationItem.query.filter_by(
            source_module=source_module,
            source_record_id=str(source_record_id),
            status='Pending'
        ).first()
        if existing:
            return existing.id

        tmpl = _AVI_TEMPLATES.get(source_module, _AVI_TEMPLATES['action'])
        cur_year = datetime.utcnow().year
        cur_q    = (datetime.utcnow().month - 1) // 3 + 1
        next_q   = cur_q + 1 if cur_q < 4 else 1
        next_q_year = cur_year if cur_q < 4 else cur_year + 1
        due_cycle = f'Q{next_q}-{next_q_year}'
        due_date  = f'{next_q_year}-{(next_q * 3):02d}-28'

        avi = AuditVerificationItem(
            source_module       = source_module,
            source_record_id    = str(source_record_id),
            source_description  = source_description[:500] if source_description else '',
            department_id       = department_id,
            linked_report_id    = linked_report_id,
            linked_hazard_id    = linked_hazard_id,
            linked_investigation_id = linked_investigation_id,
            linked_spi_id       = linked_spi_id,
            linked_action_id    = linked_action_id,
            linked_audit_id     = linked_audit_id,
            linked_finding_id   = linked_finding_id,
            linked_risk_id      = linked_risk_id,
            verification_area       = tmpl['area'],
            verification_objective  = override_objective or tmpl['objective'],
            required_evidence       = override_evidence  or tmpl['evidence'],
            effectiveness_criteria  = override_criteria  or tmpl['criteria'],
            operational_risk        = operational_risk   or tmpl['risk'],
            due_audit_cycle         = due_cycle,
            due_date                = due_date,
            status                  = 'Pending',
            created_by              = session.get('user', 'system'),
        )
        db.session.add(avi)
        db.session.flush()
        return avi.id
    except Exception as _avi_err:
        try:
            db.session.rollback()
        except Exception:
            pass
        return None


def _avi_check_recurrence(source_module, verification_area, department_id=None):
    """
    Detect if the same type of issue has recurred (3+ AVIs same area/module).
    Returns (is_recurring: bool, count: int, existing_ids: list).
    """
    try:
        q = AuditVerificationItem.query.filter_by(
            source_module=source_module,
            verification_area=verification_area,
        )
        if department_id:
            q = q.filter_by(department_id=department_id)
        existing = q.order_by(AuditVerificationItem.created_at.desc()).all()
        count = len(existing)
        return count >= 3, count, [e.id for e in existing[:5]]
    except Exception:
        return False, 0, []


def _avi_purge(source_record_id=None, linked_report_id=None, linked_hazard_id=None,
               linked_investigation_id=None, linked_spi_id=None,
               linked_action_id=None, linked_audit_id=None, linked_finding_id=None,
               spi_indicator_id=None):
    """
    Delete all AuditVerificationItems linked to a source record that is being deleted.
    Call this inside any deletion route BEFORE db.session.commit() to keep the
    Assurance Engine clean.  Safe to call even if AuditVerificationItem is not imported.
    Returns the number of rows deleted.
    """
    if not AuditVerificationItem:
        return 0
    try:
        from sqlalchemy import or_
        conditions = []
        if source_record_id is not None:
            conditions.append(
                AuditVerificationItem.source_record_id == str(source_record_id))
        if linked_report_id is not None:
            conditions.append(
                AuditVerificationItem.linked_report_id == str(linked_report_id))
        if linked_hazard_id is not None:
            conditions.append(
                AuditVerificationItem.linked_hazard_id == str(linked_hazard_id))
        if linked_investigation_id is not None:
            conditions.append(
                AuditVerificationItem.linked_investigation_id == str(linked_investigation_id))
        if linked_spi_id is not None:
            conditions.append(
                AuditVerificationItem.linked_spi_id == int(linked_spi_id))
        if spi_indicator_id is not None:
            # SPI source_record_id is stored as "{indicator_id}-{month}-{rule}"
            # We match all AVIs whose source_record_id starts with this indicator id
            conditions.append(
                AuditVerificationItem.source_record_id.like(f'{spi_indicator_id}-%'))
            conditions.append(
                AuditVerificationItem.linked_spi_id == int(spi_indicator_id))
        if linked_action_id is not None:
            conditions.append(
                AuditVerificationItem.linked_action_id == str(linked_action_id))
        if linked_audit_id is not None:
            conditions.append(
                AuditVerificationItem.linked_audit_id == str(linked_audit_id))
        if linked_finding_id is not None:
            conditions.append(
                AuditVerificationItem.linked_finding_id == str(linked_finding_id))
        if not conditions:
            return 0
        deleted = AuditVerificationItem.query.filter(
            or_(*conditions)
        ).delete(synchronize_session=False)
        return deleted
    except Exception:
        return 0


# ═══════════════════════════════════════════════════════════════════════════════
#  ASSURANCE ENGINE ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route('/audit-assurance/clear-all', methods=['POST'])
@require_login
def avi_clear_all():
    """Admin: wipe ALL AuditVerificationItems — testing / reset only."""
    try:
        n = AuditVerificationItem.query.delete(synchronize_session=False)
        db.session.commit()
        flash(f'✓ Cleared {n} verification items from the Assurance Engine.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'⚠ Could not clear: {str(e)[:120]}', 'error')
    return redirect(url_for('audit_assurance_dashboard'))


@app.route('/audit-assurance')
@require_login
def audit_assurance_dashboard():
    """Enterprise Continuous Compliance & Assurance Dashboard."""
    try:
        total_avis      = AuditVerificationItem.query.count()
        pending_avis    = AuditVerificationItem.query.filter_by(status='Pending').count()
        in_verify       = AuditVerificationItem.query.filter(
                            AuditVerificationItem.status.in_(['Scheduled','In Verification'])).count()
        effective       = AuditVerificationItem.query.filter_by(effectiveness_result='Effective').count()
        ineffective     = AuditVerificationItem.query.filter_by(effectiveness_result='Ineffective').count()
        escalated       = AuditVerificationItem.query.filter_by(escalation_required=True).count()
        systemic        = AuditVerificationItem.query.filter_by(is_systemic=True).count()
        recurring       = AuditVerificationItem.query.filter_by(recurrence_flag=True).count()

        from datetime import date
        today_str = date.today().isoformat()
        overdue   = AuditVerificationItem.query.filter(
                        AuditVerificationItem.due_date < today_str,
                        AuditVerificationItem.status.in_(['Pending','Scheduled'])).count()

        # Recent AVIs grouped by source
        recent_avis = AuditVerificationItem.query.order_by(
                        AuditVerificationItem.created_at.desc()).limit(20).all()

        # Breakdown by source module
        from sqlalchemy import func
        module_counts = db.session.query(
            AuditVerificationItem.source_module,
            func.count(AuditVerificationItem.id)
        ).group_by(AuditVerificationItem.source_module).all()

        # Open assurance loops (pending + overdue)
        open_loops = AuditVerificationItem.query.filter(
            AuditVerificationItem.status.in_(['Pending','Scheduled','In Verification'])
        ).order_by(AuditVerificationItem.operational_risk.desc(),
                   AuditVerificationItem.due_date.asc()).limit(50).all()

        # Systemic/recurring issues
        systemic_items = AuditVerificationItem.query.filter(
            (AuditVerificationItem.is_systemic == True) |
            (AuditVerificationItem.recurrence_flag == True)
        ).order_by(AuditVerificationItem.recurrence_count.desc()).limit(20).all()

        # Ineffective implementations requiring action
        ineffective_items = AuditVerificationItem.query.filter_by(
            effectiveness_result='Ineffective'
        ).order_by(AuditVerificationItem.verified_at.desc()).limit(20).all()

    except Exception as _e:
        db.session.rollback()
        total_avis = pending_avis = in_verify = effective = 0
        ineffective = escalated = systemic = recurring = overdue = 0
        recent_avis = open_loops = systemic_items = ineffective_items = []
        module_counts = []

    return render_template('audit/assurance_dashboard.html',
        total_avis=total_avis, pending_avis=pending_avis,
        in_verify=in_verify, effective=effective, ineffective=ineffective,
        escalated=escalated, systemic=systemic, recurring=recurring,
        overdue=overdue, recent_avis=recent_avis, open_loops=open_loops,
        systemic_items=systemic_items, ineffective_items=ineffective_items,
        module_counts=module_counts,
        now=datetime.utcnow())


@app.route('/audit-assurance/items')
@require_login
def avi_list():
    """All Audit Verification Items with filtering."""
    status_f = request.args.get('status', '')
    module_f = request.args.get('module', '')
    risk_f   = request.args.get('risk', '')

    q = AuditVerificationItem.query
    if status_f:
        q = q.filter_by(status=status_f)
    if module_f:
        q = q.filter_by(source_module=module_f)
    if risk_f:
        q = q.filter_by(operational_risk=risk_f)

    items = q.order_by(AuditVerificationItem.operational_risk.desc(),
                       AuditVerificationItem.created_at.desc()).all()

    departments = Department.query.order_by(Department.name).all()
    return render_template('audit/avi_list.html',
        items=items, status_f=status_f, module_f=module_f, risk_f=risk_f,
        departments=departments,
        all_statuses=['Pending','Scheduled','In Verification',
                      'Verified Effective','Verified Ineffective','Escalated','Closed'],
        all_modules=list(_AVI_TEMPLATES.keys()))


@app.route('/audit-assurance/items/<int:avi_id>', methods=['GET','POST'])
@require_login
def avi_detail(avi_id):
    """AVI detail view + auditor verification form."""
    avi = AuditVerificationItem.query.get_or_404(avi_id)

    if request.method == 'POST':
        f = request.form
        action = f.get('action', 'verify')

        if action == 'verify':
            avi.status               = 'In Verification'
            avi.verified_by          = f.get('verified_by', session.get('user',''))
            avi.verified_at          = datetime.utcnow()
            avi.effectiveness_result = f.get('effectiveness_result', '')
            avi.effectiveness_notes  = f.get('effectiveness_notes', '')
            avi.evidence_collected   = f.get('evidence_collected', '')
            avi.scheduled_audit_id   = f.get('scheduled_audit_id', '')

            if avi.effectiveness_result == 'Effective':
                avi.status = 'Verified Effective'
            elif avi.effectiveness_result == 'Ineffective':
                avi.status = 'Verified Ineffective'
                avi.followup_required    = True
                avi.escalation_required  = True
                # Re-open the source finding if applicable
                if avi.linked_finding_id:
                    try:
                        finding = AuditFinding.query.get(avi.linked_finding_id)
                        if finding and finding.status == 'Closed':
                            finding.status = 'Open'
                            finding.revision_reason = f'Effectiveness verification FAILED: {avi.effectiveness_notes}'
                            db.session.add(finding)
                    except Exception:
                        pass
            elif avi.effectiveness_result == 'Partially Effective':
                avi.status            = 'In Verification'
                avi.followup_required = True

        elif action == 'schedule':
            avi.status             = 'Scheduled'
            new_cycle = f.get('due_audit_cycle', '').strip()
            new_date  = f.get('due_date', '').strip()
            new_auditor = f.get('verified_by', '').strip()
            sched_id    = f.get('scheduled_audit_id', '').strip()
            sched_notes = f.get('effectiveness_notes', '').strip()
            if new_cycle:
                avi.due_audit_cycle = new_cycle
            if new_date:
                avi.due_date = new_date
            if new_auditor:
                avi.verified_by = new_auditor
            if sched_id:
                avi.scheduled_audit_id = sched_id
            if sched_notes and not avi.effectiveness_notes:
                avi.effectiveness_notes = sched_notes

        elif action == 'escalate':
            avi.status               = 'Escalated'
            avi.escalation_required  = True
            avi.escalation_date      = date.today().isoformat()
            avi.recurrence_notes     = f.get('escalation_reason', '')
            # Check recurrence
            is_rec, count, _ = _avi_check_recurrence(
                avi.source_module, avi.verification_area, avi.department_id)
            if is_rec:
                avi.recurrence_flag  = True
                avi.recurrence_count = count
                avi.is_systemic      = True

        elif action == 'close':
            avi.status = 'Closed'

        db.session.add(avi)
        db.session.commit()
        flash(f'✓ Verification Item {avi_id} updated — status: {avi.status}', 'success')
        return redirect(url_for('avi_detail', avi_id=avi_id))

    # Related context for intelligence panel
    related_findings = []
    related_actions  = []
    related_spi      = None
    try:
        if avi.linked_finding_id:
            related_findings = [AuditFinding.query.get(avi.linked_finding_id)]
            related_findings = [f for f in related_findings if f]
        if avi.linked_action_id:
            related_actions = [Action.query.get(avi.linked_action_id)]
            related_actions = [a for a in related_actions if a]
        if avi.linked_spi_id:
            related_spi = SPIIndicator.query.get(avi.linked_spi_id)
    except Exception:
        pass

    # Recurrence check
    is_rec, rec_count, rec_ids = _avi_check_recurrence(
        avi.source_module, avi.verification_area, avi.department_id)

    audit_schedules = AuditSchedule.query.filter(
        AuditSchedule.status.in_(['Planned','In Progress'])
    ).order_by(AuditSchedule.scheduled_date).all()

    return render_template('audit/avi_detail.html',
        avi=avi, related_findings=related_findings,
        related_actions=related_actions, related_spi=related_spi,
        is_recurring=is_rec, recurrence_count=rec_count,
        audit_schedules=audit_schedules,
        now=datetime.utcnow())


# ─── AUDIT DASHBOARD (summary view) ──────────────────────────────────────────
@app.route('/audit-dashboard')
@require_login
def audit_dashboard():
    total_plans     = AuditPlan.query.count()
    total_scheduled = AuditSchedule.query.count()
    in_progress     = AuditSchedule.query.filter_by(status='In Progress').count()
    completed       = AuditSchedule.query.filter_by(status='Completed').count()
    planned         = AuditSchedule.query.filter_by(status='Planned').count()
    total_findings  = AuditFinding.query.count()
    major           = AuditFinding.query.filter_by(severity='Major').count()
    minor           = AuditFinding.query.filter_by(severity='Minor').count()
    obs             = AuditFinding.query.filter_by(severity='Observation').count()
    open_actions    = AuditAction.query.filter_by(status='Open').count()
    overdue_actions = AuditAction.query.filter_by(status='Overdue').count()

    recent_audits   = AuditSchedule.query.order_by(
        AuditSchedule.scheduled_date.desc()).limit(5).all()
    recent_findings = AuditFinding.query.order_by(
        AuditFinding.created_at.desc()).limit(5).all()

    return render_template('audit/audit_dashboard.html',
        total_plans=total_plans, total_scheduled=total_scheduled,
        in_progress=in_progress, completed=completed, planned=planned,
        total_findings=total_findings, major=major, minor=minor, obs=obs,
        open_actions=open_actions, overdue_actions=overdue_actions,
        recent_audits=recent_audits, recent_findings=recent_findings)

# ─── Init ─────────────────────────────────────────────────────────────────────

# ═══════════════════════════════════════════════════════════════════════════════
#  SAFETY POLICY & OBJECTIVES — COMPONENT 1 OF SMS
#  ICAO Annex 19 §3 / Doc 9859 Ch.3 — Extension only, no existing code changed
# ═══════════════════════════════════════════════════════════════════════════════

# ── Document ID generator ────────────────────────────────────────────────────
def gen_doc_id(doc_type, dept_code, year, seq, rev=0):
    return f"{doc_type}-{dept_code}-{year}-{seq:03d}-REV{rev}"

def next_seq(doc_type, dept_id, year):
    existing = SMSDocument.query.filter_by(
        doc_type=doc_type, department_id=dept_id).filter(
        SMSDocument.id.like(f"%-{year}-%")).count()
    return existing + 1

# ─── SAFETY POLICY ───────────────────────────────────────────────────────────
@app.route('/safety-policy')
@require_login
def safety_policy():
    active  = SafetyPolicy.query.filter_by(status='Active').first()
    history = SafetyPolicy.query.filter_by(status='Archived').order_by(
              SafetyPolicy.version_num.desc()).all()
    drafts  = SafetyPolicy.query.filter_by(status='Draft').all()
    return render_template('safety_policy/policy.html',
                           active=active, history=history, drafts=drafts)

@app.route('/safety-policy/new', methods=['GET','POST'])
@require_login
def new_safety_policy():
    if request.method == 'POST':
        f = request.form
        # Get current max version
        latest = SafetyPolicy.query.order_by(
                 SafetyPolicy.version_num.desc()).first()
        ver_num = (latest.version_num + 1) if latest else 0
        pid = new_id('POL')
        p = SafetyPolicy(
            id=pid, version=f'REV{ver_num}', version_num=ver_num,
            title=f['title'], content=f['content'],
            approved_by=f['approved_by'],
            approved_by_title=f.get('approved_by_title','Accountable Manager'),
            effective_date=f['effective_date'],
            review_date=f.get('review_date',''),
            status='Draft',
            change_summary=f.get('change_summary','Initial issue') if ver_num==0 else f.get('change_summary','')
        )
        db.session.add(p)
        db.session.commit()
        flash(f'✓ Safety Policy {pid} created as REV{ver_num} (Draft).', 'success')
        return redirect(url_for('safety_policy'))
    latest = SafetyPolicy.query.order_by(SafetyPolicy.version_num.desc()).first()
    next_ver = (latest.version_num + 1) if latest else 0
    return render_template('safety_policy/policy_form.html', next_ver=next_ver, latest=latest)

@app.route('/safety-policy/<pid>/activate', methods=['POST'])
@require_login
def activate_policy(pid):
    # Archive current active
    current = SafetyPolicy.query.filter_by(status='Active').first()
    if current:
        current.status = 'Archived'
    policy = SafetyPolicy.query.get_or_404(pid)
    policy.status = 'Active'
    db.session.commit()
    flash(f'✓ Policy {policy.version} is now Active. Previous version archived.', 'success')
    return redirect(url_for('safety_policy'))

@app.route('/safety-policy/<pid>/edit', methods=['POST'])
@require_login
def edit_policy(pid):
    p = SafetyPolicy.query.get_or_404(pid)
    f = request.form
    if p.status == 'Archived':
        flash('Cannot edit archived policy.', 'error')
        return redirect(url_for('safety_policy'))
    p.content          = f.get('content', p.content)
    p.approved_by      = f.get('approved_by', p.approved_by)
    p.effective_date   = f.get('effective_date', p.effective_date)
    p.review_date      = f.get('review_date', p.review_date)
    p.change_summary   = f.get('change_summary', p.change_summary)
    db.session.commit()
    flash('✓ Policy updated.', 'success')
    return redirect(url_for('safety_policy'))

# ─── SAFETY ACCOUNTABILITY (ROLES) ───────────────────────────────────────────
@app.route('/safety-roles')
@require_login
def safety_roles():
    roles = SafetyRole.query.filter_by(active=True).order_by(SafetyRole.role_type).all()
    return render_template('safety_policy/roles.html', roles=roles)

@app.route('/safety-roles/new', methods=['GET','POST'])
@require_login
def new_safety_role():
    if request.method == 'POST':
        f = request.form
        r = SafetyRole(
            id=new_id('ROLE'),
            role_name=f['role_name'],
            role_type=f['role_type'],
            person_name=f['person_name'],
            department_id=int(f['department_id']) if f.get('department_id') else None,
            responsibilities=f.get('responsibilities',''),
            authority=f.get('authority',''),
            contact_email=f.get('contact_email',''),
            contact_phone=f.get('contact_phone',''),
            effective_from=f.get('effective_from',''),
            active=True
        )
        db.session.add(r)
        db.session.commit()
        flash(f'✓ Role {r.role_name} assigned to {r.person_name}.', 'success')
        return redirect(url_for('safety_roles'))
    return render_template('safety_policy/role_form.html')

@app.route('/safety-roles/<rid>/update', methods=['POST'])
@require_login
def update_safety_role(rid):
    r = SafetyRole.query.get_or_404(rid)
    f = request.form
    r.person_name      = f.get('person_name', r.person_name)
    r.responsibilities = f.get('responsibilities', r.responsibilities)
    r.authority        = f.get('authority', r.authority)
    r.contact_email    = f.get('contact_email', r.contact_email)
    r.contact_phone    = f.get('contact_phone', r.contact_phone)
    r.active           = f.get('active','true') == 'true'
    db.session.commit()
    flash('✓ Role updated.', 'success')
    return redirect(url_for('safety_roles'))

# ─── KEY SAFETY PERSONNEL ─────────────────────────────────────────────────────
@app.route('/safety-personnel')
@require_login
def safety_personnel():
    personnel = SafetyPersonnel.query.filter_by(active=True).order_by(
                SafetyPersonnel.sms_role).all()
    return render_template('safety_policy/personnel.html', personnel=personnel)

@app.route('/safety-personnel/new', methods=['GET','POST'])
@require_login
def new_safety_personnel():
    if request.method == 'POST':
        f = request.form
        p = SafetyPersonnel(
            id=new_id('PERS'),
            name=f['name'], position=f['position'],
            department_id=int(f['department_id']) if f.get('department_id') else None,
            sms_role=f.get('sms_role',''),
            qualifications=f.get('qualifications',''),
            contact_email=f.get('contact_email',''),
            contact_phone=f.get('contact_phone',''),
            sms_trained=f.get('sms_trained') == 'yes',
            training_date=f.get('training_date',''),
            active=True
        )
        db.session.add(p)
        db.session.commit()
        flash(f'✓ Personnel record created for {p.name}.', 'success')
        return redirect(url_for('safety_personnel'))
    return render_template('safety_policy/personnel_form.html')

@app.route('/safety-personnel/<pid>/update', methods=['POST'])
@require_login
def update_personnel(pid):
    p = SafetyPersonnel.query.get_or_404(pid)
    f = request.form
    p.position       = f.get('position', p.position)
    p.sms_role       = f.get('sms_role', p.sms_role)
    p.qualifications = f.get('qualifications', p.qualifications)
    p.contact_email  = f.get('contact_email', p.contact_email)
    p.contact_phone  = f.get('contact_phone', p.contact_phone)
    p.sms_trained    = f.get('sms_trained') == 'yes'
    p.training_date  = f.get('training_date', p.training_date)
    p.active         = f.get('active','true') == 'true'
    db.session.commit()
    flash('✓ Personnel record updated.', 'success')
    return redirect(url_for('safety_personnel'))

# ─── EMERGENCY RESPONSE PLANNING ─────────────────────────────────────────────
@app.route('/erp')
@require_login
def erp_list():
    plans = ERPlan.query.filter_by(status='Active').order_by(ERPlan.scenario_type).all()
    archived = ERPlan.query.filter_by(status='Archived').all()
    return render_template('safety_policy/erp.html', plans=plans, archived=archived)

@app.route('/erp-dashboard')
@require_login
def erp_dashboard():
    from datetime import date as _date, timedelta
    plans     = ERPlan.query.filter_by(status='Active').order_by(ERPlan.scenario_type).all()
    all_drills= ERPDrill.query.order_by(ERPDrill.drill_date.desc()).all()
    all_acts  = ERPActivation.query.order_by(ERPActivation.activated_at.desc()).all()
    today     = _date.today().isoformat()
    cutoff    = (_date.today() - timedelta(days=365)).isoformat()
    # ERPs with no drill in the last 12 months
    def _last_drill_date(p):
        return p.drills[0].drill_date if p.drills else None
    overdue_erps = [p for p in plans if (_last_drill_date(p) or '0000') < cutoff]
    active_acts  = sum(1 for a in all_acts if a.status == 'Active')
    return render_template('safety_policy/erp_dashboard.html',
                           plans=plans,
                           total_plans=len(plans),
                           total_drills=len(all_drills),
                           active_acts=active_acts,
                           overdue_count=len(overdue_erps),
                           overdue_erps=overdue_erps,
                           all_drills=all_drills,
                           all_acts=all_acts)

@app.route('/erp/new', methods=['GET','POST'])
@require_login
def new_erp():
    if request.method == 'POST':
        f   = request.form
        count = ERPlan.query.count() + 1
        e = ERPlan(
            id=new_id('ERP'),
            erp_ref=f'ERP-{count:03d}',
            scenario_type=f['scenario_type'],
            title=f['title'],
            description=f.get('description',''),
            activation_criteria=f.get('activation_criteria',''),
            response_procedures=f.get('response_procedures',''),
            responsible_roles=f.get('responsible_roles',''),
            emergency_contacts=f.get('emergency_contacts',''),
            resources_required=f.get('resources_required',''),
            notification_list=f.get('notification_list',''),
            review_date=f.get('review_date',''),
            version='REV0', status='Active'
        )
        db.session.add(e)
        db.session.commit()
        flash(f'✓ ERP {e.erp_ref} created: {e.title}', 'success')
        return redirect(url_for('erp_list'))
    return render_template('safety_policy/erp_form.html')

@app.route('/erp/<eid>')
@require_login
def erp_detail(eid):
    e = ERPlan.query.get_or_404(eid)
    return render_template('safety_policy/erp_detail.html', e=e)

@app.route('/erp/<eid>/update', methods=['POST'])
@require_login
def update_erp(eid):
    e = ERPlan.query.get_or_404(eid)
    f = request.form
    e.response_procedures = f.get('response_procedures', e.response_procedures)
    e.emergency_contacts  = f.get('emergency_contacts', e.emergency_contacts)
    e.responsible_roles   = f.get('responsible_roles', e.responsible_roles)
    e.notification_list   = f.get('notification_list', e.notification_list)
    e.resources_required  = f.get('resources_required', e.resources_required)
    e.review_date         = f.get('review_date', e.review_date)
    e.status              = f.get('status', e.status)
    db.session.commit()
    flash('✓ ERP updated.', 'success')
    return redirect(url_for('erp_detail', eid=eid))

# ─── DOCUMENT CONTROL ─────────────────────────────────────────────────────────
@app.route('/documents')
@require_login
def documents():
    type_f   = request.args.get('type','')
    dept_f   = request.args.get('dept','')
    status_f = request.args.get('status','')
    q = SMSDocument.query
    if type_f:   q = q.filter_by(doc_type=type_f)
    if dept_f:   q = q.filter_by(department_id=int(dept_f))
    if status_f: q = q.filter_by(status=status_f)
    docs = q.order_by(SMSDocument.created_at.desc()).all()
    doc_types = ['POL','MAN','SOP','RA','AUD','MOC','INV','TRN','NEWS']
    return render_template('document/document_list.html', docs=docs, doc_types=doc_types,
                           type_f=type_f, dept_f=dept_f, status_f=status_f)

@app.route('/documents/new', methods=['GET','POST'])
@require_login
def new_document():
    if request.method == 'POST':
        f        = request.form
        doc_type = f['doc_type']
        dept_id  = int(f['department_id'])
        year     = datetime.now().year
        dept     = Department.query.get(dept_id)
        dept_code = dept.code if dept else 'XX'
        seq      = next_seq(doc_type, dept_id, year)
        doc_id   = gen_doc_id(doc_type, dept_code, year, seq, 0)
        d = SMSDocument(
            id=doc_id, doc_type=doc_type,
            department_id=dept_id,
            title=f['title'],
            description=f.get('description',''),
            content=f.get('content',''),
            version='REV0', version_num=0, seq_num=seq,
            status='Draft',
            created_by=f.get('created_by',''),
            effective_date=f.get('effective_date',''),
            review_due=f.get('review_due',''),
            change_summary='Initial issue'
        )
        db.session.add(d)
        db.session.commit()
        flash(f'✓ Document {doc_id} created as Draft.', 'success')
        return redirect(url_for('documents'))
    doc_types = ['POL','MAN','SOP','RA','AUD','MOC','INV','TRN','NEWS']
    return render_template('document/document_form.html', doc_types=doc_types)

@app.route('/documents/<did>')
@require_login
def document_detail(did):
    doc = SMSDocument.query.get_or_404(did)
    # Get all versions (parent chain)
    versions = []
    current = doc
    while current:
        versions.append(current)
        if current.parent_doc_id:
            current = SMSDocument.query.get(current.parent_doc_id)
        else:
            break
    return render_template('document/document_detail.html', doc=doc, versions=versions)

@app.route('/documents/<did>/advance', methods=['POST'])
@require_login
def advance_document(did):
    """Draft → Under Review → Approved → Archived"""
    doc = SMSDocument.query.get_or_404(did)
    f   = request.form
    transitions = {
        'Draft':        'Under Review',
        'Under Review': 'Approved',
        'Approved':     'Archived',
    }
    if doc.status in transitions:
        doc.status = transitions[doc.status]
        if doc.status == 'Approved':
            doc.approved_by    = f.get('approved_by', doc.approved_by)
            doc.effective_date = f.get('effective_date', doc.effective_date)
            doc.review_due     = f.get('review_due', doc.review_due)
            doc.reviewed_by    = f.get('reviewed_by', doc.reviewed_by)
        db.session.commit()
        flash(f'✓ Document status updated to {doc.status}.', 'success')
    return redirect(url_for('document_detail', did=did))

@app.route('/documents/<did>/revise', methods=['POST'])
@require_login
def revise_document(did):
    """Create a new revision — old version becomes archived."""
    old = SMSDocument.query.get_or_404(did)
    if old.status != 'Approved':
        flash('Only Approved documents can be revised.', 'error')
        return redirect(url_for('document_detail', did=did))
    f       = request.form
    new_ver = old.version_num + 1
    dept    = Department.query.get(old.department_id)
    dept_code = dept.code if dept else 'XX'
    year    = datetime.now().year
    new_id_str = gen_doc_id(old.doc_type, dept_code, year, old.seq_num, new_ver)
    new_doc = SMSDocument(
        id=new_id_str, doc_type=old.doc_type,
        department_id=old.department_id,
        title=old.title,
        description=old.description,
        content=f.get('content', old.content),
        version=f'REV{new_ver}', version_num=new_ver,
        seq_num=old.seq_num,
        status='Draft',
        created_by=f.get('created_by', old.created_by),
        change_summary=f.get('change_summary',''),
        parent_doc_id=old.id
    )
    old.status = 'Archived'
    db.session.add(new_doc)
    db.session.commit()
    flash(f'✓ New revision {new_id_str} created. {old.id} archived.', 'success')
    return redirect(url_for('document_detail', did=new_id_str))


# ═══════════════════════════════════════════════════════════════════════════════
#  DOCUMENT TRACEABILITY BACKBONE — FULL INTEGRATION
#  Documents ↔ Hazards ↔ Risks ↔ Audits ↔ Actions ↔ MOC ↔ Training
#  Added as extension — all existing routes unchanged
# ═══════════════════════════════════════════════════════════════════════════════

# ── Entity resolver — fetch any entity by type + id ─────────────────────────
ENTITY_RESOLVERS = {
    'hazard':         lambda eid: Hazard.query.get(eid),
    'risk':           lambda eid: Risk.query.get(eid),
    'action':         lambda eid: Action.query.get(eid),
    'audit_schedule': lambda eid: AuditSchedule.query.get(eid),
    'audit_finding':  lambda eid: AuditFinding.query.get(eid),
    'audit_action':   lambda eid: AuditAction.query.get(eid),
    'moc':            lambda eid: MOC.query.get(eid),
    'investigation':  lambda eid: Investigation.query.get(eid),
    'training':       lambda eid: Training.query.get(eid),
    'erp':            lambda eid: ERPlan.query.get(eid),
}

ENTITY_LABELS = {
    'hazard':         ('Hazard',           '/hazard-log/{}'),
    'risk':           ('Risk',             '/hazard-log/{}'),   # risks shown via hazard
    'action':         ('Action',           '/actions'),
    'audit_schedule': ('Audit',            '/audit-schedule/{}'),
    'audit_finding':  ('Audit Finding',    '/audit-findings/{}'),
    'audit_action':   ('Audit Action',     '/audit-actions'),
    'moc':            ('MOC',              '/moc'),
    'investigation':  ('Investigation',    '/investigations/{}'),
    'training':       ('Training Record',  '/safety-promotion'),
    'erp':            ('ERP',              '/erp/{}'),
}

def resolve_entity_label(entity_type, entity_id):
    """Return display name and URL for a linked entity."""
    if entity_type not in ENTITY_LABELS:
        return entity_type, '#'
    label, url_tpl = ENTITY_LABELS[entity_type]
    try:
        url = url_tpl.format(entity_id)
    except Exception:
        url = url_tpl
    return label, url

def get_doc_links_for_entity(entity_type, entity_id):
    """Return all documents linked to a given entity."""
    links = DocumentLink.query.filter_by(
        entity_type=entity_type, entity_id=str(entity_id)).all()
    docs = []
    for lnk in links:
        doc = SMSDocument.query.get(lnk.document_id)
        if doc:
            docs.append({'doc': doc, 'link': lnk})
    return docs

def build_traceability(doc):
    """Build full traceability map for a document — all linked entities."""
    links = DocumentLink.query.filter_by(document_id=doc.id).all()
    result = []
    for lnk in links:
        label, url = resolve_entity_label(lnk.entity_type, lnk.entity_id)
        obj = None
        resolver = ENTITY_RESOLVERS.get(lnk.entity_type)
        if resolver:
            try:
                obj = resolver(lnk.entity_id)
            except Exception:
                pass
        result.append({
            'link': lnk,
            'label': label,
            'url': url,
            'entity_type': lnk.entity_type,
            'entity_id': lnk.entity_id,
            'obj': obj,
        })
    return result

# ─── DOCUMENT LINKING API ─────────────────────────────────────────────────────
@app.route('/documents/<did>/link', methods=['POST'])
@require_login
def link_document(did):
    doc = SMSDocument.query.get_or_404(did)
    f   = request.form
    entity_type = f['entity_type']
    entity_id   = f['entity_id'].strip()
    reason      = f.get('link_reason', '')

    if not entity_type or not entity_id:
        flash('Entity type and ID are required.', 'error')
        return redirect(url_for('document_detail', did=did))

    # Validate entity exists
    resolver = ENTITY_RESOLVERS.get(entity_type)
    if resolver:
        obj = resolver(entity_id)
        if not obj:
            flash(f'No {entity_type} found with ID: {entity_id}', 'error')
            return redirect(url_for('document_detail', did=did))

    # Check duplicate
    existing = DocumentLink.query.filter_by(
        document_id=did, entity_type=entity_type, entity_id=entity_id).first()
    if existing:
        flash(f'Document already linked to this {entity_type}.', 'error')
        return redirect(url_for('document_detail', did=did))

    lnk = DocumentLink(
        document_id=did,
        entity_type=entity_type,
        entity_id=entity_id,
        link_reason=reason
    )
    db.session.add(lnk)
    db.session.commit()
    flash(f'✓ Document linked to {entity_type} {entity_id}.', 'success')
    return redirect(url_for('document_detail', did=did))

@app.route('/documents/<did>/unlink/<int:link_id>', methods=['POST'])
@require_login
def unlink_document(did, link_id):
    lnk = DocumentLink.query.get_or_404(link_id)
    db.session.delete(lnk)
    db.session.commit()
    flash('✓ Link removed.', 'success')
    return redirect(url_for('document_detail', did=did))

# ─── DOCUMENT DETAIL (override — add traceability) ───────────────────────────
@app.route('/documents/<did>/trace')
@require_login
def document_trace(did):
    doc   = SMSDocument.query.get_or_404(did)
    trace = build_traceability(doc)
    versions = []
    current = doc
    while current:
        versions.append(current)
        current = SMSDocument.query.get(current.parent_doc_id) if current.parent_doc_id else None
    return render_template('document/document_trace.html', doc=doc, trace=trace, versions=versions)

# ─── ENTITY TRACEABILITY VIEWS ────────────────────────────────────────────────
@app.route('/hazard-log/<hid>/documents')
@require_login
def hazard_documents(hid):
    hazard = Hazard.query.get_or_404(hid)
    # Documents linked to hazard
    haz_docs  = get_doc_links_for_entity('hazard', hid)
    # Documents linked to any risk of this hazard
    risk_docs = []
    for risk in hazard.risks:
        for item in get_doc_links_for_entity('risk', risk.id):
            item['risk'] = risk
            risk_docs.append(item)
    # Actions
    action_docs = []
    for action in hazard.actions:
        for item in get_doc_links_for_entity('action', action.id):
            item['action'] = action
            action_docs.append(item)
    return render_template('hazard/hazard_documents.html',
        hazard=hazard, haz_docs=haz_docs,
        risk_docs=risk_docs, action_docs=action_docs)

@app.route('/audit-schedule/<sid>/documents')
@require_login
def audit_documents(sid):
    schedule = AuditSchedule.query.get_or_404(sid)
    audit_docs   = get_doc_links_for_entity('audit_schedule', sid)
    finding_docs = []
    for finding in schedule.findings:
        for item in get_doc_links_for_entity('audit_finding', finding.id):
            item['finding'] = finding
            finding_docs.append(item)
    return render_template('audit/audit_documents.html',
        schedule=schedule, audit_docs=audit_docs, finding_docs=finding_docs)

# ─── TRACEABILITY DASHBOARD ───────────────────────────────────────────────────
@app.route('/traceability')
@require_login
def traceability_dashboard():
    total_docs  = SMSDocument.query.count()
    total_links = DocumentLink.query.count()
    approved    = SMSDocument.query.filter_by(status='Approved').count()
    draft       = SMSDocument.query.filter_by(status='Draft').count()
    archived    = SMSDocument.query.filter_by(status='Archived').count()
    review      = SMSDocument.query.filter_by(status='Under Review').count()

    # Documents with no links (orphans)
    linked_ids = db.session.query(DocumentLink.document_id).distinct().all()
    linked_ids = [x[0] for x in linked_ids]
    orphan_docs = SMSDocument.query.filter(
        ~SMSDocument.id.in_(linked_ids)).all() if linked_ids else SMSDocument.query.all()

    # Hazards with no linked RA document
    all_hazards = Hazard.query.filter_by(status='Open').all()
    unlinked_hazards = []
    for h in all_hazards:
        ra_links = DocumentLink.query.filter_by(entity_type='hazard', entity_id=h.id).join(
            SMSDocument, DocumentLink.document_id == SMSDocument.id).filter(
            SMSDocument.doc_type == 'RA').first()
        if not ra_links:
            unlinked_hazards.append(h)

    # Recent links
    recent_links = DocumentLink.query.order_by(DocumentLink.created_at.desc()).limit(10).all()

    # Link stats by entity type
    link_stats = db.session.query(
        DocumentLink.entity_type,
        db.func.count(DocumentLink.id).label('cnt')
    ).group_by(DocumentLink.entity_type).all()

    return render_template('document/traceability.html',
        total_docs=total_docs, total_links=total_links,
        approved=approved, draft=draft, archived=archived, review=review,
        orphan_docs=orphan_docs, unlinked_hazards=unlinked_hazards,
        recent_links=recent_links, link_stats=link_stats,
        resolve_entity_label=resolve_entity_label)

# ─── QUICK LINK API (from any module page) ───────────────────────────────────
@app.route('/quick-link', methods=['POST'])
@require_login
def quick_link():
    """Link a document to an entity from any page in the system."""
    f           = request.form
    doc_id      = f['document_id']
    entity_type = f['entity_type']
    entity_id   = f['entity_id']
    reason      = f.get('link_reason', '')
    return_url  = f.get('return_url', '/documents')

    doc = SMSDocument.query.get(doc_id)
    if not doc:
        flash(f'Document {doc_id} not found.', 'error')
        return redirect(return_url)

    existing = DocumentLink.query.filter_by(
        document_id=doc_id, entity_type=entity_type, entity_id=entity_id).first()
    if not existing:
        lnk = DocumentLink(
            document_id=doc_id, entity_type=entity_type,
            entity_id=entity_id, link_reason=reason)
        db.session.add(lnk)
        db.session.commit()
        flash(f'✓ {doc_id} linked to {entity_type} {entity_id}.', 'success')
    else:
        flash('Already linked.', 'error')
    return redirect(return_url)

# ─── AUTO-LINK HELPERS (called internally when creating objects) ──────────────
def auto_link_document(doc_id, entity_type, entity_id, reason='Auto-linked'):
    """Safe auto-link — skips if already exists or entity not found."""
    if not doc_id or not entity_id:
        return
    existing = DocumentLink.query.filter_by(
        document_id=doc_id, entity_type=entity_type,
        entity_id=str(entity_id)).first()
    if not existing:
        lnk = DocumentLink(
            document_id=doc_id, entity_type=entity_type,
            entity_id=str(entity_id), link_reason=reason)
        db.session.add(lnk)

# ─── SEED TRACEABILITY DATA (called from seed()) ─────────────────────────────
def seed_traceability():
    """Traceability seed — skipped in production (no demo data)."""
    pass


# ═══════════════════════════════════════════════════════════════════════════════
#  SAFETY RISK MANAGEMENT (SRM) MODULE
#  ICAO Annex 19 §5 / Doc 9859 Ch.5
#  Hazard → Risk(s) → Control(s) → Residual Risk → Action(s) → Monitoring
#  Extension only — all existing routes unchanged
# ═══════════════════════════════════════════════════════════════════════════════

TREND_ICONS = {'Increasing': '↑', 'Stable': '→', 'Decreasing': '↓', 'New': '●'}
TREND_COLORS = {'Increasing': '#dc2626', 'Stable': '#d97706', 'Decreasing': '#15803d', 'New': '#1e40af'}

def calculate_trend(hazard_id):
    """Calculate trend based on occurrence count and recency."""
    occurrences = RiskOccurrence.query.filter_by(hazard_id=hazard_id).order_by(
        RiskOccurrence.created_at.desc()).all()
    count = len(occurrences)
    if count == 0:
        return 'New', 0
    if count == 1:
        return 'Stable', count
    # Compare recent 3 vs previous 3
    recent = len([o for o in occurrences[:3]])
    older  = len([o for o in occurrences[3:6]])
    if recent > older:
        return 'Increasing', count
    elif recent < older:
        return 'Decreasing', count
    return 'Stable', count

def get_srm_status(hazard):
    """Derive SRM status from risks and controls."""
    if not hazard.risks:
        return 'Open'
    all_controlled = all(
        r.residual_risk_index and r.residual_tolerance in ('ACCEPTABLE','TOLERABLE')
        for r in hazard.risks
    )
    any_intolerable = any(
        r.initial_tolerance == 'INTOLERABLE' for r in hazard.risks
    )
    has_controls = any(len(r.controls) > 0 for r in hazard.risks)
    if all_controlled:
        return 'Controlled'
    if has_controls:
        return 'Under Assessment'
    return 'Open'

# ─── RISK REGISTER (central view — risks, not hazards) ───────────────────────
@app.route('/risk-register')
@require_login
def risk_register():
    dept_f  = request.args.get('dept','')
    tol_f   = request.args.get('tolerance','')
    stat_f  = request.args.get('status','')
    src_f   = request.args.get('source','')

    q = Risk.query.join(Hazard, Risk.hazard_id == Hazard.id)
    if dept_f: q = q.filter(Hazard.department_id == int(dept_f))
    if tol_f:  q = q.filter(Risk.initial_tolerance == tol_f)

    risks = q.order_by(Risk.created_at.desc()).all()

    # Filter by hazard source
    if src_f:
        risks = [r for r in risks if r.hazard and r.hazard.source == src_f]

    # Stats
    total       = len(risks)
    intolerable = sum(1 for r in risks if r.initial_tolerance == 'INTOLERABLE')
    tolerable   = sum(1 for r in risks if r.initial_tolerance == 'TOLERABLE')
    acceptable  = sum(1 for r in risks if r.initial_tolerance == 'ACCEPTABLE')
    no_controls = sum(1 for r in risks if len(r.controls) == 0)
    no_residual = sum(1 for r in risks if not r.residual_risk_index)

    return render_template('risk/risk_register.html',
        risks=risks, dept_f=dept_f, tol_f=tol_f, stat_f=stat_f, src_f=src_f,
        total=total, intolerable=intolerable, tolerable=tolerable,
        acceptable=acceptable, no_controls=no_controls, no_residual=no_residual,
        get_srm_status=get_srm_status, calculate_trend=calculate_trend,
        TREND_ICONS=TREND_ICONS, TREND_COLORS=TREND_COLORS)

# ─── RISK DETAIL ──────────────────────────────────────────────────────────────
@app.route('/risk/<rid>')
@require_login
def risk_detail(rid):
    risk    = Risk.query.get_or_404(rid)
    hazard  = risk.hazard
    # Linked documents via DocumentLink
    doc_links = DocumentLink.query.filter_by(entity_type='risk', entity_id=rid).all()
    docs    = [SMSDocument.query.get(lnk.document_id) for lnk in doc_links if SMSDocument.query.get(lnk.document_id)]
    # Risk actions
    r_actions = RiskAction.query.filter_by(risk_id=rid).order_by(RiskAction.created_at.desc()).all()
    # Audit findings linked to same hazard
    audit_findings = AuditFinding.query.filter_by(hazard_id=hazard.id).all() if hazard else []
    return render_template('risk/risk_detail.html',
        risk=risk, hazard=hazard, docs=docs, r_actions=r_actions,
        audit_findings=audit_findings)

# ─── UPDATE RISK STATUS / RESIDUAL ────────────────────────────────────────────
@app.route('/risk/<rid>/update', methods=['POST'])
@require_login
def update_risk(rid):
    risk = Risk.query.get_or_404(rid)
    f    = request.form
    risk.description = f.get('description', risk.description)
    rl = f.get('residual_likelihood','')
    rs = f.get('residual_severity','')
    if rl and rs:
        risk.residual_likelihood = int(rl)
        risk.residual_severity   = rs
        rri = f'{rl}{rs}'
        risk.residual_risk_index  = rri
        risk.residual_tolerance   = get_tolerance(rri)
    if f.get('consequence'):
        risk.description = f.get('consequence')
    db.session.commit()
    flash('✓ Risk updated.', 'success')
    return redirect(url_for('risk_detail', rid=rid))

# ─── RISK → ACTION (direct risk-level action) ─────────────────────────────────
@app.route('/risk/<rid>/add-action', methods=['POST'])
@require_login
def add_risk_action(rid):
    risk = Risk.query.get_or_404(rid)
    f    = request.form
    ra   = RiskAction(
        id=new_id('RACT'),
        risk_id=rid,
        hazard_id=risk.hazard_id,
        description=f['description'],
        owner=f['owner'],
        due_date=f['due_date'],
        priority=f.get('priority','Medium'),
        status='Open'
    )
    db.session.add(ra)
    # Also add to unified Action table
    unified = Action(
        id=new_id('ACT'),
        source='Risk Assessment',
        hazard_id=risk.hazard_id,
        linked_ref_id=rid,
        description=f['description'],
        owner=f['owner'],
        due_date=f['due_date'],
        priority=f.get('priority','Medium'),
        status='Open'
    )
    db.session.add(unified)
    db.session.commit()
    sync_report_status(unified.hazard_id)
    db.session.commit()
    flash(f'✓ Action created for risk {rid}.', 'success')
    return redirect(url_for('risk_detail', rid=rid))

@app.route('/risk-action/<aid>/update', methods=['POST'])
@require_login
def update_risk_action(aid):
    ra = RiskAction.query.get_or_404(aid)
    f  = request.form
    ra.status        = f.get('status', ra.status)
    ra.owner         = f.get('owner', ra.owner)
    ra.due_date      = f.get('due_date', ra.due_date)
    ra.effectiveness = f.get('effectiveness', ra.effectiveness)
    if ra.status == 'Closed':
        ra.closed_date = date.today().isoformat()
    db.session.commit()
    sync_report_status(a.hazard_id)
    db.session.commit()
    flash('✓ Action updated.', 'success')
    return redirect(url_for('risk_detail', rid=ra.risk_id))

# ─── CONTROL MANAGEMENT (enhanced) ───────────────────────────────────────────
@app.route('/control/<cid>/update', methods=['POST'])
@require_login
def update_control(cid):
    ctrl = Control.query.get_or_404(cid)
    f    = request.form
    ctrl.control_type  = f.get('control_type', ctrl.control_type)
    ctrl.description   = f.get('description', ctrl.description)
    ctrl.owner         = f.get('owner', ctrl.owner)
    ctrl.effectiveness = f.get('effectiveness', ctrl.effectiveness)
    ctrl.review_date   = f.get('review_date', ctrl.review_date)
    db.session.commit()
    flash('✓ Control updated.', 'success')
    return redirect(url_for('risk_detail', rid=ctrl.risk_id))

@app.route('/delete/risk/<rid>', methods=['POST'])
@require_login
def delete_risk_record(rid):
    """Safe delete a Risk row and its controls."""
    r = Risk.query.get_or_404(rid)
    try:
        # RiskAction.risk_id is NOT NULL FK — must delete before deleting Risk
        RiskAction.query.filter_by(risk_id=rid).delete(synchronize_session=False)
        db.session.flush()
        # Nullify ra_rows referencing this risk
        db.session.execute(
            db.text("UPDATE ra_rows SET risk_id = NULL WHERE risk_id = :rid"),
            {'rid': rid}
        )
        db.session.flush()
        Control.query.filter_by(risk_id=rid).delete(synchronize_session=False)
        db.session.flush()
        db.session.delete(r)
        db.session.commit()
        flash(f'✓ Risk {rid} deleted.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'⚠ Could not delete risk: {str(e)[:120]}', 'error')
    return redirect(request.form.get('return_url', '/risk-register'))


@app.route('/control/<cid>/delete', methods=['POST'])
@require_login
def delete_control(cid):
    ctrl = Control.query.get_or_404(cid)
    rid  = ctrl.risk_id
    db.session.delete(ctrl)
    db.session.commit()
    flash('✓ Control removed.', 'success')
    return redirect(url_for('risk_detail', rid=rid))

# ─── OCCURRENCE TRACKING ──────────────────────────────────────────────────────
@app.route('/hazard-log/<hid>/occurrence', methods=['POST'])
@require_login
def add_occurrence(hid):
    hazard = Hazard.query.get_or_404(hid)
    f      = request.form
    occ    = RiskOccurrence(
        hazard_id=hid,
        occurrence_date=f.get('occurrence_date', date.today().isoformat()),
        description=f.get('description',''),
        source=f.get('source','Report'),
        linked_report_id=f.get('linked_report_id','')
    )
    db.session.add(occ)
    db.session.commit()
    flash('✓ Occurrence logged. Trend updated.', 'success')
    return redirect(url_for('hazard_detail', hid=hid))

# ─── SRM DASHBOARD ────────────────────────────────────────────────────────────
@app.route('/srm-dashboard')
@require_login
def srm_dashboard():
    all_risks   = Risk.query.all()
    total_risks = len(all_risks)
    intol_risks = [r for r in all_risks if r.initial_tolerance == 'INTOLERABLE']
    no_ctrl     = [r for r in all_risks if len(r.controls) == 0]
    no_resid    = [r for r in all_risks if not r.residual_risk_index]
    reduced     = [r for r in all_risks if r.residual_risk_index and r.residual_tolerance != r.initial_tolerance]

    # Trend analysis per hazard
    all_hazards = Hazard.query.filter_by(status='Open').all()
    trend_data  = []
    for h in all_hazards:
        trend, count = calculate_trend(h.id)
        if count > 0 or h.risks:
            trend_data.append({
                'hazard': h,
                'trend': trend,
                'count': count,
                'risks': len(h.risks),
                'intol': sum(1 for r in h.risks if r.initial_tolerance == 'INTOLERABLE')
            })
    trend_data.sort(key=lambda x: (x['intol'], x['count']), reverse=True)

    # Source breakdown
    sources = {}
    for h in Hazard.query.all():
        sources[h.source] = sources.get(h.source, 0) + 1

    # Classification breakdown
    classifications = {}
    for h in Hazard.query.all():
        c = h.classification or 'Unclassified'
        classifications[c] = classifications.get(c, 0) + 1

    return render_template('risk/srm_dashboard.html',
        total_risks=total_risks, intol_risks=intol_risks,
        no_ctrl=no_ctrl, no_resid=no_resid, reduced=reduced,
        trend_data=trend_data[:10],
        sources=sources, classifications=classifications,
        TREND_ICONS=TREND_ICONS, TREND_COLORS=TREND_COLORS)


# ═══════════════════════════════════════════════════════════════════════════════
#  RISK ASSESSMENT MODULE — AviaS/SMS/001 Rev 01
#  Converts the AviaS RA form into a full system module
#  Connected to: Hazard Log, Risk Register, Actions, Documents
# ═══════════════════════════════════════════════════════════════════════════════

def gen_control_number(dept_code):
    """Generate RA control number: JAV/RA/DEPT/YEAR/SEQ"""
    year = datetime.now().year
    count = RiskAssessment.query.count() + 1
    return f"JAV-RA-{dept_code}-{year}-{count:03d}"

def compute_ra_summary(ra):
    """Compute overall risk level before and after controls for page 2."""
    if not ra.rows:
        return None, None
    # Worst initial risk
    order = ['INTOLERABLE','TOLERABLE','ACCEPTABLE']
    initial_levels  = [r.risk_tolerance_initial  for r in ra.rows if r.risk_tolerance_initial]
    residual_levels = [r.risk_tolerance_residual for r in ra.rows if r.risk_tolerance_residual]
    worst_initial  = min(initial_levels,  key=lambda x: order.index(x) if x in order else 99) if initial_levels else None
    worst_residual = min(residual_levels, key=lambda x: order.index(x) if x in order else 99) if residual_levels else None
    return worst_initial, worst_residual

# ─── LIST ALL RISK ASSESSMENTS ───────────────────────────────────────────────
@app.route('/risk-assessments')
@require_login
def ra_list():
    check_ra_review_dates()  # Auto-set Under Review when due date reached
    dept_f  = request.args.get('dept','')
    stat_f  = request.args.get('status','')
    q = RiskAssessment.query
    if dept_f: q = q.filter_by(department_id=int(dept_f))
    if stat_f: q = q.filter_by(status=stat_f)
    ras = q.order_by(RiskAssessment.created_at.desc()).all()
    return render_template('risk/ra_list.html', ras=ras, dept_f=dept_f, stat_f=stat_f)

# ─── CREATE NEW RA (linked to hazard or standalone) ──────────────────────────
@app.route('/risk-assessments/new', methods=['GET','POST'])
@require_login
def new_ra():
    hid = request.args.get('hazard_id','')
    hazard = Hazard.query.get(hid) if hid else None

    if request.method == 'POST':
        f       = request.form
        dept_id = int(f['department_id'])
        dept    = Department.query.get(dept_id)
        ra_id   = new_id('RA')
        ctrl_no = gen_control_number(dept.code if dept else 'XX')

        ra = RiskAssessment(
            id=ra_id,
            control_number=f.get('control_number') or ctrl_no,
            responsible_name=f['responsible_name'],
            assessors_names=f.get('assessors_names',''),
            assessment_date=f['assessment_date'],
            next_review_date=f.get('next_review_date',''),
            title=f['title'],
            hazard_id=f.get('hazard_id') or None,
            department_id=dept_id,
            general_description=f.get('general_description',''),
            reasons=f.get('reasons',''),
            management_acceptance=f.get('management_acceptance',''),
            acceptance_date=f.get('acceptance_date',''),
            prepared_by_name=f.get('prepared_by_name',''),
            prepared_by_position=f.get('prepared_by_position',''),
            reviewed_by_name=f.get('reviewed_by_name',''),
            reviewed_by_position=f.get('reviewed_by_position',''),
            approved_by_name=f.get('approved_by_name',''),
            approved_by_position=f.get('approved_by_position',''),
            status='Draft'
        )
        db.session.add(ra)
        db.session.flush()

        # Page 3 — parse risk rows from the form
        seq = 1
        while f.get(f'activity_{seq}'):
            lik_i = int(f.get(f'lik_i_{seq}', 3))
            sev_i = f.get(f'sev_i_{seq}', 'C')
            ri_i  = f'{lik_i}{sev_i}'
            tol_i = get_tolerance(ri_i)

            lik_r = f.get(f'lik_r_{seq}','')
            sev_r = f.get(f'sev_r_{seq}','')
            ri_r  = f'{lik_r}{sev_r}' if lik_r and sev_r else None
            tol_r = get_tolerance(ri_r) if ri_r else None

            # Create/link a Risk record in the existing risks table
            risk_rec = Risk(
                id=new_id('RSK'),
                hazard_id=ra.hazard_id or '',
                description=f.get(f'consequences_{seq}',''),
                initial_likelihood=lik_i, initial_severity=sev_i,
                initial_risk_index=ri_i, initial_tolerance=tol_i,
                residual_likelihood=int(lik_r) if lik_r else None,
                residual_severity=sev_r or None,
                residual_risk_index=ri_r, residual_tolerance=tol_r
            )
            if ra.hazard_id:
                db.session.add(risk_rec)
                db.session.flush()

            row = RARow(
                assessment_id=ra_id, seq_num=seq,
                risk_id=risk_rec.id if ra.hazard_id else None,
                type_of_activity=f.get(f'activity_{seq}',''),
                generic_hazard=f.get(f'generic_hazard_{seq}',''),
                specific_components=f.get(f'specific_{seq}',''),
                consequences=f.get(f'consequences_{seq}',''),
                likelihood_initial=lik_i, severity_initial=sev_i,
                risk_index_initial=ri_i, risk_tolerance_initial=tol_i,
                current_defenses=f.get(f'defenses_{seq}',''),
                further_mitigations=f.get(f'mitigations_{seq}',''),
                likelihood_residual=int(lik_r) if lik_r else None,
                severity_residual=sev_r or None,
                risk_index_residual=ri_r, risk_tolerance_residual=tol_r
            )
            db.session.add(row)

            # Page 4 — auto-create mitigation + action if mitigation text exists
            mit_text = f.get(f'mitigations_{seq}','')
            resp_mgr = f.get(f'resp_manager_{seq}','')
            due_dt   = f.get(f'due_date_{seq}','')
            if mit_text:
                act_id = new_id('ACT')
                mit = RAMitigation(
                    assessment_id=ra_id,
                    hazard_seq=str(seq),
                    mitigation=mit_text,
                    responsible_manager=resp_mgr,
                    due_date=due_dt,
                    action_id=act_id,
                    status='Open'
                )
                db.session.add(mit)
                # Create unified Action
                action = Action(
                    id=act_id,
                    source='Risk Assessment',
                    hazard_id=ra.hazard_id,
                    linked_ref_id=ra_id,
                    description=f'[RA {ra.control_number}] Seq {seq}: {mit_text}',
                    owner=resp_mgr,
                    due_date=due_dt,
                    priority='High' if tol_i=='INTOLERABLE' else 'Medium',
                    status='Open'
                )
                db.session.add(action)
            seq += 1

        # Update page 2 summary levels
        worst_i, worst_r = compute_ra_summary(ra)
        ra.risk_level_prior = worst_i or ''
        ra.risk_level_after = worst_r or ''

        # Auto-link RA document to hazard in traceability
        if ra.hazard_id:
            auto_link_document(None, 'hazard', ra.hazard_id, f'Risk Assessment {ra.control_number}')

        db.session.commit()
        flash(f'✓ Risk Assessment {ra.control_number} created. {seq-1} risk row(s) added. Assign it to the Safety Action Group below.', 'success')
        return redirect(url_for('ra_assign_sag', ra_id=ra_id))

    # GET — pre-populate from hazard if provided
    return render_template('risk/ra_form.html', hazard=hazard,
                           today=date.today().isoformat())

# ─── RA DETAIL (all 5 pages in one view) ────────────────────────────────────
@app.route('/risk-assessments/<ra_id>')
@require_login
def ra_detail(ra_id):
    ra = RiskAssessment.query.get_or_404(ra_id)
    worst_i, worst_r = compute_ra_summary(ra)
    return render_template('risk/ra_detail.html', ra=ra,
                           worst_initial=worst_i, worst_residual=worst_r,
                           get_tolerance=get_tolerance)

# ─── RA ASSIGN SAG REVIEW ─────────────────────────────────────────────────────
@app.route('/risk-assessments/<ra_id>/assign-sag', methods=['GET', 'POST'])
@require_login
def ra_assign_sag(ra_id):
    """Step after RA creation: assign actions to SAG members for review."""
    ra = RiskAssessment.query.get_or_404(ra_id)
    actions = Action.query.filter_by(linked_ref_id=ra_id, source='Risk Assessment').all()
    sag_members = User.query.filter_by(is_active=True).order_by(User.full_name).all()
    departments = Department.query.order_by(Department.name).all()

    if request.method == 'POST':
        f = request.form
        assigned_count = 0
        for action in actions:
            sag_key  = f'sag_{action.id}'
            dept_key = f'dept_{action.id}'
            due_key  = f'due_{action.id}'
            if f.get(sag_key):
                old = action.sag_member or 'Unassigned'
                action.sag_member    = f[sag_key]
                action.department_id = int(f[dept_key]) if f.get(dept_key) else action.department_id
                action.due_date      = f.get(due_key) or action.due_date
                if action.status == 'Open':
                    action.status = 'Assigned'
                log_action_history(action.id, session.get('admin_name', 'Admin'),
                                   old, action.sag_member,
                                   f'SAG assigned via RA {ra.control_number}', 'assignment')
                push_notify_by_name(action.sag_member,
                    f'Action Assigned — {ra.control_number}',
                    f'You have been assigned a corrective action for Risk Assessment {ra.control_number}.',
                    'action_assigned', 'action', action.id)
                assigned_count += 1
        # Advance RA to Submitted
        if assigned_count and ra.status == 'Draft':
            ra.status = 'Submitted'
            ra.submitted_date = datetime.utcnow().strftime('%Y-%m-%d')
        db.session.commit()
        flash(f'✓ {assigned_count} action(s) assigned to SAG. Risk Assessment submitted for review.', 'success')
        return redirect(url_for('ra_detail', ra_id=ra_id))

    return render_template('risk/ra_assign_sag.html',
                           ra=ra, actions=actions,
                           sag_members=sag_members, departments=departments)

# ─── ADD ROW to existing RA ──────────────────────────────────────────────────
@app.route('/risk-assessments/<ra_id>/add-row', methods=['POST'])
@require_login
def ra_add_row(ra_id):
    ra = RiskAssessment.query.get_or_404(ra_id)
    f  = request.form
    seq = len(ra.rows) + 1

    lik_i = int(f.get('likelihood_initial', 3))
    sev_i = f.get('severity_initial','C')
    ri_i  = f'{lik_i}{sev_i}'
    tol_i = get_tolerance(ri_i)
    lik_r = f.get('likelihood_residual','')
    sev_r = f.get('severity_residual','')
    ri_r  = f'{lik_r}{sev_r}' if lik_r and sev_r else None
    tol_r = get_tolerance(ri_r) if ri_r else None

    # Create Risk record
    if ra.hazard_id:
        risk_rec = Risk(
            id=new_id('RSK'), hazard_id=ra.hazard_id,
            description=f.get('consequences',''),
            initial_likelihood=lik_i, initial_severity=sev_i,
            initial_risk_index=ri_i, initial_tolerance=tol_i,
            residual_likelihood=int(lik_r) if lik_r else None,
            residual_severity=sev_r or None,
            residual_risk_index=ri_r, residual_tolerance=tol_r
        )
        db.session.add(risk_rec)
        db.session.flush()
        risk_id = risk_rec.id
    else:
        risk_id = None

    row = RARow(
        assessment_id=ra_id, seq_num=seq, risk_id=risk_id,
        type_of_activity=f.get('type_of_activity',''),
        generic_hazard=f.get('generic_hazard',''),
        specific_components=f.get('specific_components',''),
        consequences=f.get('consequences',''),
        likelihood_initial=lik_i, severity_initial=sev_i,
        risk_index_initial=ri_i, risk_tolerance_initial=tol_i,
        current_defenses=f.get('current_defenses',''),
        further_mitigations=f.get('further_mitigations',''),
        likelihood_residual=int(lik_r) if lik_r else None,
        severity_residual=sev_r or None,
        risk_index_residual=ri_r, risk_tolerance_residual=tol_r
    )
    db.session.add(row)

    # Create mitigation + action if provided
    mit_text = f.get('further_mitigations','')
    resp_mgr = f.get('responsible_manager','')
    due_dt   = f.get('due_date','')
    if mit_text:
        act_id = new_id('ACT')
        mit = RAMitigation(
            assessment_id=ra_id, hazard_seq=str(seq),
            mitigation=mit_text, responsible_manager=resp_mgr,
            due_date=due_dt, action_id=act_id, status='Open'
        )
        db.session.add(mit)
        action = Action(
            id=act_id, source='Risk Assessment',
            hazard_id=ra.hazard_id, linked_ref_id=ra_id,
            description=f'[{ra.control_number}] Seq {seq}: {mit_text}',
            owner=resp_mgr, due_date=due_dt,
            priority='High' if tol_i=='INTOLERABLE' else 'Medium',
            status='Open'
        )
        db.session.add(action)

    # Refresh summary
    worst_i, worst_r = compute_ra_summary(ra)
    ra.risk_level_prior = worst_i or ra.risk_level_prior
    ra.risk_level_after = worst_r or ra.risk_level_after
    db.session.commit()
    flash(f'✓ Risk row {seq} added. Risk index: {ri_i} ({tol_i}).', 'success')
    return redirect(url_for('ra_detail', ra_id=ra_id))

# ─── ADD REVIEW (Page 5) ─────────────────────────────────────────────────────
@app.route('/risk-assessments/<ra_id>/add-review', methods=['POST'])
@require_login
def ra_add_review(ra_id):
    ra = RiskAssessment.query.get_or_404(ra_id)
    f  = request.form
    rev = RAReview(
        assessment_id=ra_id,
        risk_mitigation=f.get('risk_mitigation',''),
        review_of_effectiveness=f.get('review_of_effectiveness',''),
        effectiveness_rating=f.get('effectiveness_rating',''),
        date_completed=f.get('date_completed',''),
        actioner=f.get('actioner','')
    )
    db.session.add(rev)
    # Update linked mitigation status if effectiveness is set
    if f.get('effectiveness_rating') == 'Effective':
        mit = RAMitigation.query.filter_by(
            assessment_id=ra_id,
            hazard_seq=f.get('hazard_seq','')).first()
        if mit:
            mit.status = 'Completed'
    db.session.commit()
    flash('✓ Mitigation review recorded.', 'success')
    return redirect(url_for('ra_detail', ra_id=ra_id))

# ─── UPDATE RA HEADER (approval / status) ────────────────────────────────────
@app.route('/risk-assessments/<ra_id>/update', methods=['POST'])
@require_login
def ra_update(ra_id):
    ra = RiskAssessment.query.get_or_404(ra_id)
    f  = request.form
    # All General Information fields — now fully editable
    ra.status                = f.get('status', ra.status)
    ra.management_acceptance = f.get('management_acceptance', ra.management_acceptance)
    ra.acceptance_date       = f.get('acceptance_date', ra.acceptance_date)
    ra.next_review_date      = f.get('next_review_date', ra.next_review_date)
    ra.risk_level_prior      = f.get('risk_level_prior', ra.risk_level_prior)
    ra.risk_level_after      = f.get('risk_level_after', ra.risk_level_after)
    # General Info
    if f.get('responsible_name'):
        ra.responsible_name  = f.get('responsible_name')
    if f.get('assessors_names') is not None:
        ra.assessors_names   = f.get('assessors_names')
    if f.get('assessment_date'):
        ra.assessment_date   = f.get('assessment_date')
    if f.get('reasons') is not None:
        ra.reasons           = f.get('reasons')
    if f.get('general_description') is not None:
        ra.general_description = f.get('general_description')
    if f.get('department_id'):
        ra.department_id     = int(f.get('department_id'))
    # Signatories — all three rows
    if f.get('prepared_by_name') is not None:
        ra.prepared_by_name      = f.get('prepared_by_name')
    if f.get('prepared_by_position') is not None:
        ra.prepared_by_position  = f.get('prepared_by_position')
    if f.get('reviewed_by_name') is not None:
        ra.reviewed_by_name      = f.get('reviewed_by_name')
    if f.get('reviewed_by_position') is not None:
        ra.reviewed_by_position  = f.get('reviewed_by_position')
    if f.get('approved_by_name') is not None:
        ra.approved_by_name      = f.get('approved_by_name')
    if f.get('approved_by_position') is not None:
        ra.approved_by_position  = f.get('approved_by_position')
    db.session.commit()
    # ── AVI Hook: RA closed/approved → verify mitigation effectiveness ────────
    if ra.status in ('Closed', 'Approved') and f.get('status') in ('Closed', 'Approved'):
        try:
            _avi_generate(
                source_module='risk', source_record_id=ra_id,
                source_description=f'Risk Assessment {ra.status}: {(ra.general_description or ra.title or "")[:200]}',
                department_id=ra.department_id,
                linked_risk_id=ra_id,
                linked_hazard_id=getattr(ra, 'hazard_id', None),
                operational_risk='High' if (ra.risk_level_after or '') in ('High','Critical','Catastrophic') else 'Medium',
                override_objective=f'Verify that risk mitigations in RA "{ra_id}" remain adequate and residual risk has not been re-elevated.',
            )
            db.session.commit()
        except Exception:
            pass
    flash('✓ Risk Assessment updated successfully.', 'success')
    return redirect(url_for('ra_detail', ra_id=ra_id))

# ─── TRIGGER RA FROM HAZARD LOG ──────────────────────────────────────────────
@app.route('/hazard-log/<hid>/start-ra')
@require_login
def start_ra_from_hazard(hid):
    """Redirect to new RA form pre-populated from hazard."""
    hazard = Hazard.query.get_or_404(hid)
    # Check if RA already exists
    existing = RiskAssessment.query.filter_by(hazard_id=hid).first()
    if existing:
        return redirect(url_for('ra_detail', ra_id=existing.id))
    return redirect(url_for('new_ra', hazard_id=hid))


# ═══════════════════════════════════════════════════════════════════════════════
#  GUIDED RISK ASSESSMENT WIZARD — 6-STEP WORKFLOW
#  Triggered automatically after Hazard Report / ASR submission
#  ICAO Annex 19 §5 / Doc 9859 Ch.5
# ═══════════════════════════════════════════════════════════════════════════════

WIZARD_STEPS = [
    (1, 'Hazard Review',        'Review the reported hazard'),
    (2, 'Risk Identification',  'Identify consequences and risk scenarios'),
    (3, 'Initial Risk Rating',  'Rate likelihood and severity'),
    (4, 'Current Controls',     'Check existing defences and controls'),
    (5, 'Further Mitigations',  'Define additional mitigation actions'),
    (6, 'Residual Risk',        'Recalculate risk after controls'),
]

CONTROL_CHECKLIST = [
    ('SOP',       'Standard Operating Procedure (SOP) available and current'),
    ('SOP',       'Crew / staff briefed on relevant SOP'),
    ('Training',  'Specific training programme exists for this hazard type'),
    ('Training',  'Personnel have completed required training and are current'),
    ('Monitoring','Regular monitoring / inspection process in place'),
    ('Monitoring','Safety data collected and reviewed for this hazard'),
    ('Equipment', 'Technical safeguards or equipment controls installed'),
    ('Equipment', 'Equipment is serviceable and within maintenance cycle'),
    ('Procedure', 'Emergency / contingency procedure defined'),
    ('Procedure', 'Supervisory checks / sign-off required before operation'),
    ('Reporting', 'Hazard reporting culture promoted in department'),
    ('Reporting', 'Occurrence data analysed and fed back to department'),
]

def get_or_create_ra(hid):
    """Get existing RA for hazard or create a new draft one."""
    hazard = Hazard.query.get_or_404(hid)
    ra = RiskAssessment.query.filter_by(hazard_id=hid).first()
    if not ra:
        dept = hazard.department
        ctrl_no = gen_control_number(dept.code if dept else 'XX')
        ra = RiskAssessment(
            id=new_id('RA'),
            control_number=ctrl_no,
            responsible_name='',
            assessment_date=date.today().isoformat(),
            title=hazard.generic_hazard or 'Risk Assessment',
            hazard_id=hid,
            department_id=hazard.department_id,
            general_description=hazard.specific_components or '',
            reasons=f'Hazard reported from {hazard.source}',
            status='Draft'
        )
        db.session.add(ra)
        db.session.commit()
    return ra

# ─── WIZARD ENTRY POINT ───────────────────────────────────────────────────────
@app.route('/ra-wizard/<hid>')
@require_login
def ra_wizard_start(hid):
    hazard = Hazard.query.get_or_404(hid)
    ra     = get_or_create_ra(hid)
    return redirect(url_for('ra_wizard_step', hid=hid, step=1))

# ─── STEP ROUTER ─────────────────────────────────────────────────────────────
@app.route('/ra-wizard/<hid>/step/<int:step>', methods=['GET','POST'])
@require_login
def ra_wizard_step(hid, step):
    hazard = Hazard.query.get_or_404(hid)
    ra     = get_or_create_ra(hid)
    if step < 1 or step > 6:
        return redirect(url_for('ra_wizard_step', hid=hid, step=1))

    # ── POST: save current step data ─────────────────────────────────────────
    if request.method == 'POST':
        f = request.form

        if step == 1:
            # Save ALL admin + general info + signatory fields
            ra.responsible_name      = f.get('responsible_name', ra.responsible_name)
            ra.assessors_names       = f.get('assessors_names', ra.assessors_names)
            ra.assessment_date       = f.get('assessment_date', ra.assessment_date)
            ra.next_review_date      = f.get('next_review_date', ra.next_review_date)
            ra.title                 = f.get('title', ra.title)
            ra.reasons               = f.get('reasons', ra.reasons)
            ra.general_description   = f.get('general_description', ra.general_description)
            # Signatories
            ra.prepared_by_name      = f.get('prepared_by_name', ra.prepared_by_name)
            ra.prepared_by_position  = f.get('prepared_by_position', ra.prepared_by_position)
            ra.reviewed_by_name      = f.get('reviewed_by_name', ra.reviewed_by_name)
            ra.reviewed_by_position  = f.get('reviewed_by_position', ra.reviewed_by_position)
            ra.approved_by_name      = f.get('approved_by_name', ra.approved_by_name)
            ra.approved_by_position  = f.get('approved_by_position', ra.approved_by_position)
            db.session.commit()

        elif step == 2:
            # Save risk rows (one or more risk scenarios)
            activities  = f.getlist('type_of_activity[]')
            hazards_g   = f.getlist('generic_hazard[]')
            components  = f.getlist('specific_components[]')
            consequences= f.getlist('consequences[]')
            # Remove existing rows first if re-doing step 2
            # Only add new rows that don't already exist (by seq)
            existing_seqs = {r.seq_num for r in ra.rows}
            for i, cons in enumerate(consequences):
                if not cons.strip():
                    continue
                seq = i + 1
                if seq not in existing_seqs:
                    row = RARow(
                        assessment_id=ra.id,
                        seq_num=seq,
                        type_of_activity=activities[i] if i < len(activities) else '',
                        generic_hazard=hazards_g[i] if i < len(hazards_g) else '',
                        specific_components=components[i] if i < len(components) else '',
                        consequences=cons,
                        likelihood_initial=3, severity_initial='C',
                        risk_index_initial='3C', risk_tolerance_initial='TOLERABLE'
                    )
                    # Also create Risk record
                    if ra.hazard_id:
                        rsk = Risk(
                            id=new_id('RSK'), hazard_id=ra.hazard_id,
                            description=cons,
                            initial_likelihood=3, initial_severity='C',
                            initial_risk_index='3C', initial_tolerance='TOLERABLE'
                        )
                        db.session.add(rsk)
                        db.session.flush()
                        row.risk_id = rsk.id
                    db.session.add(row)
            db.session.commit()

        elif step == 3:
            # Save initial risk rating per row
            for row in ra.rows:
                lik = f.get(f'lik_{row.seq_num}')
                sev = f.get(f'sev_{row.seq_num}')
                if lik and sev:
                    ri = f'{lik}{sev}'
                    row.likelihood_initial     = int(lik)
                    row.severity_initial       = sev
                    row.risk_index_initial     = ri
                    row.risk_tolerance_initial = get_tolerance(ri)
                    # Update linked Risk record
                    if row.risk_id:
                        rsk = Risk.query.get(row.risk_id)
                        if rsk:
                            rsk.initial_likelihood = int(lik)
                            rsk.initial_severity   = sev
                            rsk.initial_risk_index = ri
                            rsk.initial_tolerance  = get_tolerance(ri)
                row.current_defenses = f.get(f'def_{row.seq_num}', row.current_defenses)
            # Update RA summary
            worst_i, _ = compute_ra_summary(ra)
            if worst_i:
                ra.risk_level_prior = get_tolerance(worst_i)
            db.session.commit()

        elif step == 4:
            # Save checklist responses + compile into current_defenses text
            for row in ra.rows:
                # Delete existing checklist for this row
                RAChecklistItem.query.filter_by(
                    assessment_id=ra.id, row_seq=row.seq_num).delete()
                checked_labels = []
                for idx, (cat, desc) in enumerate(CONTROL_CHECKLIST):
                    key        = f'ctrl_{row.seq_num}_{idx}'
                    notes_k    = f'notes_{row.seq_num}_{idx}'
                    is_checked = key in f
                    notes_val  = f.get(notes_k, '')
                    item = RAChecklistItem(
                        assessment_id=ra.id,
                        row_seq=row.seq_num,
                        category=cat,
                        description=desc,
                        checked=is_checked,
                        notes=notes_val
                    )
                    db.session.add(item)
                    if is_checked:
                        label = desc
                        if notes_val:
                            label += f' ({notes_val})'
                        checked_labels.append(f'[{cat}] {label}')
                # Write compiled controls back to the RARow.current_defenses column
                row.current_defenses = '; '.join(checked_labels) if checked_labels else ''
                db.session.add(row)
            db.session.commit()

        elif step == 5:
            # Save further mitigations + auto-create actions
            for row in ra.rows:
                mit_text = f.get(f'mitigation_{row.seq_num}','')
                resp_mgr = f.get(f'manager_{row.seq_num}','')
                due_dt   = f.get(f'due_{row.seq_num}','')
                if mit_text:
                    row.further_mitigations = mit_text
                    # Check if mitigation already exists
                    existing = RAMitigation.query.filter_by(
                        assessment_id=ra.id, hazard_seq=str(row.seq_num)).first()
                    if not existing:
                        act_id = new_id('ACT')
                        mit = RAMitigation(
                            assessment_id=ra.id,
                            hazard_seq=str(row.seq_num),
                            mitigation=mit_text,
                            responsible_manager=resp_mgr,
                            due_date=due_dt,
                            action_id=act_id, status='Open'
                        )
                        db.session.add(mit)
                        action = Action(
                            id=act_id, source='Risk Assessment',
                            hazard_id=ra.hazard_id, linked_ref_id=ra.id,
                            description=f'[{ra.control_number}] Seq {row.seq_num}: {mit_text}',
                            owner=resp_mgr, due_date=due_dt,
                            priority='High' if row.risk_tolerance_initial=='INTOLERABLE' else 'Medium',
                            status='Open'
                        )
                        db.session.add(action)
            db.session.commit()

        elif step == 6:
            # Save residual risk per row — final step
            for row in ra.rows:
                lik_r = f.get(f'res_lik_{row.seq_num}')
                sev_r = f.get(f'res_sev_{row.seq_num}')
                if lik_r and sev_r:
                    ri_r = f'{lik_r}{sev_r}'
                    row.likelihood_residual    = int(lik_r)
                    row.severity_residual      = sev_r
                    row.risk_index_residual    = ri_r
                    row.risk_tolerance_residual = get_tolerance(ri_r)
                    if row.risk_id:
                        rsk = Risk.query.get(row.risk_id)
                        if rsk:
                            rsk.residual_likelihood = int(lik_r)
                            rsk.residual_severity   = sev_r
                            rsk.residual_risk_index = ri_r
                            rsk.residual_tolerance  = get_tolerance(ri_r)
            # Finalise assessment
            _, worst_r = compute_ra_summary(ra)
            if worst_r:
                ra.risk_level_after = get_tolerance(worst_r)
            ra.status = 'Under Review'
            # Update hazard status
            if ra.hazard_id:
                h = Hazard.query.get(ra.hazard_id)
                if h:
                    h.status = 'Under Assessment'
            ra.management_acceptance = f.get('acceptance','')
            ra.prepared_by_name      = f.get('prepared_by','')
            ra.prepared_by_position  = f.get('prepared_position','')
            db.session.commit()

            flash(f'✓ Risk Assessment {ra.control_number} completed. Review and approve below.', 'success')
            return redirect(url_for('ra_detail', ra_id=ra.id))

        # Advance to next step
        if step < 6:
            return redirect(url_for('ra_wizard_step', hid=hid, step=step+1))

    # ── GET: render current step ──────────────────────────────────────────────
    checklist_items = {}
    if step == 4:
        for row in ra.rows:
            items = RAChecklistItem.query.filter_by(
                assessment_id=ra.id, row_seq=row.seq_num).all()
            checklist_items[row.seq_num] = {
                item.description: item for item in items
            }

    # Compute progress
    completed_steps = 0
    if ra.responsible_name:                    completed_steps = max(completed_steps, 1)
    if ra.rows:                                completed_steps = max(completed_steps, 2)
    if ra.rows and ra.rows[0].risk_index_initial: completed_steps = max(completed_steps, 3)
    if RAChecklistItem.query.filter_by(assessment_id=ra.id).first(): completed_steps = max(completed_steps, 4)
    if ra.mitigations:                         completed_steps = max(completed_steps, 5)
    if ra.rows and ra.rows[0].risk_index_residual: completed_steps = max(completed_steps, 6)

    return render_template('risk/ra_wizard.html',
        hazard=hazard, ra=ra, step=step,
        steps=WIZARD_STEPS,
        completed_steps=completed_steps,
        checklist=CONTROL_CHECKLIST,
        checklist_items=checklist_items,
        get_tolerance=get_tolerance)

# ─── RESUME wizard from hazard log ───────────────────────────────────────────


# ═══════════════════════════════════════════════════════════════════════════════
#  PRINT & EXPORT MODULE
#  1. Risk Assessment → Print-ready HTML (browser prints as PDF)
#  2. Hazard Log      → Excel (.xlsx) download
#  3. Hazard Report   → Print-ready HTML (browser prints as PDF)
# ═══════════════════════════════════════════════════════════════════════════════

# ─── HELPER: Excel styling ────────────────────────────────────────────────────
NAVY   = "0F1C3F"
GOLD   = "C9A84C"
WHITE  = "FFFFFF"
RED    = "DC2626"
YELLOW = "D97706"
GREEN  = "15803D"
GRAY   = "F4F6FB"
LGRAY  = "E5E7EB"

def hdr_font(bold=True, color=WHITE, size=11):
    return Font(bold=bold, color=color, size=size, name='Calibri')

def cell_font(bold=False, color="111827", size=10):
    return Font(bold=bold, color=color, size=size, name='Calibri')

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border():
    s = Side(style='thin', color=LGRAY)
    return Border(left=s, right=s, top=s, bottom=s)

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def tol_color(tol):
    if tol == 'INTOLERABLE': return RED
    if tol == 'TOLERABLE':   return YELLOW
    return GREEN

# ─── 1. RISK ASSESSMENT PRINT ─────────────────────────────────────────────────
@app.route('/risk-assessments/<ra_id>/print')
@require_login
def ra_print(ra_id):
    """Returns a print-ready HTML page that users print as PDF from the browser."""
    ra = RiskAssessment.query.get_or_404(ra_id)
    return render_template('risk/ra_print.html', ra=ra, get_tolerance=get_tolerance,
                           now=datetime.utcnow())

# ─── 2. HAZARD LOG → EXCEL ────────────────────────────────────────────────────
@app.route('/hazard-log/export-excel')
@require_login
def hazard_log_excel():
    dept_f = request.args.get('dept','')
    stat_f = request.args.get('status','')
    cls_f  = request.args.get('classification','')

    q = Hazard.query
    if dept_f: q = q.filter_by(department_id=int(dept_f))
    if stat_f: q = q.filter_by(status=stat_f)
    if cls_f:  q = q.filter_by(classification=cls_f)
    hazards = q.order_by(Hazard.created_at.desc()).all()

    wb = Workbook()

    # ── Sheet 1: Hazard Log ───────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Hazard Log"

    # Title row
    ws1.merge_cells('A1:L1')
    ws1['A1'] = 'AviaS — HAZARD LOG'
    ws1['A1'].font = Font(bold=True, size=14, color=WHITE, name='Calibri')
    ws1['A1'].fill = fill(NAVY)
    ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 28

    ws1.merge_cells('A2:L2')
    ws1['A2'] = f'Generated: {datetime.now().strftime("%d %b %Y %H:%M")} | Total Hazards: {len(hazards)} | Ref: AviaS/SMS/001'
    ws1['A2'].font = Font(size=9, color="6B7280", name='Calibri')
    ws1['A2'].fill = fill(GRAY)
    ws1['A2'].alignment = Alignment(horizontal='center')

    # Headers
    headers = ['Hazard ID','Source','Department','Classification','Generic Hazard',
               'Specific Components','Consequences','Initial Risk','Tolerance',
               'Residual Risk','Res. Tolerance','Status']
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=3, column=col, value=h)
        cell.font  = hdr_font()
        cell.fill  = fill(NAVY)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border()
    ws1.row_dimensions[3].height = 20

    # Data rows
    for row_num, haz in enumerate(hazards, 4):
        dept  = haz.department.name if haz.department else '—'
        risk  = haz.risks[0] if haz.risks else None
        ri_i  = risk.initial_risk_index  if risk else '—'
        tol_i = risk.initial_tolerance   if risk else '—'
        ri_r  = risk.residual_risk_index if risk else '—'
        tol_r = risk.residual_tolerance  if risk else '—'

        row_data = [
            haz.id, haz.source, dept, haz.classification or '—',
            haz.generic_hazard or '—', haz.specific_components or '—',
            haz.consequences or '—', ri_i, tol_i,
            ri_r or 'Pending', tol_r or 'Pending', haz.status
        ]
        bg = "FFFFFF" if row_num % 2 == 0 else GRAY
        for col, val in enumerate(row_data, 1):
            cell = ws1.cell(row=row_num, column=col, value=val)
            cell.font   = cell_font()
            cell.fill   = fill(bg)
            cell.border = border()
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            # Color code tolerance columns
            if col == 9 and val in ('INTOLERABLE','TOLERABLE','ACCEPTABLE'):
                cell.font = Font(bold=True, color=tol_color(val), size=10, name='Calibri')
            if col == 11 and val in ('INTOLERABLE','TOLERABLE','ACCEPTABLE'):
                cell.font = Font(bold=True, color=tol_color(val), size=10, name='Calibri')
            if col == 12:  # Status
                if val == 'Open':   cell.font = Font(bold=True, color=YELLOW, size=10, name='Calibri')
                elif val == 'Closed': cell.font = Font(bold=True, color=GREEN, size=10, name='Calibri')
        ws1.row_dimensions[row_num].height = 32

    set_col_widths(ws1, [18, 14, 18, 16, 24, 32, 28, 10, 14, 10, 14, 10])

    # Auto-filter
    ws1.auto_filter.ref = f'A3:L{3 + len(hazards)}'
    ws1.freeze_panes    = 'A4'

    # ── Sheet 2: Risk Register ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Risk Register")
    ws2.merge_cells('A1:K1')
    ws2['A1'] = 'AviaS — RISK REGISTER'
    ws2['A1'].font = Font(bold=True, size=14, color=WHITE, name='Calibri')
    ws2['A1'].fill = fill(NAVY)
    ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 28

    ws2.merge_cells('A2:K2')
    ws2['A2'] = f'Generated: {datetime.now().strftime("%d %b %Y %H:%M")} | ICAO Annex 19 / AviaS/SMS/001'
    ws2['A2'].font = Font(size=9, color="6B7280", name='Calibri')
    ws2['A2'].fill = fill(GRAY)
    ws2['A2'].alignment = Alignment(horizontal='center')

    rsk_headers = ['Risk ID','Hazard ID','Risk Description','Department',
                   'Initial L','Initial S','Initial Index','Initial Tolerance',
                   'Controls','Residual Index','Residual Tolerance']
    for col, h in enumerate(rsk_headers, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font  = hdr_font()
        cell.fill  = fill(NAVY)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border()
    ws2.row_dimensions[3].height = 20

    all_risks = Risk.query.join(Hazard, Risk.hazard_id == Hazard.id)
    if dept_f: all_risks = all_risks.filter(Hazard.department_id == int(dept_f))
    all_risks = all_risks.order_by(Risk.created_at.desc()).all()

    for row_num, rsk in enumerate(all_risks, 4):
        haz   = rsk.hazard
        dept  = haz.department.name if haz and haz.department else '—'
        ctrl_summary = '; '.join([c.description[:40] for c in rsk.controls[:3]]) if rsk.controls else 'None'
        bg = "FFFFFF" if row_num % 2 == 0 else GRAY
        row_data = [
            rsk.id, rsk.hazard_id, rsk.description or '—', dept,
            rsk.initial_likelihood or '—', rsk.initial_severity or '—',
            rsk.initial_risk_index or '—', rsk.initial_tolerance or '—',
            ctrl_summary,
            rsk.residual_risk_index or 'Pending',
            rsk.residual_tolerance or 'Pending'
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_num, column=col, value=val)
            cell.font   = cell_font()
            cell.fill   = fill(bg)
            cell.border = border()
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if col == 8 and val in ('INTOLERABLE','TOLERABLE','ACCEPTABLE'):
                cell.font = Font(bold=True, color=tol_color(val), size=10, name='Calibri')
            if col == 11 and val in ('INTOLERABLE','TOLERABLE','ACCEPTABLE'):
                cell.font = Font(bold=True, color=tol_color(val), size=10, name='Calibri')
        ws2.row_dimensions[row_num].height = 30

    set_col_widths(ws2, [18, 18, 32, 18, 8, 8, 12, 16, 36, 12, 16])
    ws2.auto_filter.ref = f'A3:K{3 + len(all_risks)}'
    ws2.freeze_panes    = 'A4'

    # ── Sheet 3: Actions ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Open Actions")
    ws3.merge_cells('A1:H1')
    ws3['A1'] = 'AviaS — OPEN ACTIONS'
    ws3['A1'].font = Font(bold=True, size=14, color=WHITE, name='Calibri')
    ws3['A1'].fill = fill(NAVY)
    ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws3.row_dimensions[1].height = 28

    act_headers = ['Action ID','Source','Linked Hazard','Description','Owner','Due Date','Priority','Status']
    for col, h in enumerate(act_headers, 1):
        cell = ws3.cell(row=2, column=col, value=h)
        cell.font  = hdr_font()
        cell.fill  = fill(NAVY)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border()

    open_actions = Action.query.filter(Action.status != 'Closed').order_by(Action.due_date).all()
    for row_num, act in enumerate(open_actions, 3):
        bg = "FFFFFF" if row_num % 2 == 0 else GRAY
        row_data = [
            act.id, act.source, act.hazard_id or '—',
            (act.description or '—')[:80], act.owner or '—',
            act.due_date or '—', act.priority or '—', act.status
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws3.cell(row=row_num, column=col, value=val)
            cell.font   = cell_font()
            cell.fill   = fill(bg)
            cell.border = border()
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if col == 8:
                color_map = {'Overdue': RED, 'Open': YELLOW, 'In Progress': '1D4ED8'}
                c = color_map.get(val, "111827")
                cell.font = Font(bold=True, color=c, size=10, name='Calibri')
        ws3.row_dimensions[row_num].height = 24

    set_col_widths(ws3, [18, 16, 18, 40, 20, 12, 10, 12])
    ws3.freeze_panes = 'A3'

    # Save to buffer and send
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'JAV_Hazard_Log_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    return send_file(buf, as_attachment=True,
                     download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ─── 3. HAZARD REPORT PRINT ───────────────────────────────────────────────────
@app.route('/hazard-reports/<rep_id>/print')
@require_login
def hazard_report_print(rep_id):
    rep    = HazardReport.query.get_or_404(rep_id)
    hazard = Hazard.query.get(rep.hazard_id) if rep.hazard_id else None
    ra     = RiskAssessment.query.filter_by(hazard_id=rep.hazard_id).first() if rep.hazard_id else None
    return render_template('reporting/hazard_report_print.html', rep=rep, hazard=hazard, ra=ra,
                           get_tolerance=get_tolerance)

# ─── 4. ASR PRINT ─────────────────────────────────────────────────────────────
@app.route('/asr/<asr_id>/print')
@require_login
def asr_print(asr_id):
    asr = ASRReport.query.get_or_404(asr_id)
    return render_template('reporting/asr_print.html', asr=asr)

# ─── 5. VOLUNTARY REPORT PRINT ────────────────────────────────────────────────
@app.route('/admin/reports/voluntary/<int:rid>/print')
@require_login
def voluntary_report_print(rid):
    rpt = VoluntaryReport.query.get_or_404(rid)
    return render_template('reporting/voluntary_print.html', rpt=rpt)

# ─── 6. CONFIDENTIAL REPORT PRINT ─────────────────────────────────────────────
@app.route('/admin/reports/confidential/<int:rid>/print')
@require_login
def confidential_report_print(rid):
    rpt = ConfidentialReport.query.get_or_404(rid)
    return render_template('reporting/confidential_print.html', rpt=rpt)

# ─── 7. INVESTIGATION PRINT ───────────────────────────────────────────────────
@app.route('/investigations/<iid>/print')
@require_login
def investigation_print(iid):
    inv = Investigation.query.get_or_404(iid)
    return render_template('investigation/investigation_print.html', inv=inv)


# ═══════════════════════════════════════════════════════════════════════════════
#  RISK ASSESSMENT LIFECYCLE — STATUS ENGINE
#  Draft → Submitted → Active → Under Review → Closed
#  ICAO Annex 19 §5 / Doc 9859 Ch.5 — added without changing existing routes
# ═══════════════════════════════════════════════════════════════════════════════

def ra_can_edit(ra):
    """RA is editable in all states except Closed."""
    return ra.status != 'Closed'

def ra_closure_checks(ra):
    """
    Returns list of blocking reasons. Empty list = can close.
    Rules (ICAO): residual risk assessed + all linked actions closed + effectiveness verified.
    """
    blocks = []
    # 1. Every row must have residual risk assessed
    rows_no_residual = [r for r in ra.rows if not r.risk_index_residual]
    if rows_no_residual:
        blocks.append(f'{len(rows_no_residual)} risk row(s) still have no residual risk assessment.')

    # 2. Residual risk must not be INTOLERABLE
    rows_intol = [r for r in ra.rows if r.risk_tolerance_residual == 'INTOLERABLE']
    if rows_intol:
        blocks.append(f'{len(rows_intol)} risk row(s) have INTOLERABLE residual risk — must be reduced before closing.')

    # 3. All linked actions must be Closed
    if ra.hazard:
        open_actions = [a for a in ra.hazard.actions if a.status not in ('Closed',)]
        if open_actions:
            blocks.append(f'{len(open_actions)} linked action(s) are still open or in progress.')

    # 4. At least one review recorded with effectiveness
    if ra.reviews:
        unverified = [r for r in ra.reviews if not r.effectiveness_rating]
        if unverified:
            blocks.append(f'{len(unverified)} review(s) have no effectiveness rating.')

    return blocks

def check_ra_review_dates():
    """Auto-set status = Under Review when next_review_date is reached."""
    today = date.today().isoformat()
    due = RiskAssessment.query.filter(
        RiskAssessment.status == 'Active',
        RiskAssessment.next_review_date != None,
        RiskAssessment.next_review_date <= today
    ).all()
    for ra in due:
        ra.status = 'Under Review'
    if due:
        db.session.commit()

# ─── SUBMIT (Draft → Submitted → Active) ────────────────────────────────────
@app.route('/risk-assessments/<ra_id>/submit', methods=['POST'])
@require_login
def ra_submit(ra_id):
    ra = RiskAssessment.query.get_or_404(ra_id)
    if ra.status != 'Draft':
        flash(f'Cannot submit — current status is {ra.status}.', 'error')
        return redirect(url_for('ra_detail', ra_id=ra_id))
    if not ra.rows:
        flash('Cannot submit — no risk rows added. Complete Steps 2–3 first.', 'error')
        return redirect(url_for('ra_detail', ra_id=ra_id))
    today = date.today().isoformat()
    ra.status          = 'Active'
    ra.submitted_date  = today
    ra.activated_date  = today
    # Update linked hazard status
    if ra.hazard:
        ra.hazard.status = 'Under Assessment'
    db.session.commit()
    flash(f'✓ Risk Assessment {ra.control_number} submitted and is now Active.', 'success')
    return redirect(url_for('ra_detail', ra_id=ra_id))

# ─── SEND FOR REVIEW (Active → Under Review) ─────────────────────────────────
@app.route('/risk-assessments/<ra_id>/send-review', methods=['POST'])
@require_login
def ra_send_review(ra_id):
    ra = RiskAssessment.query.get_or_404(ra_id)
    if ra.status not in ('Active', 'Submitted'):
        flash(f'Cannot send for review — status is {ra.status}.', 'error')
        return redirect(url_for('ra_detail', ra_id=ra_id))
    ra.status = 'Under Review'
    db.session.commit()
    flash('✓ Risk Assessment sent for review.', 'success')
    return redirect(url_for('ra_detail', ra_id=ra_id))

# ─── REACTIVATE (Under Review → Active) ──────────────────────────────────────
@app.route('/risk-assessments/<ra_id>/reactivate', methods=['POST'])
@require_login
def ra_reactivate(ra_id):
    ra = RiskAssessment.query.get_or_404(ra_id)
    if ra.status != 'Under Review':
        flash(f'Can only reactivate from Under Review status.', 'error')
        return redirect(url_for('ra_detail', ra_id=ra_id))
    ra.status = 'Active'
    db.session.commit()
    flash('✓ Risk Assessment reactivated to Active.', 'success')
    return redirect(url_for('ra_detail', ra_id=ra_id))

# ─── CLOSE (validated) ────────────────────────────────────────────────────────
@app.route('/risk-assessments/<ra_id>/close', methods=['POST'])
@require_login
def ra_close(ra_id):
    ra = RiskAssessment.query.get_or_404(ra_id)
    if ra.status == 'Closed':
        flash('Already closed.', 'error')
        return redirect(url_for('ra_detail', ra_id=ra_id))
    blocks = ra_closure_checks(ra)
    if blocks:
        for b in blocks:
            flash(f'✗ {b}', 'error')
        return redirect(url_for('ra_detail', ra_id=ra_id))
    ra.status       = 'Closed'
    ra.closed_date  = date.today().isoformat()
    ra.management_acceptance = request.form.get('management_acceptance', ra.management_acceptance)
    ra.acceptance_date       = date.today().isoformat()
    # Update hazard status
    if ra.hazard:
        ra.hazard.status = 'Controlled'
    db.session.commit()
    flash(f'✓ Risk Assessment {ra.control_number} closed successfully.', 'success')
    return redirect(url_for('ra_detail', ra_id=ra_id))

# ─── CLOSURE CHECK (AJAX / page query) ───────────────────────────────────────
@app.route('/risk-assessments/<ra_id>/closure-check')
@require_login
def ra_closure_check(ra_id):
    ra = RiskAssessment.query.get_or_404(ra_id)
    blocks = ra_closure_checks(ra)
    return render_template('risk/ra_closure_check.html', ra=ra, blocks=blocks)

# ─── REASSESS (create new revision) ──────────────────────────────────────────
@app.route('/risk-assessments/<ra_id>/reassess', methods=['POST'])
@require_login
def ra_reassess(ra_id):
    old_ra = RiskAssessment.query.get_or_404(ra_id)
    if old_ra.status == 'Closed':
        flash('Cannot reassess a closed Risk Assessment — create a new one.', 'error')
        return redirect(url_for('ra_detail', ra_id=ra_id))

    # Archive the current one
    old_ra.status = 'Archived'
    new_rev = (old_ra.revision or 0) + 1
    new_ctrl = f"{old_ra.control_number}-REV{new_rev}" if old_ra.control_number else gen_control_number(
        old_ra.department.code if old_ra.department else 'XX')

    # Create new revision — copy header, fresh status
    new_ra = RiskAssessment(
        id=new_id('RA'),
        control_number=new_ctrl,
        responsible_name=old_ra.responsible_name,
        assessors_names=old_ra.assessors_names,
        assessment_date=date.today().isoformat(),
        next_review_date=old_ra.next_review_date,
        title=old_ra.title,
        hazard_id=old_ra.hazard_id,
        department_id=old_ra.department_id,
        general_description=old_ra.general_description,
        reasons=f'Reassessment of {old_ra.control_number}',
        risk_level_prior=old_ra.risk_level_prior,
        prepared_by_name=old_ra.prepared_by_name,
        prepared_by_position=old_ra.prepared_by_position,
        reviewed_by_name=old_ra.reviewed_by_name,
        reviewed_by_position=old_ra.reviewed_by_position,
        approved_by_name=old_ra.approved_by_name,
        approved_by_position=old_ra.approved_by_position,
        status='Draft',
        revision=new_rev,
        parent_ra_id=old_ra.id
    )
    db.session.add(new_ra)
    db.session.flush()

    # Copy rows from old RA into new RA (user will update them)
    for old_row in old_ra.rows:
        new_row = RARow(
            assessment_id=new_ra.id,
            seq_num=old_row.seq_num,
            type_of_activity=old_row.type_of_activity,
            generic_hazard=old_row.generic_hazard,
            specific_components=old_row.specific_components,
            consequences=old_row.consequences,
            likelihood_initial=old_row.likelihood_initial,
            severity_initial=old_row.severity_initial,
            risk_index_initial=old_row.risk_index_initial,
            risk_tolerance_initial=old_row.risk_tolerance_initial,
            current_defenses=old_row.current_defenses,
            further_mitigations=old_row.further_mitigations,
            # Residual cleared — user must re-evaluate
            likelihood_residual=None,
            severity_residual=None,
            risk_index_residual=None,
            risk_tolerance_residual=None,
        )
        db.session.add(new_row)

    # Move hazard_id from old RA to new RA (unique constraint requires clearing old first)
    haz_id = old_ra.hazard_id
    old_ra.hazard_id = None
    db.session.flush()
    new_ra.hazard_id = haz_id

    db.session.commit()
    flash(f'✓ New revision {new_ctrl} (REV{new_rev}) created. Please update controls and residual risk.', 'success')
    return redirect(url_for('ra_wizard_step', hid=haz_id, step=4))


# ═══════════════════════════════════════════════════════════════════════════════
#  ENTERPRISE PDF EXPORT ROUTES
#  ICAO Annex 19 / IOSA ISM — Controlled Document Generation
#  All routes require login.  PDF bytes served inline (opens in browser tab).
#  Powered by reports.py (ReportLab — no system dependencies).
# ═══════════════════════════════════════════════════════════════════════════════

def _pdf_unavailable():
    """Return a plain-text 503 when the PDF engine failed to import."""
    return Response(
        'PDF engine unavailable. Ensure reportlab is installed (pip install reportlab).',
        status=503, mimetype='text/plain'
    )

def _pdf_response(pdf_bytes, filename):
    """Wrap raw PDF bytes in a Flask Response that opens inline in the browser."""
    return Response(
        pdf_bytes,
        mimetype='application/pdf',
        headers={'Content-Disposition': f'inline; filename="{filename}"'}
    )

def _gen_by():
    """Current logged-in user's display name for 'Generated by' field."""
    return session.get('admin_name', session.get('admin_username', 'Safety Department'))


# ── 1. HAZARD REPORT PDF ────────────────────────────────────────────────────
@app.route('/pdf/hazard-report/<rid>')
@require_login
def pdf_route_hazard_report(rid):
    if not _PDF_ENGINE:
        return _pdf_unavailable()
    hr         = HazardReport.query.get_or_404(rid)
    hazard     = Hazard.query.get(hr.hazard_id) if hr.hazard_id else None
    actions    = Action.query.filter_by(hazard_id=hr.hazard_id).all() if hr.hazard_id else []
    # ActionHistory is per-action (no hazard_id col) — collect via action IDs
    _aids      = [a.id for a in actions]
    history    = (ActionHistory.query
                  .filter(ActionHistory.action_id.in_(_aids))
                  .order_by(ActionHistory.changed_at.asc()).all()) if _aids else []
    risks      = Risk.query.filter_by(hazard_id=hr.hazard_id).all() if hr.hazard_id else []
    inv        = Investigation.query.filter_by(hazard_id=hr.hazard_id).first() if hr.hazard_id else None
    ra         = RiskAssessment.query.filter_by(hazard_id=hr.hazard_id).first() if hr.hazard_id else None
    pdf_bytes  = pdf_hazard_report(hr, hazard, actions, history, risks, inv, ra, generated_by=_gen_by())
    return _pdf_response(pdf_bytes, f'HR-{rid}.pdf')


# ── 2. AIR SAFETY REPORT PDF ────────────────────────────────────────────────
@app.route('/pdf/asr/<asr_id>')
@require_login
def pdf_route_asr(asr_id):
    if not _PDF_ENGINE:
        return _pdf_unavailable()
    asr       = ASRReport.query.get_or_404(asr_id)
    # ASRReport.hazard_id links directly to Hazard (no separate report_id field)
    hazard    = Hazard.query.get(asr.hazard_id) if asr.hazard_id else None
    hr        = HazardReport.query.filter_by(hazard_id=asr.hazard_id).first() if asr.hazard_id else None
    actions   = Action.query.filter_by(hazard_id=asr.hazard_id).all() if asr.hazard_id else []
    pdf_bytes = pdf_asr_report(asr, hazard, hr, actions, generated_by=_gen_by())
    return _pdf_response(pdf_bytes, f'ASR-{asr_id}.pdf')


# ── 3. INVESTIGATION PDF ────────────────────────────────────────────────────
@app.route('/pdf/investigation/<inv_id>')
@require_login
def pdf_route_investigation(inv_id):
    if not _PDF_ENGINE:
        return _pdf_unavailable()
    inv       = Investigation.query.get_or_404(inv_id)
    hazard    = Hazard.query.get(inv.hazard_id) if inv.hazard_id else None
    actions   = Action.query.filter_by(hazard_id=inv.hazard_id).all() if inv.hazard_id else []
    pdf_bytes = pdf_investigation(inv, hazard, actions, generated_by=_gen_by())
    return _pdf_response(pdf_bytes, f'INV-{inv_id}.pdf')


# ── 4. RISK ASSESSMENT PDF ──────────────────────────────────────────
@app.route('/pdf/risk-assessment/<ra_id>')
@require_login
def pdf_route_risk_assessment(ra_id):
    if not _PDF_ENGINE:
        return _pdf_unavailable()
    ra         = RiskAssessment.query.get_or_404(ra_id)
    hazard     = Hazard.query.get(ra.hazard_id) if ra.hazard_id else None
    rows       = RARow.query.filter_by(assessment_id=ra_id).order_by(RARow.seq_num).all()
    mitigations= RAMitigation.query.filter_by(assessment_id=ra_id).all()
    reviews    = RAReview.query.filter_by(assessment_id=ra_id).order_by(RAReview.review_date).all()
    pdf_bytes  = pdf_risk_assessment(ra, hazard, rows, mitigations, reviews, generated_by=_gen_by())
    ctrl = getattr(ra, 'control_number', ra_id) or ra_id
    return _pdf_response(pdf_bytes, f'RA-{ctrl}.pdf')


# ── 5. ACTION / CORRECTIVE ACTION PDF ───────────────────────────────────────
@app.route('/pdf/action/<aid>')
@require_login
def pdf_route_action(aid):
    if not _PDF_ENGINE:
        return _pdf_unavailable()
    action    = Action.query.get_or_404(aid)
    history   = ActionHistory.query.filter_by(action_id=aid).order_by(
                    ActionHistory.changed_at.asc()).all()
    pdf_bytes = pdf_action(action, history, generated_by=_gen_by())
    return _pdf_response(pdf_bytes, f'ACT-{aid}.pdf')


# ── 6. MANAGEMENT OF CHANGE PDF ─────────────────────────────────────────────
@app.route('/pdf/moc/<mid>')
@require_login
def pdf_route_moc(mid):
    if not _PDF_ENGINE:
        return _pdf_unavailable()
    moc       = MOC.query.get_or_404(mid)
    pdf_bytes = pdf_moc(moc, generated_by=_gen_by())
    return _pdf_response(pdf_bytes, f'MOC-{mid}.pdf')


# ── 7. AUDIT PDF ────────────────────────────────────────────────────────────
@app.route('/pdf/audit/<sid>')
@require_login
def pdf_route_audit(sid):
    if not _PDF_ENGINE:
        return _pdf_unavailable()
    schedule  = AuditSchedule.query.get_or_404(sid)
    findings  = AuditFinding.query.filter_by(schedule_id=sid).all()
    checklist = AuditChecklist.query.filter_by(schedule_id=sid).all()
    pdf_bytes = pdf_audit(schedule, findings, checklist, generated_by=_gen_by())
    return _pdf_response(pdf_bytes, f'AUDIT-{sid}.pdf')


# ── 8. EMERGENCY RESPONSE PLAN PDF ────────────────────────────────────────────
@app.route('/pdf/erp/<eid>')
@require_login
def pdf_route_erp(eid):
    if not _PDF_ENGINE:
        return _pdf_unavailable()
    erp       = ERPlan.query.get_or_404(eid)
    pdf_bytes = pdf_erp(erp, generated_by=_gen_by())
    return _pdf_response(pdf_bytes, f'ERP-{eid}.pdf')


# ── 9. VOLUNTARY REPORT PDF ───────────────────────────────────────────────────
@app.route('/pdf/voluntary/<vid>')
@require_login
def pdf_route_voluntary(vid):
    if not _PDF_ENGINE:
        return _pdf_unavailable()
    report    = VoluntaryReport.query.get_or_404(vid)
    pdf_bytes = pdf_voluntary(report, generated_by=_gen_by())
    return _pdf_response(pdf_bytes, f'VOL-{vid}.pdf')


# ── 10. CONFIDENTIAL REPORT PDF ───────────────────────────────────────────────────
@app.route('/pdf/confidential/<cid>')
@require_login
def pdf_route_confidential(cid):
    if not _PDF_ENGINE:
        return _pdf_unavailable()
    # Confidential reports: admin-only access
    if not session.get('admin_logged_in'):
        flash('Confidential reports require administrator access.', 'error')
        return redirect(url_for('index'))
    report    = ConfidentialReport.query.get_or_404(cid)
    pdf_bytes = pdf_confidential(report, generated_by=_gen_by())
    return _pdf_response(pdf_bytes, f'CONF-{cid}.pdf')


# ── 11. TRAINING RECORD PDF ───────────────────────────────────────────────────────
@app.route('/pdf/training/<tid>')
@require_login
def pdf_route_training(tid):
    if not _PDF_ENGINE:
        return _pdf_unavailable()
    training  = Training.query.get_or_404(tid)
    pdf_bytes = pdf_training(training, generated_by=_gen_by())
    return _pdf_response(pdf_bytes, f'TRN-{tid}.pdf')


# ── 12. AUDIT FINDING / CAP PDF ───────────────────────────────────────────────────
@app.route('/pdf/audit-finding/<fid>')
@require_login
def pdf_route_audit_finding(fid):
    if not _PDF_ENGINE:
        return _pdf_unavailable()
    finding   = AuditFinding.query.get_or_404(fid)
    audit     = AuditSchedule.query.get(finding.schedule_id) if finding.schedule_id else None
    pdf_bytes = pdf_audit_finding(finding, audit, generated_by=_gen_by())
    return _pdf_response(pdf_bytes, f'CAP-{fid}.pdf')


# ── 13. SPI DASHBOARD SUMMARY PDF ────────────────────────────────────────────────────────
@app.route('/pdf/spi-summary')
@require_login
def pdf_route_spi_summary():
    if not _PDF_ENGINE:
        return _pdf_unavailable()
    indicators  = SPIIndicator.query.filter_by(active=True).all()
    pdf_bytes   = pdf_spi_summary(indicators, generated_by=_gen_by())
    return _pdf_response(pdf_bytes, 'SPI-Summary.pdf')


# ─────────────────────────────────────────────────────────────────────────────────
#  APPLICATION STARTUP — migrations + seed
# ─────────────────────────────────────────────────────────────────────────────────
with app.app_context():
    db.create_all()

    _migrations = {
        'accountable_executives': [
            ('user_id', 'INTEGER'),
            ('full_name', 'VARCHAR(120)'),
            ('title', 'VARCHAR(100)'),
            ('email', 'VARCHAR(120)'),
            ('phone', 'VARCHAR(50)'),
            ('employee_number', 'VARCHAR(30)'),
            ('authority_scope', 'TEXT'),
            ('appointment_ref', 'VARCHAR(50)'),
            ('effective_from', 'VARCHAR(20)'),
            ('effective_to', 'VARCHAR(20)'),
            ('is_current', 'BOOLEAN DEFAULT FALSE'),
            ('appointment_doc', 'VARCHAR(200)'),
            ('notes', 'TEXT'),
            ('created_by', 'VARCHAR(100)'),
            ('created_at', 'TIMESTAMP'),
        ],
        'action_history': [
            ('action_id', 'VARCHAR(30)'),
            ('changed_by', 'VARCHAR(100)'),
            ('changed_at', 'TIMESTAMP'),
            ('from_status', 'VARCHAR(50)'),
            ('to_status', 'VARCHAR(50)'),
            ('notes', 'TEXT'),
            ('field_changed', 'VARCHAR(50)'),
        ],
        'actions': [
            ('source', 'VARCHAR(40)'),
            ('hazard_id', 'VARCHAR(30)'),
            ('linked_ref_id', 'VARCHAR(30)'),
            ('description', 'TEXT'),
            ('owner', 'VARCHAR(100)'),
            ('due_date', 'VARCHAR(20)'),
            ('priority', 'VARCHAR(20)'),
            ('status', 'VARCHAR(50)'),
            ('effectiveness', 'VARCHAR(30)'),
            ('effectiveness_review', 'TEXT'),
            ('closed_date', 'VARCHAR(20)'),
            ('closure_evidence_required', 'BOOLEAN DEFAULT FALSE'),
            ('auto_reopened_at', 'TIMESTAMP'),
            ('auto_reopened_reason', 'TEXT'),
            ('spi_id', 'INTEGER'),
            ('spi_alert_level', 'VARCHAR(5)'),
            ('spi_trigger_rule', 'VARCHAR(2)'),
            ('spi_alert_month', 'INTEGER'),
            ('spi_alert_year', 'INTEGER'),
            ('spi_escalation_id', 'INTEGER'),
            ('evidence', 'TEXT'),
            ('evidence_filename', 'VARCHAR(200)'),
            ('mitigation_description', 'TEXT'),
            ('corrective_description', 'TEXT'),
            ('safety_notes', 'TEXT'),
            ('follow_up_notes', 'TEXT'),
            ('mitigation_status', 'VARCHAR(30)'),
            ('safety_review_notes', 'TEXT'),
            ('safety_reviewer', 'VARCHAR(100)'),
            ('safety_review_date', 'VARCHAR(20)'),
            ('implementation_date', 'VARCHAR(20)'),
            ('assigned_by', 'VARCHAR(100)'),
            ('closure_by', 'VARCHAR(100)'),
            ('verified_by', 'VARCHAR(100)'),
            ('verified_date', 'VARCHAR(20)'),
            ('created_at', 'TIMESTAMP'),
            ('sag_member', 'VARCHAR(100)'),
            ('department_id', 'INTEGER'),
            ('root_cause', 'TEXT'),
            ('rejection_notes', 'TEXT'),
            ('reopen_count', 'INTEGER'),
            ('action_type', 'VARCHAR(20)'),
            ('linked_audit_id', 'VARCHAR(30)'),
            ('linked_ra_id', 'VARCHAR(30)'),
            ('linked_risk_id', 'VARCHAR(30)'),
            ('safety_culture_notes', 'TEXT'),
        ],
        'api_tokens': [
            ('token', 'VARCHAR(64)'),
            ('user_id', 'VARCHAR(30)'),
            ('username', 'VARCHAR(80)'),
            ('expires_at', 'TIMESTAMP'),
            ('created_at', 'TIMESTAMP'),
        ],
        'asr_reports': [
            ('report_type', 'VARCHAR(20)'),
            ('occurrence_type', 'VARCHAR(50)'),
            ('captain', 'VARCHAR(100)'),
            ('captain_staff_no', 'VARCHAR(20)'),
            ('copilot', 'VARCHAR(100)'),
            ('copilot_staff_no', 'VARCHAR(20)'),
            ('date', 'VARCHAR(20)'),
            ('time_local', 'VARCHAR(10)'),
            ('time_utc', 'VARCHAR(10)'),
            ('flight_no', 'VARCHAR(20)'),
            ('route_from', 'VARCHAR(10)'),
            ('route_to', 'VARCHAR(10)'),
            ('diverted_to', 'VARCHAR(10)'),
            ('squawk', 'VARCHAR(10)'),
            ('aircraft_type', 'VARCHAR(30)'),
            ('registration', 'VARCHAR(20)'),
            ('pax', 'INTEGER'),
            ('crew', 'INTEGER'),
            ('altitude_ft', 'INTEGER'),
            ('flight_phase', 'VARCHAR(30)'),
            ('weather_wind', 'VARCHAR(20)'),
            ('weather_vis_rvr', 'VARCHAR(20)'),
            ('weather_clouds', 'VARCHAR(30)'),
            ('weather_temp_c', 'INTEGER'),
            ('weather_qnh', 'INTEGER'),
            ('runway', 'VARCHAR(10)'),
            ('runway_state', 'VARCHAR(20)'),
            ('event_description', 'TEXT'),
            ('action_taken', 'TEXT'),
            ('severity', 'VARCHAR(2)'),
            ('likelihood', 'INTEGER'),
            ('risk_index', 'VARCHAR(5)'),
            ('hazard_id', 'VARCHAR(30)'),
            ('created_at', 'TIMESTAMP'),
        ],
        'audit_actions': [
            ('finding_id', 'VARCHAR(30)'),
            ('hazard_id', 'VARCHAR(30)'),
            ('description', 'TEXT'),
            ('action_type', 'VARCHAR(30)'),
            ('owner', 'VARCHAR(100)'),
            ('due_date', 'VARCHAR(20)'),
            ('priority', 'VARCHAR(20)'),
            ('status', 'VARCHAR(20)'),
            ('implementation_notes', 'TEXT'),
            ('closed_date', 'VARCHAR(20)'),
            ('verified_by', 'VARCHAR(100)'),
            ('verification_date', 'VARCHAR(20)'),
            ('effectiveness', 'VARCHAR(30)'),
            ('effectiveness_notes', 'TEXT'),
            ('reopen_reason', 'TEXT'),
            ('created_at', 'TIMESTAMP'),
        ],
        'audit_checklists': [
            ('schedule_id', 'VARCHAR(30)'),
            ('category', 'VARCHAR(100)'),
            ('item_ref', 'VARCHAR(30)'),
            ('question', 'TEXT'),
            ('response', 'VARCHAR(10)'),
            ('comment', 'TEXT'),
            ('evidence', 'TEXT'),
            ('evidence_filename', 'VARCHAR(200)'),
            ('sequence', 'INTEGER'),
            ('linked_finding_id', 'VARCHAR(30)'),
        ],
        'audit_findings': [
            ('schedule_id', 'VARCHAR(30)'),
            ('finding_ref', 'VARCHAR(30)'),
            ('finding_title', 'VARCHAR(200)'),
            ('description', 'TEXT'),
            ('category', 'VARCHAR(50)'),
            ('severity', 'VARCHAR(20)'),
            ('standard_ref', 'VARCHAR(100)'),
            ('requirement', 'TEXT'),
            ('evidence', 'TEXT'),
            ('status', 'VARCHAR(40)'),
            ('assigned_to', 'VARCHAR(100)'),
            ('assigned_dept', 'VARCHAR(100)'),
            ('assigned_date', 'VARCHAR(20)'),
            ('root_cause', 'TEXT'),
            ('investigation_notes', 'TEXT'),
            ('contributing_factors', 'TEXT'),
            ('root_cause_submitted_at', 'TIMESTAMP'),
            ('immediate_action', 'TEXT'),
            ('longterm_action', 'TEXT'),
            ('cap_responsible', 'VARCHAR(100)'),
            ('cap_due_date', 'VARCHAR(20)'),
            ('cap_status', 'VARCHAR(30)'),
            ('cap_completion_pct', 'INTEGER'),
            ('cap_submitted_at', 'TIMESTAMP'),
            ('evidence_files', 'TEXT'),
            ('review_notes', 'TEXT'),
            ('reviewed_by', 'VARCHAR(100)'),
            ('review_date', 'VARCHAR(20)'),
            ('revision_reason', 'TEXT'),
            ('closure_verified_by', 'VARCHAR(100)'),
            ('closure_date', 'VARCHAR(20)'),
            ('closure_notes', 'TEXT'),
            ('sig_dept_manager', 'VARCHAR(100)'),
            ('sig_auditor', 'VARCHAR(100)'),
            ('sig_safety_manager', 'VARCHAR(100)'),
            ('sig_date', 'VARCHAR(20)'),
            ('hazard_id', 'VARCHAR(30)'),
            ('linked_action_id', 'VARCHAR(30)'),
            ('created_at', 'TIMESTAMP'),
        ],
        'audit_plans': [
            ('year', 'INTEGER'),
            ('month', 'INTEGER'),
            ('department_id', 'INTEGER'),
            ('audit_type', 'VARCHAR(50)'),
            ('frequency', 'VARCHAR(30)'),
            ('responsible_manager', 'VARCHAR(100)'),
            ('scope', 'TEXT'),
            ('objectives', 'TEXT'),
            ('iosa_reference', 'VARCHAR(100)'),
            ('auditor_name', 'VARCHAR(100)'),
            ('planned_week', 'INTEGER'),
            ('status', 'VARCHAR(20)'),
            ('created_at', 'TIMESTAMP'),
        ],
        'audit_schedules': [
            ('plan_id', 'VARCHAR(30)'),
            ('department_id', 'INTEGER'),
            ('audit_type', 'VARCHAR(50)'),
            ('scheduled_date', 'VARCHAR(20)'),
            ('actual_date', 'VARCHAR(20)'),
            ('lead_auditor', 'VARCHAR(100)'),
            ('audit_team', 'VARCHAR(200)'),
            ('scope', 'TEXT'),
            ('objectives', 'TEXT'),
            ('status', 'VARCHAR(20)'),
            ('opening_meeting', 'VARCHAR(20)'),
            ('closing_meeting', 'VARCHAR(20)'),
            ('summary', 'TEXT'),
            ('closure_date', 'VARCHAR(20)'),
            ('closed_by', 'VARCHAR(100)'),
            ('final_remarks', 'TEXT'),
            ('audit_result', 'VARCHAR(80)'),
            ('followup_required', 'VARCHAR(10)'),
            ('created_at', 'TIMESTAMP'),
        ],
        'audit_verification_items': [
            ('source_module', 'VARCHAR(50)'),
            ('source_record_id', 'VARCHAR(50)'),
            ('source_description', 'TEXT'),
            ('department_id', 'INTEGER'),
            ('linked_report_id', 'VARCHAR(50)'),
            ('linked_hazard_id', 'VARCHAR(50)'),
            ('linked_investigation_id', 'VARCHAR(50)'),
            ('linked_spi_id', 'INTEGER'),
            ('linked_action_id', 'VARCHAR(50)'),
            ('linked_audit_id', 'INTEGER'),
            ('linked_finding_id', 'INTEGER'),
            ('linked_risk_id', 'VARCHAR(50)'),
            ('scheduled_audit_id', 'INTEGER'),
            ('verification_area', 'VARCHAR(100)'),
            ('verification_objective', 'TEXT'),
            ('required_evidence', 'TEXT'),
            ('effectiveness_criteria', 'TEXT'),
            ('operational_risk', 'VARCHAR(20)'),
            ('due_audit_cycle', 'VARCHAR(20)'),
            ('due_date', 'VARCHAR(20)'),
            ('status', 'VARCHAR(30)'),
            ('recurrence_count', 'INTEGER'),
            ('escalation_level', 'INTEGER'),
            ('completed_at', 'TIMESTAMP'),
            ('completed_by', 'VARCHAR(100)'),
            ('effectiveness_rating', 'VARCHAR(30)'),
            ('notes', 'TEXT'),
            ('created_by', 'VARCHAR(100)'),
            ('created_at', 'TIMESTAMP'),
        ],
        'audits': [
            ('title', 'VARCHAR(200)'),
            ('audit_type', 'VARCHAR(50)'),
            ('department_id', 'INTEGER'),
            ('planned_date', 'VARCHAR(20)'),
            ('actual_date', 'VARCHAR(20)'),
            ('lead_auditor', 'VARCHAR(100)'),
            ('status', 'VARCHAR(20)'),
            ('summary', 'TEXT'),
            ('created_at', 'TIMESTAMP'),
        ],
        'checklist_template_items': [
            ('template_id', 'INTEGER'),
            ('category', 'VARCHAR(100)'),
            ('item_ref', 'VARCHAR(30)'),
            ('question', 'TEXT'),
            ('iosa_ref', 'VARCHAR(100)'),
            ('sequence', 'INTEGER'),
        ],
        'checklist_templates': [
            ('department_id', 'INTEGER'),
            ('audit_type', 'VARCHAR(50)'),
            ('name', 'VARCHAR(100)'),
            ('version', 'INTEGER'),
            ('is_active', 'BOOLEAN DEFAULT FALSE'),
            ('created_at', 'TIMESTAMP'),
            ('updated_at', 'TIMESTAMP'),
        ],
        'compliance_obligations': [
            ('ref_number', 'VARCHAR(30)'),
            ('regulation_body', 'VARCHAR(50)'),
            ('standard_ref', 'VARCHAR(100)'),
            ('requirement_title', 'VARCHAR(200)'),
            ('requirement_text', 'TEXT'),
            ('applicability', 'VARCHAR(200)'),
            ('obligation_type', 'VARCHAR(30)'),
            ('compliance_status', 'VARCHAR(30)'),
            ('priority', 'VARCHAR(20)'),
            ('evidence_description', 'TEXT'),
            ('evidence_ref', 'VARCHAR(200)'),
            ('evidence_location', 'VARCHAR(500)'),
            ('finding_ref', 'VARCHAR(100)'),
            ('linked_action_id', 'VARCHAR(30)'),
            ('responsible_dept', 'VARCHAR(100)'),
            ('responsible_person', 'VARCHAR(100)'),
            ('department_id', 'INTEGER'),
            ('review_frequency', 'VARCHAR(30)'),
            ('next_review_date', 'VARCHAR(20)'),
            ('next_review_due', 'VARCHAR(20)'),
            ('last_reviewed_date', 'VARCHAR(20)'),
            ('last_reviewed', 'VARCHAR(20)'),
            ('last_reviewed_by', 'VARCHAR(100)'),
            ('non_compliance_risk', 'VARCHAR(20)'),
            ('notes', 'TEXT'),
            ('created_by', 'VARCHAR(100)'),
            ('created_at', 'TIMESTAMP'),
        ],
        'confidential_access_logs': [
            ('report_ref', 'VARCHAR(30)'),
            ('report_type', 'VARCHAR(30)'),
            ('accessed_by', 'VARCHAR(100)'),
            ('accessor_role', 'VARCHAR(50)'),
            ('access_type', 'VARCHAR(20)'),
            ('justification', 'TEXT'),
            ('ip_address', 'VARCHAR(45)'),
            ('accessed_at', 'TIMESTAMP'),
            ('flagged_for_review', 'BOOLEAN DEFAULT FALSE'),
            ('review_notes', 'TEXT'),
        ],
        'confidential_reports': [
            ('ref_number', 'VARCHAR(30)'),
            ('position', 'VARCHAR(100)'),
            ('department_id', 'INTEGER'),
            ('date', 'VARCHAR(20)'),
            ('location', 'VARCHAR(200)'),
            ('report_type', 'VARCHAR(50)'),
            ('description', 'TEXT'),
            ('consequences', 'TEXT'),
            ('suggestion', 'TEXT'),
            ('status', 'VARCHAR(20)'),
            ('created_at', 'TIMESTAMP'),
        ],
        'controls': [
            ('risk_id', 'VARCHAR(30)'),
            ('control_type', 'VARCHAR(20)'),
            ('description', 'TEXT'),
            ('owner', 'VARCHAR(100)'),
            ('effectiveness', 'VARCHAR(30)'),
            ('review_date', 'VARCHAR(20)'),
            ('created_at', 'TIMESTAMP'),
        ],
        'departments': [
            ('code', 'VARCHAR(10)'),
            ('name', 'VARCHAR(100)'),
            ('color', 'VARCHAR(20)'),
        ],
        'device_tokens': [
            ('user_id', 'VARCHAR(30)'),
            ('fcm_token', 'TEXT'),
            ('updated_at', 'TIMESTAMP'),
        ],
        'distribution_lists': [
            ('name', 'VARCHAR(100)'),
            ('email', 'VARCHAR(200)'),
            ('department_id', 'INTEGER'),
            ('position', 'VARCHAR(100)'),
            ('is_active', 'BOOLEAN DEFAULT FALSE'),
            ('sag_role', 'VARCHAR(80)'),
            ('created_at', 'TIMESTAMP'),
        ],
        'document_links': [
            ('document_id', 'VARCHAR(50)'),
            ('entity_type', 'VARCHAR(30)'),
            ('entity_id', 'VARCHAR(50)'),
            ('link_reason', 'VARCHAR(200)'),
            ('created_at', 'TIMESTAMP'),
        ],
        'email_logs': [
            ('subject', 'VARCHAR(300)'),
            ('content_type', 'VARCHAR(30)'),
            ('content_ref', 'VARCHAR(50)'),
            ('sent_by', 'VARCHAR(100)'),
            ('sent_at', 'TIMESTAMP'),
            ('recipient_count', 'INTEGER'),
            ('dept_filter', 'VARCHAR(200)'),
            ('status', 'VARCHAR(20)'),
            ('error_message', 'TEXT'),
        ],
        'employees': [
            ('employee_id', 'VARCHAR(30)'),
            ('username', 'VARCHAR(80)'),
            ('password_hash', 'VARCHAR(200)'),
            ('full_name', 'VARCHAR(120)'),
            ('email', 'VARCHAR(120)'),
            ('mobile', 'VARCHAR(30)'),
            ('department_id', 'INTEGER'),
            ('role', 'VARCHAR(50)'),
            ('is_active', 'BOOLEAN DEFAULT FALSE'),
            ('created_at', 'TIMESTAMP'),
            ('last_login', 'TIMESTAMP'),
            ('profile_image', 'VARCHAR(200)'),
            ('base_station', "VARCHAR(10) DEFAULT 'AMM'"),
            ('join_date', 'VARCHAR(20)'),
            ('employment_status', "VARCHAR(30) DEFAULT 'Active'"),
            ('position', 'VARCHAR(100)'),
            ('language_preference', "VARCHAR(10) DEFAULT 'en'"),
            ('dark_mode', 'BOOLEAN DEFAULT FALSE'),
            ('notification_prefs', "TEXT DEFAULT '{}'"),
            ('privacy_settings', "TEXT DEFAULT '{}'"),
            ('password_changed_at', 'TIMESTAMP'),
        ],
        'employee_notification_log': [
            ('employee_user_id', 'VARCHAR(30)'),
            ('title', 'VARCHAR(200)'),
            ('body', 'TEXT'),
            ('notification_type', 'VARCHAR(50)'),
            ('content_type', 'VARCHAR(30)'),
            ('content_id', 'VARCHAR(50)'),
            ('is_read', 'BOOLEAN DEFAULT FALSE'),
            ('sent_at', 'TIMESTAMP'),
        ],
        'erp': [
            ('erp_ref', 'VARCHAR(30)'),
            ('scenario_type', 'VARCHAR(50)'),
            ('title', 'VARCHAR(200)'),
            ('description', 'TEXT'),
            ('activation_criteria', 'TEXT'),
            ('response_procedures', 'TEXT'),
            ('responsible_roles', 'TEXT'),
            ('emergency_contacts', 'TEXT'),
            ('resources_required', 'TEXT'),
            ('notification_list', 'TEXT'),
            ('review_date', 'VARCHAR(20)'),
            ('version', 'VARCHAR(10)'),
            ('status', 'VARCHAR(20)'),
            ('created_at', 'TIMESTAMP'),
        ],
        'erp_activations': [
            ('erp_id', 'VARCHAR(30)'),
            ('activation_ref', 'VARCHAR(30)'),
            ('investigation_id', 'VARCHAR(30)'),
            ('activated_at', 'TIMESTAMP'),
            ('activated_by', 'VARCHAR(100)'),
            ('activation_reason', 'TEXT'),
            ('caa_notified', 'BOOLEAN DEFAULT FALSE'),
            ('caa_notified_at', 'TIMESTAMP'),
            ('caa_ref', 'VARCHAR(50)'),
            ('icao_notified', 'BOOLEAN DEFAULT FALSE'),
            ('icao_notified_at', 'TIMESTAMP'),
            ('media_statement', 'BOOLEAN DEFAULT FALSE'),
            ('nok_notified', 'BOOLEAN DEFAULT FALSE'),
            ('deactivated_at', 'TIMESTAMP'),
            ('deactivated_by', 'VARCHAR(100)'),
            ('duration_hours', 'FLOAT'),
            ('actions_taken', 'TEXT'),
            ('effectiveness', 'VARCHAR(30)'),
            ('lessons_learned', 'TEXT'),
            ('erp_update_required', 'BOOLEAN DEFAULT FALSE'),
            ('status', 'VARCHAR(20)'),
            ('created_at', 'TIMESTAMP'),
        ],
        'erp_drills': [
            ('erp_id', 'VARCHAR(30)'),
            ('drill_ref', 'VARCHAR(30)'),
            ('drill_type', 'VARCHAR(30)'),
            ('drill_date', 'VARCHAR(20)'),
            ('duration_min', 'INTEGER'),
            ('facilitator', 'VARCHAR(100)'),
            ('participants', 'TEXT'),
            ('participant_count', 'INTEGER'),
            ('scenario_brief', 'TEXT'),
            ('objectives', 'TEXT'),
            ('observations', 'TEXT'),
            ('strengths', 'TEXT'),
            ('deficiencies', 'TEXT'),
            ('recommendations', 'TEXT'),
            ('action_items', 'TEXT'),
            ('erp_update_required', 'BOOLEAN DEFAULT FALSE'),
            ('erp_updated_date', 'VARCHAR(20)'),
            ('outcome', 'VARCHAR(20)'),
            ('next_drill_due', 'VARCHAR(20)'),
            ('created_by', 'VARCHAR(100)'),
            ('created_at', 'TIMESTAMP'),
        ],
        'findings': [
            ('audit_id', 'VARCHAR(30)'),
            ('description', 'TEXT'),
            ('severity', 'VARCHAR(20)'),
            ('root_cause', 'TEXT'),
            ('corrective_action', 'TEXT'),
            ('status', 'VARCHAR(20)'),
            ('hazard_id', 'VARCHAR(30)'),
            ('action_id', 'VARCHAR(30)'),
            ('created_at', 'TIMESTAMP'),
        ],
        'governance_audit_log': [
            ('entity_type', 'VARCHAR(50)'),
            ('entity_id', 'VARCHAR(50)'),
            ('action', 'VARCHAR(50)'),
            ('performed_by', 'VARCHAR(100)'),
            ('detail', 'TEXT'),
            ('ip_address', 'VARCHAR(45)'),
            ('created_at', 'TIMESTAMP'),
        ],
        'hazard_reports': [
            ('department_id', 'INTEGER'),
            ('location', 'VARCHAR(200)'),
            ('date', 'VARCHAR(20)'),
            ('description', 'TEXT'),
            ('classification', 'VARCHAR(50)'),
            ('generic_hazard', 'VARCHAR(200)'),
            ('consequences', 'TEXT'),
            ('immediate_action', 'TEXT'),
            ('suggested_mitigation', 'TEXT'),
            ('reporter_severity', 'VARCHAR(20)'),
            ('severity', 'VARCHAR(2)'),
            ('likelihood', 'INTEGER'),
            ('risk_index', 'VARCHAR(5)'),
            ('reporter', 'VARCHAR(100)'),
            ('report_type', 'VARCHAR(30)'),
            ('status', 'VARCHAR(30)'),
            ('hazard_id', 'VARCHAR(30)'),
            ('created_at', 'TIMESTAMP'),
            ('reporter_user_id', 'VARCHAR(30)'),
            ('acknowledged_at', 'TIMESTAMP'),
            ('acknowledged_by', 'VARCHAR(100)'),
            ('acknowledgment_due', 'TIMESTAMP'),
            ('requires_caa_notification', 'BOOLEAN DEFAULT FALSE'),
            ('caa_notification_due', 'TIMESTAMP'),
            ('caa_notified_at', 'TIMESTAMP'),
            ('caa_notification_ref', 'VARCHAR(50)'),
            ('triage_severity', 'VARCHAR(20)'),
            ('triage_by', 'VARCHAR(100)'),
            ('triage_at', 'TIMESTAMP'),
        ],
        'hazards': [
            ('source', 'VARCHAR(30)'),
            ('linked_report_id', 'VARCHAR(30)'),
            ('department_id', 'INTEGER'),
            ('classification', 'VARCHAR(30)'),
            ('type_of_activity', 'VARCHAR(100)'),
            ('generic_hazard', 'VARCHAR(200)'),
            ('specific_components', 'TEXT'),
            ('consequences', 'TEXT'),
            ('status', 'VARCHAR(20)'),
            ('owner', 'VARCHAR(100)'),
            ('created_at', 'TIMESTAMP'),
        ],
        'investigation_events': [
            ('investigation_id', 'VARCHAR(30)'),
            ('event_type', 'VARCHAR(30)'),
            ('from_stage', 'VARCHAR(40)'),
            ('to_stage', 'VARCHAR(40)'),
            ('note', 'TEXT'),
            ('performed_by', 'VARCHAR(100)'),
            ('created_at', 'TIMESTAMP'),
        ],
        'investigation_timelines': [
            ('investigation_id', 'VARCHAR(30)'),
            ('preliminary_due_date', 'VARCHAR(20)'),
            ('preliminary_submitted', 'BOOLEAN DEFAULT FALSE'),
            ('preliminary_submitted_at', 'TIMESTAMP'),
            ('preliminary_submitted_by', 'VARCHAR(100)'),
            ('preliminary_ref', 'VARCHAR(50)'),
            ('preliminary_overdue_notified', 'BOOLEAN DEFAULT FALSE'),
            ('interim_due_date', 'VARCHAR(20)'),
            ('interim_submitted', 'BOOLEAN DEFAULT FALSE'),
            ('interim_submitted_at', 'TIMESTAMP'),
            ('interim_overdue_notified', 'BOOLEAN DEFAULT FALSE'),
            ('final_due_date', 'VARCHAR(20)'),
            ('final_submitted', 'BOOLEAN DEFAULT FALSE'),
            ('final_submitted_at', 'TIMESTAMP'),
            ('final_submitted_by', 'VARCHAR(100)'),
            ('final_ref', 'VARCHAR(50)'),
            ('final_overdue_notified', 'BOOLEAN DEFAULT FALSE'),
            ('caa_notified', 'BOOLEAN DEFAULT FALSE'),
            ('caa_notified_at', 'TIMESTAMP'),
            ('caa_notification_ref', 'VARCHAR(50)'),
            ('caa_authority_name', 'VARCHAR(100)'),
            ('escalation_count', 'INTEGER'),
            ('last_escalated_at', 'TIMESTAMP'),
            ('last_escalated_to', 'VARCHAR(100)'),
            ('created_at', 'TIMESTAMP'),
            ('updated_at', 'TIMESTAMP'),
        ],
        'investigations': [
            ('title', 'VARCHAR(200)'),
            ('linked_report_id', 'VARCHAR(30)'),
            ('hazard_id', 'VARCHAR(30)'),
            ('department_id', 'INTEGER'),
            ('date_of_occurrence', 'VARCHAR(20)'),
            ('investigator', 'VARCHAR(100)'),
            ('description', 'TEXT'),
            ('why1', 'TEXT'),
            ('why2', 'TEXT'),
            ('why3', 'TEXT'),
            ('why4', 'TEXT'),
            ('why5', 'TEXT'),
            ('root_cause', 'TEXT'),
            ('human_factors', 'TEXT'),
            ('technical_factors', 'TEXT'),
            ('organizational_factors', 'TEXT'),
            ('environmental_factors', 'TEXT'),
            ('recommendations', 'TEXT'),
            ('classification', 'VARCHAR(30)'),
            ('severity_index', 'VARCHAR(5)'),
            ('occurrence_category', 'VARCHAR(50)'),
            ('phase_of_flight', 'VARCHAR(50)'),
            ('aircraft_type', 'VARCHAR(50)'),
            ('aircraft_reg', 'VARCHAR(20)'),
            ('location', 'VARCHAR(200)'),
            ('authority_notified', 'BOOLEAN DEFAULT FALSE'),
            ('notification_date', 'VARCHAR(20)'),
            ('notification_ref', 'VARCHAR(50)'),
            ('lifecycle_stage', 'VARCHAR(40)'),
            ('assigned_date', 'VARCHAR(20)'),
            ('target_close_date', 'VARCHAR(20)'),
            ('final_findings', 'TEXT'),
            ('closed_date', 'VARCHAR(20)'),
            ('closed_at', 'TIMESTAMP'),
            ('closed_by', 'VARCHAR(100)'),
            ('status', 'VARCHAR(20)'),
            ('icao_occurrence_category', 'VARCHAR(100)'),
            ('contributing_factors', 'TEXT'),
            ('findings', 'TEXT'),
            ('regulatory_ref', 'VARCHAR(200)'),
            ('notified_authority', 'VARCHAR(100)'),
            ('erp_activated', 'BOOLEAN DEFAULT FALSE'),
            ('created_at', 'TIMESTAMP'),
            ('updated_at', 'TIMESTAMP'),
            ('preliminary_report_due', 'VARCHAR(20)'),
            ('preliminary_report_submitted', 'BOOLEAN DEFAULT FALSE'),
            ('final_report_due', 'VARCHAR(20)'),
            ('final_report_submitted', 'BOOLEAN DEFAULT FALSE'),
            ('caa_notification_due_verbal', 'TIMESTAMP'),
            ('caa_notification_due_written', 'VARCHAR(20)'),
            ('reg_notification_id', 'INTEGER'),
        ],
        'just_culture_policies': [
            ('version', 'VARCHAR(10)'),
            ('version_num', 'INTEGER'),
            ('title', 'VARCHAR(200)'),
            ('content', 'TEXT'),
            ('ae_name', 'VARCHAR(100)'),
            ('ae_title', 'VARCHAR(100)'),
            ('ae_signed_at', 'TIMESTAMP'),
            ('ae_user_id', 'INTEGER'),
            ('effective_date', 'VARCHAR(20)'),
            ('review_date', 'VARCHAR(20)'),
            ('status', 'VARCHAR(20)'),
            ('distribution_sent_at', 'TIMESTAMP'),
            ('acknowledgment_rate_pct', 'FLOAT'),
            ('change_summary', 'VARCHAR(300)'),
            ('created_by', 'VARCHAR(100)'),
            ('created_at', 'TIMESTAMP'),
        ],
        'leading_indicator_configs': [
            ('spi_id', 'INTEGER'),
            ('indicator_source', 'VARCHAR(50)'),
            ('training_type_filter', 'VARCHAR(200)'),
            ('department_filter', 'VARCHAR(100)'),
            ('lookback_months', 'INTEGER'),
            ('minimum_acceptable_pct', 'FLOAT'),
            ('last_auto_calculated_at', 'TIMESTAMP'),
            ('auto_calculate', 'BOOLEAN DEFAULT FALSE'),
            ('created_at', 'TIMESTAMP'),
        ],
        'lessons_learned': [
            ('ref_number', 'VARCHAR(30)'),
            ('title', 'VARCHAR(200)'),
            ('category', 'VARCHAR(50)'),
            ('department_id', 'INTEGER'),
            ('date', 'VARCHAR(20)'),
            ('author', 'VARCHAR(100)'),
            ('description', 'TEXT'),
            ('lesson', 'TEXT'),
            ('recommendations', 'TEXT'),
            ('status', 'VARCHAR(20)'),
            ('attachment', 'VARCHAR(200)'),
            ('linked_hazard_id', 'VARCHAR(30)'),
            ('created_at', 'TIMESTAMP'),
        ],
        'moc': [
            ('title', 'VARCHAR(200)'),
            ('description', 'TEXT'),
            ('department_id', 'INTEGER'),
            ('change_type', 'VARCHAR(50)'),
            ('initiator', 'VARCHAR(100)'),
            ('planned_date', 'VARCHAR(20)'),
            ('pre_change_risk', 'TEXT'),
            ('approval_status', 'VARCHAR(30)'),
            ('approved_by', 'VARCHAR(100)'),
            ('implementation_status', 'VARCHAR(30)'),
            ('post_change_review', 'TEXT'),
            ('hazard_id', 'VARCHAR(30)'),
            ('created_at', 'TIMESTAMP'),
            ('moc_number', 'VARCHAR(30)'),
            ('change_category', 'VARCHAR(50)'),
            ('date_raised', 'VARCHAR(20)'),
            ('current_situation', 'TEXT'),
            ('proposed_change', 'TEXT'),
            ('reason_for_change', 'TEXT'),
            ('expected_benefits', 'TEXT'),
            ('impact_aircraft_ops', 'BOOLEAN DEFAULT FALSE'),
            ('impact_flight_crew', 'BOOLEAN DEFAULT FALSE'),
            ('impact_cabin_crew', 'BOOLEAN DEFAULT FALSE'),
            ('impact_ground_ops', 'BOOLEAN DEFAULT FALSE'),
            ('impact_maintenance', 'BOOLEAN DEFAULT FALSE'),
            ('impact_occ', 'BOOLEAN DEFAULT FALSE'),
            ('impact_training', 'BOOLEAN DEFAULT FALSE'),
            ('impact_safety_reporting', 'BOOLEAN DEFAULT FALSE'),
            ('impact_erp', 'BOOLEAN DEFAULT FALSE'),
            ('impact_security', 'BOOLEAN DEFAULT FALSE'),
            ('impact_regulatory', 'BOOLEAN DEFAULT FALSE'),
            ('impact_contractor', 'BOOLEAN DEFAULT FALSE'),
            ('safety_impact_level', 'VARCHAR(20)'),
            ('risk_assessment_required', 'BOOLEAN DEFAULT FALSE'),
            ('ra_status', 'VARCHAR(30)'),
            ('linked_ra_id', 'VARCHAR(30)'),
            ('icao_impact', 'BOOLEAN DEFAULT FALSE'),
            ('iosa_impact', 'BOOLEAN DEFAULT FALSE'),
            ('easa_impact', 'BOOLEAN DEFAULT FALSE'),
            ('national_authority_impact', 'BOOLEAN DEFAULT FALSE'),
            ('company_manual_impact', 'BOOLEAN DEFAULT FALSE'),
            ('regulatory_approval_required', 'BOOLEAN DEFAULT FALSE'),
            ('regulatory_approval_ref', 'VARCHAR(100)'),
            ('regulatory_approval_date', 'VARCHAR(20)'),
            ('regulatory_evidence', 'TEXT'),
            ('implementation_start_date', 'VARCHAR(20)'),
            ('target_completion_date', 'VARCHAR(20)'),
            ('training_required', 'BOOLEAN DEFAULT FALSE'),
            ('documentation_update_required', 'BOOLEAN DEFAULT FALSE'),
            ('sop_revision_required', 'BOOLEAN DEFAULT FALSE'),
            ('erp_update_required', 'BOOLEAN DEFAULT FALSE'),
            ('stakeholder_summary', 'TEXT'),
            ('dept_manager_status', 'VARCHAR(20)'),
            ('dept_manager_name', 'VARCHAR(100)'),
            ('dept_manager_date', 'VARCHAR(20)'),
            ('dept_manager_comments', 'TEXT'),
            ('safety_review_status', 'VARCHAR(20)'),
            ('safety_reviewer_name', 'VARCHAR(100)'),
            ('safety_review_date', 'VARCHAR(20)'),
            ('safety_review_comments', 'TEXT'),
            ('sm_approval_status', 'VARCHAR(20)'),
            ('sm_name', 'VARCHAR(100)'),
            ('sm_date', 'VARCHAR(20)'),
            ('sm_comments', 'TEXT'),
            ('ae_approval_required', 'BOOLEAN DEFAULT FALSE'),
            ('ae_approval_status', 'VARCHAR(20)'),
            ('ae_name', 'VARCHAR(100)'),
            ('ae_date', 'VARCHAR(20)'),
            ('ae_comments', 'TEXT'),
            ('status', 'VARCHAR(40)'),
            ('submitted_date', 'VARCHAR(20)'),
            ('approved_date', 'VARCHAR(20)'),
            ('implemented_date', 'VARCHAR(20)'),
            ('closed_date', 'VARCHAR(20)'),
            ('pir_date', 'VARCHAR(20)'),
            ('pir_reviewer', 'VARCHAR(100)'),
            ('pir_actual_outcome', 'TEXT'),
            ('pir_new_hazards', 'TEXT'),
            ('pir_effectiveness', 'VARCHAR(30)'),
            ('pir_additional_actions', 'TEXT'),
            ('pir_lessons_learned', 'TEXT'),
        ],
        'moc_hazards': [
            ('moc_id', 'VARCHAR(30)'),
            ('hazard_description', 'TEXT'),
            ('potential_consequence', 'TEXT'),
            ('existing_controls', 'TEXT'),
            ('proposed_controls', 'TEXT'),
            ('initial_risk', 'VARCHAR(20)'),
            ('residual_risk', 'VARCHAR(20)'),
            ('acceptance_status', 'VARCHAR(30)'),
            ('acceptance_authority', 'VARCHAR(100)'),
            ('linked_hazard_id', 'VARCHAR(30)'),
            ('created_at', 'TIMESTAMP'),
        ],
        'moc_milestones': [
            ('moc_id', 'VARCHAR(30)'),
            ('description', 'VARCHAR(300)'),
            ('responsible_person', 'VARCHAR(100)'),
            ('target_date', 'VARCHAR(20)'),
            ('status', 'VARCHAR(20)'),
            ('completed_date', 'VARCHAR(20)'),
            ('notes', 'TEXT'),
            ('created_at', 'TIMESTAMP'),
        ],
        'moc_stakeholders': [
            ('moc_id', 'VARCHAR(30)'),
            ('department_name', 'VARCHAR(100)'),
            ('contact_name', 'VARCHAR(100)'),
            ('consultation_date', 'VARCHAR(20)'),
            ('comments', 'TEXT'),
            ('reviewed', 'BOOLEAN DEFAULT FALSE'),
            ('created_at', 'TIMESTAMP'),
        ],
        'moc_updates': [
            ('moc_id', 'VARCHAR(30)'),
            ('update_text', 'TEXT'),
            ('update_by', 'VARCHAR(100)'),
            ('update_type', 'VARCHAR(30)'),
            ('created_at', 'TIMESTAMP'),
        ],
        'ra_checklist_items': [
            ('assessment_id', 'VARCHAR(30)'),
            ('row_seq', 'INTEGER'),
            ('category', 'VARCHAR(50)'),
            ('description', 'VARCHAR(200)'),
            ('checked', 'BOOLEAN DEFAULT FALSE'),
            ('notes', 'TEXT'),
        ],
        'ra_mitigations': [
            ('assessment_id', 'VARCHAR(30)'),
            ('hazard_seq', 'VARCHAR(10)'),
            ('mitigation', 'TEXT'),
            ('responsible_manager', 'VARCHAR(100)'),
            ('due_date', 'VARCHAR(20)'),
            ('action_id', 'VARCHAR(30)'),
            ('status', 'VARCHAR(20)'),
        ],
        'ra_review_cycles': [
            ('ra_id', 'VARCHAR(30)'),
            ('cycle_number', 'INTEGER'),
            ('due_date', 'VARCHAR(20)'),
            ('status', 'VARCHAR(20)'),
            ('warning_sent_at', 'TIMESTAMP'),
            ('overdue_notified_at', 'TIMESTAMP'),
            ('escalated_at', 'TIMESTAMP'),
            ('escalated_to', 'VARCHAR(100)'),
            ('completed_at', 'TIMESTAMP'),
            ('completed_by', 'VARCHAR(100)'),
            ('review_outcome', 'TEXT'),
            ('new_review_date', 'VARCHAR(20)'),
            ('ra_revised', 'BOOLEAN DEFAULT FALSE'),
            ('new_ra_id', 'VARCHAR(30)'),
            ('waived_by', 'VARCHAR(100)'),
            ('waiver_reason', 'TEXT'),
            ('waived_until', 'VARCHAR(20)'),
            ('created_at', 'TIMESTAMP'),
        ],
        'ra_reviews': [
            ('assessment_id', 'VARCHAR(30)'),
            ('risk_mitigation', 'VARCHAR(200)'),
            ('review_of_effectiveness', 'TEXT'),
            ('effectiveness_rating', 'VARCHAR(30)'),
            ('date_completed', 'VARCHAR(20)'),
            ('actioner', 'VARCHAR(100)'),
        ],
        'ra_rows': [
            ('assessment_id', 'VARCHAR(30)'),
            ('risk_id', 'VARCHAR(30)'),
            ('seq_num', 'INTEGER'),
            ('type_of_activity', 'VARCHAR(200)'),
            ('generic_hazard', 'VARCHAR(200)'),
            ('specific_components', 'TEXT'),
            ('consequences', 'TEXT'),
            ('likelihood_initial', 'INTEGER'),
            ('severity_initial', 'VARCHAR(2)'),
            ('risk_index_initial', 'VARCHAR(5)'),
            ('risk_tolerance_initial', 'VARCHAR(20)'),
            ('current_defenses', 'TEXT'),
            ('further_mitigations', 'TEXT'),
            ('likelihood_residual', 'INTEGER'),
            ('severity_residual', 'VARCHAR(2)'),
            ('risk_index_residual', 'VARCHAR(5)'),
            ('risk_tolerance_residual', 'VARCHAR(20)'),
        ],
        'regulatory_notifications': [
            ('ref_number', 'VARCHAR(30)'),
            ('source_type', 'VARCHAR(30)'),
            ('source_ref', 'VARCHAR(30)'),
            ('occurrence_date', 'VARCHAR(20)'),
            ('occurrence_description', 'TEXT'),
            ('notification_class', 'VARCHAR(30)'),
            ('requires_notification', 'BOOLEAN DEFAULT FALSE'),
            ('notification_trigger', 'TEXT'),
            ('authority_name', 'VARCHAR(100)'),
            ('authority_ref', 'VARCHAR(50)'),
            ('verbal_notification_due', 'TIMESTAMP'),
            ('verbal_notification_done', 'BOOLEAN DEFAULT FALSE'),
            ('verbal_notification_at', 'TIMESTAMP'),
            ('verbal_notification_by', 'VARCHAR(100)'),
            ('verbal_notification_method', 'VARCHAR(50)'),
            ('written_notification_due', 'VARCHAR(20)'),
            ('written_notification_done', 'BOOLEAN DEFAULT FALSE'),
            ('written_notification_at', 'TIMESTAMP'),
            ('written_notification_by', 'VARCHAR(100)'),
            ('written_ref', 'VARCHAR(50)'),
            ('verbal_overdue_notified', 'BOOLEAN DEFAULT FALSE'),
            ('written_overdue_notified', 'BOOLEAN DEFAULT FALSE'),
            ('status', 'VARCHAR(30)'),
            ('exempt', 'BOOLEAN DEFAULT FALSE'),
            ('exempt_reason', 'TEXT'),
            ('exempt_by', 'VARCHAR(100)'),
            ('notes', 'TEXT'),
            ('department_id', 'INTEGER'),
            ('created_by', 'VARCHAR(100)'),
            ('created_at', 'TIMESTAMP'),
            ('updated_at', 'TIMESTAMP'),
        ],
        'report_feedback': [
            ('report_ref', 'VARCHAR(30)'),
            ('report_type', 'VARCHAR(30)'),
            ('reporter_user_id', 'VARCHAR(30)'),
            ('stage_num', 'INTEGER'),
            ('stage_label', 'VARCHAR(50)'),
            ('submitted_at', 'TIMESTAMP'),
            ('acknowledged_at', 'TIMESTAMP'),
            ('acknowledged_by', 'VARCHAR(100)'),
            ('acknowledgment_due', 'TIMESTAMP'),
            ('review_started_at', 'TIMESTAMP'),
            ('ra_initiated_at', 'TIMESTAMP'),
            ('investigation_at', 'TIMESTAMP'),
            ('actions_created_at', 'TIMESTAMP'),
            ('closed_at', 'TIMESTAMP'),
            ('closed_by', 'VARCHAR(100)'),
            ('outcome_shared', 'BOOLEAN DEFAULT FALSE'),
            ('outcome_summary', 'TEXT'),
            ('outcome_actions_taken', 'TEXT'),
            ('outcome_risk_level', 'VARCHAR(20)'),
            ('lessons_learned_ref', 'VARCHAR(30)'),
            ('current_guidance', 'TEXT'),
            ('push_sent', 'BOOLEAN DEFAULT FALSE'),
            ('push_sent_at', 'TIMESTAMP'),
            ('created_at', 'TIMESTAMP'),
            ('updated_at', 'TIMESTAMP'),
        ],
        'risk_acceptances': [
            ('ref_number', 'VARCHAR(30)'),
            ('risk_id', 'VARCHAR(30)'),
            ('hazard_id', 'VARCHAR(30)'),
            ('risk_tolerance', 'VARCHAR(20)'),
            ('risk_index', 'VARCHAR(5)'),
            ('risk_description', 'TEXT'),
            ('justification', 'TEXT'),
            ('mitigations_in_place', 'TEXT'),
            ('conditions', 'TEXT'),
            ('valid_until', 'VARCHAR(20)'),
            ('review_date', 'VARCHAR(20)'),
            ('submitted_by', 'VARCHAR(100)'),
            ('submitted_date', 'VARCHAR(20)'),
            ('safety_mgr_review', 'TEXT'),
            ('safety_mgr_by', 'VARCHAR(100)'),
            ('safety_mgr_date', 'VARCHAR(20)'),
            ('ae_id', 'INTEGER'),
            ('ae_decision', 'VARCHAR(20)'),
            ('ae_decision_by', 'VARCHAR(100)'),
            ('ae_decision_date', 'VARCHAR(20)'),
            ('ae_notes', 'TEXT'),
            ('status', 'VARCHAR(30)'),
            ('created_at', 'TIMESTAMP'),
        ],
        'risk_actions': [
            ('risk_id', 'VARCHAR(30)'),
            ('hazard_id', 'VARCHAR(30)'),
            ('description', 'TEXT'),
            ('owner', 'VARCHAR(100)'),
            ('due_date', 'VARCHAR(20)'),
            ('priority', 'VARCHAR(20)'),
            ('status', 'VARCHAR(20)'),
            ('effectiveness', 'VARCHAR(30)'),
            ('closed_date', 'VARCHAR(20)'),
            ('created_at', 'TIMESTAMP'),
        ],
        'risk_assessments': [
            ('control_number', 'VARCHAR(50)'),
            ('responsible_name', 'VARCHAR(100)'),
            ('assessors_names', 'VARCHAR(300)'),
            ('assessment_date', 'VARCHAR(20)'),
            ('next_review_date', 'VARCHAR(20)'),
            ('title', 'VARCHAR(200)'),
            ('hazard_id', 'VARCHAR(30)'),
            ('department_id', 'INTEGER'),
            ('general_description', 'TEXT'),
            ('reasons', 'TEXT'),
            ('risk_level_prior', 'VARCHAR(20)'),
            ('risk_level_after', 'VARCHAR(20)'),
            ('management_acceptance', 'VARCHAR(20)'),
            ('acceptance_date', 'VARCHAR(20)'),
            ('prepared_by_name', 'VARCHAR(100)'),
            ('prepared_by_position', 'VARCHAR(100)'),
            ('reviewed_by_name', 'VARCHAR(100)'),
            ('reviewed_by_position', 'VARCHAR(100)'),
            ('approved_by_name', 'VARCHAR(100)'),
            ('approved_by_position', 'VARCHAR(100)'),
            ('status', 'VARCHAR(20)'),
            ('submitted_date', 'VARCHAR(20)'),
            ('activated_date', 'VARCHAR(20)'),
            ('closed_date', 'VARCHAR(20)'),
            ('revision', 'INTEGER'),
            ('parent_ra_id', 'VARCHAR(30)'),
            ('created_at', 'TIMESTAMP'),
        ],
        'risk_occurrences': [
            ('hazard_id', 'VARCHAR(30)'),
            ('occurrence_date', 'VARCHAR(20)'),
            ('description', 'TEXT'),
            ('source', 'VARCHAR(30)'),
            ('linked_report_id', 'VARCHAR(30)'),
            ('created_at', 'TIMESTAMP'),
        ],
        'risks': [
            ('hazard_id', 'VARCHAR(30)'),
            ('description', 'TEXT'),
            ('initial_likelihood', 'INTEGER'),
            ('initial_severity', 'VARCHAR(2)'),
            ('initial_risk_index', 'VARCHAR(5)'),
            ('initial_tolerance', 'VARCHAR(20)'),
            ('residual_likelihood', 'INTEGER'),
            ('residual_severity', 'VARCHAR(2)'),
            ('residual_risk_index', 'VARCHAR(5)'),
            ('residual_tolerance', 'VARCHAR(20)'),
            ('created_at', 'TIMESTAMP'),
        ],
        'safety_bulletins': [
            ('ref_number', 'VARCHAR(30)'),
            ('title', 'VARCHAR(200)'),
            ('bulletin_type', 'VARCHAR(30)'),
            ('severity', 'VARCHAR(20)'),
            ('department_id', 'INTEGER'),
            ('issue_date', 'VARCHAR(20)'),
            ('content', 'TEXT'),
            ('recommendations', 'TEXT'),
            ('issued_by', 'VARCHAR(100)'),
            ('status', 'VARCHAR(20)'),
            ('attachment', 'VARCHAR(200)'),
            ('linked_hazard_id', 'VARCHAR(30)'),
            ('created_at', 'TIMESTAMP'),
        ],
        'safety_campaigns': [
            ('title', 'VARCHAR(200)'),
            ('campaign_type', 'VARCHAR(50)'),
            ('department_id', 'INTEGER'),
            ('start_date', 'VARCHAR(20)'),
            ('end_date', 'VARCHAR(20)'),
            ('description', 'TEXT'),
            ('objectives', 'TEXT'),
            ('status', 'VARCHAR(20)'),
            ('attachment', 'VARCHAR(200)'),
            ('created_at', 'TIMESTAMP'),
        ],
        'safety_newsletters': [
            ('ref_number', 'VARCHAR(30)'),
            ('title', 'VARCHAR(200)'),
            ('issue_number', 'VARCHAR(20)'),
            ('department_id', 'INTEGER'),
            ('issue_date', 'VARCHAR(20)'),
            ('author', 'VARCHAR(100)'),
            ('summary', 'TEXT'),
            ('content', 'TEXT'),
            ('status', 'VARCHAR(20)'),
            ('attachment', 'VARCHAR(200)'),
            ('created_at', 'TIMESTAMP'),
        ],
        'safety_personnel': [
            ('name', 'VARCHAR(100)'),
            ('position', 'VARCHAR(100)'),
            ('department_id', 'INTEGER'),
            ('sms_role', 'VARCHAR(100)'),
            ('qualifications', 'TEXT'),
            ('contact_email', 'VARCHAR(100)'),
            ('contact_phone', 'VARCHAR(50)'),
            ('sms_trained', 'BOOLEAN DEFAULT FALSE'),
            ('training_date', 'VARCHAR(20)'),
            ('active', 'BOOLEAN DEFAULT FALSE'),
            ('created_at', 'TIMESTAMP'),
        ],
        'safety_policies': [
            ('version', 'VARCHAR(10)'),
            ('version_num', 'INTEGER'),
            ('title', 'VARCHAR(200)'),
            ('content', 'TEXT'),
            ('approved_by', 'VARCHAR(100)'),
            ('approved_by_title', 'VARCHAR(100)'),
            ('effective_date', 'VARCHAR(20)'),
            ('review_date', 'VARCHAR(20)'),
            ('status', 'VARCHAR(20)'),
            ('change_summary', 'TEXT'),
            ('created_at', 'TIMESTAMP'),
        ],
        'safety_promo_acks': [
            ('user_id', 'VARCHAR(30)'),
            ('full_name', 'VARCHAR(100)'),
            ('content_type', 'VARCHAR(30)'),
            ('content_id', 'VARCHAR(50)'),
            ('acked_at', 'TIMESTAMP'),
            ('device_info', 'VARCHAR(200)'),
        ],
        'safety_promo_reads': [
            ('user_id', 'VARCHAR(30)'),
            ('content_type', 'VARCHAR(30)'),
            ('content_id', 'VARCHAR(50)'),
            ('read_at', 'TIMESTAMP'),
        ],
        'safety_recommendations': [
            ('sr_number', 'VARCHAR(30)'),
            ('source_type', 'VARCHAR(30)'),
            ('source_id', 'VARCHAR(30)'),
            ('investigation_id', 'VARCHAR(30)'),
            ('audit_finding_id', 'VARCHAR(30)'),
            ('title', 'VARCHAR(200)'),
            ('description', 'TEXT'),
            ('safety_issue', 'TEXT'),
            ('addressee_type', 'VARCHAR(20)'),
            ('addressee_name', 'VARCHAR(200)'),
            ('addressee_email', 'VARCHAR(200)'),
            ('department_id', 'INTEGER'),
            ('priority', 'VARCHAR(20)'),
            ('icao_reference', 'VARCHAR(100)'),
            ('issued_date', 'VARCHAR(20)'),
            ('issued_by', 'VARCHAR(100)'),
            ('response_due_date', 'VARCHAR(20)'),
            ('status', 'VARCHAR(30)'),
            ('response_text', 'TEXT'),
            ('response_date', 'VARCHAR(20)'),
            ('response_by', 'VARCHAR(100)'),
            ('linked_action_id', 'VARCHAR(30)'),
            ('closure_verified_by', 'VARCHAR(100)'),
            ('closure_date', 'VARCHAR(20)'),
            ('closure_notes', 'TEXT'),
            ('effectiveness_rating', 'VARCHAR(30)'),
            ('created_by', 'VARCHAR(100)'),
            ('created_at', 'TIMESTAMP'),
            ('updated_at', 'TIMESTAMP'),
        ],
        'safety_roles': [
            ('role_name', 'VARCHAR(100)'),
            ('role_type', 'VARCHAR(50)'),
            ('person_name', 'VARCHAR(100)'),
            ('department_id', 'INTEGER'),
            ('responsibilities', 'TEXT'),
            ('authority', 'TEXT'),
            ('contact_email', 'VARCHAR(100)'),
            ('contact_phone', 'VARCHAR(50)'),
            ('effective_from', 'VARCHAR(20)'),
            ('active', 'BOOLEAN DEFAULT FALSE'),
            ('created_at', 'TIMESTAMP'),
        ],
        'safety_surveys': [
            ('title', 'VARCHAR(200)'),
            ('survey_type', 'VARCHAR(50)'),
            ('department_id', 'INTEGER'),
            ('start_date', 'VARCHAR(20)'),
            ('end_date', 'VARCHAR(20)'),
            ('description', 'TEXT'),
            ('questions', 'TEXT'),
            ('status', 'VARCHAR(20)'),
            ('target_count', 'INTEGER'),
            ('response_count', 'INTEGER'),
            ('target_audience', 'VARCHAR(200)'),
            ('created_at', 'TIMESTAMP'),
        ],
        'sms_documents': [
            ('doc_type', 'VARCHAR(10)'),
            ('department_id', 'INTEGER'),
            ('title', 'VARCHAR(200)'),
            ('description', 'TEXT'),
            ('content', 'TEXT'),
            ('version', 'VARCHAR(10)'),
            ('version_num', 'INTEGER'),
            ('seq_num', 'INTEGER'),
            ('status', 'VARCHAR(20)'),
            ('created_by', 'VARCHAR(100)'),
            ('reviewed_by', 'VARCHAR(100)'),
            ('approved_by', 'VARCHAR(100)'),
            ('effective_date', 'VARCHAR(20)'),
            ('review_due', 'VARCHAR(20)'),
            ('parent_doc_id', 'VARCHAR(50)'),
            ('change_summary', 'TEXT'),
            ('created_at', 'TIMESTAMP'),
        ],
        'sod_violation_blocks': [
            ('attempted_by', 'VARCHAR(100)'),
            ('attempted_role', 'VARCHAR(50)'),
            ('entity_type', 'VARCHAR(50)'),
            ('entity_id', 'VARCHAR(50)'),
            ('attempted_action', 'VARCHAR(50)'),
            ('violation_rule', 'VARCHAR(100)'),
            ('violation_detail', 'TEXT'),
            ('original_submitter', 'VARCHAR(100)'),
            ('ip_address', 'VARCHAR(45)'),
            ('blocked_at', 'TIMESTAMP'),
        ],
        'spi_data': [
            ('spi_id', 'INTEGER'),
            ('year', 'INTEGER'),
            ('month', 'INTEGER'),
            ('events', 'INTEGER'),
            ('exposure', 'FLOAT'),
            ('total_events', 'INTEGER'),
            ('value', 'FLOAT'),
            ('source', 'VARCHAR(20)'),
            ('notes', 'TEXT'),
            ('flights', 'INTEGER'),
            ('rate', 'FLOAT'),
            ('mean_at_time', 'FLOAT'),
            ('sd_at_time', 'FLOAT'),
        ],
        'spi_escalations': [
            ('spi_id', 'INTEGER'),
            ('trigger_month', 'INTEGER'),
            ('trigger_year', 'INTEGER'),
            ('trigger_rule', 'VARCHAR(2)'),
            ('alert_level', 'VARCHAR(5)'),
            ('spi_value', 'FLOAT'),
            ('threshold_value', 'FLOAT'),
            ('mean_value', 'FLOAT'),
            ('sd_value', 'FLOAT'),
            ('description', 'TEXT'),
            ('detected_at', 'TIMESTAMP'),
            ('status', 'VARCHAR(20)'),
            ('linked_action_id', 'VARCHAR(30)'),
        ],
        'spi_indicators': [
            ('code', 'VARCHAR(20)'),
            ('name', 'VARCHAR(200)'),
            ('department_ids', 'VARCHAR(50)'),
            ('category', 'VARCHAR(50)'),
            ('description', 'TEXT'),
            ('calc_type', 'VARCHAR(10)'),
            ('exposure_type', 'VARCHAR(30)'),
            ('unit', 'VARCHAR(50)'),
            ('frequency', 'VARCHAR(20)'),
            ('spt_target', 'FLOAT'),
            ('alert_l1', 'FLOAT'),
            ('alert_l2', 'FLOAT'),
            ('alert_l3', 'FLOAT'),
            ('auto_source', 'VARCHAR(50)'),
            ('auto_category', 'VARCHAR(50)'),
            ('indicator_type', 'VARCHAR(10)'),
            ('industry_benchmark', 'FLOAT'),
            ('baseline_months', 'INTEGER'),
            ('improvement_pct', 'FLOAT'),
            ('stat_mode', 'BOOLEAN DEFAULT FALSE'),
            ('active', 'BOOLEAN DEFAULT FALSE'),
            ('created_at', 'TIMESTAMP'),
        ],
        'srb_agenda_items': [
            ('meeting_id', 'VARCHAR(30)'),
            ('item_number', 'INTEGER'),
            ('title', 'VARCHAR(200)'),
            ('description', 'TEXT'),
            ('item_type', 'VARCHAR(30)'),
            ('source_type', 'VARCHAR(30)'),
            ('source_id', 'VARCHAR(50)'),
            ('presenter', 'VARCHAR(100)'),
            ('time_allocated', 'INTEGER'),
            ('status', 'VARCHAR(20)'),
            ('discussion_notes', 'TEXT'),
            ('decision', 'TEXT'),
            ('action_required', 'BOOLEAN DEFAULT FALSE'),
            ('deferred_to_meeting_id', 'VARCHAR(30)'),
            ('created_at', 'TIMESTAMP'),
        ],
        'srb_attendees': [
            ('meeting_id', 'VARCHAR(30)'),
            ('person_name', 'VARCHAR(100)'),
            ('role_title', 'VARCHAR(100)'),
            ('department', 'VARCHAR(100)'),
            ('is_required', 'BOOLEAN DEFAULT FALSE'),
            ('attended', 'BOOLEAN DEFAULT FALSE'),
            ('apology_given', 'BOOLEAN DEFAULT FALSE'),
            ('proxy_for', 'VARCHAR(100)'),
            ('sort_order', 'INTEGER'),
            ('created_at', 'TIMESTAMP'),
        ],
        'srb_decisions': [
            ('meeting_id', 'VARCHAR(30)'),
            ('agenda_item_id', 'INTEGER'),
            ('decision_ref', 'VARCHAR(30)'),
            ('decision_text', 'TEXT'),
            ('decision_type', 'VARCHAR(20)'),
            ('responsible_party', 'VARCHAR(100)'),
            ('due_date', 'VARCHAR(20)'),
            ('linked_action_id', 'VARCHAR(30)'),
            ('status', 'VARCHAR(20)'),
            ('closed_date', 'VARCHAR(20)'),
            ('closure_notes', 'TEXT'),
            ('created_at', 'TIMESTAMP'),
        ],
        'srb_meetings': [
            ('meeting_type', 'VARCHAR(20)'),
            ('title', 'VARCHAR(200)'),
            ('scheduled_date', 'VARCHAR(20)'),
            ('actual_date', 'VARCHAR(20)'),
            ('start_time', 'VARCHAR(10)'),
            ('end_time', 'VARCHAR(10)'),
            ('venue', 'VARCHAR(200)'),
            ('chair_person', 'VARCHAR(100)'),
            ('secretary', 'VARCHAR(100)'),
            ('status', 'VARCHAR(20)'),
            ('quorum_met', 'BOOLEAN DEFAULT FALSE'),
            ('quorum_count', 'INTEGER'),
            ('objectives', 'TEXT'),
            ('minutes_text', 'TEXT'),
            ('key_outcomes', 'TEXT'),
            ('next_meeting_date', 'VARCHAR(20)'),
            ('minutes_approved_by', 'VARCHAR(100)'),
            ('minutes_approved_date', 'VARCHAR(20)'),
            ('ae_id', 'INTEGER'),
            ('created_by', 'VARCHAR(100)'),
            ('created_at', 'TIMESTAMP'),
            ('updated_at', 'TIMESTAMP'),
        ],
        'survey_responses': [
            ('survey_id', 'INTEGER'),
            ('respondent_name', 'VARCHAR(100)'),
            ('respondent_email', 'VARCHAR(200)'),
            ('department_id', 'INTEGER'),
            ('is_anonymous', 'BOOLEAN DEFAULT FALSE'),
            ('answers', 'TEXT'),
            ('submitted_at', 'TIMESTAMP'),
            ('ip_address', 'VARCHAR(50)'),
        ],
        'trainings': [
            ('employee_name', 'VARCHAR(100)'),
            ('employee_id', 'VARCHAR(50)'),
            ('department_id', 'INTEGER'),
            ('position', 'VARCHAR(100)'),
            ('training_type', 'VARCHAR(50)'),
            ('training_program', 'VARCHAR(200)'),
            ('course_code', 'VARCHAR(50)'),
            ('instructor', 'VARCHAR(100)'),
            ('location', 'VARCHAR(100)'),
            ('scheduled_date', 'VARCHAR(20)'),
            ('training_date', 'VARCHAR(20)'),
            ('completion_date', 'VARCHAR(20)'),
            ('expiry_date', 'VARCHAR(20)'),
            ('duration_hours', 'FLOAT'),
            ('status', 'VARCHAR(20)'),
            ('certificate', 'VARCHAR(200)'),
            ('evidence', 'VARCHAR(200)'),
            ('is_recurrent', 'BOOLEAN DEFAULT FALSE'),
            ('recurrence_months', 'INTEGER'),
            ('notes', 'TEXT'),
            ('updated_at', 'TIMESTAMP'),
            ('created_at', 'TIMESTAMP'),
        ],
        'users': [
            ('username', 'VARCHAR(80)'),
            ('password_hash', 'VARCHAR(200)'),
            ('full_name', 'VARCHAR(100)'),
            ('role', 'VARCHAR(30)'),
            ('department_id', 'INTEGER'),
            ('is_active', 'BOOLEAN DEFAULT FALSE'),
            ('sag_role', 'VARCHAR(80)'),
            ('created_at', 'TIMESTAMP'),
            ('last_login', 'TIMESTAMP'),
        ],
        'voluntary_reports': [
            ('ref_number', 'VARCHAR(30)'),
            ('reporter_name', 'VARCHAR(100)'),
            ('position', 'VARCHAR(100)'),
            ('department_id', 'INTEGER'),
            ('date', 'VARCHAR(20)'),
            ('location', 'VARCHAR(200)'),
            ('report_type', 'VARCHAR(50)'),
            ('description', 'TEXT'),
            ('consequences', 'TEXT'),
            ('suggestion', 'TEXT'),
            ('status', 'VARCHAR(20)'),
            ('is_confidential', 'BOOLEAN DEFAULT FALSE'),
            ('created_at', 'TIMESTAMP'),
        ],
    }

    # Run column migrations using a raw psycopg2 connection so they are
    # guaranteed to execute before any ORM query — bypasses SQLAlchemy pool.
    import psycopg2 as _pg2
    _db_url = os.environ.get('DATABASE_URL', '')
    if _db_url.startswith('postgres://'):
        _db_url = _db_url.replace('postgres://', 'postgresql://', 1)
    try:
        _raw = _pg2.connect(_db_url)
        _raw.autocommit = True
        _cur = _raw.cursor()
        for _tbl, _cols in _migrations.items():
            for _col, _ctype in _cols:
                try:
                    _cur.execute(
                        f"ALTER TABLE {_tbl} ADD COLUMN IF NOT EXISTS {_col} {_ctype}"
                    )
                except Exception as _ce:
                    print(f"[MIGRATION] {_tbl}.{_col}: {_ce}", flush=True)
        # Widen profile_image to VARCHAR(500) for Cloudinary URLs
        try:
            _cur.execute(
                "ALTER TABLE employees ALTER COLUMN profile_image TYPE VARCHAR(500)"
            )
        except Exception:
            pass
        # Add dedicated Cloudinary URL column
        try:
            _cur.execute(
                "ALTER TABLE employees ADD COLUMN IF NOT EXISTS profile_photo_url VARCHAR(500)"
            )
        except Exception:
            pass
        _cur.close()
        _raw.close()
        print("[MIGRATION] Column migrations complete", flush=True)
    except Exception as _mig_err:
        print(f"[MIGRATION] Could not run migrations: {_mig_err}", flush=True)

    # spi_event_links DDL
    try:
        with db.engine.connect() as conn:
            conn = conn.execution_options(isolation_level="AUTOCOMMIT")
            conn.execute(_sa_text(
                "CREATE TABLE IF NOT EXISTS spi_event_links ("
                "  id SERIAL PRIMARY KEY,"
                "  spi_id INTEGER NOT NULL REFERENCES spi_indicators(id) ON DELETE CASCADE,"
                "  event_type VARCHAR(50) NOT NULL,"
                "  event_id VARCHAR(50) NOT NULL,"
                "  event_title VARCHAR(300),"
                "  event_date VARCHAR(20),"
                "  severity VARCHAR(20),"
                "  department_id INTEGER,"
                "  category VARCHAR(100),"
                "  linked_at TIMESTAMP DEFAULT NOW()"
                ")"
            ))
    except Exception: pass

    # moc_updates DDL
    try:
        with db.engine.connect() as conn:
            conn = conn.execution_options(isolation_level="AUTOCOMMIT")
            conn.execute(_sa_text(
                "CREATE TABLE IF NOT EXISTS moc_updates ("
                "  id SERIAL PRIMARY KEY,"
                "  moc_id VARCHAR(30) NOT NULL REFERENCES moc(id) ON DELETE CASCADE,"
                "  update_text TEXT,"
                "  update_by VARCHAR(100),"
                "  update_type VARCHAR(30) DEFAULT 'Progress',"
                "  created_at TIMESTAMP DEFAULT NOW()"
                ")"
            ))
    except Exception: pass

    # moc_stakeholders DDL
    try:
        with db.engine.connect() as conn:
            conn = conn.execution_options(isolation_level="AUTOCOMMIT")
            conn.execute(_sa_text(
                "CREATE TABLE IF NOT EXISTS moc_stakeholders ("
                "  id SERIAL PRIMARY KEY,"
                "  moc_id VARCHAR(30) NOT NULL REFERENCES moc(id) ON DELETE CASCADE,"
                "  name VARCHAR(100),"
                "  role VARCHAR(100),"
                "  department VARCHAR(100),"
                "  consulted_date VARCHAR(20),"
                "  feedback TEXT,"
                "  reviewed BOOLEAN DEFAULT FALSE,"
                "  created_at TIMESTAMP DEFAULT NOW()"
                ")"
            ))
    except Exception: pass

def seed():
    try:
        if not Department.query.first():
            for code, name in [
                ('FLT','Flight Operations'),('CAB','Cabin Crew'),
                ('GND','Ground Operations'),('MNT','Maintenance'),
                ('OCC','Operations Control'),('TRN','Training'),
                ('SMS','Safety Management'),('SEC','Security'),
                ('MGT','Management'),('QA','Quality Assurance'),
            ]:
                db.session.add(Department(code=code, name=name))
            db.session.commit()
    except Exception:
        db.session.rollback()

with app.app_context():
    seed()


# ═══════════════════════════════════════════════════════════════════════════════
#  EXCEL REGISTER EXPORTS
# ═══════════════════════════════════════════════════════════════════════════════
from excel_exports import (
    export_moc_register, export_investigation_register, export_hazard_register,
    export_risk_register, export_action_register, export_audit_register,
    export_finding_register, export_training_register, export_employee_register,
    export_survey_register, export_bulletin_register, export_newsletter_register,
    export_spi_register, export_document_register, export_ra_register,
)

@app.route('/moc/export-excel')
@require_login
def moc_export_excel():
    mocs = MOC.query.order_by(MOC.created_at.desc()).all()
    return export_moc_register(mocs)

@app.route('/investigations/export-excel')
@require_login
def investigation_export_excel():
    invs = Investigation.query.order_by(Investigation.created_at.desc()).all()
    return export_investigation_register(invs)

@app.route('/hazards/export-excel')
@require_login
def hazard_export_excel():
    hazards = Hazard.query.order_by(Hazard.created_at.desc()).all()
    return export_hazard_register(hazards)

@app.route('/risks/export-excel')
@require_login
def risk_export_excel():
    risks = Risk.query.order_by(Risk.id.desc()).all()
    return export_risk_register(risks)

@app.route('/actions/export-excel')
@require_login
def action_export_excel():
    actions = Action.query.order_by(Action.created_at.desc()).all()
    return export_action_register(actions)

@app.route('/audits/export-excel')
@require_login
def audit_export_excel():
    audits = Audit.query.order_by(Audit.id.desc()).all()
    return export_audit_register(audits)

@app.route('/findings/export-excel')
@require_login
def finding_export_excel():
    findings = AuditFinding.query.order_by(AuditFinding.id.desc()).all()
    return export_finding_register(findings)

@app.route('/training/export-excel')
@require_login
def training_export_excel():
    trainings = Training.query.order_by(Training.id.desc()).all()
    return export_training_register(trainings)

@app.route('/employees/export-excel')
@require_login
def employee_export_excel():
    employees = Employee.query.order_by(Employee.id.desc()).all()
    return export_employee_register(employees)

@app.route('/surveys/export-excel')
@require_login
def survey_export_excel():
    surveys = SafetySurvey.query.order_by(SafetySurvey.id.desc()).all()
    return export_survey_register(surveys)

@app.route('/bulletins/export-excel')
@require_login
def bulletin_export_excel():
    bulletins = SafetyBulletin.query.order_by(SafetyBulletin.id.desc()).all()
    return export_bulletin_register(bulletins)

@app.route('/newsletters/export-excel')
@require_login
def newsletter_export_excel():
    newsletters = SafetyNewsletter.query.order_by(SafetyNewsletter.id.desc()).all()
    return export_newsletter_register(newsletters)

@app.route('/spi/export-excel')
@require_login
def spi_export_excel():
    spis = SPIIndicator.query.order_by(SPIIndicator.id.desc()).all()
    return export_spi_register(spis)

@app.route('/documents/export-excel')
@require_login
def document_export_excel():
    docs = SMSDocument.query.order_by(SMSDocument.id.desc()).all()
    return export_document_register(docs)

@app.route('/risk-assessments/export-excel')
@require_login
def ra_export_excel():
    ras = RiskAssessment.query.order_by(RiskAssessment.id.desc()).all()
    return export_ra_register(ras)


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
