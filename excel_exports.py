"""
AviaS SMS — Excel Register Exports
All exports follow consistent formatting: AviaS branding, column headers,
filters, reference numbers, dates, status, owner, export timestamp.
"""
import io
from datetime import datetime

import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter
from flask import send_file

# ── Brand colours ────────────────────────────────────────────────────────────
NAVY   = "0A1628"
GOLD   = "C9A84C"
WHITE  = "FFFFFF"
LIGHT  = "F0F4F8"
MUTED  = "64748B"
GREEN  = "15803D"
RED    = "DC2626"
AMBER  = "B45309"

# ── Style factories ───────────────────────────────────────────────────────────
def _fill(hex_colour):
    return PatternFill("solid", fgColor=hex_colour)

def _border():
    thin = Side(style="thin", color="D1D5DB")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _title_font():   return Font(name="Calibri", bold=True, size=14, color=WHITE)
def _sub_font():     return Font(name="Calibri", size=10, color="CBD5E1")
def _header_font():  return Font(name="Calibri", bold=True, size=10, color=WHITE)
def _cell_font():    return Font(name="Calibri", size=10, color="1E293B")
def _muted_font():   return Font(name="Calibri", size=9, color=MUTED)

def _center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def _left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _make_workbook(sheet_title):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    return wb, ws


def _write_header(ws, title, subtitle, columns):
    """Write branded title rows + column headers. Returns the row after headers."""
    # Row 1: Title
    ws.merge_cells(f"A1:{get_column_letter(len(columns))}1")
    c = ws["A1"]
    c.value = f"AviaS Safety Management System  |  {title}"
    c.font  = _title_font()
    c.fill  = _fill(NAVY)
    c.alignment = _center()
    ws.row_dimensions[1].height = 28

    # Row 2: Subtitle / timestamp
    ws.merge_cells(f"A2:{get_column_letter(len(columns))}2")
    c = ws["A2"]
    c.value = f"{subtitle}  ·  Exported {datetime.utcnow().strftime('%d %b %Y %H:%M')} UTC"
    c.font  = _sub_font()
    c.fill  = _fill(NAVY)
    c.alignment = _center()
    ws.row_dimensions[2].height = 18

    # Row 3: blank divider
    ws.row_dimensions[3].height = 6

    # Row 4: column headers
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        c = ws.cell(row=4, column=col_idx, value=col_name)
        c.font      = _header_font()
        c.fill      = _fill(GOLD)
        c.alignment = _center()
        c.border    = _border()
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[4].height = 22
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(columns))}4"
    return 5   # first data row


def _write_row(ws, row_num, values, shade=False):
    bg = LIGHT if shade else WHITE
    for col_idx, val in enumerate(values, start=1):
        c = ws.cell(row=row_num, column=col_idx, value=val)
        c.font      = _cell_font()
        c.fill      = _fill(bg)
        c.alignment = _left()
        c.border    = _border()
    ws.row_dimensions[row_num].height = 16


def _send(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ─────────────────────────────────────────────────────────────────────────────
# MOC Register
# ─────────────────────────────────────────────────────────────────────────────
def export_moc_register(mocs):
    wb, ws = _make_workbook("MOC Register")
    cols = [
        ("MOC Number", 18), ("Title", 40), ("Category", 20), ("Status", 18),
        ("Safety Impact", 14), ("Department", 20), ("Initiator", 20),
        ("Date Raised", 14), ("Target Completion", 18), ("Approved Date", 14),
        ("Implemented Date", 18), ("Closed Date", 14),
    ]
    row = _write_header(ws, "Management of Change Register", "MOC Register", cols)
    for i, m in enumerate(mocs):
        _write_row(ws, row, [
            m.moc_number or m.id,
            m.title or "—",
            m.change_category or "—",
            m.status or "—",
            m.safety_impact_level or "—",
            m.department.name if m.department else "—",
            m.initiator or "—",
            m.date_raised or (m.created_at.strftime("%Y-%m-%d") if m.created_at else "—"),
            m.target_completion_date or "—",
            m.approved_date or "—",
            m.implemented_date or "—",
            m.closed_date or "—",
        ], shade=i % 2 == 1)
        row += 1
    return _send(wb, f"MOC_Register_{datetime.utcnow().strftime('%Y%m%d')}.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# Investigation Register
# ─────────────────────────────────────────────────────────────────────────────
def export_investigation_register(investigations):
    wb, ws = _make_workbook("Investigation Register")
    cols = [
        ("Reference", 18), ("Title", 40), ("Type", 18), ("Status", 16),
        ("Severity", 12), ("Department", 20), ("Lead Investigator", 22),
        ("Occurrence Date", 16), ("Opened Date", 14), ("Closed Date", 14),
        ("Root Cause", 30),
    ]
    row = _write_header(ws, "Investigation Register", "Safety Investigation Register", cols)
    for i, inv in enumerate(investigations):
        _write_row(ws, row, [
            inv.ref_number or inv.id,
            inv.title or "—",
            inv.investigation_type or "—",
            inv.status or "—",
            inv.severity or "—",
            inv.department.name if hasattr(inv, "department") and inv.department else "—",
            inv.lead_investigator or "—",
            inv.occurrence_date or "—",
            inv.created_at.strftime("%Y-%m-%d") if inv.created_at else "—",
            inv.closed_date or "—",
            (inv.root_cause or "—")[:100],
        ], shade=i % 2 == 1)
        row += 1
    return _send(wb, f"Investigation_Register_{datetime.utcnow().strftime('%Y%m%d')}.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# Hazard Register
# ─────────────────────────────────────────────────────────────────────────────
def export_hazard_register(hazards):
    wb, ws = _make_workbook("Hazard Register")
    cols = [
        ("Hazard ID", 16), ("Description", 40), ("Status", 16),
        ("Source", 18), ("Department", 20), ("Reporter", 20),
        ("Reported Date", 14), ("Risk Level", 14),
    ]
    row = _write_header(ws, "Hazard Register", "Identified Hazards", cols)
    for i, h in enumerate(hazards):
        _write_row(ws, row, [
            h.id,
            h.description or "—",
            h.status or "—",
            h.source or "—",
            h.department.name if hasattr(h, "department") and h.department else "—",
            h.reporter_name or "—",
            h.created_at.strftime("%Y-%m-%d") if hasattr(h, "created_at") and h.created_at else "—",
            h.risk_level or "—",
        ], shade=i % 2 == 1)
        row += 1
    return _send(wb, f"Hazard_Register_{datetime.utcnow().strftime('%Y%m%d')}.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# Risk Register
# ─────────────────────────────────────────────────────────────────────────────
def export_risk_register(risks):
    wb, ws = _make_workbook("Risk Register")
    cols = [
        ("Risk ID", 16), ("Title", 40), ("Status", 16), ("Category", 18),
        ("Likelihood", 14), ("Severity", 12), ("Risk Level", 14),
        ("Owner", 20), ("Department", 20), ("Raised Date", 14),
        ("Review Date", 14),
    ]
    row = _write_header(ws, "Risk Register", "Operational Risk Register", cols)
    for i, r in enumerate(risks):
        _write_row(ws, row, [
            r.id,
            r.title or r.description or "—",
            r.status or "—",
            r.category or "—",
            r.likelihood or "—",
            r.severity or "—",
            r.risk_level or "—",
            r.owner or "—",
            r.department.name if hasattr(r, "department") and r.department else "—",
            r.created_at.strftime("%Y-%m-%d") if hasattr(r, "created_at") and r.created_at else "—",
            r.review_date or "—",
        ], shade=i % 2 == 1)
        row += 1
    return _send(wb, f"Risk_Register_{datetime.utcnow().strftime('%Y%m%d')}.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# Action Register
# ─────────────────────────────────────────────────────────────────────────────
def export_action_register(actions):
    wb, ws = _make_workbook("Action Register")
    cols = [
        ("Action ID", 18), ("Description", 40), ("Status", 16),
        ("Priority", 12), ("Source", 18), ("Owner / SAG Member", 22),
        ("Department", 20), ("Due Date", 14), ("Assigned Date", 14),
        ("Closed Date", 14), ("Effectiveness", 18),
    ]
    row = _write_header(ws, "Action Register", "Corrective & Preventive Actions", cols)
    for i, a in enumerate(actions):
        _write_row(ws, row, [
            a.id,
            (a.description or "—")[:100],
            a.status or "—",
            a.priority or "—",
            a.source or "—",
            a.sag_member or a.owner or "—",
            a.department.name if hasattr(a, "department") and a.department else "—",
            a.due_date or "—",
            a.created_at.strftime("%Y-%m-%d") if hasattr(a, "created_at") and a.created_at else "—",
            a.closed_date or "—",
            a.effectiveness or "—",
        ], shade=i % 2 == 1)
        row += 1
    return _send(wb, f"Action_Register_{datetime.utcnow().strftime('%Y%m%d')}.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# Audit Register
# ─────────────────────────────────────────────────────────────────────────────
def export_audit_register(audits):
    wb, ws = _make_workbook("Audit Register")
    cols = [
        ("Audit ID", 16), ("Title", 40), ("Type", 18), ("Status", 16),
        ("Department", 20), ("Lead Auditor", 22), ("Planned Date", 14),
        ("Completed Date", 16), ("Findings", 12),
    ]
    row = _write_header(ws, "Audit Register", "Internal Safety Audits", cols)
    for i, a in enumerate(audits):
        findings_count = len(a.findings) if hasattr(a, "findings") else "—"
        _write_row(ws, row, [
            a.id,
            a.title or "—",
            a.audit_type or "—",
            a.status or "—",
            a.department.name if hasattr(a, "department") and a.department else "—",
            a.lead_auditor or "—",
            a.planned_date or "—",
            a.completed_date or "—",
            findings_count,
        ], shade=i % 2 == 1)
        row += 1
    return _send(wb, f"Audit_Register_{datetime.utcnow().strftime('%Y%m%d')}.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# Finding Register
# ─────────────────────────────────────────────────────────────────────────────
def export_finding_register(findings):
    wb, ws = _make_workbook("Finding Register")
    cols = [
        ("Finding ID", 16), ("Description", 40), ("Type", 18),
        ("Status", 16), ("Severity", 12), ("Audit", 20),
        ("Department", 20), ("Owner", 20), ("Due Date", 14),
        ("Raised Date", 14), ("Closed Date", 14),
    ]
    row = _write_header(ws, "Audit Finding Register", "Audit Findings & Observations", cols)
    for i, f in enumerate(findings):
        _write_row(ws, row, [
            f.id,
            (f.description or "—")[:100],
            f.finding_type or "—",
            f.status or "—",
            f.severity or "—",
            f.audit.title if hasattr(f, "audit") and f.audit else "—",
            f.department.name if hasattr(f, "department") and f.department else "—",
            f.owner or "—",
            f.due_date or "—",
            f.created_at.strftime("%Y-%m-%d") if hasattr(f, "created_at") and f.created_at else "—",
            f.closed_date or "—",
        ], shade=i % 2 == 1)
        row += 1
    return _send(wb, f"Finding_Register_{datetime.utcnow().strftime('%Y%m%d')}.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# Training Register
# ─────────────────────────────────────────────────────────────────────────────
def export_training_register(trainings):
    wb, ws = _make_workbook("Training Register")
    cols = [
        ("Training ID", 16), ("Title", 36), ("Type", 18), ("Status", 16),
        ("Department", 20), ("Trainer", 20), ("Target Group", 20),
        ("Training Date", 14), ("Expiry Date", 14), ("Attendees", 12),
    ]
    row = _write_header(ws, "Training Register", "Safety Training Records", cols)
    for i, t in enumerate(trainings):
        _write_row(ws, row, [
            t.id,
            t.title or "—",
            t.training_type or "—",
            t.status or "—",
            t.department.name if hasattr(t, "department") and t.department else "—",
            t.trainer or "—",
            t.target_group or "—",
            t.training_date or "—",
            t.expiry_date or "—",
            t.attendee_count or "—",
        ], shade=i % 2 == 1)
        row += 1
    return _send(wb, f"Training_Register_{datetime.utcnow().strftime('%Y%m%d')}.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# Employee Register
# ─────────────────────────────────────────────────────────────────────────────
def export_employee_register(employees):
    wb, ws = _make_workbook("Employee Register")
    cols = [
        ("Employee ID", 14), ("Name", 28), ("Email", 32), ("Phone", 18),
        ("Position", 24), ("Department", 20), ("Status", 14),
        ("Registered Date", 16),
    ]
    row = _write_header(ws, "Employee Register", "Mobile App Employee Accounts", cols)
    for i, e in enumerate(employees):
        _write_row(ws, row, [
            e.id,
            e.name or "—",
            e.email or "—",
            e.phone or "—",
            e.position or "—",
            e.department or "—",
            "Active" if e.is_active else "Disabled",
            e.created_at.strftime("%Y-%m-%d") if hasattr(e, "created_at") and e.created_at else "—",
        ], shade=i % 2 == 1)
        row += 1
    return _send(wb, f"Employee_Register_{datetime.utcnow().strftime('%Y%m%d')}.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# Survey Register
# ─────────────────────────────────────────────────────────────────────────────
def export_survey_register(surveys):
    wb, ws = _make_workbook("Survey Register")
    cols = [
        ("Survey ID", 14), ("Title", 36), ("Status", 16),
        ("Start Date", 14), ("End Date", 14),
        ("Responses", 14), ("Created By", 20), ("Created Date", 14),
    ]
    row = _write_header(ws, "Survey Register", "Safety Culture Surveys", cols)
    for i, s in enumerate(surveys):
        response_count = len(s.responses) if hasattr(s, "responses") else "—"
        _write_row(ws, row, [
            s.id,
            s.title or "—",
            s.status or "—",
            s.start_date or "—",
            s.end_date or "—",
            response_count,
            s.created_by or "—",
            s.created_at.strftime("%Y-%m-%d") if hasattr(s, "created_at") and s.created_at else "—",
        ], shade=i % 2 == 1)
        row += 1
    return _send(wb, f"Survey_Register_{datetime.utcnow().strftime('%Y%m%d')}.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# Safety Bulletin Register
# ─────────────────────────────────────────────────────────────────────────────
def export_bulletin_register(bulletins):
    wb, ws = _make_workbook("Bulletin Register")
    cols = [
        ("Bulletin ID", 16), ("Reference", 18), ("Title", 36), ("Category", 18),
        ("Status", 16), ("Author", 20), ("Department", 20),
        ("Issue Date", 14), ("Expiry Date", 14),
    ]
    row = _write_header(ws, "Safety Bulletin Register", "Safety Bulletins & Notices", cols)
    for i, b in enumerate(bulletins):
        _write_row(ws, row, [
            b.id,
            b.ref_number or "—",
            b.title or "—",
            b.category or "—",
            b.status or "—",
            b.author or "—",
            b.department.name if hasattr(b, "department") and b.department else "—",
            b.issue_date or "—",
            b.expiry_date or "—",
        ], shade=i % 2 == 1)
        row += 1
    return _send(wb, f"Bulletin_Register_{datetime.utcnow().strftime('%Y%m%d')}.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# Newsletter Register
# ─────────────────────────────────────────────────────────────────────────────
def export_newsletter_register(newsletters):
    wb, ws = _make_workbook("Newsletter Register")
    cols = [
        ("ID", 12), ("Title", 40), ("Status", 16),
        ("Issue Date", 14), ("Author", 22), ("Created Date", 14),
    ]
    row = _write_header(ws, "Safety Newsletter Register", "Safety Newsletters", cols)
    for i, n in enumerate(newsletters):
        _write_row(ws, row, [
            n.id,
            n.title or "—",
            n.status or "—",
            n.issue_date or "—",
            n.author or "—",
            n.created_at.strftime("%Y-%m-%d") if hasattr(n, "created_at") and n.created_at else "—",
        ], shade=i % 2 == 1)
        row += 1
    return _send(wb, f"Newsletter_Register_{datetime.utcnow().strftime('%Y%m%d')}.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# SPI Register
# ─────────────────────────────────────────────────────────────────────────────
def export_spi_register(spis):
    wb, ws = _make_workbook("SPI Register")
    cols = [
        ("SPI ID", 14), ("Name", 36), ("Category", 18), ("Status", 16),
        ("Unit", 14), ("Target", 14), ("Alert Threshold", 16),
        ("Frequency", 16), ("Owner", 20), ("Department", 20),
    ]
    row = _write_header(ws, "SPI Register", "Safety Performance Indicators", cols)
    for i, s in enumerate(spis):
        _write_row(ws, row, [
            s.id,
            s.name or "—",
            s.category or "—",
            s.status or "—",
            s.unit or "—",
            s.target_value if hasattr(s, "target_value") else "—",
            s.alert_threshold if hasattr(s, "alert_threshold") else "—",
            s.measurement_frequency or "—",
            s.owner or "—",
            s.department.name if hasattr(s, "department") and s.department else "—",
        ], shade=i % 2 == 1)
        row += 1
    return _send(wb, f"SPI_Register_{datetime.utcnow().strftime('%Y%m%d')}.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# Document Register
# ─────────────────────────────────────────────────────────────────────────────
def export_document_register(documents):
    wb, ws = _make_workbook("Document Register")
    cols = [
        ("Doc ID", 14), ("Reference", 18), ("Title", 36), ("Type", 16),
        ("Status", 14), ("Version", 10), ("Owner", 22), ("Department", 20),
        ("Issue Date", 14), ("Review Date", 14),
    ]
    row = _write_header(ws, "Document Register", "SMS Document Control Register", cols)
    for i, d in enumerate(documents):
        _write_row(ws, row, [
            d.id,
            d.doc_number or "—",
            d.title or "—",
            d.doc_type or "—",
            d.status or "—",
            d.version or "—",
            d.owner or "—",
            d.department.name if hasattr(d, "department") and d.department else "—",
            d.issue_date or "—",
            d.review_date or "—",
        ], shade=i % 2 == 1)
        row += 1
    return _send(wb, f"Document_Register_{datetime.utcnow().strftime('%Y%m%d')}.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# Risk Assessment Register
# ─────────────────────────────────────────────────────────────────────────────
def export_ra_register(ras):
    wb, ws = _make_workbook("RA Register")
    cols = [
        ("RA Reference", 18), ("Title", 36), ("Status", 16),
        ("Department", 20), ("Assessor", 20), ("Assessment Date", 16),
        ("Next Review Date", 16), ("Rows", 10), ("Mitigations", 14),
    ]
    row = _write_header(ws, "Risk Assessment Register", "Risk Assessment Register", cols)
    for i, r in enumerate(ras):
        _write_row(ws, row, [
            r.control_number or r.id,
            r.title or "—",
            r.status or "—",
            r.department.name if hasattr(r, "department") and r.department else "—",
            r.responsible_name or "—",
            r.assessment_date or "—",
            r.next_review_date or "—",
            len(r.rows) if hasattr(r, "rows") else "—",
            len(r.mitigations) if hasattr(r, "mitigations") else "—",
        ], shade=i % 2 == 1)
        row += 1
    return _send(wb, f"RA_Register_{datetime.utcnow().strftime('%Y%m%d')}.xlsx")
